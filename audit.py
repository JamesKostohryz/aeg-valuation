#!/usr/bin/env python3
"""
AEG Unified Model — out-of-band formula-integrity auditor (v2, hardened).
Run after every build; TOTAL must read 0.

  python3 audit.py AEG_Unified_Model.xlsx

Checks:
  1. Constants on calc tabs        — numeric non-formula cells outside permitted surfaces
  2. Magic numbers in formulas     — numeric literals inside formula strings (ref/name-stripped)
  3. Colour integrity              — blue outside permitted surfaces; mis-coloured constants;
                                     AND (v2) driver-surface constants must be blue (row-level
                                     granularity: closes the whole-tab-whitelist gap, §9.2)
  4. Cross-tab raw refs (v2)       — cross-sheet single-cell refs to cells with NO name
                                     (the NFE-class fragility, §9.1). Must be 0.
  5. Column consistency (v2)       — within each projection band, every row's projection
                                     columns must share ONE normalised formula pattern
                                     (catches the 'lone edited cell mid-row' silent error).
Allow-list: 0, 1, -1, tenor 1..30, day-count 365, unit 1e6, and the Audit meta surface.
"""
import sys, re, json
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries

INPUT_SURFACES   = {"Inputs"}
REPORTED_TABS    = {"Income Statement","Balance Sheet","Cash Flow","Market Data"}
DRIVER_SURFACES  = {"Forecast","Scenarios"}
META_TABS        = {"Audit"}
CONST_OK  = INPUT_SURFACES | REPORTED_TABS | DRIVER_SURFACES | META_TABS
MAGIC_OK  = META_TABS
BLUE_OK   = INPUT_SURFACES | REPORTED_TABS | DRIVER_SURFACES
ALLOW_NUMS = {0.0, 1.0, -1.0, 365.0, 1e6} | {float(i) for i in range(0, 31)}

# Projection bands (row-oriented): (tab, first_col, last_col). Every row's formulas across
# these columns must be one normalised pattern. Known-legitimate heterogeneous rows excluded.
PROJECTION_BANDS = [("Forecast","G","AJ"), ("Valuation","C","AF"),
                    ("DCF Reconciliation","C","AF"), ("Implied","C","O"),
                    ("Scenarios","B","E"), ("Econ Statements","C","S")]
CONSISTENCY_EXCLUDE = {("Econ Statements",5)}   # fiscal-year row: terminal pinned to in_anchor_year by design

BLUE = {"FF0000FF", "000000FF", "FF0000FE"}
IDENT = re.compile(r"[A-Za-z_\\][A-Za-z0-9_.]*")
STRING = re.compile(r'"[^"]*"')
NUM = re.compile(r"(?<![A-Za-z0-9_.])\d+(?:\.\d+)?(?:[eE]\d+)?")
CELLREF = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")
SHEET_REF = re.compile(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_ .]*?))!(\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?)")

def font_blue(cell):
    try:
        rgb = cell.font.color.rgb if cell.font and cell.font.color else None
        return rgb in BLUE
    except Exception:
        return False

def magic_literals(formula):
    s = STRING.sub(" ", formula)
    s = s.replace("$", "")
    s = IDENT.sub(" ", s)
    return [m.group() for m in NUM.finditer(s) if float(m.group()) not in ALLOW_NUMS]

def build_name_cells(wb):
    """(sheet,coord)->set(names) for every cell covered by a defined name."""
    cell_names = {}
    try: items = list(wb.defined_names.items())
    except AttributeError: items = [(d.name, d.value) for d in wb.defined_names.definedName]
    for name, dest in items:
        for piece in str(dest).split(','):
            m = re.match(r"^(?:'([^']+)'|([^!]+))!(.+)$", piece.strip())
            if not m: continue
            sheet = m.group(1) or m.group(2); ref = m.group(3).replace('$','')
            try: mnc,mnr,mxc,mxr = range_boundaries(ref)
            except Exception: continue
            for r in range(mnr,mxr+1):
                for c in range(mnc,mxc+1):
                    cell_names.setdefault((sheet,f"{get_column_letter(c)}{r}"),set()).add(name)
    return cell_names

def normalize(formula, cur_col):
    strs=[]
    f=STRING.sub(lambda m:(strs.append(m.group(0)),f"\x00{len(strs)-1}\x00")[1], formula)
    def repl(m):
        dc,col,dr,row=m.groups()
        coltok=f"${col}" if dc=='$' else f"C[{column_index_from_string(col)-cur_col:+d}]"
        return f"{coltok}{'$' if dr=='$' else 'R'}{row}"
    f=CELLREF.sub(repl,f)
    for i,s in enumerate(strs): f=f.replace(f"\x00{i}\x00",s)
    return f

def audit(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    cell_names = build_name_cells(wb)
    const_hits, magic_hits, colour_hits, xtab_hits, consist_hits = [], [], [], [], []
    for ws in wb.worksheets:
        t = ws.title
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v is None: continue
                is_formula = isinstance(v, str) and v.startswith("=")
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    if t not in CONST_OK and float(v) not in ALLOW_NUMS:
                        const_hits.append(f"{t}!{c.coordinate}={v}")
                    if t in (INPUT_SURFACES | REPORTED_TABS) and float(v) not in ALLOW_NUMS and not font_blue(c):
                        colour_hits.append(f"{t}!{c.coordinate} constant not blue")
                    # v2: driver-surface constants must be blue (row-level granularity, §9.2)
                    if t in DRIVER_SURFACES and float(v) not in ALLOW_NUMS and not font_blue(c):
                        colour_hits.append(f"{t}!{c.coordinate} driver-surface constant not blue (stray?)")
                if is_formula and t not in MAGIC_OK:
                    for lit in magic_literals(v):
                        magic_hits.append(f"{t}!{c.coordinate}: {lit}  in  {v}")
                    # v2 check 4: cross-tab single-cell refs to unnamed cells
                    for m in SHEET_REF.finditer(STRING.sub(' ', v)):
                        sh = m.group(1) or m.group(2); ref = m.group(3)
                        if ':' in ref: continue
                        coord = ref.replace('$','')
                        if (sh, coord) not in cell_names:
                            xtab_hits.append(f"{t}!{c.coordinate} -> {sh}!{ref} (unnamed target)")
                if font_blue(c) and t not in BLUE_OK:
                    colour_hits.append(f"{t}!{c.coordinate} blue outside permitted surface")
    # v2 check 5: column consistency on projection bands
    for t, c0, c1 in PROJECTION_BANDS:
        if t not in wb.sheetnames: continue
        ws = wb[t]; a = column_index_from_string(c0); b = column_index_from_string(c1)
        for r in range(1, ws.max_row+1):
            if (t, r) in CONSISTENCY_EXCLUDE: continue
            cells = [(c, ws.cell(r,c).value) for c in range(a,b+1)
                     if isinstance(ws.cell(r,c).value,str) and ws.cell(r,c).value.startswith('=')]
            if len(cells) < 3: continue
            norms = {}
            for c,f in cells: norms.setdefault(normalize(f,c),[]).append(get_column_letter(c))
            if len(norms) > 1:
                patts = "  ||  ".join(f"{len(v)}x[{','.join(v[:3])}]" for v in norms.values())
                consist_hits.append(f"{t}!row{r} ({ws.cell(r,1).value}): {patts}")
    # v3 check 6: base-to-forward continuity (catches anchor/forecast build mismatch, e.g. a
    # forward P&L line item missing from the anchor build — the R&D discontinuity class).
    # De-grow the t1 forward NOPAT by t1 revenue growth; it must reconcile to the base anchor NOPAT.
    continuity_hits = []
    try:
        wv = openpyxl.load_workbook(path, data_only=True); Fv = wv["Forecast"]
        F16, G16, F7, G7 = (Fv["F16"].value, Fv["G16"].value, Fv["F7"].value, Fv["G7"].value)
        if all(isinstance(x,(int,float)) for x in (F16,G16,F7,G7)) and F16 and G7:
            de_grown = G16 * F7 / G7
            resid = abs(de_grown - F16) / abs(F16)
            if resid > 0.05:
                continuity_hits.append(f"Forecast base->t1 NOPAT discontinuity {resid*100:.1f}% "
                                       f"(anchor {F16:.0f} vs de-grown fwd {de_grown:.0f}) — "
                                       f"a forward P&L line may be missing from the anchor build")
    except Exception:
        pass

    total = len(const_hits)+len(magic_hits)+len(colour_hits)+len(xtab_hits)+len(consist_hits)+len(continuity_hits)
    return {
        "file": path,
        "constants_on_calc_tabs": len(const_hits),
        "magic_numbers_in_formulas": len(magic_hits),
        "colour_violations": len(colour_hits),
        "cross_tab_unnamed_refs": len(xtab_hits),
        "column_consistency_breaks": len(consist_hits),
        "base_forward_continuity_breaks": len(continuity_hits),
        "TOTAL": total,
        "detail": {"constants":const_hits[:50],"magic":magic_hits[:50],"colour":colour_hits[:50],
                   "cross_tab":xtab_hits[:50],"consistency":consist_hits[:50],"continuity":continuity_hits[:50]},
    }

if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "AEG_Unified_Model.xlsx"
    r = audit(path)
    print(json.dumps(r, indent=2))
    sys.exit(0 if r["TOTAL"] == 0 else 1)

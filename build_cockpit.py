#!/usr/bin/env python3
"""build_cockpit.py — assemble the Google Sheets cockpit workbook from the Work Area.

The cockpit is the ONLY thing the user opens. It holds: a Control tab (ticker/repo/price
+ toggles), the live Forecast Work Area (drivers + forecast display), a Summary tab that
pulls the pipeline's published outputs via IMPORTDATA, and the hidden data-landing helpers.
The valuation engine is NOT here — it runs headless in Actions; the cockpit sends inputs
and reads results.

IMPORTDATA/VLOOKUP formulas are written as text; they go inert in Excel/LibreOffice but
become live the moment the file is imported to Google Sheets. The user sets ticker + repo
once on the Control tab and every IMPORTDATA URL builds from those two cells.
"""
import copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = "Forecast_WorkArea.xlsx"
OUT = "AEG_Cockpit.xlsx"

BLUE_FILL = PatternFill("solid", fgColor="DDEBF7")
BLUE_FONT = Font(color="1F4E78", bold=True)
HDR_FONT = Font(size=14, bold=True, color="1F4E78")
SUB_FONT = Font(size=11, italic=True, color="595959")
LBL_FONT = Font(bold=True)
SECT_FILL = PatternFill("solid", fgColor="1F4E78")
SECT_FONT = Font(bold=True, color="FFFFFF")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(bottom=thin)


def url(kind):
    """Build an IMPORTDATA URL from Control!B2 (ticker) and Control!B3 (repo)."""
    base = '"https://raw.githubusercontent.com/"&Control!$B$3&"/main/outputs/"&Control!$B$2&"'
    return f'=IMPORTDATA({base}_{kind}.csv")'


def vlookup(field, block):
    return f'=IFERROR(VLOOKUP("{field}",_Data!{block},2,FALSE),"—")'


def build():
    wb = openpyxl.load_workbook(SRC)

    # ---- Control tab (front) --------------------------------------------------
    ctrl = wb.create_sheet("Control", 0)
    ctrl["B1"] = "AEG Valuation — Cockpit"; ctrl["B1"].font = HDR_FONT
    ctrl["B2"] = "Ticker"; ctrl["B2"].font = LBL_FONT
    ctrl["C2"] = "AAPL"; ctrl["C2"].fill = BLUE_FILL; ctrl["C2"].font = BLUE_FONT
    ctrl["B3"] = "GitHub repo (user/name)"; ctrl["B3"].font = LBL_FONT
    ctrl["C3"] = "JamesKostohryz/aeg-valuation"; ctrl["C3"].fill = BLUE_FILL; ctrl["C3"].font = BLUE_FONT
    ctrl["B4"] = "Current price /sh"; ctrl["B4"].font = LBL_FONT
    ctrl["C4"] = 0; ctrl["C4"].fill = BLUE_FILL; ctrl["C4"].font = BLUE_FONT
    ctrl["B6"] = "Set the two blue cells above once; every result tab reads from them."
    ctrl["B6"].font = SUB_FONT
    ctrl["B8"] = "Forecast view toggles live on the 'Forecast Work Area' tab (mode, financing,"
    ctrl["B9"] = "payout, target FLEV). Per-company judgments (R&D, leases, minority interest)"
    ctrl["B10"] = "live in companies/<TICKER>.yaml in the repo, versioned there."
    for r in (8, 9, 10):
        ctrl.cell(r, 2).font = SUB_FONT
    ctrl.column_dimensions["B"].width = 30
    ctrl.column_dimensions["C"].width = 32

    # ---- _Data helper (hidden): IMPORTDATA landing for published outputs ------
    data = wb.create_sheet("_Data")
    data["A1"] = url("summary")      # A:B
    data["D1"] = url("valuation")    # D:E
    data["G1"] = url("anchors")      # G:H
    data["A1"].font = data["D1"].font = data["G1"].font = SUB_FONT
    data.sheet_state = "hidden"

    # ---- Summary tab: the research note, fed by the pipeline ------------------
    s = wb.create_sheet("Summary", 1)
    def section(r, title):
        s.cell(r, 2, title); s.cell(r, 2).font = SECT_FONT; s.cell(r, 2).fill = SECT_FILL
        s.cell(r, 3).fill = SECT_FILL
    def row(r, label, field, block, fmt=None):
        s.cell(r, 2, label).font = LBL_FONT
        c = s.cell(r, 3, vlookup(field, block))
        c.alignment = Alignment(horizontal="right")
        if fmt:
            c.number_format = fmt
        s.cell(r, 2).border = s.cell(r, 3).border = BORDER

    s["B1"] = "Valuation Summary"; s["B1"].font = HDR_FONT
    s["B2"] = "AEG / Ohlson–Juettner, real-terms, economically restated"; s["B2"].font = SUB_FONT
    section(4, "VALUATION")
    row(5, "Intrinsic value /sh", "intrinsic_value_ps", "$A:$B", "0.00")
    row(6, "Current price /sh (real)", "current_price_real_ps", "$A:$B", "0.00")
    row(7, "Upside / (downside)", "upside_downside", "$A:$B", "0.0%")
    row(8, "Adjusted equity (market-debt + idio) /sh", "adjusted_equity_ps", "$D:$E", "0.00")
    row(9, "Equity↔Enterprise tie", "equity_enterprise_tie", "$A:$B", "0.0000000")
    row(10, "Reconciliation check", "tie_check", "$D:$E")
    section(12, "IMPLIED EXPECTATIONS")
    row(13, "Normal (no-growth) value /sh", "normal_no_growth_value_ps", "$A:$B", "0.00")
    row(14, "PVGO /sh", "pvgo_ps", "$A:$B", "0.00")
    row(15, "Implied real EPS growth", "implied_real_eps_growth", "$A:$B", "0.00%")
    row(16, "Growth premium (turns)", "growth_premium_turns", "$A:$B", "0.00")
    section(18, "ECONOMIC REFORMULATION (anchor)")
    row(19, "Economic RNOA", "economic_rnoa", "$A:$B", "0.0%")
    row(20, "Real COE ρE (long-run)", "real_coe_longrun", "$A:$B", "0.00%")
    row(21, "Economic NOA", "economic_noa", "$A:$B", "0.000")
    section(23, "DISCLOSURES (V1-Plus)")
    row(24, "Debt capital gain /sh", "debt_capital_gain_ps", "$D:$E", "0.00")
    row(25, "Idiosyncratic haircut /sh", "idiosyncratic_haircut_ps", "$D:$E", "0.00")
    s.column_dimensions["B"].width = 40
    s.column_dimensions["C"].width = 18
    s["B27"] = "Live from the latest gated pipeline run (IMPORTDATA). Green tie-check = the four"
    s["B28"] = "methods reconcile. Blank cells mean the repo/ticker on Control isn't set yet."
    for r in (27, 28):
        s.cell(r, 2).font = SUB_FONT

    # ---- financial statement tabs (original + restated), IMPORTDATA-fed ---------
    #   values arrive in engine units ($ trillions; ×1000 = $B). Format / scale here.
    STMTS = [
        ("Reported IS", "reported_is"),
        ("Reported BS", "reported_bs"),
        ("Reported CF", "reported_cf"),
        ("Restated Financials", "restated"),
    ]
    for title, kind in STMTS:
        st = wb.create_sheet(title)
        st["A1"] = url(kind)                       # spills the full statement grid
        note = st["A1"]  # keep formula in A1; the grid spills from here
        st.freeze_panes = "B2"
        st.column_dimensions["A"].width = 34
        # a light header note above is not possible over a spill; leave A1 as the anchor.

    # order: Control, Summary, Forecast Work Area, financials, then helpers
    order = ["Control", "Summary", "Forecast Work Area",
             "Reported IS", "Reported BS", "Reported CF", "Restated Financials",
             "_Actuals", "_Rates", "_Data"]
    wb._sheets.sort(key=lambda ws: order.index(ws.title) if ws.title in order else 99)
    wb.save(OUT)
    print("wrote", OUT, "| tabs:", wb.sheetnames)


if __name__ == "__main__":
    build()

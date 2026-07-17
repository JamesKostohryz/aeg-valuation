#!/usr/bin/env python3
"""
company_setup.py — per-company repointing the statement loader does NOT cover:
  1. set_cost_of_debt() — write the nominal cost-of-debt YTW curve (Market Data row 27,
     28 tenor points) from the company's issued bonds. Falls back to a statement-implied
     flat rate (interest expense / total debt) — LOUDLY flagged as an assumption.
  2. set_company_labels() — replace every base-company ("AT&T") label so the workbook reads
     as the target company and the provenance gate passes. Header cells get the company name;
     company-specific guidance notes are genericized (never blindly renamed, which would make
     e.g. an R&D note wrong for a different firm).

Both are pure openpyxl; call them on the workbook before recalc.
"""
import openpyxl

COD_ROW = 27
COD_TENOR_ROW = 20
COD_FIRST_COL = 2          # tenor 1yr
COD_N = 28                 # tenors 1..28


def _tenors(ws):
    out = []
    for c in range(COD_FIRST_COL, COD_FIRST_COL + COD_N):
        t = ws.cell(COD_TENOR_ROW, c).value
        out.append(float(t) if t not in (None, "") else (c - COD_FIRST_COL + 1))
    return out


def _interp(tenors, points):
    """Linear interpolation of (tenor, yield) points onto `tenors`, flat-extrapolated."""
    pts = sorted((float(t), float(y)) for t, y in points)
    xs = [p[0] for p in pts]; ys = [p[1] for p in pts]
    out = []
    for t in tenors:
        if t <= xs[0]:
            out.append(ys[0])
        elif t >= xs[-1]:
            out.append(ys[-1])
        else:
            for i in range(1, len(xs)):
                if t <= xs[i]:
                    f = (t - xs[i - 1]) / (xs[i] - xs[i - 1])
                    out.append(ys[i - 1] + f * (ys[i] - ys[i - 1]))
                    break
    return out


def set_cost_of_debt(wb, ytw_points=None, single_ytw=None,
                     interest_expense=None, total_debt=None, ticker="CO"):
    """Write the nominal cost-of-debt curve to Market Data row 27.

    Priority: ytw_points (list of (tenor_yrs, yield)) > single_ytw (flat) >
    statement-implied interest_expense/total_debt (flat, FLAGGED as assumption).
    Returns a report dict; raises if no usable input at all.
    """
    md = wb["Market Data"]
    tenors = _tenors(md)
    flagged = False
    if ytw_points:
        curve = _interp(tenors, ytw_points); source = f"issued-bond YTW ({len(ytw_points)} pts, interpolated)"
    elif single_ytw is not None:
        curve = [float(single_ytw)] * COD_N; source = f"single YTW {single_ytw:.4f} (flat)"
    elif interest_expense not in (None, 0) and total_debt not in (None, 0):
        r = abs(float(interest_expense)) / abs(float(total_debt))
        curve = [r] * COD_N; source = f"STATEMENT-IMPLIED interest/debt = {r:.4f} (flat)"
        flagged = True
    else:
        raise ValueError("set_cost_of_debt: supply ytw_points, single_ytw, or interest_expense+total_debt")
    for i, v in enumerate(curve):
        md.cell(COD_ROW, COD_FIRST_COL + i).value = round(float(v), 6)
    # relabel the row so it no longer says AT&T
    md.cell(COD_ROW, 1).value = f"ρD nominal — SPOT ({ticker} YTW ladder)"
    return {"source": source, "flagged": flagged,
            "curve_1_5_10_28": [round(curve[0], 5), round(curve[4], 5),
                                round(curve[9], 5), round(curve[-1], 5)]}


def set_company_labels(wb, company, ticker, fy0):
    """Rewrite every base-company label cell. Headers -> company; notes -> genericized."""
    C = f"{company} ({ticker})"
    a2 = (f"Filed data ({company}). $mm unless noted · per-share $ · ratios frac · "
          f"shares mm. Blue = filed input. Loaded by the AEG data pipeline.")
    repl = {
        ("Inputs", "A4"):  f"COMPANY BASE — {C} · FY{fy0} · nominal $mm unless noted",
        ("Inputs", "D43"): "R&D reserve life (yr); 0 if R&D immaterial",
        ("Inputs", "A45"): "R&D expense, period-0 ($mm)",
        ("Inputs", "D45"): "Reported R&D read here; 0 if none reported.",
        ("Income Statement", "A1"): f"INCOME STATEMENT — {C}",
        ("Income Statement", "A2"): a2,
        ("Balance Sheet", "A1"): f"BALANCE SHEET — {C}",
        ("Balance Sheet", "A2"): a2,
        ("Cash Flow", "A1"): f"CASH FLOW — {C}",
        ("Cash Flow", "A2"): a2,
        ("Market Data", "A16"): f"{company} year-end close (contemp., $/sh)  [buyback reserve]",
        ("Market Data", "A19"): ("Treasury nominal par (CMT) & TIPS real par → breakeven infl "
                                 "= nominal−TIPS; 1–4yr infl stub = Cleveland Fed EXPINF (TIPS "
                                 "knots ≥5yr). Company ρD from issued-bond YTW ladder. Refresh each session."),
        ("Market Data", "A27"): f"ρD nominal — SPOT ({ticker} YTW ladder)",
        ("Cap Engine", "A58"): "R&D CAPITALIZATION (general)",
        ("Implied", "A2"): ("PVGO decomposition + implied constant real EPS growth "
                            "(Gordon-equivalent). Re-bracketed to admit negative growth "
                            "(g > −100%, guarded g < ρE_LR)."),
        ("Presentation", "A1"): f"RESEARCH NOTE — {C}",
    }
    changed = 0
    for (tab, coord), text in repl.items():
        if tab in wb.sheetnames:
            wb[tab][coord].value = text; changed += 1
    return {"cells_relabeled": changed}

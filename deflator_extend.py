#!/usr/bin/env python3
"""deflator_extend.py — keep the real-terms deflator tables covering the anchor year.

WHY THIS EXISTS
---------------
The engine deflates every historical statement to real base-year dollars using two
hand-seeded tables on the Market Data tab:
    CPI-U block   : row 7 years, row 8 CPI-U (annual avg), row 9 deflator = md_cpi_base/CPI
    BEA PP&E block: row 13 years, row 14 BEA nonres-FI, row 15 deflator = md_bea_base/BEA
Those tables were seeded only through the last COMPLETE CALENDAR year (2025), but a
company whose newest fiscal year ends in early 2026 anchors on 2026. Every real-terms
lookup `INDEX(md_deflator, MATCH(in_anchor_year, md_years))` then misses and returns
#N/A, which cascades into the audit/mode cells (audit_ok/mode_ok fail while the four-
method identity still ties). AAPL/T (fiscal 2025) are unaffected; HD/POOL (fiscal 2026)
fail. See the tie-diagnostic dump.

WHAT THIS DOES (per James: use the monthly CPI data)
----------------------------------------------------
1. Pulls MONTHLY CPI-U (FRED CPIAUCNS, NSA — its calendar-year mean IS the BLS annual
   average) and, for every year the tables are missing up to the anchor year, appends a
   column with CPI = the calendar-year mean of the available monthly prints (a partial
   year firms up as more months publish). Deflator formula stays md_cpi_base/CPI.
2. Extends the BEA PP&E deflator the same way, carrying the PP&E price index forward by
   realised CPI inflation (BEA_y = BEA_{y-1} * CPI_y/CPI_{y-1}). BEA nonres-FI publishes
   quarterly with a lag; the CPI carry is a documented one-year proxy and only feeds the
   PP&E vintage restatement (a second-order input). ERP can later supply the exact series.
3. Extends the six named ranges (md_years/md_cpi/md_deflator and the _bea/_ppe trio) to
   include the new column so MATCH/INDEX see it.
4. Wraps the Cap Engine capex lookups in IFERROR(...,0): some issuers' cap-engine vintage
   history starts a year before their reported cash-flow data (e.g. HD 1989 vs 1990),
   so MATCH(vintage, rep_years_cf) misses on the earliest year -> #N/A. A no-op when the
   lookup succeeds, so AAPL/T are unchanged.

Fail-closed: if CPI cannot be fetched the build must stop, never ship a real valuation on
a missing deflator.
"""
import os
import io
import re
import csv
import json
import datetime
import urllib.request

import openpyxl
from openpyxl.utils import get_column_letter

FRED_CPI_SERIES = "CPIAUCNS"   # CPI-U, all items, U.S. city avg, NSA, monthly
_CPI_NAMES = ("md_years", "md_cpi", "md_deflator")            # rows 7,8,9
_BEA_NAMES = ("md_years_bea", "md_bea", "md_deflator_ppe")    # rows 13,14,15


class DeflatorError(Exception):
    pass


def _parse_fred_csv(text):
    out = {}
    rd = csv.reader(io.StringIO(text))
    next(rd, None)  # header: observation_date,<SERIES>
    for row in rd:
        if len(row) < 2:
            continue
        d, v = row[0].strip(), row[1].strip()
        if v in (".", "", None):
            continue
        try:
            out[datetime.date.fromisoformat(d)] = float(v)
        except Exception:
            continue
    return out


def _fetch_fred_api(series_id, api_key, start, timeout, retries=2):
    url = (f"https://api.stlouisfed.org/fred/series/observations?series_id={series_id}"
           f"&api_key={api_key}&file_type=json&observation_start={start}")
    last = None
    for _ in range(retries + 1):
        try:
            with urllib.request.urlopen(url, timeout=timeout) as r:
                data = json.load(r)
            out = {}
            for o in data.get("observations", []):
                v = o.get("value")
                if v in (None, ".", ""):
                    continue
                try:
                    out[datetime.date.fromisoformat(o["date"])] = float(v)
                except Exception:
                    continue
            if out:
                return out
            last = "keyed API returned no observations"
        except Exception as e:
            last = f"keyed API failed: {e}"
    return None, last


def _fetch_fred_csv(series_id, start, timeout, retries=3):
    csv_url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}&cosd={start}"
    last = None
    for _ in range(retries + 1):
        try:
            req = urllib.request.Request(csv_url, headers={"User-Agent": "aeg-valuation/1.0"})
            with urllib.request.urlopen(req, timeout=timeout) as r:
                out = _parse_fred_csv(r.read().decode("utf-8"))
            if out:
                return out
            last = "keyless CSV returned no rows"
        except Exception as e:
            last = f"keyless CSV failed: {e}"
    return None, last


def _fetch_bls(series_id="CUUR0000SA0", timeout=40, retries=2):
    """Keyless BLS Public Data API v1. CUUR0000SA0 == CPI-U, US city avg, all items, NSA
    (same index/base as FRED CPIAUCNS). Returns {date: value} or (None, err)."""
    url = f"https://api.bls.gov/publicAPI/v1/timeseries/data/{series_id}"
    last = None
    for _ in range(retries + 1):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "aeg-valuation/1.0"})
            with urllib.request.urlopen(req, timeout=timeout) as r:
                data = json.load(r)
            out = {}
            for s in data.get("Results", {}).get("series", []):
                for d in s.get("data", []):
                    per = d.get("period", "")
                    if not (per.startswith("M") and per != "M13"):
                        continue
                    try:
                        out[datetime.date(int(d["year"]), int(per[1:]), 1)] = float(d["value"])
                    except Exception:
                        continue
            if out:
                return out
            last = "BLS returned no monthly rows"
        except Exception as e:
            last = f"BLS failed: {e}"
    return None, last


def fetch_cpi_monthly(api_key=None, series_id=FRED_CPI_SERIES, start="2015-01-01", timeout=40):
    """Return {date: value} monthly CPI-U from FRED.

    The authenticated FRED API (api.stlouisfed.org) is the RELIABLE path from CI and is
    tried first when a FRED_API_KEY is available; the keyless public CSV (fredgraph) is a
    fallback but FRED throttles cloud IPs on it, so it may time out on GitHub runners.
    Raises DeflatorError only if every source fails — the build must never proceed on a
    missing CPI.
    """
    api_key = api_key if api_key is not None else os.environ.get("FRED_API_KEY")
    errors = []
    if api_key:
        res = _fetch_fred_api(series_id, api_key, start, timeout)
        if isinstance(res, dict):
            return res
        errors.append(res[1])
    res = _fetch_bls(timeout=timeout)          # keyless, CI-reliable
    if isinstance(res, dict):
        return res
    errors.append(res[1])
    res = _fetch_fred_csv(series_id, start, timeout)   # keyless, may be throttled
    if isinstance(res, dict):
        return res
    errors.append(res[1])
    raise DeflatorError("could not fetch monthly CPI-U (" + "; ".join(e for e in errors if e) + ")")


def _calendar_year_mean(monthly, year):
    vals = [v for d, v in monthly.items() if d.year == year]
    return (sum(vals) / len(vals)) if vals else None


def _last_year_col(ws, row):
    last = 1
    for c in range(2, ws.max_column + 1):
        if ws.cell(row, c).value not in (None, ""):
            last = c
    return last


def wrap_capengine_capex(wb):
    """IFERROR-guard the Cap Engine capex lookups so a vintage year with no reported
    cash-flow (cap-engine history starting before CF data) contributes 0, not #N/A.
    No-op when the lookup already resolves. Returns count wrapped."""
    if "Cap Engine" not in wb.sheetnames:
        return 0
    CE = wb["Cap Engine"]
    n = 0
    for r in range(1, CE.max_row + 1):
        c = CE.cell(r, 2)  # column B = nominal capex
        v = c.value
        if isinstance(v, str) and v.startswith("=-INDEX(rep_capex") and "IFERROR" not in v:
            c.value = f"=IFERROR({v[1:]},0)"
            n += 1
    return n


def _extend_named_ranges(wb, names, prev_letter, new_letter):
    for nm in names:
        dn = wb.defined_names.get(nm)
        if dn is None:
            raise DeflatorError(f"expected defined name missing: {nm}")
        dn.value = re.sub(r":\$%s\$" % prev_letter, ":$%s$" % new_letter, dn.value)


def ensure_deflator_covers_anchor(path, anchor_year, fy_end_month=12,
                                  api_key=None, cpi_monthly=None):
    """Extend the CPI-U and BEA deflator tables (and their named ranges) to cover
    anchor_year, and IFERROR-guard the cap-engine capex. Idempotent. Returns a report."""
    api_key = api_key if api_key is not None else os.environ.get("FRED_API_KEY")
    anchor_year = int(anchor_year)
    wb = openpyxl.load_workbook(path, data_only=False)
    if "Market Data" not in wb.sheetnames:
        raise DeflatorError("workbook has no 'Market Data' tab")
    MD = wb["Market Data"]

    cap_wrapped = wrap_capengine_capex(wb)

    last_col = _last_year_col(MD, 7)
    last_year = MD.cell(7, last_col).value
    try:
        last_year_i = int(str(last_year).strip())
    except Exception:
        raise DeflatorError(f"could not read last deflator year (row7 col {last_col}={last_year!r})")

    report = {"anchor_year": anchor_year, "last_covered": last_year_i,
              "cap_capex_wrapped": cap_wrapped, "added": []}

    if last_year_i >= anchor_year:
        wb.save(path)                       # still persist the cap-engine guard
        report["extended"] = False
        return report

    if cpi_monthly is None:
        cpi_monthly = fetch_cpi_monthly(api_key)

    prev_cpi = MD.cell(8, last_col).value
    prev_bea = MD.cell(14, last_col).value
    if not isinstance(prev_cpi, (int, float)) or not isinstance(prev_bea, (int, float)):
        raise DeflatorError(f"non-numeric last CPI/BEA ({prev_cpi!r}/{prev_bea!r})")

    col = last_col
    prev_letter = get_column_letter(col)
    for yr in range(last_year_i + 1, anchor_year + 1):
        cpi = _calendar_year_mean(cpi_monthly, yr)
        if cpi is None:
            raise DeflatorError(f"no monthly CPI available for {yr}; cannot deflate the anchor")
        bea = prev_bea * (cpi / prev_cpi)           # carry PP&E deflator by realised CPI inflation
        col += 1
        L = get_column_letter(col)
        # CPI block
        MD.cell(7, col).value = str(yr)
        MD.cell(8, col).value = round(cpi, 6)
        MD.cell(9, col).value = f"=md_cpi_base/{L}8"
        # BEA block
        MD.cell(13, col).value = str(yr)
        MD.cell(14, col).value = round(bea, 6)
        MD.cell(15, col).value = f"=md_bea_base/{L}14"
        report["added"].append({"year": yr, "cpi": round(cpi, 4), "bea": round(bea, 4),
                                 "cpi_months": sum(1 for d in cpi_monthly if d.year == yr)})
        prev_cpi, prev_bea = cpi, bea

    new_letter = get_column_letter(col)
    _extend_named_ranges(wb, _CPI_NAMES + _BEA_NAMES, prev_letter, new_letter)
    wb.save(path)
    report["extended"] = True
    return report


if __name__ == "__main__":
    import sys
    pth = sys.argv[1]
    ay = int(sys.argv[2]) if len(sys.argv) > 2 else 2026
    print(json.dumps(ensure_deflator_covers_anchor(pth, ay), indent=2))

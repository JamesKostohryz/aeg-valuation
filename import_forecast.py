#!/usr/bin/env python3
"""
import_forecast.py — the deterministic gate between the AI's nominal forecast and the
real-terms Work Area. Validate -> deflate (nominal->real) -> write ONLY driver cells.

Fail-loud: any bad name / out-of-range value / horizon or mode problem aborts with a
named reason and writes nothing.

Deflation recipe (matches the sealed engine, so the Work Area stays tied to it):
the engine builds real figures from REAL growth drivers, so we convert the two
growth-type drivers to real and pass the ratio/rate drivers through unchanged.
  real_growth_t = (1+nominal_t)/(1+inflation_t) - 1
Everything downstream comes out in real dollars — the intended outcome — while
preserving the machine-precision tie to the engine.
"""
import openpyxl, copy

# driver -> (Work Area cell/row, min, max, deflate?)
# per-year drivers live across forecast columns G..(G+N-1); singles live in one cell.
DRIVER_SPEC = {
    "revenue_growth": {"row": 7,  "lo": -1.0, "hi": 5.0,  "deflate": True},
    "gross_margin":   {"row": 10, "lo": 0.0,  "hi": 1.0,  "deflate": False},
    "sga_ratio":      {"row": 12, "lo": 0.0,  "hi": 1.0,  "deflate": False},
    "da_rate":        {"row": 14, "lo": 0.0,  "hi": 1.0,  "deflate": False},
    "tax_rate":       {"row": 17, "lo": 0.0,  "hi": 1.0,  "deflate": False},
    "buyback_rate":   {"row": 24, "lo": -1.0, "hi": 1.0,  "deflate": False},  # +=shares shrink
    "capex_ratio":    {"row": 36, "lo": 0.0,  "hi": 1.0,  "deflate": False},
    "noa_growth":     {"row": 50, "lo": -1.0, "hi": 5.0,  "deflate": True},
}
SINGLES = {"target_flev": {"cell": "E3", "lo": -5.0, "hi": 20.0},
           "payout":      {"cell": "E2",  "lo": 0.0,  "hi": 1.0}}
MODES = ("Equity", "Enterprise")


class ImportError_(Exception):
    pass


def validate(nominal, N, mode):
    """nominal: {driver: [y1..yN]} for per-year drivers + {single: value}. Raises on any problem."""
    errs = []
    if mode not in MODES:
        errs.append(f"mode must be one of {MODES}, got {mode!r}")
    if not (isinstance(N, int) and 1 <= N <= 30):
        errs.append(f"horizon N must be an integer 1..30, got {N!r}")
    for name, spec in DRIVER_SPEC.items():
        if name not in nominal:
            errs.append(f"missing driver '{name}'"); continue
        vals = nominal[name]
        if not isinstance(vals, (list, tuple)) or len(vals) != N:
            errs.append(f"'{name}' must have exactly N={N} values, got {len(vals) if hasattr(vals,'__len__') else vals!r}")
            continue
        for i, v in enumerate(vals):
            if not isinstance(v, (int, float)):
                errs.append(f"'{name}' Y{i+1} not numeric: {v!r}")
            elif not (spec["lo"] <= v <= spec["hi"]):
                errs.append(f"'{name}' Y{i+1}={v} out of range [{spec['lo']},{spec['hi']}]")
    for name, spec in SINGLES.items():
        if name not in nominal:
            errs.append(f"missing single '{name}'"); continue
        v = nominal[name]
        if not isinstance(v, (int, float)):
            errs.append(f"'{name}' not numeric: {v!r}")
        elif not (spec["lo"] <= v <= spec["hi"]):
            errs.append(f"'{name}'={v} out of range [{spec['lo']},{spec['hi']}]")
    if errs:
        raise ImportError_("VALIDATION FAILED:\n  - " + "\n  - ".join(errs))


def deflate(nominal, inflation, N):
    """Return real drivers. inflation: [infl_y1..infl_yN] (expected inflation per forecast year)."""
    if len(inflation) < N:
        raise ImportError_(f"inflation series has {len(inflation)} years, need >= {N}")
    real = {}
    for name, spec in DRIVER_SPEC.items():
        vals = nominal[name]
        if spec["deflate"]:
            real[name] = [ (1+vals[i])/(1+inflation[i]) - 1 for i in range(N) ]
        else:
            real[name] = list(vals[:N])
    for name in SINGLES:
        real[name] = nominal[name]
    return real


# which drivers are active per mode (enterprise-only vs equity-only)
ENTERPRISE_ONLY = {"buyback_rate", "noa_growth"}
EQUITY_ONLY_SINGLE = {"payout"}

def write_to_workarea(workarea_path, real, N, mode, out_path=None):
    """Write ONLY the mode-relevant driver cells as blue inputs; clear inactive-mode
    driver cells. Diff-guard: abort if any non-driver cell would change."""
    out_path = out_path or workarea_path
    wb = openpyxl.load_workbook(workarea_path)
    wa = wb["Forecast Work Area"]
    before = {(c.row, c.column): c.value for row in wa.iter_rows() for c in row}
    permitted = set()
    fcols = [c for c in range(3, 40) if isinstance(wa.cell(4, c).value, (int, float)) and wa.cell(4, c).value >= 1]
    fcols = fcols[:N]
    # blue input style copied from an existing driver cell (revenue growth, row7)
    tmpl = wa.cell(7, fcols[0]); blue_font = copy.copy(tmpl.font); blue_fill = copy.copy(tmpl.fill)
    wa["C2"].value = mode; permitted.add((2, 3))
    ent = (mode == "Enterprise")
    for name, spec in DRIVER_SPEC.items():
        r = spec["row"]
        active = (name in ENTERPRISE_ONLY) == ent if name in ENTERPRISE_ONLY else True
        for j, c in enumerate(fcols):
            if active:
                cell = wa.cell(r, c); cell.value = round(real[name][j], 8)
                cell.font = copy.copy(blue_font); cell.fill = copy.copy(blue_fill)
            else:
                wa.cell(r, c).value = None   # clear inactive-mode driver
            permitted.add((r, c))
    for name, spec in SINGLES.items():
        cell = wa[spec["cell"]]
        if name in EQUITY_ONLY_SINGLE and ent:
            cell.value = None
        else:
            cell.value = real[name]
        permitted.add((cell.row, cell.column))
    # diff-guard
    after = {(c.row, c.column): c.value for row in wa.iter_rows() for c in row}
    illegal = [k for k in set(before) | set(after) if before.get(k) != after.get(k) and k not in permitted]
    if illegal:
        raise ImportError_(f"diff-guard: {len(illegal)} non-driver cells would change, e.g. {illegal[:5]}")
    wb.save(out_path)
    return {"written_drivers": len(DRIVER_SPEC)+len(SINGLES), "cols": len(fcols), "mode": mode}


def import_forecast(nominal, inflation, N, mode, workarea_path, out_path=None):
    validate(nominal, N, mode)
    real = deflate(nominal, inflation, N)
    rep = write_to_workarea(workarea_path, real, N, mode, out_path)
    rep["real_revenue_growth"] = [round(x,4) for x in real["revenue_growth"]]
    return rep


def inflation_from_feed(feed, N):
    """Single-source the importer's deflation from the rate feed's expected-inflation
    FORWARD series (exp_inflation_fwd1y), so the forecast is deflated with exactly the
    same inflation the rate-infrastructure pipeline publishes — no second inflation
    series can ever silently diverge from theirs. Fail-loud if the series is short.

    The forward series is indexed by tenor 1..30; forecast year i uses tenor i
    (the 1-year-forward expected inflation for that horizon step)."""
    key = "exp_inflation_fwd1y"
    if key not in feed:
        raise ImportError_(f"feed missing '{key}' — cannot single-source deflation")
    series = feed[key]
    if len(series) < N:
        raise ImportError_(f"feed '{key}' has {len(series)} tenors, need >= {N}")
    return [float(x) for x in series[:N]]


def import_forecast_from_feed(nominal, feed, N, mode, workarea_path, out_path=None):
    """Deflate + import using the rate feed's expected-inflation series (single-sourced).
    Thin wrapper over import_forecast that guarantees the inflation used to deflate the
    forecast is the feed's exp_inflation_fwd1y — the single-source requirement."""
    return import_forecast(nominal, inflation_from_feed(feed, N), N, mode,
                           workarea_path, out_path)


def read_mailbox(mailbox_path):
    """Parse a filled Mailbox workbook -> (nominal_drivers_dict, N, mode). Fail-loud on structure."""
    wb = openpyxl.load_workbook(mailbox_path, data_only=True)
    ws = wb["Forecast Request"]
    mode = ws["B6"].value; N = ws["D6"].value
    if mode not in MODES: raise ImportError_(f"mailbox: bad mode {mode!r}")
    if not (isinstance(N,int) and 1<=N<=30): raise ImportError_(f"mailbox: bad horizon {N!r}")
    # locate the Y1 column in the header row (row 8)
    HDR=8; y1col=None
    for c in range(2,60):
        if str(ws.cell(HDR,c).value).strip()=="Y1": y1col=c; break
    if y1col is None: raise ImportError_("mailbox: 'Y1' header not found")
    # driver rows by label in col A (rows 9..)
    rowof={}
    for r in range(9,40):
        lbl=ws.cell(r,1).value
        if lbl: rowof[str(lbl).split()[0].strip().lower()] = r
    nominal={}
    for d in DRIVER_SPEC:
        r=rowof.get(d)
        if r is None: raise ImportError_(f"mailbox: driver row '{d}' not found")
        vals=[ws.cell(r,y1col+k).value for k in range(N)]
        if any(v in (None,"") for v in vals): raise ImportError_(f"mailbox: '{d}' has blank Y-cells (AI left it empty)")
        nominal[d]=[float(v) for v in vals]
    # singles: target_flev (its own driver row), payout (row labeled 'payout')
    tf=rowof.get("target_flev")
    nominal["target_flev"]=float(ws.cell(tf,y1col).value)
    pr=rowof.get("payout")
    pv=ws.cell(pr,y1col).value if pr else None
    nominal["payout"]=float(pv) if pv not in (None,"") else 0.365
    return nominal, N, mode

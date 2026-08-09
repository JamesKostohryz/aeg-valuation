#!/usr/bin/env python3
"""loader_core.py — deterministic ingestion + auto-derivation for the AEG model.

This is the engine the Colab notebook drives. It is a hardened superset of the
hand-typed setup_company.py:

  * parse_statement / norm / NOSCALE  -> reused verbatim (same parsing & unit rules)
  * populate_raw_tabs                 -> same label-match + unit-convert + blue font,
                                         plus a fail-loud gate on missing critical lines
                                         and on ambiguous (duplicated) labels
  * derive_inputs                     -> NEW: replaces the hand-typed CFG["inputs"] block;
                                         every scalar is computed from the filed statements,
                                         each carrying its source line for the report
  * apply_judgments                   -> overlays the handful of genuine judgment calls
  * snapshot_layer / diff_guard       -> same permitted-cell diff-guard

Nothing here recalculates or audits — that is done by recalc_lo.recalc and audit.py,
which the notebook calls after this module has written the workbook.
"""
import csv, re, copy
import openpyxl

# --- unit / label rules (identical to setup_company.py) ---
NOSCALE = re.compile(r"\b(EPS|PER SHARE|RATE|MARGIN|RATIO|YIELD)\b|%", re.I)


def norm(lbl):
    return re.sub(r"\s+", " ", str(lbl).strip()).lower()


def detect_fy_end_month(path):
    """Read the statement header and return the fiscal-year-end month if the year
    columns are full dates (e.g. '09/30/2024' -> 9, '2024-09-28' -> 9). Returns
    None when the headers are bare years (e.g. '2024'), so the caller can default."""
    with open(path, newline="") as fh:
        header = next(csv.reader(fh))
    for cell in header[1:]:
        c = str(cell).strip()
        if "ttm" in c.lower():
            continue
        m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", c)
        if m:
            return int(m.group(1))
        m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", c)
        if m:
            return int(m.group(2))
    return None


def _parsed_fy0(parsed, key, *labels):
    """FY0 (latest-year) value of the first matching label in a parsed statement.
    Used for lines the model's reported tabs don't carry (e.g. R&D). Returns None
    if no label matches or the FY0 cell is blank."""
    if not parsed or key not in parsed:
        return None
    years, rows = parsed[key]
    if not years:
        return None
    idx = years.index(max(years))
    normed = {norm(k): v for k, v in rows.items()}
    for lbl in labels:
        series = normed.get(norm(lbl))
        if series is not None and idx < len(series) and series[idx] is not None:
            return series[idx]
    return None


def parse_statement(path):
    """Parse a Yahoo-format CSV: label in col 1, fiscal years across the top.
    Drops any 'ttm' column. Returns (years:list[int], rows:dict[label->list])."""
    with open(path, newline="") as fh:
        rd = csv.reader(fh)
        header = next(rd)
        raw_years = header[1:]
        keep = [i for i, y in enumerate(raw_years) if "ttm" not in y.lower()]
        years = [int(re.search(r"(\d{4})", raw_years[i]).group(1)) for i in keep]
        rows = {}
        seen_norm = {}
        for rec in rd:
            if not rec or not rec[0].strip():
                continue
            label = rec[0].strip()
            n = norm(label)
            if n in seen_norm and seen_norm[n] != label:
                raise ValueError(
                    f"ambiguous labels normalize identically: "
                    f"'{seen_norm[n]}' and '{label}' in {path}")
            if n in seen_norm and seen_norm[n] == label:
                raise ValueError(f"duplicate label '{label}' appears twice in {path}")
            seen_norm[n] = label
            vals = []
            for i in keep:
                x = rec[i + 1].strip().replace(",", "") if i + 1 < len(rec) else ""
                vals.append(float(x) if x not in ("", "-", "--", "N/A") else None)
            rows[label] = vals
    return years, rows


# --- tab layout ---
TAB_OF = {"is_csv": "Income Statement", "bs_csv": "Balance Sheet", "cf_csv": "Cash Flow"}

# Critical source lines: if a Yahoo CSV lacks any of these labels the model cannot
# be built correctly, so we abort loudly rather than silently blanking the row.
# (tab, exact model label).  Chosen because each feeds an anchor / reconciliation.
CRITICAL_LINES = [
    ("Income Statement", "Total Revenue"),
    ("Income Statement", "Cost of Revenue"),
    ("Income Statement", "Operating Income"),
    ("Income Statement", "Net Income Common Stockholders"),
    ("Income Statement", "Diluted EPS"),
    ("Income Statement", "Tax Rate for Calcs"),
    ("Income Statement", "Reconciled Depreciation"),
    ("Balance Sheet", "Total Assets"),
    ("Balance Sheet", "Cash And Cash Equivalents"),
    ("Balance Sheet", "Common Stock Equity"),
    ("Balance Sheet", "Ordinary Shares Number"),
    ("Balance Sheet", "Total Debt"),
    ("Balance Sheet", "Gross PPE"),
    ("Balance Sheet", "Net PPE"),
]


def _last_year_col(ws):
    for c in range(ws.max_column, 1, -1):
        if ws.cell(3, c).value is not None:
            return c
    return None


def populate_raw_tabs(wb, parsed):
    """Populate the three reported tabs from parsed CSVs.

    parsed: {"is_csv": (years, rows), "bs_csv": ..., "cf_csv": ...}
    Returns (permitted:set[(tab,coord)], match_report:dict, anchor_year:int).
    Raises ValueError (fail-loud) on missing critical line, ambiguous label,
    or mis-aligned fiscal years.
    """
    permitted = set()
    match_report = {}
    blue_font = copy.copy(wb["Income Statement"]["B4"].font)  # filed-input blue
    latest_years = {}

    for key, tab in TAB_OF.items():
        years, rows = parsed[key]
        ws = wb[tab]
        norm_csv = {}
        dup = set()
        for lbl, v in rows.items():
            k = norm(lbl)
            if k in norm_csv:
                dup.add(lbl)
            norm_csv[k] = v
        ncols = len(years)
        latest_years[tab] = years[-1]

        # ambiguity gate: a model row label that appears more than once in the CSV
        model_labels = [ws.cell(r, 1).value for r in range(4, ws.max_row + 1)
                        if ws.cell(r, 1).value is not None]
        ambiguous = sorted({lbl for lbl in model_labels if lbl in dup})
        if ambiguous:
            raise ValueError(
                f"[{tab}] ambiguous Yahoo labels (appear >once, cannot resolve): {ambiguous}")

        # year header row 3 -> STRINGS (model MATCHes against TEXT()); blank beyond span
        for j in range(41):  # cols B..AP
            cell = ws.cell(3, 2 + j)
            newv = str(years[j]) if j < ncols else None
            if cell.value != newv:
                cell.value = newv
                permitted.add((tab, cell.coordinate))

        matched = 0
        matched_norms = set()
        for r in range(4, ws.max_row + 1):
            lbl = ws.cell(r, 1).value
            if lbl is None:
                continue
            key_n = norm(lbl)
            scale = 1.0 if NOSCALE.search(str(lbl)) else 1e6
            series = norm_csv.get(key_n)
            for j in range(41):
                cell = ws.cell(r, 2 + j)
                newv = (round(series[j] / scale, 6)
                        if (series is not None and j < ncols and series[j] is not None)
                        else None)
                if cell.value != newv:
                    cell.value = newv
                    permitted.add((tab, cell.coordinate))
                    if newv is not None:
                        cell.font = copy.copy(blue_font)
            if series is not None:
                matched += 1
                matched_norms.add(key_n)
        match_report[tab] = {"matched": matched, "years": (years[0], years[-1]),
                             "n_years": ncols}

        # critical-line gate for this tab (compare on normalized labels)
        missing = [lbl for (t, lbl) in CRITICAL_LINES
                   if t == tab and norm(lbl) not in matched_norms]
        if missing:
            raise ValueError(
                f"[{tab}] missing critical line(s) not found in the CSV: {missing}")

    # --- derived fill: noncontrolling interest when the feed leaves it blank ---------
    # EODHD intermittently ships a blank 'Minority Interest' row (AT&T: 2008-09, 2023-25).
    # NCI then lands in neither CSE (excluded by the minority_include judgment) nor NFO
    # (design: MI+pension in NFO), so the per-year partition NOA-NFO-CSE breaks by exactly
    # the missing NCI. It is recoverable as the balance-sheet plug, and that plug
    # reproduces the reported NCI to the dollar in every year the feed does populate it.
    BSws = wb["Balance Sheet"]

    def _row_of(ws, label):
        for r in range(4, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v is not None and norm(v) == norm(label):
                return r
        return None

    r_mi = _row_of(BSws, "Minority Interest")
    r_ta = _row_of(BSws, "Total Assets")
    r_tl = _row_of(BSws, "Total Liabilities Net Minority Interest")
    r_te = _row_of(BSws, "Total Equity Gross Minority Interest")
    r_cse = _row_of(BSws, "Common Stock Equity")
    if all(x is not None for x in (r_mi, r_ta, r_tl, r_te)):
        filled = []          # blank / hard-zero MI filled from the balance plug (as before)
        reconciled = []      # filed MI nudged by a SUB-materiality residual so the BS articulates
        warned = []          # small-but-notable imbalance: ABSORBED so the tie holds, WARNED loudly
        grossnet = []        # gross-reported equity: NCI netted out of common equity for that year
        material = []        # imbalance > the warn margin: flagged loudly, left for the tie to catch
        MATL_FRAC = 0.001    # sub-materiality threshold = 0.1% of total assets (absorbed quietly)
        # "warn but allow within a relatively small margin" (James, 2026-08-03): a source-data
        # balance-sheet articulation error between MATL_FRAC and WARN_FRAC is absorbed into the
        # excluded Minority-Interest line (so the four-method tie stays exact and the error is
        # quarantined OUT of common equity) but is WARNED loudly and recorded for disclosure.
        # Above WARN_FRAC we still refuse — that size of gap is a real data problem, not noise.
        WARN_FRAC = 0.005    # 0.5% of total assets — the "relatively small margin" ceiling
        for j in range(41):
            c_mi = BSws.cell(r_mi, 2 + j)
            ta = BSws.cell(r_ta, 2 + j).value
            tl = BSws.cell(r_tl, 2 + j).value
            te = BSws.cell(r_te, 2 + j).value
            if not all(isinstance(x, (int, float)) for x in (ta, tl, te)):
                continue                      # no full balance sheet that year -> leave as filed
            close = round(ta - (tl + te), 6)  # the MI that makes the reported BS articulate
            filed = c_mi.value if isinstance(c_mi.value, (int, float)) else None
            yr = BSws.cell(3, 2 + j).value
            if filed is None or filed == 0:
                # Feed left it blank, or filed a hard 0 that does NOT close (AT&T 2011: MI=0
                # yet TA-(TL+TE)=263). Recover it as the balance plug; genuinely-zero NCI stays.
                if abs(close) < 1e-9:
                    continue
                c_mi.value = close
                c_mi.font = copy.copy(blue_font)
                permitted.add(("Balance Sheet", c_mi.coordinate))
                filled.append((yr, close))
            else:
                # GROSS-equity case (e.g. Walmart): the reported totals ALREADY articulate
                # (TA = TL + TE) because "Total Equity Gross Minority Interest" already INCLUDES
                # the noncontrolling interest. The filed NCI is then a disclosed component that
                # lives INSIDE TE, not a missing plug — so the sheet is balanced regardless of the
                # NCI's size. BUT "Common Stock Equity" is mapped to the same (gross) total, so it
                # over-states COMMON equity by the NCI. Net the disclosed NCI out of the common-
                # equity row for this year so the economic common-equity book (minority_include=
                # False) excludes minority interest and the four-method tie closes. Names with ~0
                # NCI never reach here, so they are untouched. (Anchor years with a blank NCI are
                # the NET case and are handled by the plug branch above.)
                if abs(close) <= 1e-6 * abs(ta):
                    if r_cse is not None:
                        c_cse = BSws.cell(r_cse, 2 + j)
                        if isinstance(c_cse.value, (int, float)) and abs(round(c_cse.value - te, 6)) < 1e-6:
                            # CSE row == gross TE -> it carries the NCI; net it out to get common.
                            c_cse.value = round(te - filed, 6)
                            c_cse.font = copy.copy(blue_font)
                            permitted.add(("Balance Sheet", c_cse.coordinate))
                            grossnet.append((yr, filed, (filed / ta) if ta else 0.0))
                    continue
                # Otherwise the equity total is NET of MI: the reported BS articulates only when
                # TA = TL + TE + MI, i.e. filed NCI should equal `close`. A gap is source-data
                # noise — rounding, merger/spinoff-year reclassifications (e.g. MRK 2009).
                resid = round(close - filed, 6)
                if abs(resid) < 1e-9:
                    continue                  # already closes -> filed NCI untouched
                if abs(resid) <= MATL_FRAC * abs(ta):
                    # sub-materiality: absorb into the NCI balancing line so every ECONOMIC tie
                    # stays machine-precise. Logged below; never silent.
                    c_mi.value = close
                    c_mi.font = copy.copy(blue_font)
                    permitted.add(("Balance Sheet", c_mi.coordinate))
                    reconciled.append((yr, resid, (resid / ta) if ta else 0.0))
                elif abs(resid) <= WARN_FRAC * abs(ta):
                    # small-but-notable (0.1%-0.5% of assets): ABSORB so the tie holds, but WARN
                    # loudly and record for disclosure. Same plug as sub-materiality — the gap goes
                    # into the excluded NCI line, so it never touches common-equity value.
                    c_mi.value = close
                    c_mi.font = copy.copy(blue_font)
                    permitted.add(("Balance Sheet", c_mi.coordinate))
                    warned.append((yr, resid, (resid / ta) if ta else 0.0))
                else:
                    # Beyond the warn margin: do NOT swallow. Flag loudly and leave it; the standing
                    # partition tie fails loud, forcing investigation of the source data.
                    material.append((yr, resid, (resid / ta) if ta else 0.0))
        if filled:
            print("  [loader] 'Minority Interest' blank in feed; derived as balance plug "
                  "TA-(TL+TE) for: " + ", ".join(f"{y}={v:,.0f}" for y, v in filled))
        if reconciled:
            print("  [loader] reported BS did not articulate; absorbed sub-materiality residual "
                  "(<0.1% assets) into Minority Interest for: "
                  + ", ".join(f"{y}={r:+,.1f} ({p*100:+.4f}% TA)" for y, r, p in reconciled))
        if grossnet:
            print("  [loader] equity reported GROSS of minority interest; netted the disclosed NCI "
                  "out of Common Stock Equity (common = gross - NCI) for: "
                  + ", ".join(f"{y}=-{v:,.0f} ({p*100:.3f}% TA)" for y, v, p in grossnet))
        for y, r, p in warned:
            print(f"  [loader] *** WARNING reported-BS imbalance {y}: {r:+,.1f} "
                  f"({p*100:+.3f}% of assets) exceeds 0.1% but is within the {WARN_FRAC*100:.1f}% "
                  f"margin — ABSORBED into the excluded Minority-Interest line so the tie holds "
                  f"(does not touch common-equity value). Disclosed data-quality adjustment; "
                  f"verify the source data for {y} if this name is high-stakes.")
        for y, r, p in material:
            print(f"  [loader] *** MATERIAL reported-BS imbalance {y}: {r:+,.1f} "
                  f"({p*100:+.3f}% of assets) exceeds the {WARN_FRAC*100:.1f}% margin — NOT "
                  f"absorbed; the standing tie will fail. Investigate the source data for {y}.")
        # expose the disclosed adjustments so downstream (status.csv / audit) can surface them
        globals()["_LAST_BS_WARNINGS"] = [
            {"year": y, "residual": r, "pct_assets": p} for y, r, p in warned]

    # fiscal-year alignment gate: latest (FY0) year must agree across the three tabs
    yrs = set(latest_years.values())
    if len(yrs) != 1:
        raise ValueError(
            f"fiscal years do not align across statements (latest per tab): {latest_years}")
    anchor_year = latest_years["Income Statement"]
    return permitted, match_report, anchor_year


def stabilize_cost_boundary(wb, tol_gpm_swing=0.15):
    """Make the operating-cost decomposition BOUNDARY-INDEPENDENT.

    Some issuers (AT&T and other function-cost filers) don't report a genuine
    Cost-of-Revenue / Gross-Profit split; the feed (EODHD) fabricates one, and it
    reshuffles the COGS<->OpEx boundary across years (e.g. AT&T's gross margin jumps
    43%->80% in 2025). Revenue, Operating Income and Reconciled Depreciation are all
    genuinely reported and stable, so the reliable non-D&A operating cost is
    (Revenue - OI - D&A) regardless of where the feed draws the COGS/OpEx line.

    When a filer's gross margin is UNSTABLE (fabricated split), rebuild Cost of Revenue
    from that stable spine so the forecast decomposition, the row-61 opex wedge, and the
    valuation stop depending on the fabricated boundary. Filers with a genuinely stable
    gross margin (e.g. AAPL) are left EXACTLY as filed. Engine formulas are untouched, so
    the four-method tie is preserved by construction.

    Reconstruction (per year, for a flagged filer): COGS := (Rev - OI - D&A) - SG&A,
    keeping the reported (stable) SG&A. Then GP = Rev - COGS = OI + D&A + SG&A, so
    GP - SG&A - D&A = OI and the wedge collapses to the genuine R&D adjustment (~0 for a
    no-R&D filer). Fail-loud/logged: the reconstruction is recorded in the returned report.

    Rows in the Income Statement tab: 3=years, 4=Revenue, 6=Cost of Revenue, 9=SG&A,
    13=Operating Income, 54=Reconciled Depreciation. Values are in engine units.
    """
    IS = wb["Income Statement"]
    R_REV, R_COGS, R_SGA, R_OI, R_DA, R_YR = 4, 6, 9, 13, 54, 3
    cols = [c for c in range(2, IS.max_column + 1) if IS.cell(R_YR, c).value not in (None, "")]

    def num(c, r):
        v = IS.cell(r, c).value
        return float(v) if isinstance(v, (int, float)) else None

    # gross-margin instability detector (fabricated-boundary symptom)
    gpm = {}
    for c in cols:
        rev, cg = num(c, R_REV), num(c, R_COGS)
        if rev:
            gpm[c] = (rev - (cg or 0.0)) / rev
    # detect over the RECENT window only: the forecast anchors on FY0 and the wedge
    # rides recent margins, so a recent reclassification (AT&T 2021/23/25) is what matters;
    # old-history margin volatility (e.g. 1990s Apple) is irrelevant to today's valuation.
    ordered = sorted(cols, key=lambda c: str(IS.cell(R_YR, c).value))[-6:]
    swings = [abs(gpm[ordered[i]] - gpm[ordered[i - 1]])
              for i in range(1, len(ordered)) if ordered[i] in gpm and ordered[i - 1] in gpm]
    max_swing = max(swings) if swings else 0.0
    if max_swing <= tol_gpm_swing:
        return {"reconstructed": False, "max_gpm_swing": round(max_swing, 4)}

    # fabricated -> rebuild COGS from the stable spine, keep SG&A
    permitted, changed = set(), []
    blue = copy.copy(IS.cell(4, 2).font)
    for c in cols:
        rev, oi, da, sga = num(c, R_REV), num(c, R_OI), num(c, R_DA), num(c, R_SGA)
        if None in (rev, oi, da):
            continue
        sga = sga or 0.0
        non_da_opex = rev - oi - da
        if non_da_opex < 0:            # degenerate; leave the year as-is
            continue
        sga_use = min(sga, non_da_opex)
        cogs_new = round(non_da_opex - sga_use, 6)
        for r, newv in ((R_COGS, cogs_new), (R_SGA, round(sga_use, 6))):
            cell = IS.cell(r, c)
            if cell.value != newv:
                cell.value = newv
                cell.font = copy.copy(blue)
                permitted.add(("Income Statement", cell.coordinate))
        changed.append(str(IS.cell(R_YR, c).value))
    return {"reconstructed": True, "max_gpm_swing": round(max_swing, 4),
            "years": changed, "permitted": permitted,
            "note": "COGS rebuilt from Rev-OI-D&A-SG&A (fabricated gross-margin boundary)"}


# --- Inputs auto-derivation --------------------------------------------------
# Each derived scalar carries (row, name, meaning, value, source, kind).
def _fy0(ws, label):
    """Return the FY0 (latest-year) value of a raw-tab row matched by label."""
    c = _last_year_col(ws)
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 1).value is not None and norm(ws.cell(r, 1).value) == norm(label):
            return ws.cell(r, c).value
    return None


def _lastn(ws, label, n=3):
    """Return the last `n` numeric values of a raw-tab row (oldest first, blanks skipped).

    Used by the multi-year policy estimators in resolve_policy_inputs, where a single
    fiscal year is too noisy to characterise a policy: a composite depreciable life read
    off one year swings with a single large disposal or acquisition."""
    c_last = _last_year_col(ws)
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 1).value is not None and norm(ws.cell(r, 1).value) == norm(label):
            vals = [ws.cell(r, c).value for c in range(2, c_last + 1)]
            vals = [v for v in vals if isinstance(v, (int, float))]
            return vals[-n:] if vals else []
    return []


class PolicyInputError(Exception):
    """A per-company policy input could not be derived from the filings and has no
    override in the company config. Fail loud: the alternative is silently inheriting
    the template's base-company value, which is the defect this class exists to stop."""


def resolve_policy_inputs(wb, derived, *, payout_override=None, ppe_life_override=None):
    """P1/P3 — set the two Inputs scalars that used to ship as template constants.

    `in_payout_seed` (Inputs B39) and `in_ppe_life` (Inputs B42) were never written by
    the build. Every issuer therefore inherited the template base company's values
    (AT&T's 36.5% dividend payout and 18-year composite plant life), and nothing in the
    system flagged it. Both are first-order: the payout seed drives Forecast row 29 in
    Equity mode (dividends per share = seed x earnings per share) and so drives retained
    earnings, the book roll and the normal-earnings benchmark; the plant life drives tax
    depreciation, the economic depreciation restatement and the disclosure bridge's
    depreciation-anchor penalty.

    Call AFTER apply_market_data, so the payout seed is computed against the dividend
    per share that actually lands on the sheet.

    Definition note: the payout seed here is the DIVIDEND payout ratio — dividends per
    share divided by diluted earnings per share. It deliberately EXCLUDES share
    repurchases. In Equity mode the share count is currently held constant, so buybacks
    have no representation anywhere in the forecast; whether that should change, and
    whether the seed should therefore carry total shareholder distribution instead, is an
    open modelling question and is NOT settled by this function.

    Mutates `derived` (adding rows 39 and 42) and returns a report dict.
    """
    IS = wb["Income Statement"]; BS = wb["Balance Sheet"]
    report = {}

    # --- in_payout_seed (B39) -------------------------------------------------
    eps = derived.get(13, {}).get("value")
    dps = derived.get(15, {}).get("value")
    if payout_override is not None:
        payout = float(payout_override)
        p_src = f"company config judgments.payout_override = {payout}"
        p_kind = "analyst"
    elif dps is None:
        # A genuine non-payer is derivable and correct: zero dividends, full retention.
        payout, p_src, p_kind = 0.0, "no dividend found in the filings -> non-payer, payout 0.0", "derived"
    elif eps is None or eps <= 0:
        raise PolicyInputError(
            f"cannot derive the dividend payout seed: diluted EPS is {eps!r} (a loss or "
            f"missing year), so dividends {dps!r} / EPS is not a usable policy ratio. Set "
            f"judgments.payout_override in the company config to the payout ratio you "
            f"intend the forecast to carry.")
    else:
        payout = round(float(dps) / float(eps), 6)
        p_src = (f"dividends per share {dps} / diluted EPS {eps} = {payout}  "
                 f"(FY0 dividend payout; EXCLUDES buybacks)")
        p_kind = "derived"
    if payout < 0 or payout > 2.0:
        raise PolicyInputError(
            f"derived dividend payout seed {payout} is outside the plausible band [0, 2.0] "
            f"({p_src}). This usually means a trough or restated earnings year. Set "
            f"judgments.payout_override in the company config.")
    report["payout_seed"] = payout
    report["payout_review"] = payout > 1.0   # paying out more than it earns: worth a look
    derived[39] = dict(name="in_payout_seed", meaning="Dividend payout seed (equity-mode DPS driver)",
                       value=payout, source=p_src, kind=p_kind)

    # --- in_ppe_life (B42) ----------------------------------------------------
    gross = _lastn(BS, "Gross PPE", 3)
    dep = _lastn(IS, "Reconciled Depreciation", 3)
    if ppe_life_override is not None:
        life = float(ppe_life_override)
        l_src = f"company config judgments.ppe_life_override = {life}"
        l_kind = "analyst"
    elif gross and dep and sum(dep) > 0:
        n = min(len(gross), len(dep))
        gbar = sum(gross[-n:]) / n
        dbar = sum(dep[-n:]) / n
        life = round(gbar / dbar, 2)
        l_src = (f"mean Gross PPE over {n}y ({gbar:.6g}) / mean reported depreciation "
                 f"({dbar:.6g}) = {life} years")
        l_kind = "derived"
    else:
        raise PolicyInputError(
            f"cannot derive the composite depreciable life: Gross PPE series={gross!r}, "
            f"depreciation series={dep!r}. Set judgments.ppe_life_override in the company "
            f"config to the composite life you intend.")
    if life < 2.0 or life > 50.0:
        raise PolicyInputError(
            f"derived composite depreciable life {life} years is outside the plausible band "
            f"[2, 50] ({l_src}). Set judgments.ppe_life_override in the company config.")
    report["ppe_life"] = life
    derived[42] = dict(name="in_ppe_life", meaning="PP&E composite depreciable life L (yr)",
                       value=life, source=l_src, kind=l_kind)
    return report


# Inputs rows that shape the valuation, with a class for each:
#   company  — a fact or policy of THIS company; must come from its filings or config
#   control  — a run-mode switch (Equity/Enterprise, Single/Term, scenario). Chosen per
#              run by the harness, the payload or the analyst, not a company fact.
#   inert    — present but not referenced by any live formula in the default
#              configuration; kept so the register can say so out loud instead of
#              leaving someone to assume it matters.
VALUATION_RELEVANT_ROWS = {
    5:  ("in_debt", "company"),        6:  ("in_cash", "company"),
    7:  ("in_sti", "company"),         8:  ("in_finlease", "company"),
    9:  ("anchor_shares0", "company"), 10: ("anchor_cse0", "company"),
    11: ("in_intexp0", "company"),     12: ("in_oiadj0", "company"),
    13: ("anchor_eps0", "company"),    14: ("in_tax0", "company"),
    15: ("anchor_dps0", "company"),    25: ("in_price", "company"),
    26: ("cfg_N", "company"),          27: ("in_g_terminal", "inert"),
    28: ("in_erp", "company"),         29: ("cfg_coe_mode", "control"),
    30: ("cfg_rf_mode", "control"),    31: ("in_rf_single", "inert"),
    32: ("in_rf_lr", "inert"),         33: ("cfg_cod_mode", "control"),
    34: ("cfg_valread", "control"),    37: ("cfg_mode", "control"),
    39: ("in_payout_seed", "company"), 42: ("in_ppe_life", "company"),
    43: ("in_rd_life", "company"),     45: ("in_rd_expense0", "company"),
    47: ("cfg_buyback", "control"),    66: ("in_anchor_year", "company"),
    67: ("cfg_view", "control"),       68: ("cfg_funding", "control"),
    69: ("cfg_scenario", "control"),
}

# Why each 'inert' row is inert. Read directly off the template's formula graph rather
# than asserted: in_g_terminal is referenced by no formula at all; in_rf_single and
# in_rf_lr are referenced only inside IF(cfg_rf_mode="Single", ...) branches and the
# template ships in "Term" mode. If cfg_rf_mode is ever switched to Single, the last two
# stop being inert and become unset template constants that move numbers.
INERT_NOTES = {
    "in_g_terminal": "referenced by no formula in the template (dead input)",
    "in_rf_single": "only live when cfg_rf_mode='Single'; template ships 'Term'",
    "in_rf_lr": "only live when cfg_rf_mode='Single'; template ships 'Term'",
}

# Company-class inputs the BUILD does not set but a LATER pipeline stage does. The
# register is written at build time, so without this note it would report them as
# template leftovers on every live run, which would be wrong and would train the reader
# to ignore the column that matters.
LATE_BOUND_NOTES = {
    "in_erp": ("template value at build time; superseded on live runs by the rate feed, "
               "which rewrites the market equity-risk-premium term structure on "
               "Market Data row 25 (repoint_rates). Live on a fixture build."),
}


def provenance_register(wb, derived, analyst_set=()):
    """S2 — one row per valuation-relevant Inputs cell: value, provenance, source.

    Provenance is one of:
      filings   — computed from this company's own statements or market data
      analyst   — set deliberately for this company (config judgment or explicit control)
      control   — a run-mode switch, carrying the template's shipped mode
      inert     — not referenced by any live formula in this configuration
      template  — a COMPANY-level input still carrying whatever the template shipped with

    'template' is the category that matters, and after P1/P2/P3 it should normally be
    empty. It is how the template base company's dividend payout, its 18-year plant life
    and an unchosen four-year forecast horizon reached every valuation in the system
    without anyone noticing. Anything left in that category is reported, not silently
    accepted.
    """
    inp = wb["Inputs"]
    kind_map = {"auto": "filings", "derived": "filings", "judgment": "analyst",
                "analyst": "analyst"}
    out = []
    for row in sorted(VALUATION_RELEVANT_ROWS):
        name, cls = VALUATION_RELEVANT_ROWS[row]
        cell = inp.cell(row, 2)
        d = derived.get(row)
        if d is not None:
            prov = kind_map.get(d.get("kind"), "filings")
            src = d.get("source", "")
        elif name in analyst_set:
            prov, src = "analyst", "set explicitly for this run"
        elif cls == "control":
            prov = "control"
            src = "run-mode switch; carries the template's shipped mode unless overridden"
        elif cls == "inert":
            prov = "inert"
            src = INERT_NOTES.get(name, "not referenced by any live formula")
        elif name in LATE_BOUND_NOTES:
            prov, src = "late-bound", LATE_BOUND_NOTES[name]
        else:
            prov = "template"
            src = "NOT set by the build — inherited from MODEL_TEMPLATE.xlsx"
        out.append({"row": row, "cell": f"B{row}", "name": name, "class": cls,
                    "label": str(inp.cell(row, 1).value or "").strip(),
                    "value": cell.value, "provenance": prov, "source": src})
    return out


def derive_inputs(wb, anchor_year, parsed=None):
    """Auto-derive every non-judgment Inputs scalar from the populated raw tabs.

    `parsed` (optional) is the {key:(years,rows)} dict from parse_statement; it is
    used to read lines the model's reported tabs don't carry (e.g. R&D, which the
    model keeps only as the Inputs scalar B45).

    Returns dict: row -> {"value","source","meaning","kind"} where kind is
    'auto' | 'derived' | 'judgment-default' (judgment rows are placeholders that
    apply_judgments will overwrite; they are surfaced with the filed number).
    """
    IS = wb["Income Statement"]; BS = wb["Balance Sheet"]; CF = wb["Cash Flow"]

    def g(ws, lbl):
        v = _fy0(ws, lbl)
        return v

    debt   = g(BS, "Total Debt")
    cash   = g(BS, "Cash And Cash Equivalents")
    sti    = g(BS, "Other Short Term Investments") or 0.0
    shares = g(BS, "Ordinary Shares Number")          # already /1e6 in the tab
    cse    = g(BS, "Common Stock Equity")
    mi     = g(BS, "Minority Interest") or 0.0
    intexp = g(IS, "Interest Expense") or 0.0
    oi     = g(IS, "Operating Income")
    unusual = g(IS, "Total Unusual Items")
    eps    = g(IS, "Diluted EPS")
    # R&D: the model keeps this only as an Inputs scalar (no reported-tab row), so
    # read it from the parsed CSV (raw dollars) and scale to $mm like the tabs do.
    rd_raw = _parsed_fy0(parsed, "is_csv", "Research And Development",
                         "Research & Development", "Research and Development")
    rd = round(rd_raw / 1e6, 6) if rd_raw is not None else 0.0

    # tax: prefer the filed effective-rate line ("Tax Rate for Calcs"); fall back
    # to provision / pretax.  Both are reported; we surface both for confirmation.
    tax_line = g(IS, "Tax Rate for Calcs")
    prov = g(IS, "Tax Provision"); pre = g(IS, "Pretax Income")
    tax_ratio = (prov / pre) if (prov is not None and pre not in (None, 0)) else None
    if tax_line is not None:
        tax = tax_line
        tax_src = f"IS 'Tax Rate for Calcs' = {tax_line}  (provision/pretax = {tax_ratio:.5f})" \
            if tax_ratio is not None else f"IS 'Tax Rate for Calcs' = {tax_line}"
    else:
        tax = tax_ratio
        tax_src = f"provision {prov} / pretax {pre} = {tax_ratio}"

    # dividends per share: use a filed per-share dividend line if present; else
    # derive from cash dividends paid (net of preferred) / shares.
    dps_line = None
    for cand in ("Dividends Per Share", "Common Stock Dividend Per Share",
                 "Trailing Dividend Rate", "Forward Dividend Rate"):
        v = g(IS, cand) or g(CF, cand)
        if v is not None:
            dps_line = v; dps_src = f"filed per-share line '{cand}' = {v}"; break
    if dps_line is not None:
        dps = dps_line
    else:
        divpaid = g(CF, "Common Stock Dividend Paid")
        if divpaid is None:
            divpaid = g(CF, "Cash Dividends Paid")
        pref = g(IS, "Preferred Stock Dividends") or 0.0
        if divpaid is not None and shares:
            dps = round((abs(divpaid) - abs(pref)) / shares, 6)
            dps_src = (f"|Cash Dividends Paid {divpaid}| - |Pref {pref}| / shares {shares} "
                       f"= {dps}  (no filed per-share line; CONFIRM)")
        else:
            dps = None
            dps_src = "no dividend data found"

    D = {}
    D[5]  = dict(name="in_debt",   meaning="Total debt incl. finance leases",
                 value=debt,   source="BS 'Total Debt'", kind="auto")
    D[6]  = dict(name="in_cash",   meaning="Cash & equivalents",
                 value=cash,   source="BS 'Cash And Cash Equivalents'", kind="auto")
    D[7]  = dict(name="in_sti",    meaning="Short-term investments",
                 value=sti,    source="BS 'Other Short Term Investments' (blank->0)", kind="auto")
    D[9]  = dict(name="anchor_shares0", meaning="Common shares outstanding (mm)",
                 value=shares, source="BS 'Ordinary Shares Number' (/1e6)", kind="auto")
    D[11] = dict(name="in_intexp0", meaning="Interest expense",
                 value=intexp, source="IS 'Interest Expense' (blank->0)", kind="auto")
    D[13] = dict(name="anchor_eps0", meaning="Diluted EPS",
                 value=eps,    source="IS 'Diluted EPS' (unscaled)", kind="auto")
    D[14] = dict(name="in_tax0",   meaning="Effective tax rate",
                 value=tax,    source=tax_src, kind="derived")
    D[15] = dict(name="anchor_dps0", meaning="Dividends per share (FY0)",
                 value=dps,    source=dps_src, kind="derived")
    D[45] = dict(name="in_rd_expense0", meaning="R&D expense (FY0)",
                 value=rd,     source="IS 'Research & Development' (blank->0)", kind="auto")
    D[66] = dict(name="in_anchor_year", meaning="Anchor / FY0 fiscal year",
                 value=anchor_year, source="latest full fiscal year in CSVs", kind="derived")

    # Judgment rows — surfaced WITH the filed number so the user confirms vs data.
    D[8]  = dict(name="in_finlease", meaning="Finance/capital-lease obligations",
                 value=None, kind="judgment",
                 source=f"filed BS 'Capital Lease Obligations' = {g(BS, 'Capital Lease Obligations')}"
                        f" (default 0 unless you include it)")
    D[10] = dict(name="anchor_cse0", meaning="Common stock equity (CSE)",
                 value=None, kind="judgment",
                 source=f"BS 'Common Stock Equity' = {cse}; 'Minority Interest' = {mi}"
                        f"  (Exclude MI -> {cse}; Include MI -> {(cse or 0)+mi})")
    D[12] = dict(name="in_oiadj0", meaning="Operating income, adjusted",
                 value=None, kind="judgment",
                 source=f"IS 'Operating Income' = {oi}; 'Total Unusual Items' = {unusual}"
                        f"  (default = Operating Income)")
    D[25] = dict(name="in_price", meaning="Current share price",
                 value=None, kind="judgment", source="today's market price (external)")
    D[43] = dict(name="in_rd_life", meaning="R&D amortization life (years)",
                 value=None, kind="judgment", source="0 = do not capitalize")

    # stash raw pieces the judgment layer needs
    D["_raw"] = dict(cse=cse, mi=mi, oi=oi, unusual=unusual, rd=rd)
    return D


def apply_judgments(derived, *, price, minority_include, finlease, oi_adj_override,
                    rd_capitalize, rd_life, dps_override=None):
    """Fold the judgment-form values into the derived dict, filling the value of
    every judgment row.  Returns the same dict (mutated).

    dps_override: None -> keep the auto-derived DPS; a number -> use it (e.g. the
    filed per-share dividend, which a standard Yahoo cash-flow export lacks)."""
    raw = derived["_raw"]
    cse = raw["cse"] or 0.0
    mi = raw["mi"] or 0.0
    derived[8]["value"] = float(finlease)
    derived[8]["kind"] = "judgment"
    derived[10]["value"] = (cse + mi) if minority_include else cse
    derived[10]["kind"] = "judgment"
    derived[10]["source"] += f"  -> chose {'Include' if minority_include else 'Exclude'}"
    # oi_adj_override: None/blank -> use filed Operating Income; else use the override
    derived[12]["value"] = raw["oi"] if oi_adj_override in (None, "") else float(oi_adj_override)
    derived[25]["value"] = float(price)
    derived[43]["value"] = float(rd_life) if rd_capitalize else 0.0
    if dps_override is not None:
        derived[15]["value"] = float(dps_override)
        derived[15]["kind"] = "judgment"
        derived[15]["source"] += f"  -> OVERRIDDEN to {float(dps_override)}"
    return derived


def write_inputs(wb, derived):
    """Write every derived/judgment scalar into the Inputs tab. Returns permitted set."""
    inp = wb["Inputs"]
    permitted = set()
    for row, d in derived.items():
        if not isinstance(row, int):
            continue
        v = d["value"]
        if v is None:
            continue
        cell = inp.cell(row, 2)
        if cell.value != v:
            cell.value = v
            permitted.add(("Inputs", cell.coordinate))
    return permitted


# --- diff-guard --------------------------------------------------------------
def snapshot_layer(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    snap = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    snap[(ws.title, c.coordinate)] = c.value
    return snap


def diff_guard(before, after, permitted):
    """Return (changed, illegal). illegal must be empty or the caller must revert."""
    changed = [k for k in set(before) | set(after) if before.get(k) != after.get(k)]
    illegal = [k for k in changed if k not in permitted]
    return changed, illegal

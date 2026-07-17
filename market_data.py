#!/usr/bin/env python3
"""market_data.py — ingest Yahoo historical prices / dividends / splits into the
AEG model's Market Data tab and the near-term dividend input.

Two things the three financial-statement CSVs cannot supply but the model needs:

  1. Contemporaneous year-end prices (Market Data row 16, `md_yeprice`) — drive the
     repurchase-premium / price-to-book buyback adjustment on Econ Statements. The
     price lookup there is wrapped in IFERROR(...,0), so a company loaded WITHOUT
     prices silently zeroes the whole reserve. We populate it, or abort loudly.

  2. The near-term dividend (Inputs B15, `anchor_dps0` -> Valuation B8) — the
     dividend received before the forward-earnings stream begins. We compute the
     trailing-12-month dividend-per-share from the dividends file.

Contemporaneous prices: the model multiplies each year-end price by the AS-FILED
share count (rep_shares), so the price must be as-traded too. Yahoo's "Close" is
split-adjusted to today's basis, so we UN-adjust it using the splits file. Spinoff
adjustments (e.g. AT&T's x1.324 for the WBD spin) are company-specific and appear
in no Yahoo file, so they are a prompted factor — never guessed.
"""
import csv, copy, datetime
import openpyxl


# ---------------------------------------------------------------- date parsing
def _parse_date(s):
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%b %d, %Y", "%d-%b-%y", "%Y/%m/%d"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _find_col(fieldnames, *candidates):
    low = {c.lower().strip(): c for c in fieldnames}
    for want in candidates:
        if want in low:
            return low[want]
    return None


# ---------------------------------------------------------------- splits
def parse_splits(path):
    """Yahoo splits export (Date, Stock Splits). Returns [(date, ratio_float)]
    where ratio = new/old (a '2:1' split -> 2.0)."""
    if not path:
        return []
    out = []
    with open(path, newline="") as fh:
        rd = csv.DictReader(fh)
        dcol = _find_col(rd.fieldnames, "date")
        scol = _find_col(rd.fieldnames, "stock splits", "splits", "split ratio", "ratio")
        if dcol is None or scol is None:
            raise ValueError(f"splits CSV must have Date and Stock Splits columns; got {rd.fieldnames}")
        for row in rd:
            d = _parse_date(row[dcol]); raw = (row[scol] or "").strip()
            if d is None or not raw or raw in ("0", "1:1", "1.0", "1"):
                continue
            if ":" in raw:
                a, b = raw.split(":"); ratio = float(a) / float(b)
            elif "/" in raw:
                a, b = raw.split("/"); ratio = float(a) / float(b)
            else:
                ratio = float(raw)
            if ratio and abs(ratio - 1.0) > 1e-9:
                out.append((d, ratio))
    return sorted(out)


def _cum_split_factor(splits, after_date):
    """Product of split ratios strictly AFTER after_date (used to un-adjust
    Yahoo's split-adjusted Close/Dividends back to as-traded)."""
    f = 1.0
    for d, r in splits:
        if d > after_date:
            f *= r
    return f


# ---------------------------------------------------------------- prices
def parse_prices(path):
    """Yahoo daily historical prices. Returns [(date, close)] using the split-
    adjusted 'Close' column (NOT 'Adj Close', which also nets out dividends)."""
    rows = []
    with open(path, newline="") as fh:
        rd = csv.DictReader(fh)
        dcol = _find_col(rd.fieldnames, "date")
        ccol = _find_col(rd.fieldnames, "close")
        if dcol is None or ccol is None:
            raise ValueError(f"prices CSV must have Date and Close columns; got {rd.fieldnames}")
        for row in rd:
            d = _parse_date(row[dcol]); c = (row[ccol] or "").strip().replace(",", "")
            if d is None or c in ("", "null", "-", "N/A"):
                continue
            try:
                rows.append((d, float(c)))
            except ValueError:
                continue
    return sorted(rows)


def _fy_end_date(year, fy_end_month):
    """Last calendar day of the fiscal-year-end month for a given FY label."""
    if fy_end_month == 12:
        return datetime.date(year, 12, 31)
    # first day of next month, minus a day
    nm = fy_end_month % 12 + 1
    ny = year + (1 if fy_end_month == 12 else 0)
    return datetime.date(ny if nm != 1 else year + 1, nm, 1) - datetime.timedelta(days=1)


def yearend_prices(price_rows, splits, fy_end_month=12,
                   spinoff_factor=1.0, spinoff_before_year=0):
    """Contemporaneous (as-traded) year-end close per fiscal year.

    For each fiscal year, take the last trading day on/before the fiscal year-end,
    un-adjust for any splits AFTER that date, then apply the spinoff factor to
    years before `spinoff_before_year`. Returns {year:int -> price:float}.
    """
    out = {}
    if not price_rows:
        return out
    years = range(price_rows[0][0].year, price_rows[-1][0].year + 2)
    for y in years:
        fye = _fy_end_date(y, fy_end_month)
        # last trading day on/before fiscal year-end
        candidates = [(d, c) for d, c in price_rows if d <= fye]
        if not candidates:
            continue
        d0, close = candidates[-1]
        # only accept if the price is reasonably close to the fiscal year-end
        if (fye - d0).days > 10:  # no trading data near year-end -> skip that year
            continue
        contemp = close * _cum_split_factor(splits, d0)
        if spinoff_before_year and y < spinoff_before_year:
            contemp *= spinoff_factor
        out[y] = round(contemp, 6)
    return out


# ---------------------------------------------------------------- dividends
def parse_dividends(path):
    """Yahoo dividends export (Date, Dividends). Returns [(date, dps)]."""
    rows = []
    with open(path, newline="") as fh:
        rd = csv.DictReader(fh)
        dcol = _find_col(rd.fieldnames, "date")
        vcol = _find_col(rd.fieldnames, "dividends", "dividend", "amount")
        if dcol is None or vcol is None:
            raise ValueError(f"dividends CSV must have Date and Dividends columns; got {rd.fieldnames}")
        for row in rd:
            d = _parse_date(row[dcol]); v = (row[vcol] or "").strip().replace(",", "")
            if d is None or v in ("", "null", "-", "N/A"):
                continue
            try:
                rows.append((d, float(v)))
            except ValueError:
                continue
    return sorted(rows)


def fiscal_year_dps(div_rows, splits, anchor_year, fy_end_month=12):
    """Sum of the per-share dividends whose ex-date falls inside the FY0 fiscal
    year (the quarterly dividends of the anchor year) -> `anchor_dps0`.

    This is the near-term dividend the model adds before the forward stream
    (Valuation!B8), NOT a rolling 12-month figure. For a completed fiscal year it
    is the full set of that year's quarterly dividends; for an in-progress year it
    is what has gone ex so far. Amounts are un-adjusted for any splits after each
    ex-date (to as-traded terms). Returns (dps, [(date, amount)...])."""
    fy_end = _fy_end_date(anchor_year, fy_end_month)
    fy_start = _fy_end_date(anchor_year - 1, fy_end_month) + datetime.timedelta(days=1)
    comps = [(d, round(amt * _cum_split_factor(splits, d), 6))
             for d, amt in div_rows if fy_start <= d <= fy_end]
    return round(sum(a for _, a in comps), 6), comps


# ---------------------------------------------------------------- write md_yeprice
def populate_market_data(wb, yearend, blue_from=None):
    """Write the contemporaneous year-end prices into Market Data row 16
    (md_yeprice), aligned to md_years (row 7). Returns (permitted, written_years).
    Only columns whose md_years label has a price are written; others left as-is
    unless they must be cleared for a fresh company (we clear stale prices first)."""
    MD = wb["Market Data"]
    permitted = set()
    # locate md_years row (7) and md_yeprice row (16) via defined names for safety
    def _row_of(name, default):
        dn = wb.defined_names.get(name)
        if dn is None:
            return default
        ref = str(dn.attr_text).replace("$", "")
        return int("".join(ch for ch in ref.split("!")[1].split(":")[0] if ch.isdigit()))
    yr_row = _row_of("md_years", 7)
    px_row = _row_of("md_yeprice", 16)
    blue = copy.copy((blue_from or MD.cell(px_row, 25)).font)

    written = []
    for c in range(2, MD.max_column + 1):
        y = MD.cell(yr_row, c).value
        if y is None:
            continue
        try:
            yi = int(str(y))
        except ValueError:
            continue
        cell = MD.cell(px_row, c)
        newv = yearend.get(yi)
        if newv is None:
            # clear any stale price so a shorter-history company doesn't inherit
            if cell.value is not None:
                cell.value = None
                permitted.add(("Market Data", cell.coordinate))
        else:
            if cell.value != newv:
                cell.value = newv
                cell.font = copy.copy(blue)
                permitted.add(("Market Data", cell.coordinate))
            written.append(yi)
    return permitted, sorted(written)


# ---------------------------------------------------------------- one entry point
def apply_market_data(wb, derived, *, prices_path=None, dividends_path=None,
                      splits_path=None, anchor_year, fy_end_month=12,
                      spinoff_factor=1.0, spinoff_before_year=0,
                      manual_dps=None):
    """Populate md_yeprice from prices+splits and set the FY0 dividend (Inputs B15).

    Precedence for the near-term dividend: manual form override > dividends file
    (FY0 fiscal-year sum) > the cash-flow-derived fallback already in `derived[15]`.
    Returns (permitted:set, report:dict). Raises ValueError on a malformed file.
    """
    permitted = set()
    report = {"prices_written": [], "n_prices": 0, "splits": [], "dps": None,
              "dps_source": derived[15]["source"], "dps_comps": []}

    splits = parse_splits(splits_path) if splits_path else []
    report["splits"] = [(str(d), r) for d, r in splits]

    if prices_path:
        rows = parse_prices(prices_path)
        ye = yearend_prices(rows, splits, fy_end_month=fy_end_month,
                            spinoff_factor=spinoff_factor,
                            spinoff_before_year=spinoff_before_year)
        perm, written = populate_market_data(wb, ye)
        permitted |= perm
        report["prices_written"] = written
        report["n_prices"] = len(written)
        report["price_fy0"] = ye.get(anchor_year)

    # near-term dividend
    if manual_dps is not None:
        derived[15]["value"] = float(manual_dps)
        derived[15]["kind"] = "judgment"
        derived[15]["source"] = f"manual form override = {float(manual_dps)}"
    elif dividends_path:
        drows = parse_dividends(dividends_path)
        dps, comps = fiscal_year_dps(drows, splits, anchor_year, fy_end_month)
        derived[15]["value"] = dps
        derived[15]["kind"] = "derived"
        derived[15]["source"] = (f"FY{anchor_year} dividends file: "
                                 f"{len(comps)} ex-date(s) summing to {dps}")
        report["dps"] = dps
        report["dps_comps"] = [(str(d), a) for d, a in comps]
    report["dps_source"] = derived[15]["source"]
    return permitted, report


def buyback_is_on(wb):
    try:
        return str(wb["Inputs"]["B47"].value).strip().lower() == "on"
    except Exception:
        return False

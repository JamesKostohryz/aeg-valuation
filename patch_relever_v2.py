#!/usr/bin/env python3
"""relever_v2.py -- V2: re-lever the cost of equity on leverage (Modigliani-Miller,
no tax). Proposed 2026-08-12 in docs/AEG-V2-Relever-Proposal-2026-08-12.md; approved
by James the same day. GATED -- built, not yet proven. This cloud session cannot
recalculate the sealed workbook (no EODHD statement feeds here), so the four-method
tie has NOT been confirmed green with this hook on. Do not treat any number produced
with it on as real until that recalc happens and the tie is checked.

WHAT WAS FOUND, READING THE TEMPLATE DIRECTLY (not a description of it) --
MODEL_TEMPLATE.xlsx already carries two of the three ingredients this needed, unused:

  * Forecast!F33 "constant P/B = real price0 / real book0" is EXACTLY the anchor
    market-to-book multiple the proposal calls for -- already computed from
    anchor price, anchor shares and anchor real CSE, already used elsewhere in the
    sheet (row 32) to price buybacks off a constant-P/B assumption. Nothing new to
    invent here; reused as-is.
  * Forecast!G60:AJ60 "=(finrate_coe + prior FLEV x finrate_cod)/(1+prior FLEV)" is
    the no-tax MM un-levering shape itself -- referenced by NOTHING else in the
    workbook (checked every sheet). This is almost certainly the "dormant DCF
    re-lever layer" repoint_rates.py's install_idio_hook() docstring refers to. It
    uses raw BOOK FLEV, period by period, which is the wrong leverage measure for
    this fleet (see the proposal, section 2C, on why raw book leverage breaks on a
    buyback-shrunk balance sheet like Apple's) -- so it is not simply switched on.
    This module builds the corrected version alongside it rather than editing it,
    so the original dormant row is left exactly as found.

WHAT THIS MODULE ADDS -- six new Market Data rows, 32 through 37 (row 31 is already
the idiosyncratic hook; rows 32-44 were entirely blank in the base template and
referenced by no other sheet -- checked). All are LIVE Excel formulas, not
Python-precomputed values, because everything they read (Forecast rows 24/25/27/55,
F32/F33, and Market Data rows 24/25/28/29/31) was confirmed by direct search to have
no dependency on Market Data row 26 (COE) anywhere in the workbook -- so this is
provably acyclic, single-pass, and needs no iterative or two-pass recalc:

  32  V2 baseline levered rate reconstructed: rf_fwd + market_erp + idiosyncratic
      (NOT a read of row 26 itself -- row 26 will include this module's own output,
      so reading it back would be a circular reference; rebuilt from parts instead)
  33  V2 anchor leverage proxy L0 = in_flev0 / Forecast!$F$33 (book FLEV at the
      anchor, rescaled to a market-like basis by the constant P/B multiple; a single
      scalar, repeated in every column for sheet consistency)
  34  V2 unlevered real rate r_u(tenor), MM no-tax: (row32 + L0*finrate_cod)/(1+L0),
      using the FIXED anchor leverage L0 at every tenor -- per the proposal, business
      risk is solved once and held fixed; only the leverage term should move
  35  V2 period leverage proxy L(t) = Forecast FLEV(t) / Forecast!$F$33 -- the same
      market-to-book rescaling applied to the forecast's OWN driven FLEV path, so it
      moves with whatever the forecast actually does (a debt-funded buyback shows up
      here). IFERROR-guarded to 0 against a forecast period where CSE hits zero.
  36  V2 re-levered COE r_e(t) = r_u(t) + (r_u(t) - finrate_cod(t)) * L(t) --
      diagnostic; not yet consumed by anything until the hook (row 37) is turned on
  37  THE HOOK actually appended onto COE (row 26): row36 - row32, i.e. the delta
      from today's baseline to the re-levered rate. Installed at 0.0 in every column
      (a literal constant, not a formula) -- a provable no-op, exactly like the
      idiosyncratic hook's default. turn_on() overwrites it with the live formula.

Ordering: install_relever_hook() must run AFTER repoint_rates.repoint() (which
installs the idiosyncratic hook at row 31 as its own last step) -- row 32's
reconstruction of the baseline rate reads row 31, and the append onto row 26 must
happen after whatever repoint() and install_idio_hook() have already written there.

No tax adjustment (see the proposal): the enterprise-side valuation has no explicit
interest tax shield to match one, so introducing a (1-tax) term into the leverage
coefficient would price a tax effect with no matching cash flow. Cost of debt is
read straight from finrate_cod (already pre-tax, already the toggle-aware series the
rest of the model uses), matching that choice.
"""
import openpyxl

MD_SHEET = "Market Data"
COL0, COLN = 2, 31          # B..AE = tenors/periods 1..30, same convention as repoint_rates.py
FC_SHEET = "Forecast"
FC_OFFSET = 5                # Market Data col c <-> Forecast col c+5 (B<->G)

ROW = dict(rf_fwd=24, erp=25, coe=26, cod_real=28, finrate_cod=29, idio=31,
           base_rate=32, l0=33, ru=34, lt=35, re_new=36, hook=37)

FC_FLEV_ROW = 55              # Forecast!<col>55 = "FLEV = NFO / CSE"
FC_MB_CELL = "$F$33"          # Forecast!F33 = "constant P/B = real price0 / real book0"


def _colname(c):
    return openpyxl.utils.get_column_letter(c)


def _fc_colname(md_col):
    return openpyxl.utils.get_column_letter(md_col + FC_OFFSET)


def _write_row(MD, r, formula_fn, *, keep_font_from=None):
    font = None
    if keep_font_from is not None:
        import copy
        font = copy.copy(MD.cell(r, keep_font_from).font)
    for c in range(COL0, COLN + 1):
        cell = MD.cell(r, c)
        cell.value = formula_fn(c)
        if font is not None:
            import copy as _copy
            cell.font = _copy.copy(font)


def install_relever_hook(wb, *, relabel=True):
    """Install rows 32-37 (diagnostics + the hook) and append '+{col}37' onto the
    COE formula in row 26. Idempotent: safe to call more than once. The hook is
    written at a literal 0.0 in every column -- the tied base valuation is
    unaffected until turn_on() is called. Must run AFTER repoint_rates.repoint()."""
    if MD_SHEET not in wb.sheetnames or FC_SHEET not in wb.sheetnames:
        raise ValueError(f"[relever_v2] workbook missing '{MD_SHEET}' or '{FC_SHEET}' tab")
    MD = wb[MD_SHEET]

    col0 = _colname(COL0)

    _write_row(MD, ROW["base_rate"],
               lambda c: f"={_colname(c)}{ROW['rf_fwd']}+{_colname(c)}{ROW['erp']}"
                         f"+{_colname(c)}{ROW['idio']}",
               keep_font_from=COL0)

    _write_row(MD, ROW["l0"],
               lambda c: f"=IFERROR(in_flev0/{FC_SHEET}!{FC_MB_CELL},0)",
               keep_font_from=COL0)

    _write_row(MD, ROW["ru"],
               lambda c: (f"=({_colname(c)}{ROW['base_rate']}"
                          f"+{_colname(c)}{ROW['l0']}*{_colname(c)}{ROW['finrate_cod']})"
                          f"/(1+{_colname(c)}{ROW['l0']})"),
               keep_font_from=COL0)

    _write_row(MD, ROW["lt"],
               lambda c: (f"=IFERROR({FC_SHEET}!{_fc_colname(c)}{FC_FLEV_ROW}"
                          f"/{FC_SHEET}!{FC_MB_CELL},0)"),
               keep_font_from=COL0)

    _write_row(MD, ROW["re_new"],
               lambda c: (f"={_colname(c)}{ROW['ru']}"
                          f"+({_colname(c)}{ROW['ru']}-{_colname(c)}{ROW['finrate_cod']})"
                          f"*{_colname(c)}{ROW['lt']}"),
               keep_font_from=COL0)

    # hook row: default OFF (literal 0.0), only if not already a live formula
    for c in range(COL0, COLN + 1):
        cell = MD.cell(ROW["hook"], c)
        if not isinstance(cell.value, (int, float)) and not (
                isinstance(cell.value, str) and cell.value.startswith("=")):
            cell.value = 0.0

    # append the hook onto the COE row, once, after whatever's already there
    for c in range(COL0, COLN + 1):
        cell = MD.cell(ROW["coe"], c)
        f = cell.value
        if not isinstance(f, str) or not f.startswith("="):
            continue
        tag = f"{_colname(c)}{ROW['hook']}"
        if tag in f:
            continue
        cell.value = f + f"+{tag}"

    if relabel:
        MD.cell(ROW["base_rate"], 1).value = (
            "V2 baseline levered rate reconstructed = rf_fwd+market_erp+idio [diagnostic]")
        MD.cell(ROW["l0"], 1).value = (
            "V2 anchor leverage proxy L0 = in_flev0 / Forecast!F33 (constant P/B)")
        MD.cell(ROW["ru"], 1).value = (
            "V2 unlevered real rate ru (MM no-tax, L0 fixed at anchor)")
        MD.cell(ROW["lt"], 1).value = (
            "V2 period leverage proxy L(t) = Forecast FLEV(t) / Forecast!F33")
        MD.cell(ROW["re_new"], 1).value = (
            "V2 re-levered COE re(t) = ru + (ru-finrate_cod)*L(t) [diagnostic]")
        MD.cell(ROW["hook"], 1).value = (
            "V2 RELEVER HOOK [finrate_relever] -- 0.0 = OFF (no-op). "
            "turn_on_relever() switches it live.")

    _redefine_name(wb, "finrate_relever",
                   f"'{MD_SHEET}'!${col0}${ROW['hook']}:${_colname(COLN)}${ROW['hook']}")

    return {"rows_installed": list(ROW.values()), "hook_row": ROW["hook"], "state": "off"}


def turn_on_relever(wb):
    """Switch the hook live: row 37 becomes '=row36-row32' in every column, so COE
    (row 26) actually re-levers. Requires install_relever_hook() to have already run."""
    MD = wb[MD_SHEET]
    for c in range(COL0, COLN + 1):
        col = _colname(c)
        MD.cell(ROW["hook"], c).value = f"={col}{ROW['re_new']}-{col}{ROW['base_rate']}"
    return {"state": "on"}


def turn_off_relever(wb):
    """Restore the hook to a literal 0.0 no-op in every column (diagnostics rows
    32-36 are left in place and still computable/inspectable)."""
    MD = wb[MD_SHEET]
    for c in range(COL0, COLN + 1):
        MD.cell(ROW["hook"], c).value = 0.0
    return {"state": "off"}


def read_diagnostics(recalced_path):
    """Read back the diagnostic rows from an ALREADY-RECALCULATED workbook (data_only).
    Returns per-tenor L0, r_u, L(t), re(t), and the applied delta, for a plain-language
    sanity check before anyone calls the tie green with this hook on."""
    wb = openpyxl.load_workbook(recalced_path, data_only=True)
    MD = wb[MD_SHEET]

    def row(r):
        return [MD.cell(r, c).value for c in range(COL0, COLN + 1)]

    return {
        "tenor": list(range(1, 31)),
        "baseline_rate": row(ROW["base_rate"]),
        "l0_anchor_leverage_proxy": row(ROW["l0"]),
        "unlevered_rate_ru": row(ROW["ru"]),
        "lt_period_leverage_proxy": row(ROW["lt"]),
        "relevered_coe": row(ROW["re_new"]),
        "applied_delta_hook": row(ROW["hook"]),
    }


def _redefine_name(wb, name, ref):
    from openpyxl.workbook.defined_name import DefinedName
    try:
        if name in wb.defined_names:
            del wb.defined_names[name]
    except Exception:
        pass
    wb.defined_names.add(DefinedName(name, attr_text=ref))


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "MODEL_TEMPLATE.xlsx"
    wb = openpyxl.load_workbook(path)
    report = install_relever_hook(wb)
    wb.save(path)
    print(f"[relever_v2] installed rows 32-37 on '{MD_SHEET}', hook state = "
          f"{report['state']} (no-op). Run turn_on_relever(wb) to activate, then "
          f"recalc and check the four-method tie before calling anything done.")

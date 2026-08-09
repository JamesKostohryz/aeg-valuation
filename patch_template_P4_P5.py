#!/usr/bin/env python3
"""patch_template_P4_P5.py — one-shot, idempotent patch to MODEL_TEMPLATE.xlsx.

Applies two fixes to the sealed template and nothing else. Re-running it on an
already-patched template is a no-op (it detects the marker cells and exits).

P4 — Valuation row 11 gets the cfg_mode branching rows 7 and 10 already have.
     Row 11 ("NFO/sh") was the one place in the model that still froze net financial
     obligations per share at the anchor-year real value for the whole forecast,
     regardless of mode, while the Forecast tab grew debt with the equity base under a
     leverage target. Rows 12 (NFE/sh) and 26-28 (the financing AEG buildup that becomes
     V(NFE), Valuation!B37) are built off it, so the entire V(NFE)/V(OI) split ran on a
     financing-cost path out of step with the modelled financing policy. Nothing
     client-facing reads B37/B38 today, which is why it survived.

P5 — the operations wedge. `DCF Reconciliation!B40` — the disagreement between the
     direct operations valuation and the additive one — read 0.2080 against an equity
     value of 1.7379, i.e. 12.0%, and was excluded from the master gate, so the sheet
     printed "ALL SPOKES TIE" beside it.

     Cause, verified on the live engine rather than reasoned about: rho_F is built with
     BOOK leverage weights, (rhoE + FLEV*rhoD)/(1+FLEV) with FLEV = NFO/CSE. That choice
     is deliberate and load-bearing elsewhere — it is exactly what drives residual income
     to zero in the continuing period, because rhoF*(CSE+NFO) = rhoE*CSE + rhoD*NFO
     identically under book weights. But it has a second consequence: it makes residual
     OPERATING income identical to residual EQUITY income, year for year. Measured on the
     Apple golden fixture, max |ReOI - RI| across all thirty columns is exactly 0.0.

     The "direct" construction then discounts that identical stream at rhoF and
     capitalises it at rhoF_LR, while the equity construction discounts the same numbers
     at rhoE and capitalises at rhoE_LR. Since rhoF < rhoE whenever the company carries
     debt, the direct route is guaranteed to come out higher. It cannot tie, and no
     amount of care with the inputs would have made it tie.

     The fix is to discount at a VALUE-weighted rate, which is what the enterprise
     identity actually requires, while leaving the book-weighted rate alone everywhere it
     sets continuing-period earnings. New rows 48-58 on DCF Reconciliation carry the
     value-weighted construction; Audit gains CHECK 7, which is summed into the master
     gate B5.

     Audit CHECK 7 also gates flow additivity, FCFF = FCFE + FCFD. That is a genuine
     identity across three independently built rows (rows 23, 24, 25 are each built from
     a different balance-sheet roll) and it was ungated. It currently reads 6.7e-16.

Also relabels the four cells that are identically zero by construction and therefore
prove nothing — Valuation!B40, Audit!B27, Audit!B28 and DCF Reconciliation!B38 — and
stops summing the tautological pair into the master gate, so the gate stops overstating
how much it covers.
"""
import shutil
import sys

import openpyxl
from openpyxl.utils import get_column_letter as gl

C0, CN = 3, 32          # per-year columns C..AF == t=1..30
MARKER = "V_E(t) — equity value at end of t  [P5 value-weighted enterprise rate]"


def patch(path):
    wb = openpyxl.load_workbook(path)
    V = wb["Valuation"]
    D = wb["DCF Reconciliation"]
    A = wb["Audit"]

    if str(D.cell(49, 1).value or "").startswith("V_E(t)"):
        print("template already patched — no-op")
        return False

    # ---------------------------------------------------------------- P4
    V.cell(11, 1).value = ("NFO/sh — live forecast financing path (per ANCHOR share in "
                           "Enterprise mode)  <P4>")
    for c in range(C0, CN + 1):
        col = gl(c)
        V.cell(11, c).value = (
            f'=IF(cfg_mode="Enterprise",INDEX(fc_nfo,{col}4)/anchor_shares0,'
            f'INDEX(fc_nfo,{col}4)/INDEX(fc_shares,{col}4))')

    # ---------------------------------------------------------------- P5
    D.cell(48, 1).value = ("★ CHECK 7 SUPPORT — value-weighted enterprise rate. rho_F on "
                           "row 14 is BOOK-weighted, which makes ReOI identical to RI; the "
                           "enterprise identity needs VALUE weights.")
    D.cell(49, 1).value = MARKER
    D.cell(50, 1).value = "V_D(t) — debt value at end of t"
    D.cell(51, 1).value = "rho_F*(t) — VALUE-weighted = (rhoE*V_E + rhoD*V_D)/(V_E+V_D), lagged"
    D.cell(52, 1).value = "DF^F*(t) cumulative at rho_F*"
    D.cell(53, 1).value = ("ReOI*(t) = OI_at − (rho_F*(t) − pi_t)*NOA(t−1)   [Stage A: the "
                           "capital charge is at the REAL rate, because the NOA roll is "
                           "NOA_t = (1+pi)*NOA_(t−1) + OI_at − FCFF]")
    D.cell(54, 1).value = "PV(ReOI*) [t<=N]"
    D.cell(55, 1).value = ("V(ops) direct @ rho_F* = NOA0 + ΣPV(ReOI*) + PV_N(V_ops,N − NOA_N) "
                           "— the terminal is the PREMIUM over book, not the whole value")
    D.cell(56, 1).value = "★ TIE: V(ops)@rho_F* − V(ops) additive   [GATE: Audit CHECK 7]"
    D.cell(57, 1).value = "Σ |FCFF − (FCFE + FCFD)|  flow additivity   [GATE: Audit CHECK 7]"
    D.cell(58, 1).value = ("memo: rho_F row 14 stays BOOK-weighted — it is what zeroes "
                           "residual income in the continuing period. Do not 'fix' it there.")

    # value roll-forward: V(t) = V(t-1)*(1+rho_t) - flow_t. Exact, no terminal needed.
    D.cell(49, 2).value = "=B35"      # V_E,0 = V(FCFE) equity
    D.cell(50, 2).value = "=B36"      # V_D,0 = V(net debt)
    D.cell(52, 2).value = 1
    for c in range(C0, CN + 1):
        col, prev = gl(c), gl(c - 1)
        D.cell(49, c).value = f"={prev}49*(1+{col}5)-{col}24"
        D.cell(50, c).value = f"={prev}50*(1+{col}6)-{col}25"
        D.cell(51, c).value = (f"=IF(({prev}49+{prev}50)=0,0,"
                               f"({col}5*{prev}49+{col}6*{prev}50)/({prev}49+{prev}50))")
        D.cell(52, c).value = f"={prev}52/(1+{col}51)"
        D.cell(53, c).value = (f"={col}8-({col}51-INDEX(finrate_infl,{col}4))*{prev}10")
        D.cell(54, c).value = f"=IF({col}4<=cfg_N,{col}53*{col}52,0)"

    D.cell(55, 2).value = ("=B10+SUM(C54:AF54)+INDEX(C52:AF52,cfg_N)"
                           "*(INDEX(C49:AF49,cfg_N)+INDEX(C50:AF50,cfg_N)"
                           "-INDEX(C10:AF10,cfg_N))")
    D.cell(56, 2).value = "=B55-B37"
    D.cell(57, 2).value = "=SUMPRODUCT(ABS(C23:AF23-C24:AF24-C25:AF25))"

    # ------------------------------------------------- tautology relabelling
    V.cell(40, 1).value = ("decorative: V(OI)−V(NFE)−V(EPS). B38 is DEFINED as B36+B37, so "
                           "this is identically 0 for any inputs. Proves nothing.")
    V.cell(41, 1).value = ("diagnostic: V(OI) direct @ BOOK-weighted rhoF — expected to "
                           "differ from B38; the real check is Audit CHECK 7 on the DCF tab.")
    D.cell(38, 1).value = ("decorative: Equity via FCFF = B37−B36 and B37:=B35+B36, so this "
                           "is identically B35. Audit B62 therefore repeats B61.")
    D.cell(40, 1).value = ("diagnostic: direct−additive at the BOOK-weighted rhoF. Cannot "
                           "tie by construction — see row 48 and Audit CHECK 7.")
    A.cell(27, 1).value = ("  (decorative) Equity↔Enterprise tie — identically 0; NOT summed "
                           "into B5")
    A.cell(28, 1).value = ("  (decorative) Value-additivity V(OI)=V(EPS)+V(NFE) — identically "
                           "0 by definition; NOT summed into B5")

    # ------------------------------------------------------------ Audit CHECK 7
    A.cell(74, 1).value = ("CHECK 7 — operations: value-weighted direct vs additive, and flow "
                           "additivity  [both LIVE, both summed into B5]")
    A.cell(75, 1).value = "  V(ops)@rho_F* − V(ops) additive"
    A.cell(75, 2).value = "=ABS('DCF Reconciliation'!B56)"
    A.cell(76, 1).value = "  Σ|FCFF − (FCFE + FCFD)|"
    A.cell(76, 2).value = "=ABS('DCF Reconciliation'!B57)"
    A.cell(77, 1).value = "Σ |CHECK 7|   [GATE: summed into B5]"
    A.cell(77, 2).value = "=B75+B76"

    # master gate: drop the two tautologies, add CHECK 7. B29 (buyback invariance) stays.
    A.cell(5, 2).value = "=B31+B44+B50+B58+B29+B63+B72+B77"
    A.cell(5, 1).value = ("Grand total (LIVE identity residuals only; target 0) — "
                          "tautological terms excluded, see rows 27/28")

    wb.save(path)
    print("patched:", path)
    return True


if __name__ == "__main__":
    tgt = sys.argv[1] if len(sys.argv) > 1 else "MODEL_TEMPLATE.xlsx"
    if len(sys.argv) > 2:
        shutil.copy(tgt, sys.argv[2])
        tgt = sys.argv[2]
    patch(tgt)

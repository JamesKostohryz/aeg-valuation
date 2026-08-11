#!/usr/bin/env python3
"""patch_template_flev_period.py -- make target_flev a LIVE per-year lever in BOTH modes.

THE DEFECT. Forecast row 25 (net financial obligations) reads, in every forecast column:

    =IF(cfg_mode="Enterprise",
        IF(cfg_funding="Target FLEV", <col>24*$F$51/(1+$F$51), <prev>25+<col>51*(<col>24-<prev>24)),
        $F$51*<col>27)

TWO of the three branches are pinned to $F$51 -- the ANCHOR column's financial leverage,
absolutely referenced, in all thirty columns:

  * the Equity branch          ($F$51*<col>27)
  * the Enterprise "Target FLEV" funding branch  (<col>24*$F$51/(1+$F$51))

Only the Enterprise "Funding policy" branch reads the period cell <col>51.

apply_payload writes a submitted `target_flev` into row 51 columns G..(G+N-1); it never
writes column F. So in Equity mode -- and in Enterprise mode under Target FLEV, which is
the shipped default financing mode (Inputs!B68) -- the submitted leverage path is written
into cells nothing reads.

The consequence is a SILENT IGNORE, not an error: apply_payload lists target_flev in its
`written` report, the run succeeds, the gates pass, the four-method tie holds -- and the
valuation reflects the anchor's historical leverage instead of the submitted leverage.
Same failure class as the 2026-08-10 horizon bug.

THE FIX. Point both pinned branches at the period cell <col>51.

BOTH BRANCHES TOGETHER, DELIBERATELY. Fixing only the Equity branch would leave the two
forecast closures structurally unable to agree the moment anyone submits a leverage path,
which is the opposite of the requirement that Equity and Enterprise outputs must agree.

WHY THIS IS PROVABLE. Row 51's period cells ship as `=$F$51`. Any run that does not submit
a leverage path therefore evaluates <col>51 to exactly $F$51 and is bit-identical to the
unpatched model -- same value, same arithmetic, not merely close. The change can only move
an answer for a run that actually submits target_flev, which today is a run whose submitted
leverage is being discarded.

GATED. Changes valuation mechanics in both modes. Requires James's sign-off and a green
four-method tie.

Idempotent. Run:  python3 patch_template_flev_period.py [path/to/MODEL_TEMPLATE.xlsx]
"""
import sys

import openpyxl
from openpyxl.utils import get_column_letter as L

TEMPLATE = sys.argv[1] if len(sys.argv) > 1 else "MODEL_TEMPLATE.xlsx"
SHEET, NFO_ROW, FLEV_ROW, CSE_ROW, NOA_ROW = "Forecast", 25, 51, 27, 24
FIRST_COL, LAST_COL = 7, 36          # G..AJ, forecast periods 1..30


def main():
    wb = openpyxl.load_workbook(TEMPLATE, data_only=False)
    if SHEET not in wb.sheetnames:
        sys.exit(f"FAIL: workbook has no '{SHEET}' tab")
    ws = wb[SHEET]

    changed, already, bad = [], [], []
    for c in range(FIRST_COL, LAST_COL + 1):
        col = L(c)
        cell = ws.cell(row=NFO_ROW, column=c)
        f = cell.value
        if not isinstance(f, str) or not f.startswith("="):
            bad.append(col)
            continue

        subs = [
            # Equity branch
            (f"$F${FLEV_ROW}*{col}{CSE_ROW}", f"{col}{FLEV_ROW}*{col}{CSE_ROW}"),
            # Enterprise "Target FLEV" funding branch
            (f"{col}{NOA_ROW}*$F${FLEV_ROW}/(1+$F${FLEV_ROW})",
             f"{col}{NOA_ROW}*{col}{FLEV_ROW}/(1+{col}{FLEV_ROW})"),
        ]
        new = f
        hits = 0
        for old, rep in subs:
            if old in new:
                new = new.replace(old, rep)
                hits += 1
            elif rep in new:
                hits += 1            # already patched
        if hits != len(subs):
            bad.append(col)
            continue
        if new == f:
            already.append(col)
        else:
            cell.value = new
            changed.append(col)

    if bad:
        sys.exit(f"FAIL: unexpected formula shape in {SHEET}!{NFO_ROW} at columns "
                 f"{','.join(bad)} -- refusing to write a partial patch")
    if not changed:
        print(f"[flev-period] already patched ({len(already)} columns) -- no change")
        return

    wb.save(TEMPLATE)
    print(f"[flev-period] patched {len(changed)} columns ({changed[0]}..{changed[-1]}) "
          f"in {SHEET}!{NFO_ROW}: both pinned branches now read the PERIOD FLEV cell")
    print(f"[flev-period] wrote {TEMPLATE}")


if __name__ == "__main__":
    main()

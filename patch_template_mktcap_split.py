#!/usr/bin/env python3
"""patch_template_mktcap_split.py — make 'Market cap = price x shares' split-consistent.

REGISTER ITEM 12. The defect, restated correctly:

Econ Statements row 93 'Year-end price (contemp.)' is CONTEMPORANEOUS (as-traded) by
design -- market_data.yearend_prices deliberately un-adjusts Yahoo's split-adjusted
close, because the buyback reserve must value repurchases at the price actually paid.
Row 19 'Shares (mm)' is the reported share count, which is on TODAY'S split basis.
Row 94 multiplies the two, which is a product of two different bases: Apple fiscal 2013
computes to $12.4tn against an actual near $444bn, and fiscal 2019 to $4.2tn against
an actual near $1.04tn.

The register prescribed "split-adjust the price series". That remedy is WRONG and would
have corrupted the buyback reserve, which needs the contemporaneous price. The price row
is correct for its purpose. Only the market-cap product is wrong.

The fix: carry the split-ADJUSTED close alongside, in Market Data row 17 (md_yeprice_adj,
written by market_data.apply_market_data), and form market capitalization as
adjusted price x reported shares -- both on today's basis.

SAFE, not gated. Verified 2026-08-09: no formula anywhere in the workbook references
Econ Statements row 93 or 94, and no defined name covers them. They are terminal,
report-facing cells. This patch cannot move a valuation number.

FALLBACK: if md_yeprice_adj is empty for a year -- an older engine, or a run with no
price file -- the formula falls back to the previous product, so nothing breaks and no
run can fail on a missing row.

Idempotent. Run:  python3 patch_template_mktcap_split.py [path/to/MODEL_TEMPLATE.xlsx]
"""
import sys

import openpyxl
from openpyxl.workbook.defined_name import DefinedName

TEMPLATE = sys.argv[1] if len(sys.argv) > 1 else "MODEL_TEMPLATE.xlsx"

MD_SHEET = "Market Data"
ADJ_ROW = 17                     # empty and unnamed before this patch
ADJ_NAME = "md_yeprice_adj"
ADJ_REF = f"'{MD_SHEET}'!$B${ADJ_ROW}:$AP${ADJ_ROW}"
ADJ_LABEL = ("year-end close (split-adjusted, today's basis, $/sh)  "
             "[market cap only -- NOT the buyback reserve]")

ES_SHEET = "Econ Statements"
PRICE_ROW, SHARES_ROW, MKTCAP_ROW = 93, 19, 94
MKTCAP_LABEL = "Market cap = adj. price x shares  [split-consistent]"
FIRST_COL, LAST_COL = 2, 19      # B..S, the year columns


def formula(col_letter):
    """adjusted price x shares, falling back to the contemporaneous product."""
    lookup = f'INDEX({ADJ_NAME},MATCH({col_letter}$5,md_years,0))'
    return (f'=IF(IFERROR({lookup},0)=0,'
            f'{col_letter}{PRICE_ROW}*{col_letter}{SHARES_ROW},'
            f'{lookup}*{col_letter}{SHARES_ROW})')


def main():
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(TEMPLATE)
    md, es = wb[MD_SHEET], wb[ES_SHEET]

    # Guard: refuse to overwrite a row that is not empty. If a future template uses
    # row 17 for something else, fail loud rather than silently clobbering it.
    occupied = [c for c in range(1, md.max_column + 1)
                if c != 1 and md.cell(ADJ_ROW, c).value is not None]
    existing = wb.defined_names.get(ADJ_NAME)
    if occupied and existing is None:
        raise SystemExit(f"REFUSING: {MD_SHEET} row {ADJ_ROW} is not empty and "
                         f"{ADJ_NAME} is not defined. Pick another row.")

    md.cell(ADJ_ROW, 1).value = ADJ_LABEL
    if existing is None:
        wb.defined_names.add(DefinedName(ADJ_NAME, attr_text=ADJ_REF))
        added = f"added defined name {ADJ_NAME} -> {ADJ_REF}"
    else:
        added = f"defined name {ADJ_NAME} already present ({existing.attr_text})"

    es.cell(MKTCAP_ROW, 1).value = MKTCAP_LABEL
    changed = 0
    for c in range(FIRST_COL, LAST_COL + 1):
        L = get_column_letter(c)
        want = formula(L)
        if es.cell(MKTCAP_ROW, c).value != want:
            es.cell(MKTCAP_ROW, c).value = want
            changed += 1

    wb.save(TEMPLATE)
    print(f"patched {TEMPLATE}")
    print(f"  {added}")
    print(f"  {MD_SHEET}!A{ADJ_ROW} labelled")
    print(f"  {ES_SHEET} row {MKTCAP_ROW}: {changed} formula cell(s) rewritten "
          f"(columns {get_column_letter(FIRST_COL)}..{get_column_letter(LAST_COL)})")
    print(f"  example: {formula('D')}")


if __name__ == "__main__":
    main()

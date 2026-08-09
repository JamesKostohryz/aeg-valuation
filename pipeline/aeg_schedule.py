#!/usr/bin/env python3
"""aeg_schedule.py — emit <T>_aeg_schedule.csv: the year-by-year AEG build behind the headline
(COCKPIT blueprint 20260721-2203, task 3b), straight from the Valuation tab's already-materialized
per-year rows. No re-derivation of the valuation — just a tidy read of what the engine computed.

Per forecast period t (1..N): normal earnings, abnormal-earnings-growth (AEG), and the PV
contribution, for each of the three value-additive legs — EPS (=equity), NFE (=financing),
OI (=operations). cum_contrib_eps is the running PV of the equity leg.

Also emits the per-year drivers report §7 asks for (E4.1, display-only, tie-safe):
  retention_rate = retained/sh ÷ EPS/sh          (share of earnings reinvested)
  retained_eps   = retained/sh = EPS − DPS        (Valuation row 9)
  coe            = ρE per year (finrate_coe)       (Valuation row 5)
  rore           = ΔEPS_t ÷ retained_{t-1}         (return on retained equity)
The RoRE definition mirrors the canonical one in pipeline/kit_feeds.py exactly
(rore = _safe(e_t - e_p, e_p - d_p)); t=1 uses the anchor (column B, t=0) as the prior year.
Compare rore against coe: reinvestment creates value iff rore > coe (the AEG value test).

SELF-VERIFYING: normal_value + sum(contrib_eps) must equal the intrinsic value V(EPS) the
headline reports. A mismatch raises (fail-loud) rather than shipping a schedule that does not
tie to the valuation it claims to explain.

Valuation-tab rows: 5 ρE per-year; 7/8/9 EPS/DPS/retained per-share; 18 DF^F cumulative;
22/23/24 normal-EPS / AEG(EPS) / contrib-EPS; 26/27/28 NFE; 30/31/32 OI; 43 Normal value;
44 Intrinsic value (= V(EPS)).
"""
import csv
import os

PERYEAR_ROWS = {
    "df_cumulative": 18,
    "normal_eps": 22, "aeg_eps": 23, "contrib_eps": 24,
    "normal_nfe": 26, "aeg_nfe": 27, "contrib_nfe": 28,
    "normal_oi": 30, "aeg_oi": 31, "contrib_oi": 32,
    # E4.1 driver rows
    "coe": 5, "eps": 7, "dps": 8, "retained": 9,
    # The EQUITY-leg discounting. df_cumulative above is row 18, DF^F — the operating /
    # blended factor — but contrib_eps is built as AEG x A^E, the EQUITY annuity factor on
    # row 62, which starts from DF^E on row 16. Emitting only DF^F left the file unable to
    # explain its own contribution column: anyone reconciling contrib_eps by hand reached
    # for df_cumulative and got the wrong factor. Both equity factors are now emitted, so
    # contrib_eps = aeg_eps x annuity_equity is verifiable from the CSV alone.
    "df_equity": 16, "annuity_equity": 62,
    # dRI^E (row 63) is the residual-income INCREMENT the contribution is actually built
    # from, and it equals raw AEG only in year one (its own formula zeroes the correction
    # term at t=1). Beyond year one it carries the inflation and capital-charge terms the
    # term-structure correction introduced in PR #3, so without it the file reconciles at
    # t=1 and silently fails to reconcile at t=2..N. With it,
    # contrib_eps = dri_equity x annuity_equity holds for every explicit year.
    "dri_equity": 63,
    # Increment 0 Stage A: the engine's forecast path is NOMINAL. These two rows carry the
    # frame so a consumer can deflate. Absent on a pre-Stage-A engine -> emitted blank.
    "pi": 56, "infl_index": 57,
}
R_NORMAL_VALUE = 43     # Normal value = normal EPS1 / rhoE_LR
R_INTRINSIC = 44        # Intrinsic value (= V(EPS))
TIE_TOL = 1e-4          # $/share; the headline itself ties at ~1e-9, so 1e-4 is generous
C_ANCHOR = 2            # column B holds the t=0 anchor for the per-year rows


def _row_series(ws, r):
    out = []
    for c in range(3, ws.max_column + 1):     # per-year cells start at column C
        v = ws.cell(r, c).value
        out.append(v if isinstance(v, (int, float)) else None)
    while out and out[-1] is None:            # trim trailing blanks
        out.pop()
    return out


def _scalar(ws, r):
    for c in range(2, ws.max_column + 1):
        v = ws.cell(r, c).value
        if isinstance(v, (int, float)):
            return v
    return None


def _num(v):
    return v if isinstance(v, (int, float)) else None


def _safe(num, den):
    """num/den, or None if not computable (mirrors kit_feeds._safe)."""
    if num is None or den is None or den == 0:
        return None
    return num / den


def write_aeg_schedule(engine_path, ticker, out_dir):
    import openpyxl
    wb = openpyxl.load_workbook(engine_path, data_only=True)
    V = wb["Valuation"]

    cols = {k: _row_series(V, r) for k, r in PERYEAR_ROWS.items()}
    N = max((len(v) for v in cols.values()), default=0)
    if N == 0:
        raise ValueError("no per-year AEG cells found on the Valuation tab")

    # anchor (t=0) values for the first-year RoRE denominator/numerator
    anchor_eps = _num(V.cell(PERYEAR_ROWS["eps"], C_ANCHOR).value)
    anchor_retained = _num(V.cell(PERYEAR_ROWS["retained"], C_ANCHOR).value)

    # self-verify the equity-leg PV tie
    normal_value = _scalar(V, R_NORMAL_VALUE)
    intrinsic = _scalar(V, R_INTRINSIC)
    sum_contrib = sum(x for x in cols["contrib_eps"] if isinstance(x, (int, float)))
    if isinstance(normal_value, (int, float)) and isinstance(intrinsic, (int, float)):
        resid = abs((normal_value + sum_contrib) - intrinsic)
        if resid > TIE_TOL:
            raise ValueError(
                f"AEG schedule does not tie: normal_value + sum(contrib_eps) = "
                f"{normal_value + sum_contrib:.6f} vs intrinsic {intrinsic:.6f} "
                f"(resid {resid:.2e} > {TIE_TOL:.0e})")

    os.makedirs(out_dir, exist_ok=True)
    fn = f"{ticker}_aeg_schedule.csv"
    #  UNITS (Increment 0 Stage A). Existing column names are unchanged so no downstream
    #  consumer breaks, but be explicit about the frame:
    #    NOMINAL, year-t dollars : normal_eps, aeg_eps, normal_nfe, aeg_nfe, normal_oi,
    #                              aeg_oi, retained_eps, eps, dps
    #    NOMINAL rate            : coe, df_cumulative
    #    REAL 2026$ (time-0 PV)  : contrib_*, cum_contrib_eps
    #    unitless                : retention_rate, rore, pi, infl_index
    #  pi / infl_index are appended so any consumer can deflate: real = nominal / infl_index.
    #  eps_real / dps_real are provided pre-deflated for convenience.
    #  S3 — `phase` labels each year "explicit" (t <= cfg_N) or "continuing" (t > cfg_N).
    #  Without it the schedule shows an unexplained cliff: on the Apple fixture real
    #  earnings per share fall from about $8.61 in year four to about $2.35 in year five.
    #  That is not a defect and not a truncation. Forecast row 16 reads
    #      =IF(t<=cfg_N, EBIT*(1-tax) - depreciation_penalty, required_return * prior NOA)
    #  so beyond the explicit horizon the model IMPOSES that operations earn exactly their
    #  required return and residual income goes to zero. That is the competitive-advantage
    #  period doing its job. Note the correction it implies to an earlier finding: the
    #  DRIVER rows do all propagate across thirty columns, but the EARNINGS path does not —
    #  row 16 branches on cfg_N explicitly, so moving cfg_N also moves where the regime
    #  switch falls, not merely how many years of the same behaviour get booked.
    hdr = ["t", "phase", "df_cumulative", "df_equity", "annuity_equity", "dri_equity",
           "normal_eps", "aeg_eps", "contrib_eps",
           "cum_contrib_eps",
           "normal_nfe", "aeg_nfe", "contrib_nfe", "normal_oi", "aeg_oi", "contrib_oi",
           "retention_rate", "retained_eps", "coe", "rore",
           "pi", "infl_index", "eps_real", "dps_real"]
    cfg_N = None
    try:
        _dn = wb.defined_names.get("cfg_N")
        if _dn is not None:
            _sh, _cell = str(_dn.value).replace("$", "").replace("'", "").split("!")
            cfg_N = int(wb[_sh][_cell].value)
    except Exception:
        cfg_N = None
    with open(os.path.join(out_dir, fn), "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(hdr)
        cum = 0.0
        for i in range(N):
            def g(k):
                s = cols[k]
                return s[i] if i < len(s) else None
            ce = g("contrib_eps")
            if isinstance(ce, (int, float)):
                cum += ce

            # E4.1 drivers — only meaningful for the EXPLICIT forecast years (t<=N, where
            # contrib_eps is a nonzero part of the value build). Beyond N the per-year EPS row
            # enters the continuing-period regime and a cross-boundary RoRE is meaningless, so
            # the driver cells are left blank there.
            is_forecast = isinstance(ce, (int, float)) and ce != 0
            if is_forecast:
                eps_t = _num(g("eps"))
                retained_t = _num(g("retained"))
                coe_t = _num(g("coe"))
                retention_rate = _safe(retained_t, eps_t)
                if i == 0:
                    eps_prev, retained_prev = anchor_eps, anchor_retained
                else:
                    eps_prev = _num(cols["eps"][i - 1]) if i - 1 < len(cols["eps"]) else None
                    retained_prev = _num(cols["retained"][i - 1]) if i - 1 < len(cols["retained"]) else None
                rore = _safe((eps_t - eps_prev) if (eps_t is not None and eps_prev is not None) else None,
                             retained_prev)
            else:
                retention_rate = retained_t = coe_t = rore = None

            def rnd(v, n=6):
                return round(v, n) if isinstance(v, (int, float)) else None

            phase = ("explicit" if (cfg_N is not None and (i + 1) <= cfg_N)
                     else ("continuing" if cfg_N is not None else ""))
            row = [i + 1, phase, g("df_cumulative"), g("df_equity"), g("annuity_equity"),
                   g("dri_equity"),
                   g("normal_eps"), g("aeg_eps"),
                   g("contrib_eps"), round(cum, 8),
                   g("normal_nfe"), g("aeg_nfe"), g("contrib_nfe"),
                   g("normal_oi"), g("aeg_oi"), g("contrib_oi"),
                   rnd(retention_rate), rnd(retained_t), rnd(coe_t), rnd(rore),
                   rnd(_num(g("pi"))), rnd(_num(g("infl_index"))),
                   rnd(_safe(_num(g("eps")), _num(g("infl_index")))),
                   rnd(_safe(_num(g("dps")), _num(g("infl_index"))))]
            w.writerow(["" if v is None else v for v in row])
    return fn


if __name__ == "__main__":
    import sys
    eng = sys.argv[1]
    tk = sys.argv[2] if len(sys.argv) > 2 else "AAPL"
    print("wrote", write_aeg_schedule(eng, tk, "outputs"))

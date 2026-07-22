#!/usr/bin/env python3
"""apply_payload.py — write a RUN-button forecast payload into the sealed engine's
Forecast tab. This is the forecast->engine seam for the cockpit RUN loop
(COCKPIT dispatch contract 20260722-0800).

THE ONE IDEA THAT MAKES THIS SAFE
---------------------------------
Every driver already has a defined "no opinion" behaviour baked into the template's
forecast columns (G..AJ = periods 1..30):

    gross_margin  r10   =$F$10  |  sga_ratio r12 =$F$12  |  da_rate  r14 =$F$14
    tax_rate      r17   =$F$17  |  capex_ratio r36 =$F$36 |  target_flev r51 =$F$51
        -> hold flat at the PERIOD-0 ANCHOR (column F).
    revenue_growth r8, buyback_rate r21, noa_growth r50
        -> =INDEX(scen_<driver>,MATCH(cfg_scenario,scen_cols,0))  (legacy 3-scenario overlay).

So "absent driver = hold flat at the anchor" is implemented by NOT TOUCHING THE CELL.
We only overwrite rows the payload actually supplies. That makes the legacy path
bit-identical BY CONSTRUCTION rather than by careful re-derivation: send no drivers
(or only the legacy three) and every other cell is the exact formula it is today.

Corollary, and it is deliberate: a driver we DO write replaces the formula with 30
literals, so the scenario overlay no longer applies to that row. cfg_scenario keeps
driving only the legacy rows the payload left alone.

UNITS
-----
revenue_growth and noa_growth arrive NOMINAL and are deflated to real with the
engine's OWN expected-inflation forward series (Market Data row 22, finrate_infl) —
the same series the engine discounts with, so the forecast can't drift from the
discount basis. The other six are unit-free ratios/rates in decimals, pass-through.

FAIL-LOUD: bounds come from import_forecast.DRIVER_SPEC. Anything out of range,
wrong length, non-numeric, or an unknown driver name raises before a single cell is
written. A partially-applied forecast is never left on disk.
"""
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import import_forecast as IF   # DRIVER_SPEC / SINGLES bounds + deflate semantics

FORECAST_SHEET = "Forecast"
COL_P1 = 7          # column G == forecast period 1 ; period t -> column COL_P1+t-1
MAX_N = 30          # G..AJ
MD_SHEET = "Market Data"
MD_INFL_FWD_ROW = 22        # expected inflation, 1yr FORWARD (finrate_infl)
MD_COL0 = 2                 # tenor 1 == column B

# payload driver -> Forecast tab row
DRIVER_ROWS = {
    "revenue_growth": 8,
    "gross_margin":   10,
    "sga_ratio":      12,
    "da_rate":        14,
    "tax_rate":       17,
    "buyback_rate":   21,
    "capex_ratio":    36,
    "noa_growth":     50,
}
SINGLE_ROWS = {"target_flev": 51}       # written across the forecast columns
SINGLE_CELLS = {"payout": ("Inputs", "B39")}   # in_payout_seed (equity-mode DPS seed)

CFG_CELLS = {"N": ("Inputs", "B26"), "mode": ("Inputs", "B37")}
MODES = ("Equity", "Enterprise")


class PayloadError(Exception):
    """Malformed payload. Raised before anything is written."""


# ------------------------------------------------------------------ validation
def validate_payload(p):
    """Validate the dispatch payload. Only drivers that are PRESENT are checked —
    absence is legal and means 'hold at anchor'. Returns (ticker, mode, N)."""
    errs = []
    if not isinstance(p, dict):
        raise PayloadError(f"payload must be a JSON object, got {type(p).__name__}")

    ticker = p.get("ticker")
    if not isinstance(ticker, str) or not ticker.strip():
        errs.append("'ticker' must be a non-empty string")

    mode = p.get("mode")
    if mode not in MODES:
        errs.append(f"'mode' must be one of {MODES}, got {mode!r}")

    N = p.get("N")
    if not (isinstance(N, int) and not isinstance(N, bool) and 1 <= N <= MAX_N):
        errs.append(f"'N' must be an integer 1..{MAX_N}, got {N!r}")

    drivers = p.get("drivers") or {}
    if not isinstance(drivers, dict):
        errs.append("'drivers' must be an object")
        drivers = {}
    unknown = sorted(set(drivers) - set(DRIVER_ROWS))
    if unknown:
        errs.append(f"unknown driver(s): {unknown}; expected a subset of {sorted(DRIVER_ROWS)}")

    if isinstance(N, int) and 1 <= N <= MAX_N:
        for name, vals in drivers.items():
            spec = IF.DRIVER_SPEC.get(name)
            if spec is None:
                continue                       # already reported as unknown
            if not isinstance(vals, (list, tuple)):
                errs.append(f"'{name}' must be a list of {N} numbers")
                continue
            if len(vals) != N:
                errs.append(f"'{name}' must have exactly N={N} values, got {len(vals)}")
                continue
            for i, v in enumerate(vals):
                if isinstance(v, bool) or not isinstance(v, (int, float)):
                    errs.append(f"'{name}' Y{i+1} not numeric: {v!r}")
                elif not (spec["lo"] <= v <= spec["hi"]):
                    errs.append(f"'{name}' Y{i+1}={v} out of range [{spec['lo']},{spec['hi']}]")

    singles = p.get("singles") or {}
    if not isinstance(singles, dict):
        errs.append("'singles' must be an object")
        singles = {}
    for name, v in singles.items():
        if name not in IF.SINGLES:
            errs.append(f"unknown single '{name}'; expected a subset of {sorted(IF.SINGLES)}")
            continue
        spec = IF.SINGLES[name]
        if isinstance(v, bool) or not isinstance(v, (int, float)):
            errs.append(f"single '{name}' not numeric: {v!r}")
        elif not (spec["lo"] <= v <= spec["hi"]):
            errs.append(f"single '{name}'={v} out of range [{spec['lo']},{spec['hi']}]")

    if errs:
        raise PayloadError("PAYLOAD VALIDATION FAILED:\n  - " + "\n  - ".join(errs))
    return ticker.strip().upper(), mode, N


# ------------------------------------------------------------------ inflation
def engine_inflation(wb_values, N):
    """Expected inflation, 1yr FORWARD, per forecast year, read from the engine's own
    Market Data row 22. Using the engine's series (not a second copy of the feed)
    guarantees the forecast is deflated with exactly the inflation it is discounted
    against. Requires a data_only (recalculated) workbook."""
    if MD_SHEET not in wb_values.sheetnames:
        raise PayloadError(f"workbook has no '{MD_SHEET}' tab")
    MD = wb_values[MD_SHEET]
    out = []
    for t in range(N):
        v = MD.cell(MD_INFL_FWD_ROW, MD_COL0 + t).value
        if not isinstance(v, (int, float)):
            raise PayloadError(
                f"Market Data row {MD_INFL_FWD_ROW} tenor {t+1} is {v!r}; recalc the "
                f"workbook before applying a payload (need computed inflation)")
        out.append(float(v))
    return out


# ------------------------------------------------------------------ write
def apply_payload(wb, payload, inflation):
    """Write the payload into `wb` (a formulas workbook). Only supplied drivers are
    written; everything else keeps its existing formula. Returns a report dict."""
    ticker, mode, N = validate_payload(payload)
    if len(inflation) < N:
        raise PayloadError(f"inflation series has {len(inflation)} years, need >= {N}")
    if FORECAST_SHEET not in wb.sheetnames:
        raise PayloadError(f"workbook has no '{FORECAST_SHEET}' tab")
    F = wb[FORECAST_SHEET]

    drivers = payload.get("drivers") or {}
    singles = payload.get("singles") or {}

    # deflate nominal growth drivers to real, using DRIVER_SPEC's deflate flags
    real = {}
    for name, vals in drivers.items():
        if IF.DRIVER_SPEC[name]["deflate"]:
            real[name] = [(1 + vals[i]) / (1 + inflation[i]) - 1 for i in range(N)]
        else:
            real[name] = [float(v) for v in vals[:N]]

    written = {}
    for name, vals in real.items():
        row = DRIVER_ROWS[name]
        for t in range(N):
            F.cell(row, COL_P1 + t).value = float(vals[t])
        written[name] = {"row": row, "N": N,
                         "deflated": bool(IF.DRIVER_SPEC[name]["deflate"]),
                         "y1": round(vals[0], 6), "yN": round(vals[-1], 6)}

    for name, v in singles.items():
        if name in SINGLE_ROWS:
            row = SINGLE_ROWS[name]
            for t in range(N):
                F.cell(row, COL_P1 + t).value = float(v)
            written[name] = {"row": row, "N": N, "value": float(v)}
        elif name in SINGLE_CELLS:
            sh, cell = SINGLE_CELLS[name]
            wb[sh][cell] = float(v)
            written[name] = {"cell": f"{sh}!{cell}", "value": float(v)}

    # horizon + mode
    wb[CFG_CELLS["N"][0]][CFG_CELLS["N"][1]] = int(N)
    wb[CFG_CELLS["mode"][0]][CFG_CELLS["mode"][1]] = mode

    held = sorted(set(DRIVER_ROWS) - set(drivers))
    return {"ticker": ticker, "mode": mode, "N": N,
            "written": written, "held_at_anchor": held,
            "note": ("drivers absent from the payload keep their existing formula "
                     "(anchor hold, or the legacy scenario overlay)")}


def load_payload(text_or_path):
    """Accept a JSON string or a path to a JSON file (workflow passes a string)."""
    if os.path.exists(text_or_path):
        with open(text_or_path) as fh:
            return json.load(fh)
    try:
        return json.loads(text_or_path)
    except json.JSONDecodeError as e:
        raise PayloadError(f"payload is not valid JSON: {e}")


if __name__ == "__main__":
    import argparse
    import openpyxl
    ap = argparse.ArgumentParser(description="Apply a RUN payload to an engine workbook.")
    ap.add_argument("engine", help="path to the built engine .xlsx")
    ap.add_argument("payload", help="JSON string or path to a .json file")
    ap.add_argument("--out", help="write here instead of in place")
    args = ap.parse_args()

    p = load_payload(args.payload)
    vals = openpyxl.load_workbook(args.engine, data_only=True)
    infl = engine_inflation(vals, p.get("N", 1))
    wb = openpyxl.load_workbook(args.engine, data_only=False)
    rep = apply_payload(wb, p, infl)
    wb.save(args.out or args.engine)
    print(json.dumps(rep, indent=2))

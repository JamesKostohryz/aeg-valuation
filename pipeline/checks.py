#!/usr/bin/env python3
"""checks.py — the standing per-run reconciliation-tie check.

The completeness/provenance gates in validate_load do NOT check the four-method tie —
they verify the data went in, not that the model reconciles. So a formula edit, a
dependency change, or a bad rate could silently break AEG = RIV = FCFE = FCFF while the
data gates still pass. This check closes that gap: it is a hard, standing gate that runs
every job and fails loud if the reconciliation has drifted. It is the automated form of
the manual "is the tie still tiny?" verification we have been doing by hand.

Three conditions, all required:
  1. the engine's own in-sheet audit verdict is PASS (its relative in_tie_tol logic),
  2. the max identity residual is below an absolute backstop, and
  3. equity and enterprise modes still agree (mode_tie ~ 0).
"""

DEFAULT_TIE_TOL = 1e-8      # absolute backstop on max identity residual (engine ties ~1e-13)
DEFAULT_MODE_TOL = 1e-6     # equity-vs-enterprise agreement


def tie_check(results, tie_tol=DEFAULT_TIE_TOL, mode_tol=DEFAULT_MODE_TOL):
    """Evaluate the standing tie check against an aeg_engine.read_results dict.
    Returns (ok: bool, detail: dict). Never raises — the caller decides to abort."""
    reasons = []

    audit = results.get("audit_status")
    audit_ok = isinstance(audit, str) and "PASS" in audit.upper()
    if not audit_ok:
        reasons.append(f"in-sheet audit not PASS (got {audit!r})")

    tie = results.get("max_identity_tie")
    tie_ok = isinstance(tie, (int, float)) and abs(tie) <= tie_tol
    if not tie_ok:
        reasons.append(f"max identity residual {tie} exceeds backstop {tie_tol:g}")

    mode = results.get("mode_tie")
    # mode_tie may be None when only one mode is active; treat present-and-large as fail
    mode_ok = (mode is None) or (isinstance(mode, (int, float)) and abs(mode) <= mode_tol)
    if not mode_ok:
        reasons.append(f"equity/enterprise disagree (mode_tie {mode} > {mode_tol:g})")

    ok = audit_ok and tie_ok and mode_ok
    return ok, {
        "tie_check": "PASS" if ok else "FAIL",
        "audit_ok": audit_ok,
        "tie_ok": tie_ok,
        "mode_ok": mode_ok,
        "max_identity_tie": tie,
        "mode_tie": mode,
        "tie_tol": tie_tol,
        "reasons": reasons,
    }


def anchor_earnings_check(engine_path):
    """Fail-loud earnings-base guard on the FY0 operating anchor.

    The four-method tie is an ALGEBRAIC IDENTITY on the shared forecast: AEG = RIV = FCFE =
    FCFF reconcile for ANY anchor, a loss included, because all four transform the same
    restated stream. So the tie proves internal consistency, NOT that the anchor is a valid
    base — it cannot catch a cyclical trough. This guard closes that gap for the one
    bright-line, non-judgment case: a NON-POSITIVE FY0 operating-income anchor.

    An AEG going-concern valuation capitalizes the FY0 operating income and grows it. If that
    anchor is a loss (a cyclical trough, a one-off impairment year), the intrinsic value is a
    confident number on an unrepresentative base — worse than a refusal, because it looks valid
    and ties. We guard the OPERATING anchor (anchor_oi_at0), never net income: a firm with
    positive operating income but a net loss (leverage or one-offs BELOW operating income) is
    still legitimately valued at enterprise level, so net income is the wrong signal.

    Returns (ok: bool, detail: dict). Never raises — the caller decides to abort. The remedy is
    deliberately HUMAN (D1: forecasting stays human-in-the-loop): repoint FY0 to a representative
    fiscal year for the anchor (the mechanism that exists today). NOTE: a one-field normalized
    operating-income override is NOT yet wired — the `oi_adj_override` config knob sets in_oiadj0,
    which feeds only an Audit identity that REQUIRES it to equal reported OI, so setting it to a
    real normalized value breaks the tie. Averaging a cycle is a judgment, not plumbing.
    """
    import openpyxl
    wb = openpyxl.load_workbook(engine_path, data_only=True)

    def _dn(name):
        dn = wb.defined_names.get(name)
        if not dn:
            return None
        ref = str(dn.value).replace("$", "").replace("'", "")
        try:
            sh, cell = ref.split("!")
            return wb[sh][cell].value
        except Exception:
            return None

    oi = _dn("anchor_oi_at0")
    yr = _dn("in_anchor_year")
    ok = isinstance(oi, (int, float)) and oi > 0
    reason = None
    if not isinstance(oi, (int, float)):
        ok = False
        reason = ("anchor operating income (anchor_oi_at0) could not be read from the recalced "
                  "engine — cannot confirm the valuation anchor is a going concern")
    elif oi <= 0:
        reason = (
            f"FY0 anchor operating income is NON-POSITIVE (anchor_oi_at0 = {oi:.6g} for anchor "
            f"year {yr}). An AEG going-concern valuation capitalizes and grows the FY0 operating "
            f"income, so a loss year is not a valid anchor: the intrinsic value would be a confident "
            f"number on an unrepresentative base, and it would still TIE. This is the normal state of "
            f"a cyclical at a trough. REMEDY (human judgment, not automatic): repoint FY0 to a "
            f"representative fiscal year for the anchor. (A one-field normalized-OI override is not "
            f"yet wired — oi_adj_override currently only feeds an audit identity; see the engine "
            f"backlog.) Do not let the tool average the cycle for you.")
    return ok, {
        "anchor_earnings_check": "PASS" if ok else "FAIL",
        "anchor_oi_at0": oi,
        "anchor_year": yr,
        "reason": reason,
    }


def rd_wedge_report(engine_path):
    """Diagnose Forecast row 61 — the reported-vs-economic OPERATING-EXPENSE WEDGE
    (R&D plus any opex not itemized as SG&A or economic depreciation), which the engine
    scales forward by revenue. Returns a report used two ways: (1) a structural guard
    that row 61 stays revenue-proportional (catches a silent formula change), and (2) a
    visibility flag so a material wedge on a firm that should have none gets caught.

    Also surfaces the KNOWN GAP: R&D capitalization is currently INERT — the Cap Engine
    reserve (expense x life) is computed but referenced nowhere downstream, so toggling
    rd_capitalize changes neither NOA nor operating income. R&D-heavy names therefore are
    NOT getting a capitalized-R&D restatement yet. (Scoped engine sub-project; see memo.)
    """
    import openpyxl
    wb = openpyxl.load_workbook(engine_path, data_only=True)
    F = wb["Forecast"]
    ebit = F["F15"].value
    wedge = F["F61"].value
    # revenue-scaling consistency: G61/G7 == F61/F7 across the forecast columns
    from openpyxl.utils import get_column_letter as CL
    ratios = []
    for c in range(6, 20):  # F.. through ~19 forecast cols
        r61 = F.cell(61, c).value
        r7 = F.cell(7, c).value
        if isinstance(r61, (int, float)) and isinstance(r7, (int, float)) and r7:
            ratios.append(r61 / r7)
    rev_scaled_ok = bool(ratios) and (max(ratios) - min(ratios)) < 1e-9
    pct_ebit = (abs(wedge) / abs(ebit)) if isinstance(wedge, (int, float)) and ebit else None
    # R&D capitalization wired? reserve present but never consumed -> inert
    CE = wb["Cap Engine"]
    reserve = CE["B61"].value
    rd_reserve_nonzero = isinstance(reserve, (int, float)) and abs(reserve) > 0
    return {
        "opex_wedge": wedge,
        "wedge_pct_ebit": pct_ebit,
        "rev_scaled_consistent": rev_scaled_ok,
        "rd_reserve": reserve,
        "rd_capitalization_wired": False,   # confirmed: reserve referenced nowhere downstream
        "rd_reserve_nonzero_but_inert": rd_reserve_nonzero,
    }

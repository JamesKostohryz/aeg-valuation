#!/usr/bin/env python3
"""convergence.py — E2 convergence period, computed on the engine's own recalc'd numbers.

Purpose (see AEG-Engine-Spec-Convergence-Period + AEG-Methodology-Continuing-Period-Value):
the engine hard-truncates AEG at cfg_N, which extrapolates whatever level actual EPS reached —
so stopping at a cyclical peak overvalues and a trough undervalues. This module inserts a
programmatic K-year glide from actual EPS at the forecast end onto the normalized (mid-cycle)
line, books the reversion AEG with the engine's EXACT formulas, and reports a convergence-
corrected equity value plus a reconciliation guard.

It reads the already-recalc'd, already-tied engine and computes an overlay — it does NOT modify
the sealed four-method tie (that stays green at ~1e-15). The equity (EPS-leg) intrinsic is the
headline number; extending the OI/NFE legs inside the template for four-method consistency of the
convergence increment itself is a separate spreadsheet increment (see the spec).

Engine AEG formulas reproduced here (validated to 1.7e-13 against the live AAPL engine):
  normal_t   = actual_{t-1} + rho_t * retained_{t-1}      (Valuation r22)
  AEG_t      = actual_t - normal_t                        (Valuation r23)
  contrib_t  = AEG_t * DF^E_{t-1} / rho_LR                (Valuation r24, t<=cfg_N)
  intrinsic  = normal_1/rho_LR + sum contrib_t            (Valuation r43+r44)

FAITHFULNESS (verified): when the normalized level equals actual EPS at cfg_N (on-trend name),
every convergence AEG is exactly 0 and the corrected value equals today's number to the penny.
So this is safe by construction — it only ever moves a value when actual is genuinely off-trend
AND a normalized level is supplied.
"""
import csv
import os

# Reconciliation-guard thresholds. A REVIEW verdict now REFUSES the valuation in run_company.py,
# cleared only by `convergence.reviewed: true` in the company config — a human assertion, exactly
# like forecast.reviewed.
#
# THESE TWO NUMBERS ARE PROVISIONAL AND UNSTUDIED. They came from the module's first draft. They
# are the reason the escape hatch belongs to the ANALYST and not to whoever is editing this file:
# a false trip is cleared by a person looking and saying so, never by loosening a threshold here.
# Do not tune them to make a company pass. Revisit them only with evidence from many names.
# ---------------------------------------------------------------------------------------------
# 2026-08-12, on James's ruling. THE CONVERGENCE INCREMENT NO LONGER MOVES VALUE.
#
# What it was: a K-year glide from actual EPS at the forecast end onto a normalized level, whose
# booked abnormal earnings growth was added to the engine value as the headline number.
#
# Why it is gone. The tool was being read as an oracle of the business cycle. It never was one.
# Deciding whether a forecast stops at a cyclical peak is the FORECASTER'S job, and it is already
# implied by the rule that defines a legitimate horizon: the explicit forecast runs until there is
# no projected abnormal earnings growth left. A reversion from a peak back to trend necessarily
# CREATES abnormal earnings growth, so a forecast truncated at a peak cannot satisfy that rule.
# The convergence tool was only ever meant to correct a small residual inconsistency in where the
# forecaster stopped -- and a correction that only matters when it is small is not worth having,
# while one that is large means the forecast is wrong and belongs back with the forecaster rather
# than silently patched into a number nobody owns.
#
# Deleting it also closes the one hole in the correctness oracle: the increment was the only
# component of the published value sitting OUTSIDE the four-method tie. The published number is
# now entirely inside it.
#
# The glide arithmetic is preserved in docs/AEG-CONVERGENCE-RETIRED-2026-08-12.md if it is ever
# needed again. Do not resurrect it here without James.
CONVERGENCE_ADJUSTS_VALUE = False

# What remains are two GATES on the truncation point. They refuse; they never adjust.
#   A. the terminal condition -- abnormal earnings growth must be spent at year N
#   B. the neutral-level condition -- EPS at year N must be at a normalized level
# Cleared, as ever, only by a human assertion in the company config, never by loosening a number
# here. Do not tune these to make a company pass.
GAP_FRAC_WARN = 0.15      # |actual[N] - norm[N]| / actual[N]  above this => stopped off-trend
TAIL_FRAC_WARN = 0.01     # PV of the discarded AEG stream / value  above this => stopped too early

# Trend-contamination diagnostic (DISPLAY ONLY -- moves no value, refuses nothing).
# The normal line's growth is estimated from the last X forecast years, and the level it produces
# is then judged against year N. A cycle that BUILDS across those same X years is absorbed into
# the growth estimate, the peak is declared normal, and the gap guard above sees nothing. Measured
# 2026-08-11: on a synthetic path with a true trend of 5.27% and a peak 25% above trend built over
# three years, the short-window estimate reads 13.37% and the reported gap collapses from 20.0% to
# 0.5% -- a PASS on a company whose published value is then ~18% too high. A one-year spike is
# caught at every window; it is the multi-year shape that hides, and the multi-year shape is what a
# real cycle looks like.
# This diagnostic does not fix that. It makes it VISIBLE: the short-window rate is published beside
# a whole-path rate, and a wide spread means the estimator has probably eaten a cycle. It is a
# reading aid for a human, not a gate, and it must not be turned into one without the study.
TREND_SPREAD_FLAG = 0.02  # short-window minus whole-path growth, above this => estimate suspect


def _nm(wb, name):
    dn = wb.defined_names.get(name)
    if not dn:
        return None
    ref = str(dn.value).replace("$", "").replace("'", "")
    sh, cell = ref.split("!")
    try:
        return wb[sh][cell].value
    except Exception:
        return None


def _series(ws, r):
    return [(ws.cell(r, c).value if isinstance(ws.cell(r, c).value, (int, float)) else None)
            for c in range(2, ws.max_column + 1)]   # index 0 = col B = t=0 anchor


# ---------------------------------------------------------------- inflation frame
# Increment 0 Stage A: the engine may publish a NOMINAL forecast path. When it does,
# Valuation row 56 carries pi_t (expected inflation fwd) and row 57 the cumulative
# index I_t. On a real-frame engine those rows are absent, pi is 0 and every formula
# below collapses to the previous real-terms behaviour. Frame-agnostic by construction.
def _infl(V, N_max=40):
    pi = _series(V, 56)
    if not any(isinstance(x, (int, float)) and x not in (0, None) for x in pi[1:]):
        return (lambda t: 0.0), (lambda t: 1.0)
    last = max(i for i, x in enumerate(pi) if isinstance(x, (int, float)))
    def pi_at(t):
        return float(pi[t]) if (t <= last and isinstance(pi[t], (int, float))) else float(pi[last])
    def idx(t):
        v = 1.0
        for s in range(1, t + 1):
            v *= (1 + pi_at(s))
        return v
    return pi_at, idx


def _guard_terminal_eps(eps, N):
    """The glide is a RATIO of the normalized level to actual EPS at the forecast end, and the
    retention rate is retained/EPS at that same year. Both are meaningless at a zero or negative
    terminal EPS, and a bare ZeroDivisionError deep in the arithmetic tells an operator nothing.
    Refuse with the actual reason instead: a forecast that ends at or below zero earnings has no
    neutral level to hand the continuing period, and capitalizing that level is exactly the error
    the convergence period exists to prevent."""
    v = eps[N] if N < len(eps) else None
    if not isinstance(v, (int, float)) or v <= 0:
        raise ValueError(
            f"actual EPS at the forecast end (year cfg_N={N}) is {v!r}. The convergence period "
            "glides that level onto the normalized line, so a missing, zero or negative terminal "
            "EPS has no normalized ratio and the continuing period cannot be started from a "
            "neutral earnings level. Revisit the forecast or the horizon for this company — do "
            "not capitalize this level.")


def _gates_only(V, N, K, b, actualN, norm_eps_N, eng_intrinsic):
    """The post-2026-08-12 behavior: measure the truncation, judge it, adjust nothing.

    Two gates, both refusing rather than correcting:
      A. terminal condition  -- abnormal earnings growth must be spent at year N
      B. neutral level       -- EPS at year N must sit at the normalized level
    A forecast that passes both has been truncated where the framework says it may be truncated,
    and the engine value publishes unadjusted and wholly inside the four-method tie.
    """
    gap_ps = actualN - norm_eps_N
    gap_frac = abs(gap_ps) / actualN if actualN else 0.0
    term = terminal_aeg_check(V, N, eng_intrinsic)

    fails = []
    if term and term["verdict"] == "REVIEW":
        fails.append("TERMINAL CONDITION -- " + term["reason"])
    if gap_frac > GAP_FRAC_WARN:
        fails.append(
            f"NEUTRAL LEVEL -- EPS at the stop year is {actualN:.4f} against a normalized level of "
            f"{norm_eps_N:.4f}, a gap of {gap_ps:+.4f} per share ({gap_frac:.0%} of EPS). The "
            "continuing period must begin from a neutral earnings level. Move the horizon to a "
            "year that is representative, or fix the forecast drivers.")

    if fails:
        verdict = "REVIEW"
        reason = ("the truncation point does not satisfy the rule. " + " | ".join(fails) +
                  " Both conditions must hold: the explicit forecast does not end until projected "
                  "abnormal earnings growth is spent AND earnings are at a normalized level.")
    else:
        verdict = "PASS"
        reason = (f"truncation valid: gap {gap_frac:.1%} of EPS" +
                  (f", discarded AEG tail {term['tail_frac']:.2%} of value" if term and
                   term.get("tail_frac") is not None else ""))

    return {"N": N, "K": K, "retention": b, "eng_intrinsic": eng_intrinsic,
            "corrected_intrinsic": eng_intrinsic, "converge_value_ps": 0.0,
            "converge_gap_ps": gap_ps, "verdict": verdict, "verdict_reason": reason,
            "terminal": term, "schedule": []}


def converge_valuation(engine_path, K=3, norm_eps_N=None, shape="geometric"):
    """Compute the convergence-corrected equity value from a recalc'd engine.

    K          : convergence length in years (cfg_converge_K; default 3, 0 = off).
    norm_eps_N : the normalized (mid-cycle) EPS level at the forecast end (year cfg_N).
                 None  -> assume on-trend (norm = actual[N]) => no change (faithfulness path).
                 This is where the normalization engine's line feeds in (E2.1).
    Returns a dict with the corrected value, the convergence schedule, and the guard verdict.
    """
    import openpyxl
    wb = openpyxl.load_workbook(engine_path, data_only=True)
    V = wb["Valuation"]
    rho = _series(V, 5)      # rho_E per year
    eps = _series(V, 7)      # actual EPS per year
    ret = _series(V, 9)      # retained per year
    dfE = _series(V, 16)     # DF^E cumulative
    N = int(_nm(wb, "cfg_N"))
    rho_LR = V["B20"].value          # long-run REAL cost of equity (unchanged by Stage A)
    pi_at, _idx = _infl(V)
    eng_intrinsic = V["B44"].value

    if K <= 0 or norm_eps_N is None:
        # off => today's behavior, but still emit a PASS guard + the (empty) convergence block
        return {"N": N, "K": K, "retention": (ret[N] / eps[N] if eps[N] else None),
                "eng_intrinsic": eng_intrinsic, "corrected_intrinsic": eng_intrinsic,
                "converge_value_ps": 0.0, "converge_gap_ps": 0.0,
                "verdict": "PASS", "verdict_reason": "convergence off / on-trend (no glide)",
                "schedule": []}

    _guard_terminal_eps(eps, N)
    b = ret[N] / eps[N]
    actualN = eps[N]

    if not CONVERGENCE_ADJUSTS_VALUE:
        return _gates_only(V, N, K, b, actualN, norm_eps_N, eng_intrinsic)

    def rho_at(t):
        return rho[t] if (t < len(rho) and isinstance(rho[t], (int, float))) else rho_LR

    # normal (AEG=0) continuation off actual[N]
    #  real normal growth from reinvestment, then inflation on top. NOT rho_nom*b:
    #  nominal earnings grow at (1 + rho_real*b)(1 + pi) - 1.
    npath = {N: actualN}
    for t in range(N + 1, N + K + 1):
        p = pi_at(t)
        r_real = (1 + rho_at(t)) / (1 + p) - 1      # de-Fisher the per-year rate
        npath[t] = npath[t - 1] * (1 + r_real * b) * (1 + p)

    # geometric glide of the actual->normalized ratio onto that normal path;
    # ratio = norm_eps_N / actual[N] (1.0 => on-trend => glide == npath => AEG 0)
    ratio = norm_eps_N / actualN
    glide = {N: actualN}
    for t in range(N + 1, N + K + 1):
        frac = (t - N) / K
        glide[t] = npath[t] * (ratio ** frac if shape == "geometric" else 1 + (ratio - 1) * frac)

    # book convergence AEG with the engine's exact formulas; extend DF^E
    dfEx = {N: dfE[N]}
    conv_value = 0.0
    sched = []
    for t in range(N + 1, N + K + 1):
        rt = rho_at(t)
        p = pi_at(t)
        #  benchmark: inflate the prior flow, charge the REAL rate on retention
        normal_t = (1 + p) * glide[t - 1] + (rt - p) * b * glide[t - 1]
        aeg_t = glide[t] - normal_t
        dfEx[t] = dfEx[t - 1] / (1 + rt)
        #  DF sits one period behind the flow, so the cap rate carries (1 + pi_t)
        contrib = aeg_t * dfEx[t - 1] / (rho_LR * (1 + p))
        conv_value += contrib
        sched.append({"t": t, "phase": "convergence", "eps": glide[t], "normal_eps": normal_t,
                      "aeg_eps": aeg_t, "contrib_eps": contrib, "coe": rt})

    corrected = eng_intrinsic + conv_value
    gap_ps = actualN - norm_eps_N

    # reconciliation guard (E2.3)
    gap_frac = abs(gap_ps) / actualN if actualN else 0.0
    val_frac = abs(conv_value) / abs(corrected) if corrected else 0.0
    if gap_frac > GAP_FRAC_WARN or val_frac > VALUE_FRAC_WARN:
        verdict, reason = "REVIEW", (
            f"off-trend gap {gap_frac:.0%} of EPS and convergence moves {val_frac:.0%} of value "
            f"— confirm the analyst's stop year and the normalized line")
    else:
        verdict, reason = "PASS", f"gap {gap_frac:.0%} of EPS, convergence value {val_frac:.0%} of intrinsic"

    return {"N": N, "K": K, "retention": b, "eng_intrinsic": eng_intrinsic,
            "corrected_intrinsic": corrected, "converge_value_ps": conv_value,
            "converge_gap_ps": gap_ps, "verdict": verdict, "verdict_reason": reason,
            "schedule": sched}


def write_convergence_csv(result, ticker, out_dir):
    """Emit <T>_convergence.csv: the convergence block (phase-labeled) + guard header."""
    os.makedirs(out_dir, exist_ok=True)
    fn = os.path.join(out_dir, f"{ticker}_convergence.csv")
    with open(fn, "w", newline="") as fh:
        w = csv.writer(fh)
        _t = result.get("terminal") or {}
        w.writerow(["# truncation gates", f"verdict={result['verdict']}",
                    f"gap_ps={result['converge_gap_ps']:.4f}",
                    f"aeg_at_N={_t.get('aeg_N', float('nan')):.4f}",
                    f"aeg_decay={_t.get('decay', float('nan')):.4f}",
                    ("discarded_tail_frac=DIVERGES" if _t.get("tail_frac") is None else
                     f"discarded_tail_frac={_t['tail_frac']:.4f}"),
                    f"value_ps={result['eng_intrinsic']:.4f}",
                    "convergence_adjustment=RETIRED_2026-08-12"]
                   + ([] if not result.get("trend_diag") else [
                       f"trend_g_short={result['trend_diag']['g_short']:.6f}",
                       f"trend_g_full={result['trend_diag']['g_full']:.6f}",
                       f"trend_spread={result['trend_diag']['spread']:+.6f}"]))
        w.writerow(["t", "phase", "eps", "normal_eps", "aeg_eps", "contrib_eps", "coe"])
        for row in result["schedule"]:
            w.writerow([row["t"], row["phase"], round(row["eps"], 6), round(row["normal_eps"], 6),
                        round(row["aeg_eps"], 8), round(row["contrib_eps"], 8), round(row["coe"], 6)])
    return fn


# ---------------------------------------------------------------- three-period statistics
# James, 2026-08-09: "if there is any calculation of AEG prior to the continuing period, it
# obviously has to include the convergence period. So there can be statistics calculated for
# the explicit forecast period, for the convergence period and for the combination of both."
#
# The explicit block is read straight off the Valuation tab (rows 22/23/24, the engine's own
# normal / AEG / PV-contribution rows for t = 1..cfg_N). The convergence block is the schedule
# converge_valuation just computed. Combined is the two added — which is the ONLY figure that
# should ever be quoted as "abnormal earnings growth before the continuing period."
#
# Self-verifying: normal_value + explicit PV must equal the engine's intrinsic, and
# normal_value + explicit PV + convergence PV must equal the corrected intrinsic. A mismatch
# raises rather than shipping statistics that do not tie to the valuation they describe.
R_NORMAL_VALUE = 43       # Valuation: normal (no-growth) value = normal EPS1 / rho_LR
R_INTRINSIC = 44          # Valuation: intrinsic value V(EPS)
PERIOD_TIE_TOL = 1e-6     # $/share


def _annualized(entry, exit_, years):
    if not isinstance(entry, (int, float)) or not isinstance(exit_, (int, float)):
        return None
    if not years or entry <= 0 or exit_ <= 0:
        return None
    return (exit_ / entry) ** (1.0 / years) - 1.0


def _block(name, t_first, t_last, eps_entry, eps_exit, aegs, pvs, corrected):
    n = max(0, t_last - t_first + 1)
    pv = sum(pvs) if pvs else 0.0
    return {
        "period": name,
        "years": (f"{t_first}..{t_last}" if n else "none"),
        "n_years": n,
        "eps_entry": eps_entry,
        "eps_exit": eps_exit,
        "eps_growth_annualized": _annualized(eps_entry, eps_exit, n),
        "aeg_sum_nominal_ps": (sum(aegs) if aegs else 0.0),
        "aeg_first_ps": (aegs[0] if aegs else None),
        "aeg_last_ps": (aegs[-1] if aegs else None),
        "pv_contribution_ps": pv,
        "pct_of_corrected_value": (pv / corrected if corrected else None),
    }


def period_report(engine_path, result):
    """Explicit / convergence / combined statistics for one recalc'd engine + its convergence
    result. `result` is what converge_valuation or converge_auto returned."""
    import openpyxl
    wb = openpyxl.load_workbook(engine_path, data_only=True)
    V = wb["Valuation"]
    eps = _series(V, 7)          # index 0 = column B = t=0 anchor
    aeg_row = _series(V, 23)
    contrib_row = _series(V, 24)
    N = int(result["N"])
    K = int(result["K"])
    corrected = result["corrected_intrinsic"]

    exp_aeg = [aeg_row[t] for t in range(1, N + 1) if isinstance(aeg_row[t], (int, float))]
    exp_pv = [contrib_row[t] for t in range(1, N + 1) if isinstance(contrib_row[t], (int, float))]
    sched = result.get("schedule") or []
    con_aeg = [s["aeg_eps"] for s in sched]
    con_pv = [s["contrib_eps"] for s in sched]

    eps_anchor = eps[0]
    eps_N = eps[N]
    eps_end = sched[-1]["eps"] if sched else eps_N

    explicit = _block("explicit", 1, N, eps_anchor, eps_N, exp_aeg, exp_pv, corrected)
    conv = _block("convergence", N + 1, N + len(sched), eps_N, eps_end, con_aeg, con_pv, corrected)
    combined = _block("combined", 1, N + len(sched), eps_anchor, eps_end,
                      exp_aeg + con_aeg, exp_pv + con_pv, corrected)

    # --- self-verify against the two value identities
    normal_value = V.cell(R_NORMAL_VALUE, 2).value
    if not isinstance(normal_value, (int, float)):
        for c in range(2, V.max_column + 1):
            if isinstance(V.cell(R_NORMAL_VALUE, c).value, (int, float)):
                normal_value = V.cell(R_NORMAL_VALUE, c).value
                break
    checks = {}
    if isinstance(normal_value, (int, float)):
        r1 = abs(normal_value + explicit["pv_contribution_ps"] - result["eng_intrinsic"])
        r2 = abs(normal_value + combined["pv_contribution_ps"] - corrected)
        checks = {"normal_value_ps": normal_value, "explicit_identity_residual": r1,
                  "combined_identity_residual": r2}
        if r1 > PERIOD_TIE_TOL or r2 > PERIOD_TIE_TOL:
            raise ValueError(
                "convergence period statistics do not tie: normal_value + explicit PV = "
                f"{normal_value + explicit['pv_contribution_ps']:.8f} vs engine "
                f"{result['eng_intrinsic']:.8f} (resid {r1:.2e}); + convergence PV = "
                f"{normal_value + combined['pv_contribution_ps']:.8f} vs corrected "
                f"{corrected:.8f} (resid {r2:.2e})")

    return {"N": N, "K": K, "blocks": [explicit, conv, combined], "identity_checks": checks,
            "eng_intrinsic": result["eng_intrinsic"], "corrected_intrinsic": corrected,
            "norm_eps_N": result.get("norm_eps_N"), "actual_eps_N": eps_N,
            "verdict": result["verdict"], "verdict_reason": result["verdict_reason"]}


PERIOD_FIELDS = ["period", "years", "n_years", "eps_entry", "eps_exit",
                 "eps_growth_annualized", "aeg_sum_nominal_ps", "aeg_first_ps", "aeg_last_ps",
                 "pv_contribution_ps", "pct_of_corrected_value"]


def write_periods_csv(report, ticker, out_dir):
    """Emit <T>_periods.csv — the explicit / convergence / combined statistics.

    UNITS: eps_* and aeg_* are NOMINAL per-share dollars of their own year (so a sum across
    years mixes vintages and is a scale indicator, not a present value). pv_contribution_ps is
    a time-0 present value and IS additive. Quote `combined` whenever describing abnormal
    earnings growth before the continuing period."""
    os.makedirs(out_dir, exist_ok=True)
    fn = os.path.join(out_dir, f"{ticker}_periods.csv")
    with open(fn, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["# AEG by period", f"cfg_N={report['N']}", f"convergence_K={report['K']}",
                    f"actual_eps_N={report['actual_eps_N']}",
                    f"normalized_eps_N={report['norm_eps_N']}",
                    f"engine_intrinsic_ps={report['eng_intrinsic']}",
                    f"corrected_intrinsic_ps={report['corrected_intrinsic']}",
                    f"guard={report['verdict']}"])
        w.writerow(["# CAVEAT", "the convergence increment is computed on the EQUITY (EPS) leg "
                    "only and therefore sits OUTSIDE the four-method tie; the tie covers the "
                    "explicit period"])
        w.writerow(PERIOD_FIELDS)
        for b in report["blocks"]:
            w.writerow([("" if b[k] is None else
                         (round(b[k], 8) if isinstance(b[k], float) else b[k]))
                        for k in PERIOD_FIELDS])
    return f"{ticker}_periods.csv"


def _normal_line_growth(eps, N, X=4):
    """Growth rate of the NORMAL LINE, derived per company from the engine's own path.

    normalization_engine.py defines the normal line as growing at g = b_norm * rho — retention
    times normalized return on retained earnings — and describes rho = cost of equity as only
    "the value-neutral default". Read off the engine's own series, with retention taken at t-1
    and the return earned on it at t, that product reduces by identity to the realized growth
    of the line:

        b_{t-1} * RORE_t = (retained_{t-1} / E_{t-1}) * ((E_t - E_{t-1}) / retained_{t-1})
                         = (E_t - E_{t-1}) / E_{t-1}

    so the growth is read directly. That also keeps it finite as retention approaches zero, and
    under the canonical operating closure retention IS a residual that approaches zero: for the
    golden AAPL fixture it is 1.9%, where the value-neutral default put the normal line at 0.134%
    a year against earnings that in fact track 5.26%. That mismatch is what made the normalizer
    report an identical -7.0% "above normal" at years 4, 5, 6, 7 and 8 of a series with no cycle
    in it — a constant offset, not a cyclical reading, which both fabricated a correction where
    none was due and mis-sized the ones that were.

    The rate is NORMALIZED the way the engine normalizes retention: the median across the
    trailing window, so a single cyclical year perturbs one observation and is rejected.

    Year N is deliberately excluded from its own trend estimate. The entire purpose of the
    normalized level is to judge whether year N is representative, so year N must not help
    define the line against which it is judged.

    Nothing here is a constant: the rate is re-derived per company, per horizon, from that
    company's own forecast path. There is no fallback rate — if the window yields no usable
    observation the line is flat, which asserts nothing rather than inventing a drift.
    """
    import statistics
    rates = []
    for t in range(max(1, N - X + 1), N):
        a, b = eps[t - 1], eps[t]
        if isinstance(a, (int, float)) and isinstance(b, (int, float)) and a > 0:
            rates.append(b / a - 1.0)
    return statistics.median(rates) if rates else 0.0


def terminal_aeg_check(V, N, intrinsic):
    """Gate A. Is abnormal earnings growth actually spent at the truncation point?

    The engine forces AEG to zero from year N+1 onward, because that is what the continuing period
    MEANS -- no further value is created there. That truncation is only legitimate if the forecast
    had already brought AEG to zero. Forcing a small, declining residual to zero breaks nothing.
    Forcing a large or a RISING one to zero silently discards value the forecast itself says exists.

    Measured on the engine's own rows: AEG (row 23) and its present-value contribution (row 24).
    The stream's one-year factor d = AEG[N]/AEG[N-1] says which case we are in. If d >= 1 the
    stream is still growing at the stop year and the discarded tail does not converge at all --
    there is no threshold to argue about, the horizon is simply too short. If d < 1 the tail is
    the geometric continuation, contrib[N] * d/(1-d), and it must be immaterial against the value.
    """
    aeg = _series(V, 23)
    con = _series(V, 24)
    if N >= len(aeg) or not isinstance(aeg[N], (int, float)) or not isinstance(aeg[N - 1], (int, float)):
        return None
    aN, aPrev, cN = float(aeg[N]), float(aeg[N - 1]), float(con[N] or 0.0)
    d = (aN / aPrev) if aPrev else float("inf")
    if d >= 1.0:
        return {"aeg_N": aN, "decay": d, "tail_ps": None, "tail_frac": None, "verdict": "REVIEW",
                "reason": (f"abnormal earnings growth is still GROWING at the stop year "
                           f"(AEG {aPrev:.4f} -> {aN:.4f} per share, factor {d:.3f}). The "
                           "continuing period begins by forcing it to zero, so this truncation "
                           "discards a stream the forecast says is still building. Extend "
                           "forecast.horizon_N until it is spent -- up to thirty years is "
                           "available, and that is what it is for.")}
    tail = cN * d / (1.0 - d)
    frac = abs(tail) / abs(intrinsic) if intrinsic else 0.0
    if frac > TAIL_FRAC_WARN:
        return {"aeg_N": aN, "decay": d, "tail_ps": tail, "tail_frac": frac, "verdict": "REVIEW",
                "reason": (f"the abnormal earnings growth discarded at the stop year is worth "
                           f"{tail:+.2f} per share, {frac:.1%} of value (AEG decaying at a factor "
                           f"of {d:.3f} a year). The forecast has not run to the point where "
                           "abnormal growth is spent. Extend forecast.horizon_N.")}
    return {"aeg_N": aN, "decay": d, "tail_ps": tail, "tail_frac": frac, "verdict": "PASS",
            "reason": f"AEG spent at the stop year: tail {frac:.2%} of value, decaying at {d:.3f}"}


def trend_diagnostics(eps, N, X=4):
    """Compare the normal line's growth as ESTIMATED (short window, X years) against the same
    statistic taken over the whole forecast path. Both use _normal_line_growth, so this compares
    like with like and the only difference is how far back the window reaches.

    A large positive spread means the short window has absorbed a rising cycle and will report the
    terminal year as normal when it is not. A large negative spread is the mirror case on a trough.
    Small spread means the two readings agree and the normalized level can be taken at face value.

    LIMITATION, stated plainly: the whole-path rate is not a clean control. A cycle that occupies
    most of the forecast contaminates it too, and then both readings are wrong together and the
    spread is small. This flags the common case, not every case.

    Returns None when there is not enough path to compare (X >= N), rather than inventing a reading.
    """
    if not isinstance(N, int) or N <= X:
        return None
    g_short = _normal_line_growth(eps, N, X=X)
    g_full = _normal_line_growth(eps, N, X=N)
    spread = g_short - g_full
    return {"g_short": g_short, "g_full": g_full, "spread": spread, "window_X": X,
            "flag": "SUSPECT" if abs(spread) > TREND_SPREAD_FLAG else "OK"}


def normalized_eps_at_N(engine_path, X=4, g=None):
    """Model-default normalized EPS at the forecast end (year cfg_N): take EPS from each of the
    last X years, walk each forward to cfg_N along the normal line, and take the median. This is
    what feeds converge_valuation automatically, and it is the ONLY thing that decides how much
    value the convergence period books — the glide's AEG returns to zero for any level handed to
    it, including a wrong one, so nothing downstream can catch an error here.

    `g` may be supplied to override the derived normal-line growth; when it is None the rate is
    derived per company by _normal_line_growth.

    NO INFLATION RE-INDEX. An earlier version multiplied each walked anchor by the engine's
    cumulative inflation index. Removing it was right, but the reason recorded here was wrong and
    is corrected 2026-08-11: Valuation row 7 is NOMINAL, not constant-dollar. Dividing row 7 by the
    engine's cumulative inflation index reproduces the real EPS row exactly, verified on HD, PG and
    T. The re-index is nonetheless a double count, because the growth rate this function walks the
    anchors at is measured on that same nominal row and is therefore already nominal, and the
    comparator -- actual EPS at year N -- is read off that row too. Frame in, frame out.
    """
    import openpyxl
    import statistics
    wb = openpyxl.load_workbook(engine_path, data_only=True)
    V = wb["Valuation"]
    eps = _series(V, 7)
    N = int(_nm(wb, "cfg_N"))
    _guard_terminal_eps(eps, N)
    if g is None:
        g = _normal_line_growth(eps, N, X=X)
    anchors = [eps[N - a] * (1 + g) ** a
               for a in range(1, X + 1)
               if N - a >= 0 and isinstance(eps[N - a], (int, float))]
    return statistics.median(anchors) if anchors else eps[N]


def converge_auto(engine_path, K=3, X=4):
    """The automatic pipeline path: derive the normalized level from the engine (model default)
    and apply convergence. Returns the same dict as converge_valuation plus the normalized level."""
    import openpyxl
    nl = normalized_eps_at_N(engine_path, X=X)
    out = converge_valuation(engine_path, K=K, norm_eps_N=nl)
    out["norm_eps_N"] = nl
    _wb = openpyxl.load_workbook(engine_path, data_only=True)
    out["trend_diag"] = trend_diagnostics(_series(_wb["Valuation"], 7), out["N"], X=X)
    return out


if __name__ == "__main__":
    import sys
    eng = sys.argv[1]
    K = int(sys.argv[2]) if len(sys.argv) > 2 else 3
    nl = float(sys.argv[3]) if len(sys.argv) > 3 else None
    r = converge_auto(eng, K=K) if nl is None else converge_valuation(eng, K=K, norm_eps_N=nl)
    print(f"cfg_N={r['N']} K={r['K']}  engine={r['eng_intrinsic']:.4f}  "
          f"corrected={r['corrected_intrinsic']:.4f}  conv_value={r['converge_value_ps']:+.4f}  "
          f"verdict={r['verdict']} ({r['verdict_reason']})")

#!/usr/bin/env python3
"""diagnose_tie.py — generalized tie/audit-failure diagnostic.

Run AFTER a build+recalc, against the recalced engine workbook
(pipeline/_work/<TICKER>_engine.xlsx). When the standing tie check fails
(audit_ok / mode_ok / tie_ok in checks.tie_check), this pinpoints WHY without a
human having to open the recalced workbook — which the GitHub Actions API cannot
surface. It is a permanent capability: every future ticker's tie/audit failure
becomes log-diagnosable.

What it does:
  1. Reads the three headline cells the gate consumes
     (Audit!B6 = audit_status, Valuation!B55 = mode_tie, and the Audit identity-tie
     cells) and flags which are spreadsheet errors.
  2. Scans every sheet for cells whose RECALCED value is a spreadsheet error
     (#N/A, #REF!, #VALUE!, #DIV/0!, ...).
  3. Finds the ORIGIN error cells (an error whose own precedents are NOT errors —
     i.e. where the #N/A is born, not merely propagated) and prints each origin's
     formula plus the value of every cell it references, so a blank/among precedents
     shows up immediately.
  4. Dumps the Balance-Sheet 'Minority Interest' row, since an intermittently blank
     NCI line is the known cause of an unresolved equity-bridge/audit cell.
  5. Emits a human report to stdout AND a machine-readable <ticker>_diag.json.

Never raises on a bad/missing workbook — a diagnostic must not mask the failure it
is diagnosing.

Usage: python pipeline/diagnose_tie.py <engine.xlsx> [--ticker TK] [--json PATH]
"""
import sys
import os
import re
import json

ERR_PREFIXES = ("#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!", "#SPILL!")

# Headline cells the standing gate reads — keep in sync with aeg_engine.read_results.
KEY_CELLS = {
    "audit_status": ("Audit", "B6"),
    "mode_tie": ("Valuation", "B55"),
}
TIE_CELLS = ("B27", "B28", "B29", "B31", "B44", "B50", "B58", "B63")  # Audit tab

# A1 reference, optional (possibly quoted) sheet prefix. Requires col+row so bare
# function names (SUM, IF, NA) never match.
REF_RE = re.compile(
    r"(?:(?P<sheet>'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!)?"
    r"\$?(?P<col>[A-Z]{1,3})\$?(?P<row>[0-9]+)"
)


def is_err(v):
    return isinstance(v, str) and any(v.startswith(p) for p in ERR_PREFIXES)


def _val(vals_wb, sheet, coord):
    try:
        return vals_wb[sheet][coord].value
    except Exception:
        return "<unreadable>"


def _refs(formula, default_sheet, sheetnames):
    """Extract direct precedent cells from a formula string.
    Returns list of (sheet, coord). Ranges contribute their endpoints only."""
    out = []
    if not isinstance(formula, str) or not formula.startswith("="):
        return out
    for m in REF_RE.finditer(formula):
        sh = m.group("sheet")
        if sh:
            sh = sh.strip("'")
        else:
            sh = default_sheet
        if sh not in sheetnames:
            continue
        coord = f"{m.group('col')}{m.group('row')}"
        out.append((sh, coord))
    # de-dup, cap
    seen = []
    for t in out:
        if t not in seen:
            seen.append(t)
    return seen[:40]


def main():
    args = sys.argv[1:]
    if not args:
        print("[diag] usage: diagnose_tie.py <engine.xlsx> [--ticker TK] [--json PATH]")
        return 0
    path = args[0]
    ticker = None
    json_path = None
    for i, a in enumerate(args):
        if a == "--ticker" and i + 1 < len(args):
            ticker = args[i + 1]
        if a == "--json" and i + 1 < len(args):
            json_path = args[i + 1]
    if ticker is None:
        base = os.path.basename(path)
        ticker = base.split("_")[0] if "_" in base else os.path.splitext(base)[0]
    if json_path is None:
        json_path = f"{ticker}_diag.json"

    print("=" * 72)
    print(f"TIE / AUDIT DIAGNOSTIC — {ticker} — {path}")
    print("=" * 72)

    if not os.path.exists(path):
        print(f"[diag] workbook not found: {path} (build/recalc never produced it)")
        return 0

    try:
        import openpyxl
    except Exception as e:
        print(f"[diag] openpyxl unavailable: {e}")
        return 0

    try:
        vals = openpyxl.load_workbook(path, data_only=True)
        forms = openpyxl.load_workbook(path, data_only=False)
    except Exception as e:
        print(f"[diag] could not open workbook: {e}")
        return 0

    sheetnames = set(vals.sheetnames)
    report = {"workbook": path, "ticker": ticker, "sheets": sorted(sheetnames)}

    # 1) headline cells --------------------------------------------------------
    print("\n-- HEADLINE CELLS (what the gate reads)")
    key_out = {}
    for name, (sh, coord) in KEY_CELLS.items():
        v = _val(vals, sh, coord) if sh in sheetnames else "<no sheet>"
        err = is_err(v)
        key_out[name] = {"sheet": sh, "coord": coord, "value": v, "is_error": err}
        flag = "  <== ERROR" if err else ""
        print(f"   {name:14} {sh}!{coord} = {v!r}{flag}")
    report["key_cells"] = key_out

    print("\n-- AUDIT IDENTITY-TIE CELLS (Audit tab)")
    tie_out = []
    if "Audit" in sheetnames:
        for coord in TIE_CELLS:
            v = _val(vals, "Audit", coord)
            err = is_err(v)
            tie_out.append({"coord": coord, "value": v, "is_error": err})
            vs = f"{v:,.3e}" if isinstance(v, (int, float)) else repr(v)
            print(f"   Audit!{coord} = {vs}{'  <== ERROR' if err else ''}")
    report["tie_cells"] = tie_out

    # 2) scan every sheet for error cells -------------------------------------
    all_errors = []
    for sh in vals.sheetnames:
        ws = vals[sh]
        for row in ws.iter_rows():
            for c in row:
                if is_err(c.value):
                    all_errors.append((sh, c.coordinate, c.value))
    print(f"\n-- ERROR CELLS FOUND: {len(all_errors)}")
    report["error_cell_count"] = len(all_errors)
    err_set = {(s, co) for (s, co, _v) in all_errors}

    # 3) origin errors (error whose precedents are not themselves errors) ------
    origins = []
    for (sh, co, v) in all_errors:
        try:
            formula = forms[sh][co].value
        except Exception:
            formula = None
        precs = _refs(formula, sh, sheetnames)
        prec_is_err = any((ps, pc) in err_set for (ps, pc) in precs)
        if not prec_is_err:  # #N/A is BORN here, not propagated in
            prec_detail = []
            for (ps, pc) in precs:
                pv = _val(vals, ps, pc)
                prec_detail.append({
                    "ref": f"{ps}!{pc}", "value": pv,
                    "blank": pv is None, "is_error": is_err(pv),
                })
            origins.append({
                "sheet": sh, "coord": co, "value": v,
                "formula": formula, "precedents": prec_detail,
            })
    report["origin_errors"] = origins
    print(f"-- ORIGIN ERROR CELLS (where #N/A is born): {len(origins)}")
    for o in origins[:40]:
        print(f"\n   {o['sheet']}!{o['coord']} = {o['value']}")
        print(f"     formula: {o['formula']}")
        for p in o["precedents"]:
            tag = " [BLANK]" if p["blank"] else (" [ERR]" if p["is_error"] else "")
            pv = p["value"]
            pvs = f"{pv:,.3f}" if isinstance(pv, (int, float)) else repr(pv)
            print(f"       {p['ref']} = {pvs}{tag}")

    # 4) Minority Interest row on the Balance Sheet ---------------------------
    mi = {"present": False}
    if "Balance Sheet" in sheetnames:
        ws = vals["Balance Sheet"]
        r_mi = None
        for r in range(1, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if isinstance(label, str) and label.strip().lower() == "minority interest":
                r_mi = r
                break
        if r_mi:
            cells = []
            for cix in range(2, min(ws.max_column, 44) + 1):
                cell = ws.cell(r_mi, cix)
                yr = ws.cell(3, cix).value
                cells.append({
                    "coord": cell.coordinate, "year": yr,
                    "value": cell.value, "blank": cell.value is None,
                })
            n_blank = sum(1 for c in cells if c["blank"])
            mi = {"present": True, "row": r_mi, "n_blank": n_blank, "cells": cells}
            print(f"\n-- BALANCE SHEET 'Minority Interest' (row {r_mi}): "
                  f"{n_blank}/{len(cells)} year-cells BLANK")
            for c in cells:
                tag = " [BLANK]" if c["blank"] else ""
                print(f"   {c['coord']} ({c['year']}) = {c['value']!r}{tag}")
        else:
            print("\n-- 'Minority Interest' row not found on Balance Sheet")
    report["minority_interest"] = mi

    # 5) verdict --------------------------------------------------------------
    headline_err = any(key_out[k]["is_error"] for k in key_out)
    mi_blank = bool(mi.get("present") and mi.get("n_blank"))
    origin_touches_mi = False
    for o in origins:
        f = o.get("formula") or ""
        if "minority interest" in f.lower():
            origin_touches_mi = True
        if mi.get("present"):
            mi_coords = {c["coord"] for c in mi["cells"]}
            for p in o["precedents"]:
                if p["ref"].startswith("Balance Sheet!") and p["ref"].split("!", 1)[1] in mi_coords:
                    origin_touches_mi = True
    if headline_err and mi_blank and origin_touches_mi:
        verdict = ("CONFIRMED: a blank Minority Interest cell is the origin of the "
                   "#N/A that reddens audit_status/mode_tie.")
    elif headline_err and mi_blank:
        verdict = ("LIKELY: headline cells errored and Minority Interest has blanks; "
                   "inspect the origin-error formulas above to confirm the linkage.")
    elif headline_err:
        verdict = ("headline cells errored but NOT via a blank Minority Interest — "
                   "the origin-error cell above names the real cause (contingency path).")
    else:
        verdict = "no error in the headline cells — failure is elsewhere (see reasons)."
    report["verdict"] = verdict
    print(f"\n-- VERDICT: {verdict}")

    try:
        with open(json_path, "w") as fh:
            json.dump(report, fh, indent=2, default=str)
        print(f"\n[diag] wrote {json_path}")
    except Exception as e:
        print(f"[diag] could not write {json_path}: {e}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

#!/usr/bin/env python3
"""dupont_extract.py — emit a full DuPont decomposition (classic + Penman-reformulated) as a
time series, straight from a recalculated engine workbook. This is the DATA BACKEND for the
DuPont tool: the Sheet reads the CSV; a future web app reads the JSON. Reference implementation
— reuses the engine's already-computed reformulated metrics (no second implementation).

CLASSIC (reported basis, from the Income Statement / Balance Sheet tabs):
  3-step  ROE = NPM (NI/Rev) × ATO (Rev/Assets) × EM (Assets/Equity)
  5-step  ROE = TaxBurden(NI/EBT) × IntBurden(EBT/EBIT) × EBITmargin(EBIT/Rev)
                × ATO(Rev/Assets) × Leverage(Assets/Equity)

PENMAN / REFORMULATED (economic-restated basis, read from Econ Statements):
  ROCE = RNOA + FLEV × Spread,  Spread = RNOA − NBC,  RNOA = OPM × NOA-turnover
  (RNOA, FLEV, NBC, Spread, ROCE, and reported-ROE memo are already computed by the engine.)

NOTE: classic (reported) and reformulated (economic) do NOT reconcile to each other — different
bases (economic restatement lifts equity ~45% and uses economic operating income). Show both,
label the bases. Values are ratios (dimensionless), so no unit scaling needed.
"""
import csv, json, os

IS_HDR = BS_HDR = 3   # reported tabs: row 3 is 'Line item' + year columns
ES_HDR = 5            # Econ Statements: row 5 is the fiscal-year header

# reported line items -> label in the tab (classic DuPont inputs)
IS_LINES = {"revenue": "Total Revenue", "ebit": "Operating Income",
            "ebt": "Pretax Income", "net_income": "Net Income Common Stockholders"}
BS_LINES = {"assets": "Total Assets", "equity": "Common Stock Equity"}
# reformulated metrics -> Econ Statements row
ES_ROWS = {"rnoa": 59, "flev": 60, "nbc": 61, "spread": 62, "roce": 63, "reported_roe": 64}


def _year_series(ws, header_row, label):
    """{year:int -> value} for the row whose col-A label == `label`."""
    years = {c: ws.cell(header_row, c).value for c in range(2, ws.max_column + 1)}
    for r in range(header_row + 1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip().lower() == label.lower():
            out = {}
            for c, y in years.items():
                try:
                    yi = int(str(y))
                except (TypeError, ValueError):
                    continue
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)):
                    out[yi] = float(v)
            return out
    return {}


def _row_year_series(ws, header_row, row):
    """{year:int -> value} for a specific row number (used for Econ Statements metrics)."""
    out = {}
    for c in range(2, ws.max_column + 1):
        y = ws.cell(header_row, c).value
        try:
            yi = int(str(y))
        except (TypeError, ValueError):
            continue
        v = ws.cell(row, c).value
        if isinstance(v, (int, float)):
            out[yi] = float(v)
    return out


def _safe(n, d):
    return (n / d) if (isinstance(n, (int, float)) and isinstance(d, (int, float)) and d) else None


def compute_dupont(engine_path, ticker):
    import openpyxl
    wb = openpyxl.load_workbook(engine_path, data_only=True)
    IS, BS, ES = wb["Income Statement"], wb["Balance Sheet"], wb["Econ Statements"]

    # ---- classic (reported)
    isd = {k: _year_series(IS, IS_HDR, lbl) for k, lbl in IS_LINES.items()}
    bsd = {k: _year_series(BS, BS_HDR, lbl) for k, lbl in BS_LINES.items()}
    years = sorted(set(isd["revenue"]) & set(isd["net_income"]) & set(bsd["assets"]) & set(bsd["equity"]))
    classic = {"years": years, "net_profit_margin": [], "asset_turnover": [], "equity_multiplier": [],
               "roe_3step": [], "tax_burden": [], "interest_burden": [], "ebit_margin": [],
               "leverage": [], "roe_5step": []}
    for y in years:
        rev, ni = isd["revenue"].get(y), isd["net_income"].get(y)
        ebit, ebt = isd["ebit"].get(y), isd["ebt"].get(y)
        assets, eq = bsd["assets"].get(y), bsd["equity"].get(y)
        npm, ato, em = _safe(ni, rev), _safe(rev, assets), _safe(assets, eq)
        classic["net_profit_margin"].append(npm)
        classic["asset_turnover"].append(ato)
        classic["equity_multiplier"].append(em)
        classic["roe_3step"].append(npm * ato * em if None not in (npm, ato, em) else None)
        tb, ib, om = _safe(ni, ebt), _safe(ebt, ebit), _safe(ebit, rev)
        classic["tax_burden"].append(tb); classic["interest_burden"].append(ib)
        classic["ebit_margin"].append(om); classic["leverage"].append(em)
        classic["roe_5step"].append(tb * ib * om * ato * em if None not in (tb, ib, om, ato, em) else None)

    # ---- reformulated (Penman, economic) — already computed by the engine
    ref_series = {k: _row_year_series(ES, ES_HDR, r) for k, r in ES_ROWS.items()}
    ryears = sorted(set().union(*[set(s) for s in ref_series.values()]) if ref_series else [])
    reform = {"years": ryears}
    for k, s in ref_series.items():
        reform[k] = [s.get(y) for y in ryears]

    return {"ticker": ticker, "classic": classic, "reformulated": reform,
            "latest": {"year": years[-1] if years else None,
                       "roe_3step": classic["roe_3step"][-1] if years else None,
                       "roce_reformulated": reform.get("roce", [None])[-1] if ryears else None}}


def write_outputs(dupont, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    t = dupont["ticker"]
    # JSON (web-app friendly)
    with open(os.path.join(out_dir, f"dupont_{t}.json"), "w") as fh:
        json.dump(dupont, fh, indent=2)
    # CSV (Google Sheets IMPORTDATA): year-major, both blocks side by side
    c, r = dupont["classic"], dupont["reformulated"]
    ckeys = [k for k in c if k != "years"]
    rkeys = [k for k in r if k != "years"]
    with open(os.path.join(out_dir, f"dupont_{t}.csv"), "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["year"] + [f"classic_{k}" for k in ckeys] + [f"reform_{k}" for k in rkeys])
        rmap = {y: i for i, y in enumerate(r["years"])}
        for i, y in enumerate(c["years"]):
            row = [y] + [c[k][i] for k in ckeys]
            j = rmap.get(y)
            row += [(r[k][j] if j is not None else "") for k in rkeys]
            w.writerow(["" if v is None else v for v in row])
    return [f"dupont_{t}.json", f"dupont_{t}.csv"]


if __name__ == "__main__":
    import sys
    eng = sys.argv[1] if len(sys.argv) > 1 else "SLIM_OUT.xlsx"
    tk = sys.argv[2] if len(sys.argv) > 2 else "AAPL"
    d = compute_dupont(eng, tk)
    files = write_outputs(d, "outputs")
    print(f"wrote {files}")
    print("latest:", d["latest"])
    print("classic years:", d["classic"]["years"][:3], "...", d["classic"]["years"][-3:])
    print("reform years :", d["reformulated"]["years"][:3], "...", d["reformulated"]["years"][-3:])

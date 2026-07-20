#!/usr/bin/env python3
"""extract.py — pull the restated, model-ready outputs out of a recalculated engine
workbook into committed CSVs, plus a run manifest. These CSVs are what the Google Sheet
`IMPORTDATA`s: the restated real anchors, the valuation, the real statement series that
feed the Work Area historicals, and (when the rate feed is live) the Option-A disclosure
bridge. All deterministic; call only AFTER recalc + gates pass.
"""
import csv, json
import openpyxl

# scalar restated anchors + valuation names to publish (defined-name -> output field)
ANCHOR_NAMES = [
    ("anchor_shares0", "shares"),
    ("anchor_real_noa0", "real_noa"),
    ("anchor_real_nfo0", "real_nfo"),
    ("anchor_real_cse0", "real_cse"),
    ("anchor_real_gross_ppe", "real_gross_ppe"),
    ("anchor_real_net_ppe", "real_net_ppe"),
    ("anchor_real_accum_dep", "real_accum_dep"),
    ("anchor_nfo0", "book_nfo"),
    ("val_realprice", "real_price"),
]
# real statement series (defined-name -> output row label), aligned to Econ Statements yr row 5
SERIES_NAMES = [
    ("es_real_rev", "real_revenue"),
    ("es_real_gp", "real_gross_profit"),
    ("es_real_sga", "real_sga"),
    ("ce_real_gross_series", "real_gross_ppe_series"),
    ("ce_real_net_series", "real_net_ppe_series"),
]
ES_YEAR_ROW = 5   # Econ Statements year header aligned to the B:S real series

# research-note metrics (the old engine 'Presentation' tab, now emitted for the cockpit).
SUMMARY_NAMES = [
    ("val_active", "intrinsic_value_ps"),
    ("val_realprice", "current_price_real_ps"),
    ("ev_tie", "equity_enterprise_tie"),
    ("val_normalval", "normal_no_growth_value_ps"),
    ("val_pvgo", "pvgo_ps"),
    ("val_impliedg", "implied_real_eps_growth"),
    ("val_growthprem", "growth_premium_turns"),
    ("anchor_rnoa0", "economic_rnoa"),
    ("val_rhoe_lr", "real_coe_longrun"),
    ("anchor_real_noa0", "economic_noa"),
]


def _resolve(wb, name):
    dn = wb.defined_names.get(name)
    if not dn:
        return None
    return str(dn.value if hasattr(dn, "value") else dn.attr_text)


def _scalar(wb, name):
    ref = _resolve(wb, name)
    if not ref:
        return None
    ref = ref.replace("$", "").replace("'", "")
    sh, cell = ref.split("!")
    try:
        return wb[sh][cell].value
    except Exception:
        return None


def _range_vals(wb, name):
    """Return (sheet, [(col_idx, value)...]) for a 1-row named range."""
    ref = _resolve(wb, name)
    if not ref:
        return None, []
    ref = ref.replace("$", "").replace("'", "")
    sh, span = ref.split("!")
    ws = wb[sh]
    cells = ws[span]
    out = []
    for row in cells:
        for c in row:
            out.append((c.column, c.value))
    return sh, out


def _write_kv(path, rows):
    with open(path, "w", newline="") as fh:
        fh.write("field,value\n")
        for k, v in rows:
            fh.write(f"{k},{'' if v is None else v}\n")


def _write_series(path, header, rows):
    with open(path, "w", newline="") as fh:
        fh.write(",".join(header) + "\n")
        for r in rows:
            fh.write(",".join("" if x is None else str(x) for x in r) + "\n")


# statement tabs to publish for the cockpit (original + restated financials).
#   values are in ENGINE UNITS ($ trillions: 1.0 = $1,000B); the cockpit scales for display.
STATEMENT_DUMPS = [
    ("Income Statement", "reported_is", 3),   # (tab, out-suffix, header/start row)
    ("Balance Sheet", "reported_bs", 3),
    ("Cash Flow", "reported_cf", 3),
    ("Econ Statements", "restated", 5),
]


def _dump_grid(ws, path, start_row):
    """Write a worksheet's used grid to CSV from start_row down, dropping fully-empty rows
    and trailing empty columns. Values only (call on a data_only workbook)."""
    # last non-empty column across the header + data
    maxc = 1
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value not in (None, ""):
                maxc = max(maxc, c)
    rows = []
    for r in range(start_row, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, maxc + 1)]
        if all(v in (None, "") for v in vals[1:]) and (vals[0] in (None, "")):
            continue  # fully blank row
        rows.append(vals)
    with open(path, "w", newline="") as fh:
        w = csv.writer(fh)
        for row in rows:
            w.writerow(["" if v is None else v for v in row])
    return len(rows)


def extract_outputs(engine_path, ticker, out_dir, *, results, config_hash,
                    vintage, disclosure=None):
    """Write <TICKER>_anchors.csv, _valuation.csv, _restated_real.csv, and _manifest.json.
    `results` is aeg_engine.read_results(...) output; `disclosure` is disclose.disclose(...)
    output or None (rate feed not yet live). Returns the manifest dict."""
    import os
    os.makedirs(out_dir, exist_ok=True)
    wb = openpyxl.load_workbook(engine_path, data_only=True)

    # --- anchors
    anchors = [(field, _scalar(wb, name)) for name, field in ANCHOR_NAMES]
    _write_kv(os.path.join(out_dir, f"{ticker}_anchors.csv"), anchors)

    # --- valuation
    val = [
        ("equity_value", results.get("equity_value")),
        ("enterprise_value", results.get("enterprise_value")),
        ("active_value", results.get("active_value")),
        ("mode_tie", results.get("mode_tie")),
        ("max_identity_tie", results.get("max_identity_tie")),
        ("audit_status", results.get("audit_status")),
        ("tie_check", (results.get("tie_check") or {}).get("tie_check")),
    ]
    if disclosure:
        val += [
            ("adjusted_equity_ps", disclosure.get("adjusted_equity_ps")),
            ("debt_capital_gain_ps", disclosure.get("debt_capital_gain_ps")),
            ("idiosyncratic_haircut_ps", disclosure.get("idiosyncratic_haircut_ps")),
        ]
    _write_kv(os.path.join(out_dir, f"{ticker}_valuation.csv"), val)

    # --- research-note summary (headline value, implied expectations, reformulation)
    summary = [(field, _scalar(wb, name)) for name, field in SUMMARY_NAMES]
    # Headline intrinsic value = firm-specific-risk-INCLUDED value. When the Option-A
    # disclosure ran, report the adjusted equity (market-value debt + idiosyncratic
    # haircut) as intrinsic_value_ps instead of the idio-free tied base. The tied base
    # remains available as disclosure.base_equity_ps for cross-checks.
    if disclosure and isinstance(disclosure.get("adjusted_equity_ps"), (int, float)):
        summary = [(f, disclosure["adjusted_equity_ps"] if f == "intrinsic_value_ps" else v)
                   for f, v in summary]
    if "val_active" in [n for n, _ in SUMMARY_NAMES] and results.get("equity_value") is not None:
        va = dict(summary).get("intrinsic_value_ps"); rp = _scalar(wb, "val_realprice")
        if isinstance(va, (int, float)) and isinstance(rp, (int, float)) and rp:
            summary.append(("upside_downside", va / rp - 1))
    _write_kv(os.path.join(out_dir, f"{ticker}_summary.csv"), summary)

    # --- real statement series (aligned to the Econ Statements year row)
    ES = wb["Econ Statements"]
    # year header from the same columns the first series spans
    sh, first = _range_vals(wb, SERIES_NAMES[0][0])
    cols = [c for c, _ in first]
    years = [ES.cell(ES_YEAR_ROW, c).value for c in cols]
    series_rows = [["year"] + [str(y) for y in years]]
    for name, label in SERIES_NAMES:
        _, vals = _range_vals(wb, name)
        series_rows.append([label] + [v for _, v in vals])
    # transpose to year-major for a friendlier IMPORTDATA shape
    header = ["year"] + [r[0] for r in series_rows[1:]]
    trows = []
    for j, y in enumerate(years):
        trows.append([str(y)] + [series_rows[k + 1][j + 1] for k in range(len(SERIES_NAMES))])
    _write_series(os.path.join(out_dir, f"{ticker}_restated_real.csv"), header, trows)

    # --- original + restated financial statements (for the cockpit)
    stmt_files = []
    for tab, suffix, start in STATEMENT_DUMPS:
        if tab in wb.sheetnames:
            fn = f"{ticker}_{suffix}.csv"
            _dump_grid(wb[tab], os.path.join(out_dir, fn), start)
            stmt_files.append(fn)

    # --- manifest
    manifest = {
        "ticker": ticker,
        "config_hash": config_hash,
        "data_vintage": vintage,
        "gates_ok": bool(results.get("ok")),
        "tie_check": results.get("tie_check"),
        "rd_wedge": results.get("rd_wedge"),
        "audit_status": results.get("audit_status"),
        "max_identity_tie": results.get("max_identity_tie"),
        "equity_value": results.get("equity_value"),
        "enterprise_value": results.get("enterprise_value"),
        "mode_tie": results.get("mode_tie"),
        "anchor_year": results.get("anchor_year"),
        "disclosure": (None if not disclosure else {
            "base_equity_ps": disclosure.get("base_equity_ps"),
            "adjusted_equity_ps": disclosure.get("adjusted_equity_ps"),
            "debt_capital_gain_ps": disclosure.get("debt_capital_gain_ps"),
            "idiosyncratic_haircut_ps": disclosure.get("idiosyncratic_haircut_ps"),
        }),
        "outputs": [f"{ticker}_anchors.csv", f"{ticker}_valuation.csv",
                    f"{ticker}_summary.csv", f"{ticker}_restated_real.csv"] + stmt_files,
    }
    with open(os.path.join(out_dir, f"{ticker}_manifest.json"), "w") as fh:
        json.dump(manifest, fh, indent=2, default=str)
    return manifest

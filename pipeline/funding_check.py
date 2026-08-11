#!/usr/bin/env python3
"""funding_check.py — the UNFUNDED DISTRIBUTION guard.

WHY THIS EXISTS (2026-08-11). Under the canonical operating closure, net operating assets
and operating income are driven, financing absorbs, and distributions are IMPLIED. Forecast
row 29 in its Enterprise branch computes the dividend as a residual:

    DPS_t = (net income_t − change in common equity_t − repurchases_t) / shares_t

A residual can come out negative. A negative implied dividend is arithmetically valid and
economically a capital raise: the model is asserting the company issues equity in order to
fund a share repurchase it cannot afford, while simultaneously retiring shares. That is not
a forecast anyone would sign.

It is invisible to everything else on this system. Measured on the golden Apple fixture
under the default Consensus overlay (three percent buyback against 2.5 percent
net-operating-asset growth), the implied dividend is negative in every forecast year, from
−2.1748 to −2.4310 per share — and the four-method tie reads 8.4e-16, the audit reads PASS,
and the convergence reconciliation guard reads PASS. The only automated check that noticed
was test_convergence.py, and only because it happened to hard-code a directional expectation
about Apple that the unfunded plan reversed. See
AEG-REGRESSION-Third-Failure-FOUND-2026-08-11.md.

This is the same failure class as the horizon bug and the leverage bug: a number that is
silently wrong while every gate reports success. The remedy is the same one that worked for
those — refuse, and make a person say otherwise.

THE TWO-OF-THREE RULE, COMPLETED. The resolution document states that a forecaster may set
the operating plan, the distribution policy, or the financing structure — any two, never all
three, because the balance sheet has to balance. The engine enforces that by construction.
What it did not do is check that the implied third term is ECONOMICALLY ADMISSIBLE. This
module is that check.

SCOPE. The guard applies only under the canonical closure (cfg_mode = "Enterprise"), where
distributions are implied. In the Equity presentation branch row 29 is the payout seed times
a base — set, not implied — so there is no residual to police and the guard reports
NOT_APPLICABLE rather than inventing an opinion.

TOLERANCE. A dividend of exactly zero is admissible: a company that distributes nothing is
plausible. Only a genuinely negative residual trips the guard, and only beyond a small
absolute floor so that floating-point dust at 1e-15 cannot refuse a valuation.
"""
import openpyxl

# Forecast tab geometry. Anchor is column F (6); forecast periods run G.. (7..).
SHEET = "Forecast"
ROW_NI, ROW_SHARES, ROW_CSE, ROW_DPS, ROW_REPURCHASE = 19, 20, 27, 29, 31
ANCHOR_COL, FIRST_FORECAST_COL = 6, 7

# Floating-point dust floor, in per-share currency units. A residual dividend inside this
# band is treated as zero. This is a numerical tolerance, NOT a materiality threshold, and
# it must never be widened to make a company pass.
ZERO_FLOOR = 1e-9


def _num(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return v if isinstance(v, (int, float)) else None


def funding_report(engine_path):
    """Inspect the recalculated engine's implied distribution path.

    Returns a dict with a verdict of PASS, REVIEW or NOT_APPLICABLE, plus the per-year
    detail a reviewer needs to see what the plan is actually asserting. Reads the
    workbook's own row 29 rather than recomputing it, so the guard tracks the model
    instead of drifting from it; the funding arithmetic is recomputed alongside purely
    to explain the shortfall.
    """
    wb = openpyxl.load_workbook(engine_path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"funding_check: workbook has no {SHEET!r} tab")
    mode = wb["Inputs"]["B37"].value
    N = int(wb["Inputs"]["B26"].value or 0)
    ws = wb[SHEET]

    if mode != "Enterprise":
        return {"verdict": "NOT_APPLICABLE", "mode": mode, "N": N, "years": [],
                "reason": ("distributions are SET by the payout seed in the equity "
                           "presentation branch, not implied — there is no residual to check")}

    years, worst = [], None
    for i in range(N):
        c = FIRST_FORECAST_COL + i
        dps = _num(ws, ROW_DPS, c)
        if dps is None:
            continue
        ni = _num(ws, ROW_NI, c)
        cse, cse_prev = _num(ws, ROW_CSE, c), _num(ws, ROW_CSE, c - 1)
        rep = _num(ws, ROW_REPURCHASE, c)
        sh = _num(ws, ROW_SHARES, c)
        capacity = (ni - (cse - cse_prev)) if None not in (ni, cse, cse_prev) else None
        shortfall = (capacity - rep) if None not in (capacity, rep) else None
        rec = {"year": i + 1, "implied_dps": dps, "net_income": ni,
               "equity_increase": (cse - cse_prev) if None not in (cse, cse_prev) else None,
               "distribution_capacity": capacity, "repurchases": rep, "shares": sh,
               "funding_shortfall": shortfall, "ok": dps >= -ZERO_FLOOR}
        years.append(rec)
        if not rec["ok"] and (worst is None or dps < worst["implied_dps"]):
            worst = rec

    bad = [y for y in years if not y["ok"]]
    if not bad:
        return {"verdict": "PASS", "mode": mode, "N": N, "years": years, "worst": None,
                "reason": (f"implied dividend is non-negative in all {len(years)} forecast "
                           f"year(s)")}
    return {"verdict": "REVIEW", "mode": mode, "N": N, "years": years, "worst": worst,
            "n_bad": len(bad),
            "reason": (f"implied dividend is NEGATIVE in {len(bad)} of {len(years)} forecast "
                       f"year(s); worst {worst['implied_dps']:+.4f}/sh in year "
                       f"{worst['year']} — the plan implies issuing equity to fund a buyback")}


def format_report(rep):
    """A compact, human-readable rendering for the run log and the refusal message."""
    if rep["verdict"] == "NOT_APPLICABLE":
        return f"[funding] NOT_APPLICABLE (cfg_mode={rep['mode']!r}) — {rep['reason']}"
    lines = [f"[funding] guard {rep['verdict']}: {rep['reason']}"]
    for y in rep["years"]:
        flag = "" if y["ok"] else "   <-- UNFUNDED"
        cap = "n/a" if y["distribution_capacity"] is None else f"{y['distribution_capacity']:.6f}"
        rp = "n/a" if y["repurchases"] is None else f"{y['repurchases']:.6f}"
        lines.append(f"[funding]   year {y['year']:<2d} implied DPS {y['implied_dps']:+10.4f}/sh"
                     f"  capacity {cap}  repurchases {rp}{flag}")
    return "\n".join(lines)


if __name__ == "__main__":
    import sys, json
    r = funding_report(sys.argv[1])
    print(format_report(r))
    if "--json" in sys.argv:
        print(json.dumps(r, indent=1, default=float))
    raise SystemExit(0 if r["verdict"] in ("PASS", "NOT_APPLICABLE") else 2)

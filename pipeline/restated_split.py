#!/usr/bin/env python3
"""restated_split.py — emit the real-terms reformulated statements as separate IS and BS
CSVs for the cockpit's Financials-Restated tab (COCKPIT blueprint 20260721-2203, task 3a).

Splits the engine's 'Econ Statements' ECONOMIC RESTATEMENT (real 2026$) section:
  <T>_restated_bs.csv = real stocks  (rows 39-46: Real net PP&E, Real NOA/NFO/CSE + tie)
  <T>_restated_is.csv = real flows   (rows 48-56: Real EBITDA .. Comprehensive income)
Columns = fiscal years (Econ Statements row 5).

UNITS — verified, and worth stating precisely so nobody mislabels the cockpit:
Econ Statements values are in the engine's INFERRED per-ticker scale, and — verified against
the built engines — the reported IS/BS/CF tabs use the SAME scale. So restated_* is in exactly
the same units as the reported_* files the cockpit already imports: trillions for AAPL (total
debt reads 0.112377 = $112B), millions for T (173986 = $174B). NOT uniformly $millions. tab 6
(Reported) and tab 7 (Restated) are therefore consistent WITHIN a ticker; they differ ACROSS
tickers, exactly as the existing reported_* feeds already do. As a guard we CONFIRM reported and
restated share scale (BS 'Total Debt' == Econ Statements row-14 'Total debt') and fail loud if
that invariant ever breaks — the only way this file's units could silently diverge from reported.

CASH FLOW is intentionally NOT emitted. The engine computes no real-terms cash flow that ties
to the four-method identity, and a fabricated CF is exactly the silent-wrong-value we guard
against (COCKPIT 20260722-0304 decision: defer restated_cf until a tying construction exists).
"""
import csv
import os

ES_HDR = 5
ES_TOTAL_DEBT_ROW = 14   # Econ Statements 'Total debt' (nominal) — same quantity/scale as BS 'Total Debt'
# out-suffix -> (first row, last row) within Econ Statements; REAL-terms sections only.
SECTIONS = {
    "restated_bs": (39, 46),   # ECONOMIC RESTATEMENT (real 2026$): real stocks + tie row
    "restated_is": (48, 56),   # ECONOMIC FLOWS (real 2026$)
}


def _year_cols(ws):
    out = []
    for c in range(2, ws.max_column + 1):
        y = ws.cell(ES_HDR, c).value
        try:
            out.append((c, int(str(y))))
        except (TypeError, ValueError):
            pass
    return out


def _assert_scale_matches_reported(wb, ES, yrs):
    """Guard: reported (BS tab) and restated (Econ Statements) must share the engine scale.
    Confirmed by BS 'Total Debt' == ES row-14 'Total debt' (same quantity, same year)."""
    BS = wb["Balance Sheet"]
    bs_debt = {}
    for r in range(4, BS.max_row + 1):
        if str(BS.cell(r, 1).value or "").strip().lower() == "total debt":
            for c in range(2, BS.max_column + 1):
                y = BS.cell(3, c).value
                try:
                    bs_debt[int(str(y))] = BS.cell(r, c).value
                except (TypeError, ValueError):
                    continue
            break
    for c, y in reversed(yrs):
        es_d = ES.cell(ES_TOTAL_DEBT_ROW, c).value
        bs_d = bs_debt.get(y)
        if isinstance(es_d, (int, float)) and isinstance(bs_d, (int, float)) and es_d:
            if abs(bs_d / es_d - 1.0) > 1e-4:
                raise ValueError(
                    f"reported vs restated SCALE MISMATCH in FY{y}: BS Total Debt {bs_d} != "
                    f"Econ Statements Total debt {es_d} (ratio {bs_d/es_d:.4g}). restated_* would "
                    f"be mislabeled vs reported_* — aborting rather than shipping wrong units.")
            return
    raise ValueError("cannot confirm restated/reported scale: no year with both debt values")


def write_restated(engine_path, ticker, out_dir):
    import openpyxl
    wb = openpyxl.load_workbook(engine_path, data_only=True)
    ES = wb["Econ Statements"]
    yrs = _year_cols(ES)
    _assert_scale_matches_reported(wb, ES, yrs)      # fail-loud on a units divergence
    os.makedirs(out_dir, exist_ok=True)
    produced = []
    for suffix, (r0, r1) in SECTIONS.items():
        fn = f"{ticker}_{suffix}.csv"
        with open(os.path.join(out_dir, fn), "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["line_item"] + [str(y) for _, y in yrs])
            for r in range(r0, r1 + 1):
                lab = ES.cell(r, 1).value
                if lab is None or not str(lab).strip():
                    continue
                row = [str(lab).strip()]
                for c, _ in yrs:
                    v = ES.cell(r, c).value
                    row.append(v if isinstance(v, (int, float)) else "")
                w.writerow(row)
        produced.append(fn)
    return produced


if __name__ == "__main__":
    import sys
    eng = sys.argv[1]
    tk = sys.argv[2] if len(sys.argv) > 2 else "AAPL"
    print("wrote", write_restated(eng, tk, "outputs"))

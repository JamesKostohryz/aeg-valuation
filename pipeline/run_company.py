#!/usr/bin/env python3
"""run_company.py — the deterministic per-company valuation job. This is what a GitHub
Actions step invokes; it is also runnable locally in --cached mode for testing.

Pipeline (every stage deterministic, fail-loud):
  1. load + validate companies/<TICKER>.yaml            (config.py)
  2. stage raw statements: --cached DIR, or EODHD pull   (eodhd_puller.py)
  3. build the model from config                         (aeg_engine.build_model)
  4. recalc headless (LibreOffice)                       (recalc_lo)
  5. if a rate feed is available: re-point rates + install idio hook, recalc again
  6. run the completeness/provenance/tie GATES           (aeg_engine.read_results)
        -> ANY gate failure exits non-zero == a failed CI check == nothing ships
  7. if bonded + rate feed: Option-A disclosure bridge   (disclose.py)
  8. extract restated anchors / valuation / real series / manifest -> outputs/ CSVs

The engine + restatement stay in the sealed Excel; this job orchestrates and gates them.
"""
import os, sys, argparse, shutil
import csv
import openpyxl

# make the build_v2 modules importable when run from the pipeline/ dir
_HERE = os.path.dirname(os.path.abspath(__file__))
_BUILD_V2 = os.path.dirname(_HERE)
for p in (_HERE, _BUILD_V2):
    if p not in sys.path:
        sys.path.insert(0, p)

import config as CFG
import aeg_engine as AE
import extract as EX

RAW_FILES = {"is_csv": "REAL_IS.csv", "bs_csv": "REAL_BS.csv", "cf_csv": "REAL_CF.csv",
             "prices": "REAL_prices.csv", "dividends": "REAL_div.csv", "splits": "REAL_splits.csv"}


# A refusal must not leave a number behind that the engine will not stand behind.
# Before 2026-08-09 a company that tripped the horizon or convergence gate exited while its
# PREVIOUS <T>_valuation.csv, <T>_summary.csv and <T>_manifest.json sat in the output
# directory looking perfectly current -- Home Depot at $2,173.77 per share against a $360
# price, with nothing in the file to say the engine had declined to publish it. Anything
# reading the outputs directory (the cockpit, a report generator, a person at 2am) got a
# stale number with no warning. The gate protected the run; it did not protect the reader.
#
# We QUARANTINE rather than delete: the files are renamed to <T>_<name>.STALE.<ext>, so the
# previous numbers stay readable for the review the refusal is asking for, while any consumer
# globbing *_summary.csv simply does not find one -- which is the honest state.
_REFUSAL_CTX = {"ticker": None, "out_dir": None}

# Files that carry a headline valuation a reader could quote. Diagnostics the reviewer needs
# (convergence, periods, status, restated statements) are deliberately NOT quarantined.
_VALUATION_BEARING = ("{t}_valuation.csv", "{t}_summary.csv", "{t}_manifest.json",
                      "{t}_fact_sheet.csv")


def _quarantine_stale_outputs(ticker, out_dir, reason):
    """Rename valuation-bearing outputs aside and drop a REFUSED marker. Best-effort:
    a failure here must never mask the refusal that caused it."""
    moved = []
    try:
        for pat in _VALUATION_BEARING:
            fn = pat.format(t=ticker)
            src = os.path.join(out_dir, fn)
            if not os.path.exists(src):
                continue
            stem, ext = os.path.splitext(fn)
            dst = os.path.join(out_dir, f"{stem}.STALE{ext}")
            os.replace(src, dst)
            moved.append((fn, os.path.basename(dst)))
        marker = os.path.join(out_dir, f"{ticker}_REFUSED.csv")
        with open(marker, "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["field", "value"])
            w.writerow(["ticker", ticker])
            w.writerow(["verdict", "REFUSED — no valuation was produced by this run"])
            w.writerow(["reason", " ".join(str(reason).split())])
            w.writerow(["stale_files_quarantined", len(moved)])
            for was, now in moved:
                w.writerow([was, now])
            w.writerow(["note", "The .STALE files hold the PREVIOUS run's numbers. They are "
                                "not current and must not be published or quoted."])
        if moved:
            sys.stderr.write(f"[run_company] quarantined {len(moved)} stale output(s) for "
                             f"{ticker}: {', '.join(w for w, _ in moved)}\n")
        sys.stderr.write(f"[run_company] wrote {ticker}_REFUSED.csv\n")
    except Exception as e:                                  # never mask the real refusal
        sys.stderr.write(f"[run_company] WARNING: could not quarantine stale outputs "
                         f"for {ticker}: {e}\n")
    return moved


def _clear_stale_markers(ticker, out_dir):
    """On a successful run, remove a previous refusal's marker and quarantined files."""
    try:
        for pat in _VALUATION_BEARING:
            stem, ext = os.path.splitext(pat.format(t=ticker))
            f = os.path.join(out_dir, f"{stem}.STALE{ext}")
            if os.path.exists(f):
                os.remove(f)
        m = os.path.join(out_dir, f"{ticker}_REFUSED.csv")
        if os.path.exists(m):
            os.remove(m)
    except Exception as e:
        sys.stderr.write(f"[run_company] WARNING: could not clear stale markers "
                         f"for {ticker}: {e}\n")


def _peek_ticker(config_path):
    """Read just the ticker, tolerating a config that will not validate.

    The horizon gate refuses INSIDE load_config, so the quarantine has to know the ticker
    before the config is known to be good. Returns None if it cannot be read, in which case
    the quarantine simply does not run and behaviour is unchanged.
    """
    try:
        import yaml
        with open(config_path) as fh:
            doc = yaml.safe_load(fh) or {}
        t = doc.get("ticker")
        return str(t).strip() if t else None
    except Exception:
        return None


def _fail(msg, code=1):
    sys.stderr.write(f"\n[run_company] ABORT: {msg}\n")
    t, o = _REFUSAL_CTX["ticker"], _REFUSAL_CTX["out_dir"]
    if t and o and os.path.isdir(o):
        _quarantine_stale_outputs(t, o, msg)
    sys.exit(code)


def stage_raw(cfg, cached_dir, work_dir):
    """Return a files dict for build_model. Cached mode copies the six statement/market
    CSVs from cached_dir; EODHD mode pulls them (needs EODHD_API_KEY)."""
    files = {}
    if cached_dir:
        for key, fname in RAW_FILES.items():
            src = os.path.join(cached_dir, fname)
            if not os.path.exists(src):
                if key in ("is_csv", "bs_csv", "cf_csv"):
                    _fail(f"cached raw missing required {fname} in {cached_dir}")
                files[key] = None
                continue
            dst = os.path.join(work_dir, fname)
            shutil.copy(src, dst)
            files[key] = dst
        return files
    # --- EODHD live pull (production path)
    key = os.environ.get("EODHD_API_KEY")
    if not key:
        _fail("no --cached dir and EODHD_API_KEY not set; cannot stage raw statements")
    try:
        import eodhd_puller as EP
    except Exception as e:
        _fail(f"eodhd_puller import failed: {e}")
    # eodhd_puller writes the six CSVs into work_dir for this ticker (see its API)
    written = EP.pull_to_csvs(cfg["ticker"], key, work_dir)  # noqa: contract w/ puller
    for k in RAW_FILES:
        files[k] = written.get(k)
    if not all(files.get(k) for k in ("is_csv", "bs_csv", "cf_csv")):
        _fail("EODHD pull did not produce all three statement CSVs")
    return files


def resolve_price(cfg, files, cli_price):
    if cli_price is not None:
        return float(cli_price)
    if cfg["price_source"] == "override":
        return cfg["price_override"]
    # market: use the latest close in the staged prices file (production may pull live)
    pf = files.get("prices")
    if pf and os.path.exists(pf):
        import csv
        last = None
        with open(pf, newline="") as fh:
            for row in csv.DictReader(fh):
                c = row.get("Close") or row.get("close")
                if c not in (None, "", "null"):
                    last = c
        if last is not None:
            return float(last)
    _fail("could not resolve a price (no --price, no override, no prices file)")


def build_cost_of_debt(cfg):
    """Map config cost_of_debt into build_model's cod dict. For bond_list we let the
    build use the flagged statement-implied fallback, then the rate re-point overrides
    COD entirely with real_cod — so the initial value is a placeholder that never ships."""
    cod = cfg["cost_of_debt"]
    src = cod["source"]
    if src == "ytw_points":
        return {"ytw_points": cod["ytw_points"]}
    if src == "single_ytw":
        return {"single_ytw": cod["single_ytw"]}
    if src == "bond_list":
        # throwaway seed for the initial build; the rate re-point overrides COD entirely
        return {"single_ytw": cod.get("seed_ytw", 0.05)}
    return {}  # interest_implied -> statement-implied fallback (flagged), may fail if interest≈0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("config", help="path to companies/<TICKER>.yaml")
    ap.add_argument("--template", default=os.path.join(_BUILD_V2, "MODEL_TEMPLATE.xlsx"))
    ap.add_argument("--cached", help="dir with cached raw CSVs (else EODHD pull)")
    ap.add_argument("--out-dir", default=os.path.join(_HERE, "outputs"))
    ap.add_argument("--work-dir", default=os.path.join(_HERE, "_work"))
    ap.add_argument("--rate-feed-dir", help="local dir with rate CSVs (testing)")
    ap.add_argument("--rate-feed-live", action="store_true",
                    help="fetch rate CSVs from the live rate-infra repo (production)")
    ap.add_argument("--price", type=float, help="explicit price (repro/test)")
    ap.add_argument("--vintage", default="unset", help="data vintage tag for the manifest")
    ap.add_argument("--payload", help="RUN-button forecast payload: JSON string or path to .json")
    # The convergence period after cfg_N. Three years is James's specification (2026-08-09) and
    # is the module default; this flag exists for diagnostics only. 0 turns the correction off,
    # which reproduces the pre-convergence (uncorrected, and by his argument wrong) valuation.
    ap.add_argument("--converge-K", type=int, default=3,
                    help="convergence period length in years after cfg_N (default 3; 0 = off)")
    # The vendor total-debt row against Securities and Exchange Commission primary
    # source. REPORT ONLY: it writes <TICKER>_debt_feed.csv and prints a line, and it
    # cannot change in_debt or any valuation number. Off is for air-gapped runs.
    ap.add_argument("--no-debt-feed-check", action="store_true",
                    help="skip BOTH the vendor-vs-primary-source debt report AND the lease "
                         "ruling, so in_debt comes from the vendor row unchanged. For "
                         "air-gapped runs and for reproducing a pre-ruling number.")
    ap.add_argument("--sec-cache-dir", default=None,
                    help="directory to cache Securities and Exchange Commission responses in")
    args = ap.parse_args()

    # Arm the refusal quarantine BEFORE the config gate. The horizon gate refuses inside
    # load_config, which is exactly the case that would otherwise leave a stale valuation
    # sitting in the output directory looking current -- so reading the ticker has to happen
    # first, independently of whether the config validates.
    _REFUSAL_CTX["out_dir"] = args.out_dir
    _REFUSAL_CTX["ticker"] = _peek_ticker(args.config)

    # A config problem is an operator problem, not a bug: report it as a plain-language
    # refusal rather than a Python traceback. The horizon gate lands here most often.
    try:
        cfg = CFG.load_config(args.config)
    except CFG.ConfigError as e:
        _fail(f"CONFIG REJECTED ({args.config})\n{e}")
    tk = cfg["ticker"]
    _REFUSAL_CTX["ticker"] = tk        # authoritative, replaces the peeked value
    os.makedirs(args.work_dir, exist_ok=True)
    print(f"[run_company] {tk}  config_hash={cfg['config_hash']}  bonded={cfg['bonded']}")

    # --- unsupported-knob guard: fail fast rather than emit a number that ignores a stated
    #     judgment. oi_adj_override sets in_oiadj0, which feeds ONLY an Audit identity requiring
    #     it to equal reported OI (Econ Statements anchors on reported OI directly) — so a real
    #     override silently breaks the tie instead of normalizing the anchor. Refuse with the
    #     reason rather than dying mid-tie. (rd_capitalize=true is inert too, but committed
    #     configs set it; that one is a tracked engine backlog item, not a hard refusal here.)
    if cfg["judgments"].get("oi_adj_override") is not None:
        _fail(
            "oi_adj_override is set, but it is NOT a working normalization knob yet: the economic "
            "restatement anchors on REPORTED operating income (Econ Statements S20 <- rep_oi), and "
            "in_oiadj0 feeds only an Audit identity that requires it to equal reported OI — so any "
            "real override breaks the four-method tie rather than repricing the anchor. To value on "
            "a representative base, repoint FY0 to a representative fiscal year instead. Leave "
            "oi_adj_override null until the normalized-anchor increment lands.")
    if cfg["judgments"].get("finlease"):
        _fail(
            f"finlease={cfg['judgments']['finlease']} is set, but the finance-lease add-back is NOT "
            "wired: in_finlease (Inputs B8) feeds no engine formula, and NFO is built as "
            "in_debt - in_cash - in_sti — so a non-zero finlease would silently NOT change NFO or "
            "equity. EODHD 'Total Debt' already includes capitalized finance leases for most names "
            "(so 0.0 is correct there); an extra add-back is an engine-backlog item. Refusing rather "
            "than emitting a valuation that ignores the lease obligation you entered.")

    files = stage_raw(cfg, args.cached, args.work_dir)
    price = resolve_price(cfg, files, args.price)

    build_config = {
        "company": cfg["company"], "ticker": tk, "price": price, "files": files,
        "fy_end_month": cfg["fy_end_month"], "judgments": cfg["judgments"],
        "forecast_horizon_N": cfg["forecast_horizon_N"],
        "cost_of_debt": build_cost_of_debt(cfg),
    }
    out_xlsx = os.path.join(args.work_dir, f"{tk}_engine.xlsx")
    try:
        rep = AE.build_model(build_config, args.template, out_xlsx,
                             resolve_debt_basis=not args.no_debt_feed_check,
                             sec_cache_dir=args.sec_cache_dir)
    except Exception as e:
        _fail(f"build_model failed (statement adjustment): {e}")
    print(f"[build] anchor {rep['anchor_year']}  COD {rep['cost_of_debt']['source']}"
          f"{'  [FLAGGED fallback]' if rep.get('cod_flagged') else ''}")
    # THE LEASE RULING, disclosed either way. Where the anchor could not be corroborated the
    # engine is still valuing on a vendor row whose lease definition changes mid-series, and
    # saying so on every run is the whole point of not applying it silently.
    if rep.get("debt_basis"):
        import debt_feed as _DF
        print(_DF.anchor_basis_console_line(rep["debt_basis"]))
    # P1/P2/P3 — the three first-order judgments that used to be template constants.
    print(f"[policy] cfg_N={rep['forecast_horizon_N']}"
          f"{'' if cfg['horizon_reviewed'] else '  [HORIZON NOT YET REVIEWED]'}"
          f"  payout_seed={rep['policy']['payout_seed']}"
          f"{'  [PAYOUT > 1.0 — REVIEW]' if rep['policy'].get('payout_review') else ''}"
          f"  ppe_life={rep['policy']['ppe_life']}y")
    if rep["template_defaults"]:
        print(f"[inputs] {len(rep['template_defaults'])} valuation-relevant input(s) still "
              f"carrying MODEL_TEMPLATE defaults: {', '.join(rep['template_defaults'])}")

    # Keep the real-terms deflator tables (CPI-U + BEA PP&E) covering the anchor fiscal
    # year using monthly FRED CPI-U, and IFERROR-guard the cap-engine capex window. Must
    # run BEFORE recalc so the real-terms lookups resolve. Fail-closed on a missing CPI.
    import deflator_extend as DE
    try:
        _defl = DE.ensure_deflator_covers_anchor(
            out_xlsx, anchor_year=rep["anchor_year"], fy_end_month=cfg["fy_end_month"])
    except DE.DeflatorError as e:
        _fail(f"deflator extension failed (real-terms base cannot cover the anchor): {e}")
    if _defl.get("extended"):
        print(f"[deflator] extended to {_defl['anchor_year']} via monthly CPI "
              f"(cap-capex-guarded {_defl['cap_capex_wrapped']}): {_defl['added']}")
    else:
        print(f"[deflator] anchor {_defl['anchor_year']} already covered "
              f"(cap-capex-guarded {_defl['cap_capex_wrapped']})")

    # Point the reported-FY0 reconciliation anchors (Audit CHECK-4b) at the ACTUAL
    # anchor-year column. The template hardcodes them to the last column (AP), which is
    # only correct when the newest fiscal year fills AP (long histories: AAPL/HD/T). A
    # shorter-history issuer (e.g. POOL, newest year in AK) leaves AP blank, reddening the
    # audit on an otherwise-tying model. No-op for AP-anchored names. Runs before recalc.
    import repoint_fy0 as F0
    try:
        _f0 = F0.repoint_anchor_columns(out_xlsx, anchor_year=rep["anchor_year"])
    except F0.Fy0Error as e:
        _fail(f"FY0 anchor repoint failed (reconciliation cannot target the anchor year): {e}")
    if _f0["moved"]:
        print(f"[fy0] repointed reconciliation anchors to the {_f0['anchor_year']} "
              f"column: {_f0['moved']}")
    else:
        print(f"[fy0] reconciliation anchors already at the {_f0['anchor_year']} column")

    from recalc_lo import recalc
    recalc(out_xlsx)

    # --- optional payload overrides parsed early: 'bonded' (gates the rate re-point) and
    #     'erp_override' (cockpit ERP-method selector: COE = model real_rf + erp_override).
    #     Neither edits the committed company config; absent -> committed behaviour (bit-identical).
    #
    # FAIL-LOUD 2026-08-12 (AEG-D1-MECHANISM-FOUND): this block used to wrap the whole
    # peek in a blanket `except Exception: pass`, so a malformed or out-of-bounds
    # 'erp_override' -- including a stray/leftover Cockpit dropdown value -- was
    # silently dropped rather than rejected, and the value that DID parse was applied
    # with zero validation. apply_payload.validate_overrides() now bounds-checks
    # 'erp_override' and requires a co-supplied 'erp_override_reason' before a run
    # will honor it; any problem aborts the run instead of silently proceeding.
    _erp_override = None
    _scenarios = None
    if args.payload:
        import apply_payload as _APB
        try:
            _peek = _APB.load_payload(args.payload)
        except _APB.PayloadError as e:
            _fail(f"payload is not valid JSON: {e}")
        if isinstance(_peek, dict):
            try:
                _APB.validate_overrides(_peek)
            except _APB.PayloadError as e:
                _fail(str(e))
            if _peek.get("bonded") is not None:
                cfg["bonded"] = bool(_peek["bonded"])
                print(f"[payload] bonded override -> {cfg['bonded']}")
            if _peek.get("erp_override") is not None:
                _erp_override = float(_peek["erp_override"])
                print(f"[payload] erp_override requested = {_erp_override} "
                      f"(reason: {_peek.get('erp_override_reason')!r})")
            if _peek.get("scenarios") is not None:
                _scenarios = _peek["scenarios"]

    # --- optional rate re-point (only if a feed is provided/available)
    disclosure = None
    feed = None
    if args.rate_feed_dir or args.rate_feed_live:
        import rate_feed as RF, repoint_rates as RP   # openpyxl is module-level (line 19)
        wb = openpyxl.load_workbook(out_xlsx, data_only=False)
        cash = wb["Inputs"]["B6"].value or 0.0
        sti = wb["Inputs"]["B7"].value or 0.0
        try:
            feed = RF.load_all(tk, cash=cash, sti=sti, local_dir=args.rate_feed_dir,
                               bonded=cfg["bonded"])  # local_dir=None -> live repo fetch
            # Cost-of-debt provenance + generalized (unbonded) fallback (ERP ladder ratified
            # 2026-07-25): issuer_bonds -> published -> synthetic coverage rating. Unbonded keeps
            # the build-time BOOK NFO; only real_cod is supplied from the rating curve.
            if "real_cod" in feed:                       # issuer bonds present (GREEN)
                feed["cod_provenance"] = {"cod_source": "issuer_bonds", "rating": feed.get("cod_rating"),
                                          "coverage": None, "audit": "GREEN", "as_of": args.vintage,
                                          "flags": []}
            else:                                        # unbonded -> rating-curve fallback (synthetic)
                import cod_fallback as CF
                vwb = openpyxl.load_workbook(out_xlsx, data_only=True)

                def _dn(name):
                    # scalar value of a defined name; if it points at a range, take the last
                    # numeric (latest period); tolerate a missing name (-> None).
                    try:
                        sheet, coord = list(vwb.defined_names[name].destinations)[0]
                        cells = vwb[sheet][coord]
                        if isinstance(cells, tuple):
                            nums = [c.value for r in cells for c in (r if isinstance(r, tuple) else (r,))
                                    if isinstance(c.value, (int, float))]
                            return nums[-1] if nums else None
                        return cells.value
                    except Exception:
                        return None
                fundamentals = dict(ebit=_dn("in_oiadj0"), interest_expense=_dn("in_intexp0"),
                                    total_debt=_dn("in_debt"), assets=_dn("rep_total_assets"))
                curve = CF.load_credit_curve(local_dir=args.rate_feed_dir)
                real_cod, prov = CF.resolve_cod(fundamentals=fundamentals, curve=curve,
                                                real_rf=feed["real_rf_fwd1y"])
                if not prov.get("spread_nonneg", True):
                    _fail(f"cod validation gate: negative spread vs real_rf on the rating curve ({prov['flags']})")
                prov["as_of"] = args.vintage
                feed["real_cod"] = real_cod
                feed["cod_rating"] = prov["rating"]
                feed["cod_provenance"] = prov
                print(f"[cod] unbonded fallback: {prov['cod_source']} rating={prov['rating']} "
                      f"coverage={prov['coverage']} audit={prov['audit']} flags={prov['flags']}")
            RP.repoint(wb, feed)
            if _erp_override is not None:
                RP.apply_erp_override(wb, _erp_override)
                feed["erp_override"] = _erp_override
                feed["erp_override_reason"] = _peek.get("erp_override_reason") if args.payload else None
                print(f"[erp-override] *** WARNING: this run does NOT use {tk}'s own ERP curve. ***")
                print(f"[erp-override] COE = model real_rf + {_erp_override} (company ERP flat; idio zeroed) "
                      f"reason={feed['erp_override_reason']!r}")
            wb.save(out_xlsx)
            recalc(out_xlsx)
            print(f"[rates] re-pointed from feed (nfo_basis={feed['nfo_basis']})")
        except RF.RateFeedError as e:
            print(f"[rates] feed unavailable/invalid ({e}); keeping build-time rates")
            feed = None

    # --- Phase-2.1 MULTI-SCENARIO path (gated behind payload.scenarios). Values + ties
    #     base+bull+bear INDEPENDENTLY off the rate-repointed workbook and writes
    #     outputs/<TICKER>_scenarios.csv (one row per scenario + an expected-value row).
    #     Absent -> the single-scenario path below runs unchanged (bit-identical by
    #     construction). Fail-closed PER scenario: any non-tie aborts the whole dispatch.
    if _scenarios is not None:
        import run_scenarios as RS
        try:
            srep = RS.run_scenarios(out_xlsx, _scenarios, ticker=tk, price=price,
                                    out_dir=args.out_dir, recalc=recalc,
                                    commit_sha=os.environ.get("GITHUB_SHA", ""))
        except RS.ScenariosError as e:
            _fail(f"SCENARIOS FAILED: {e}")
        print(f"[done] {tk}  scenarios={srep['scenarios']} -> {tk}_scenarios.csv")
        return

    # --- optional RUN-button forecast payload (cockpit dispatch contract 20260722-0800).
    # Applied AFTER the rate re-point so nominal growth drivers are deflated with the very
    # inflation series the engine discounts against. Drivers absent from the payload are NOT
    # written, so they keep their existing formula (anchor hold / legacy scenario overlay) and
    # a payload-free run stays bit-identical. The gates below remain authoritative.
    payload_report = None
    if args.payload:
        import apply_payload as AP
        try:
            payload = AP.load_payload(args.payload)
            # FIX 2026-08-10: the cockpit's "Run valuation" button always sends a payload,
            # even when no real forecast has been loaded -- it reads N from Forecast!C7
            # unconditionally, which is not necessarily this company's reviewed horizon.
            # An empty-drivers payload means "value at anchor, no forecast" -- there is no
            # forecast whose length N is describing, so N here can only legitimately be
            # this company's own reviewed cfg_N. Found by the PEP guest forecaster
            # (2026-08-10): a plain "Run valuation" click had capitalized 30 years of
            # abnormal growth against a reviewed, authorized horizon of 12, because
            # Forecast!C7 held 30 and nothing checked it against the config. A payload
            # that DOES carry real drivers is unaffected -- N there genuinely describes
            # the forecast being submitted and is left exactly as sent.
            if not (payload.get("drivers") or {}):
                _sent_n = payload.get("N")
                if _sent_n != cfg["forecast_horizon_N"]:
                    print(f"[horizon] empty-drivers payload sent N={_sent_n!r}; overriding "
                          f"to this company's reviewed cfg_N={cfg['forecast_horizon_N']} "
                          f"(no forecast is being submitted, so N cannot mean anything else)")
                payload["N"] = cfg["forecast_horizon_N"]
            vals = openpyxl.load_workbook(out_xlsx, data_only=True)
            infl = AP.engine_inflation(vals, int(payload.get("N") or 0) or 1)
            wbp = openpyxl.load_workbook(out_xlsx, data_only=False)
            payload_report = AP.apply_payload(wbp, payload, infl)
            wbp.save(out_xlsx)
            recalc(out_xlsx)
        except AP.PayloadError as e:
            _fail(f"PAYLOAD REJECTED: {e}")
        if payload_report["ticker"] != tk:
            _fail(f"payload ticker {payload_report['ticker']!r} != config ticker {tk!r}")
        print(f"[payload] mode={payload_report['mode']} N={payload_report['N']} "
              f"wrote={sorted(payload_report['written'])} "
              f"held_at_anchor={payload_report['held_at_anchor']}")

    # --- GATES (required check): completeness/provenance, THEN the standing tie check
    results = AE.read_results(out_xlsx, price=price)
    results["anchor_year"] = rep.get("anchor_year")
    tie = results.get("max_identity_tie")
    print(f"[gates] ok={results['ok']}  audit={results['audit_status']!r}  tie={tie:.2e}"
          if isinstance(tie, float) else f"[gates] ok={results['ok']}")
    if not results["ok"]:
        _fail(f"GATES FAILED (completeness/provenance): {results.get('gates')}")

    import checks as CK
    tie_ok, tie_detail = CK.tie_check(results)
    results["tie_check"] = tie_detail
    _rel = tie_detail.get("tie_relative")
    print(f"[tie-check] {tie_detail['tie_check']}  "
          f"(audit_ok={tie_detail['audit_ok']} tie_ok={tie_detail['tie_ok']} mode_ok={tie_detail['mode_ok']})"
          + (f"  relative={_rel:.2e} of scale {tie_detail['tie_scale']:.4g}"
             f" (tol {tie_detail['tie_rel_tol']:g})" if isinstance(_rel, float) else ""))
    if not tie_ok:
        _fail("TIE CHECK FAILED: " + "; ".join(tie_detail["reasons"]))

    # --- earnings-base guard: the tie is an identity on the shared forecast, so it holds for
    #     ANY anchor including a loss. This catches the one bright-line invalid anchor the tie
    #     cannot: a non-positive FY0 OPERATING income (a cyclical trough). Placed AFTER the tie
    #     so a refusal here means the anchor itself is invalid, not a data artifact. Remedy is
    #     human (pick a representative anchor year / supply a normalized OI) — D1.
    anchor_ok, anchor_detail = CK.anchor_earnings_check(out_xlsx)
    results["anchor_earnings_check"] = anchor_detail
    print(f"[anchor-earnings] {anchor_detail['anchor_earnings_check']}  "
          f"anchor_oi_at0={anchor_detail['anchor_oi_at0']} (FY{anchor_detail['anchor_year']})")
    if not anchor_ok:
        _fail("ANCHOR EARNINGS CHECK FAILED: " + anchor_detail["reason"])

    # --- anchor representativeness: the loss guard catches OI <= 0; this catches a POSITIVE
    #     but distorted anchor — FY0 margin far below the firm's own history (impairment /
    #     restructuring / one-off embedded in OI, or a cyclical trough). The provider does not
    #     reliably tag such charges, so we detect via representativeness and defer the fix to a
    #     human (repoint / normalized anchor), rather than strip an amount we cannot trust.
    rep_ok, rep_detail = CK.anchor_representativeness_check(out_xlsx)
    results["anchor_representativeness_check"] = rep_detail
    _rv = rep_detail.get("anchor_representativeness_check")
    _am, _mn = rep_detail.get("anchor_margin"), rep_detail.get("median_prior_margin")
    print(f"[anchor-repr] {_rv}"
          + (f"  FY0 margin={_am:.1%} vs normal {_mn:.1%}" if isinstance(_am, float) and isinstance(_mn, float)
             else f"  ({rep_detail.get('reason','')[:64]})"))
    if not rep_ok:
        _fail("ANCHOR REPRESENTATIVENESS CHECK FAILED: " + rep_detail["reason"])

    # --- R&D / opex-wedge diagnostic (Forecast row 61). Visible, non-fatal — EXCEPT a
    #     firm that declares no wedge (expect_zero_rd_wedge) must actually have ~0.
    wedge = CK.rd_wedge_report(out_xlsx)
    results["rd_wedge"] = wedge
    wpct = wedge.get("wedge_pct_ebit")
    print(f"[rd-wedge] opex wedge {wedge['opex_wedge']}  "
          f"({'%.1f%% of EBIT' % (100*wpct) if wpct is not None else 'n/a'}); "
          f"rev-scaled-consistent={wedge['rev_scaled_consistent']}; "
          f"rd_capitalization_wired={wedge['rd_capitalization_wired']}")
    if wedge["rd_reserve_nonzero_but_inert"]:
        print("[rd-wedge] NOTE: R&D reserve is nonzero but INERT (capitalization not yet "
              "wired into NOA/OI) — see docs; R&D-heavy names not yet capitalized.")
    if not wedge["rev_scaled_consistent"]:
        _fail("row-61 wedge is no longer revenue-proportional (engine structure changed unexpectedly)")
    if cfg.get("expect_zero_rd_wedge") and wpct is not None and wpct > 0.005:
        _fail(f"expect_zero_rd_wedge set but Forecast row 61 wedge is {100*wpct:.2f}% of EBIT "
              f"(expected ~0 for a no-R&D / no-opex-wedge name)")

    # --- Option-A disclosure (needs the live feed + bonded issuer)
    if feed is not None and cfg["bonded"] and "company" in feed:
        try:
            import disclose as D
            disclosure = D.disclose(out_xlsx, feed, price=price, recalc=recalc,
                                    sens_path=os.path.join(args.work_dir, f"{tk}_idiosens.xlsx"))
            print(D.format_bridge(disclosure))
        except Exception as e:
            print(f"[disclose] skipped ({e})")

    # --- disclosed inflation scorecard (Increment 2; tie-safe — never enters the four-method
    #     tie). Interest-tax-shield PVs under both debt policies (Miller-excluded) + the
    #     capital-intensity-vs-leverage verdict. Computed in engine units from the recalced
    #     workbook + the rate feed; additive, no engine edit.
    if feed is not None:
        try:
            import scorecard as SCD
            _sc = SCD.compute_scorecard(out_xlsx, feed)
            SCD.write_scorecard_csv(os.path.join(args.out_dir, f"{tk}_inflation_scorecard.csv"), _sc)
            results["inflation_verdict"] = _sc.get("verdict")
            _np = _sc.get("net_inflation_position_annual")
            print(f"[scorecard] {tk}: {_sc.get('verdict')}"
                  + (f"  (net {_np:.4g}/yr = benefit {_sc['interest_benefit_annual']:.4g}"
                     f" - penalty {_sc['depreciation_penalty_annual']:.4g})" if isinstance(_np, (int, float)) else ""))
        except Exception as e:
            print(f"[scorecard] skipped ({e})")

    # --- CONVERGENCE PERIOD (James's specification 2026-08-09; see
    #     claude/AEG-CONVERGENCE-PERIOD-REQUIRED-2026-08-09.md).
    #
    #     Beyond the explicit horizon cfg_N the engine hard-gates every AEG contribution to zero,
    #     which capitalizes whatever earnings level the forecast happened to END on. The
    #     continuing period must instead BEGIN at a normalized (neutral) level with abnormal
    #     earnings growth already spent.
    #
    #     CHANGED 2026-08-12, on James's ruling. This used to glide EPS onto the normalized line
    #     and ADD the booked reversion to the value. It no longer adjusts anything. Deciding
    #     whether a forecast stops at a cyclical peak is the forecaster's job, and the horizon
    #     rule already implies it: the explicit forecast does not end until projected abnormal
    #     earnings growth is spent, and a reversion from a peak necessarily creates abnormal
    #     growth. What remains is two GATES on the truncation point -- abnormal growth spent, and
    #     earnings at a normalized level -- which REFUSE and hand the forecast back rather than
    #     silently patching a number nobody owns. See docs/AEG-CONVERGENCE-RETIRED-2026-08-12.md.
    #
    #     A consequence worth stating: the published value is now the engine value, so it is
    #     entirely INSIDE the four-method tie. The old increment was the one published component
    #     the tie could not see. That hole is closed by deletion.
    #
    #     FAIL-LOUD: a failure here must not degrade quietly into publishing an ungated number.
    import convergence as CV
    try:
        _conv = CV.converge_auto(out_xlsx, K=args.converge_K)
        _periods = CV.period_report(out_xlsx, _conv)
        CV.write_convergence_csv(_conv, tk, args.out_dir)
        _periods_fn = CV.write_periods_csv(_periods, tk, args.out_dir)
    except Exception as e:
        _fail("TRUNCATION GATES FAILED TO RUN: the stop year could not be judged against the "
              f"terminal and neutral-level conditions, so no valuation can be published ({e})")

    #     The idiosyncratic haircut is measured as a SENSITIVITY re-run of the whole engine at a
    #     higher cost of equity (disclose.py). Until 2026-08-12 the convergence increment had to
    #     be re-priced on that sensitivity workbook as well, or it would have escaped the haircut
    #     and overstated the headline -- a second full engine recalculation for one number. With
    #     the increment retired that number is identically zero, so the re-run is gone with it.
    #     The headline is the disclosure bridge's adjusted equity, or the engine value when there
    #     is no bridge. Nothing is added to either.
    _adj = (disclosure or {}).get("adjusted_equity_ps")
    if isinstance(_adj, (int, float)):
        _headline = _adj
        _headline_basis = "adjusted equity (depreciation + market debt + idiosyncratic haircut)"
    else:
        _headline = _conv["eng_intrinsic"]
        _headline_basis = "engine equity value (no disclosure bridge)"
    _pre = _headline
    _conv_sens_ps = None

    convergence = {
        "cfg_N": _conv["N"], "K": _conv["K"],
        "actual_eps_N": _periods["actual_eps_N"],
        "normalized_eps_N": _conv.get("norm_eps_N"),
        "gap_ps": _conv["converge_gap_ps"],
        "engine_intrinsic_ps": _conv["eng_intrinsic"],
        "corrected_intrinsic_ps": _conv["corrected_intrinsic"],
        "convergence_value_ps": _conv["converge_value_ps"],
        "convergence_value_idio_adjusted_ps": _conv_sens_ps,
        "headline_value_ps": _headline,
        "headline_value_pre_convergence_ps": _pre,
        "headline_basis": _headline_basis,
        "guard": _conv["verdict"], "guard_reason": _conv["verdict_reason"],
        "in_four_method_tie": False,
        "caveat": ("the convergence increment is computed on the EQUITY (EPS) leg only and "
                   "therefore sits OUTSIDE the four-method tie; the tie covers the explicit "
                   "period through cfg_N"),
        "periods": _periods["blocks"],
        "identity_checks": _periods["identity_checks"],
        "schedule": _conv["schedule"],
        "outputs": [f"{tk}_convergence.csv", _periods_fn],
    }
    convergence["reviewed"] = bool(cfg.get("convergence_reviewed"))
    convergence["review_note"] = cfg.get("convergence_note") or ""
    results["convergence"] = convergence
    print(f"[convergence] K={_conv['K']} after cfg_N={_conv['N']}: actual EPS "
          f"{_periods['actual_eps_N']:.4f} -> normalized {_conv['norm_eps_N']:.4f} "
          f"(gap {_conv['converge_gap_ps']:+.4f}/sh)")
    _t = _conv.get("terminal") or {}
    print(f"[truncation] gate A, abnormal growth spent: AEG at cfg_N "
          f"{_t.get('aeg_N', float('nan')):+.4f}/sh, year-on-year factor "
          f"{_t.get('decay', float('nan')):.3f}, discarded tail "
          + ("DIVERGES (still growing)" if _t.get("tail_frac") is None
             else f"{_t['tail_frac']:.2%} of value"))
    print(f"[truncation] value {_conv['eng_intrinsic']:.4f}/sh, unadjusted  |  HEADLINE "
          f"{_headline:.4f}/sh  ({_headline_basis})")
    for _b in _periods["blocks"]:
        print(f"[periods] {_b['period']:<12} yrs {_b['years']:<8} "
              f"PV {_b['pv_contribution_ps']:+9.4f}/sh  "
              f"({(_b['pct_of_corrected_value'] or 0):+.1%} of value)")
    print(f"[convergence] guard {_conv['verdict']}: {_conv['verdict_reason']}"
          + ("  [REVIEWED by analyst]" if convergence["reviewed"] else ""))
    print("[truncation] NOTE: the convergence increment was retired 2026-08-12. The published "
          "value is the engine value and is wholly inside the four-method tie.")

    # --- UNFUNDED DISTRIBUTION GATE (James, 2026-08-11).
    #
    #     Under the canonical operating closure distributions are IMPLIED -- Forecast row 29 is a
    #     residual. A residual can come out negative, and a negative implied dividend means the
    #     plan is asserting that the company issues equity in order to fund a buyback it cannot
    #     afford, while simultaneously retiring shares. Nobody would sign that forecast.
    #
    #     Nothing else on this system can see it. On the golden Apple fixture under the default
    #     Consensus overlay the implied dividend is negative in every forecast year while the
    #     four-method tie reads 8.4e-16, the audit reads PASS and the convergence guard reads
    #     PASS. Same failure class as the horizon bug and the leverage bug: silently wrong with
    #     every gate green. So it refuses, and only a person clears it.
    import funding_check as FCK
    _fund = FCK.funding_report(out_xlsx)
    funding = {"verdict": _fund["verdict"], "reason": _fund["reason"], "years": _fund["years"],
               "reviewed": bool(cfg.get("funding_reviewed")),
               "review_note": cfg.get("funding_note") or ""}
    results["funding"] = funding
    print(FCK.format_report(_fund))
    if funding["reviewed"] and _fund["verdict"] == "REVIEW":
        print("[funding] REVIEWED by analyst -- gate cleared by explicit assertion in the config")

    if _fund["verdict"] == "REVIEW" and not funding["reviewed"]:
        _w = _fund["worst"]
        _fail(
            "UNFUNDED DISTRIBUTION -- no valuation produced for " + tk + ".\n"
            f"  {_fund['reason']}.\n"
            f"  Worst year {_w['year']}: the operating plan and financing structure permit "
            f"distributions of {_w['distribution_capacity']:.6f} (net income less the increase in "
            f"common equity), but the plan calls for repurchases of {_w['repurchases']:.6f} -- a "
            f"shortfall of {_w['funding_shortfall']:+.6f}, leaving an implied dividend of "
            f"{_w['implied_dps']:+.4f} per share.\n"
            "  A negative implied dividend is a capital RAISE. The forecast is buying back stock "
            "with money it has not got.\n"
            "  This is the two-of-three rule biting: you have set the operating plan AND the "
            "distribution policy AND, through the target leverage, the financing structure. The "
            "balance sheet has to balance, so something had to give, and it was the dividend.\n"
            "  There are three things that cause it, and the review is deciding which:\n"
            "    1. The buyback rate is too high for the operating plan. This is the default "
            "Consensus overlay's case -- three percent of shares against 2.5 percent asset growth "
            "is not fundable for a company at this price-to-book. Remedy: set a buyback rate the "
            "plan can pay for.\n"
            "    2. Net-operating-asset growth is too low, so the business throws off more cash "
            "than the plan reinvests and the buyback overshoots what is left. Remedy: revisit the "
            "operating plan.\n"
            "    3. The target leverage is doing the work. Remedy: let financing absorb, which is "
            "what the canonical closure is for.\n"
            "  Once you have looked and either fixed it or accepted it, add to "
            f"companies/{tk}.yaml:\n"
            "      funding:\n"
            "        reviewed: true\n"
            "        note: <which of the three it was, and why the capital raise is intended>\n"
            "  Nothing else clears this gate. That is intended.")

    # --- CONVERGENCE REVIEW GATE (James, 2026-08-09).
    #
    #     A REVIEW verdict means the explicit forecast ends at an earnings level far from its own
    #     neutral line. The continuing period is then started from a level nobody has vouched for,
    #     and the four-method tie cannot see the problem: Home Depot ties at 1.2e-12 relative
    #     while valuing the equity at roughly seventeen times its market price. So REVIEW refuses
    #     to produce a valuation.
    #
    #     It is cleared by a human assertion in the company config, never by a code change and
    #     never by moving the thresholds — exactly like forecast.reviewed. The thresholds
    #     themselves (convergence.GAP_FRAC_WARN / VALUE_FRAC_WARN) are PROVISIONAL and unstudied,
    #     which is precisely why the escape hatch belongs to the analyst rather than to whoever
    #     is editing the code that day.
    if _conv["verdict"] == "REVIEW" and not convergence["reviewed"]:
        _fail(
            "TRUNCATION REVIEW REQUIRED — no valuation produced for " + tk + ".\n"
            f"  The explicit forecast ends at earnings per share of {_periods['actual_eps_N']:.4f} "
            f"in year cfg_N={_conv['N']}, against a normalized (neutral) level of "
            f"{_conv['norm_eps_N']:.4f} — a gap of {_conv['converge_gap_ps']:+.4f} per share, "
            f"with abnormal earnings growth of {(_conv.get('terminal') or {}).get('aeg_N', float('nan')):+.4f} "
            "per share still being created in that final year.\n"
            f"  Gate: {_conv['verdict_reason']}\n"
            f"  Engine value {convergence['headline_value_ps']:.2f} per share — NOT published.\n"
            "  The rule is that the explicit forecast does not end until projected abnormal "
            "earnings growth is spent AND earnings are at a normalized level. Both must hold. "
            "Nothing is corrected for you: a truncation that fails this is a forecast that has "
            "not finished, and the remedy is to extend it. There are four things that cause it, "
            "and the review is deciding which:\n"
            "    1. The forecast horizon is in the wrong place — it stops while abnormal earnings "
            "growth is still running. Remedy: move forecast.horizon_N.\n"
            "    2. The forecast drivers produce an implausible earnings path. Remedy: fix the "
            "inputs. This is the Home Depot case.\n"
            "    3. The anchor year is distorted (the anchor checks above cover most of this).\n"
            "    4. The normalized line itself misfired — it is the median of the last four "
            "forecast years walked forward at normal growth, and an unusual pattern can defeat "
            "it.\n"
            "  Once you have looked and either fixed it or accepted the gap, add to "
            f"companies/{tk}.yaml:\n"
            "      convergence:\n"
            "        reviewed: true\n"
            "        note: <which of the four it was, and what you concluded>\n"
            f"  {tk}_convergence.csv and {tk}_periods.csv WERE written to the output directory, "
            "so you can see the earnings path the review is about. No valuation file was "
            "refreshed — the previous run's numbers stand until this is resolved.\n"
            "  Nothing else clears this gate. That is intended.")

    # --- extract committed outputs + manifest
    manifest = EX.extract_outputs(out_xlsx, tk, args.out_dir, results=results,
                                  config_hash=cfg["config_hash"], vintage=args.vintage,
                                  disclosure=disclosure, convergence=convergence)
    print(f"[extract] wrote {', '.join(manifest['outputs'])} + {tk}_manifest.json to {args.out_dir}")

    # --- run-status / anchor-health CSV (cockpit feed). The guard verdicts are otherwise
    #     console-only, so the Sheet cannot show "anchor representative / run passed / by how
    #     much". Additive and tie-safe. Written only on the SUCCESS path (a failed gate aborts
    #     before here), which is exactly the "last good run" the cockpit IMPORTDATA tracks — so
    #     a rejected config leaves the prior status standing, and this file's vintage/config_hash
    #     lets the Audit tab detect that the sheet is showing an older run.
    import csv as _csv
    _ae = results.get("anchor_earnings_check") or {}
    _re = results.get("anchor_representativeness_check") or {}
    _status = [
        ("ticker", tk),
        ("run_status", "OK"),
        ("anchor_year", results.get("anchor_year")),
        ("tie_check", (results.get("tie_check") or {}).get("tie_check")),
        ("max_identity_tie", results.get("max_identity_tie")),
        ("anchor_earnings_check", _ae.get("anchor_earnings_check")),
        ("anchor_oi_at0", _ae.get("anchor_oi_at0")),
        ("anchor_representativeness", _re.get("anchor_representativeness_check")),
        ("anchor_margin", _re.get("anchor_margin")),
        ("anchor_normal_margin", _re.get("median_prior_margin")),
        ("anchor_margin_vs_normal", _re.get("margin_vs_normal")),
        ("inflation_verdict", results.get("inflation_verdict")),
        ("rd_capitalization_wired", False),
        # Truncation gates. The convergence increment was RETIRED on 2026-08-12, so
        # convergence_value_ps is identically zero and headline_value_pre_convergence_ps equals
        # the headline; both are kept only so existing readers do not break. Do not present them
        # as live inputs. The published value is the engine value, wholly inside the tie.
        ("convergence_adjustment", "RETIRED_2026-08-12 (inert)"),
        ("headline_value_ps", convergence["headline_value_ps"]),
        ("headline_value_pre_convergence_ps", convergence["headline_value_pre_convergence_ps"]),
        ("convergence_K", convergence["K"]),
        ("convergence_value_ps", convergence["convergence_value_ps"]),
        ("convergence_value_idio_adjusted_ps", convergence["convergence_value_idio_adjusted_ps"]),
        ("actual_eps_at_N", convergence["actual_eps_N"]),
        ("normalized_eps_at_N", convergence["normalized_eps_N"]),
        ("convergence_guard", convergence["guard"]),
        ("funding_guard", funding["verdict"]),
        ("funding_reviewed", funding["reviewed"]),
        ("funding_review_note", (funding["review_note"] or "").replace(",", ";")),
        ("funding_min_implied_dps", (min((y["implied_dps"] for y in funding["years"]),
                                         default=None))),
        ("convergence_reviewed", convergence["reviewed"]),
        ("convergence_review_note", (convergence["review_note"] or "").replace(",", ";")),
        ("convergence_in_four_method_tie", convergence["in_four_method_tie"]),
        # P1/P2/P3 — the first-order judgments, surfaced so the cockpit can show them.
        ("cfg_N", rep["forecast_horizon_N"]),
        ("cfg_N_reviewed", cfg["horizon_reviewed"]),
        ("payout_seed", rep["policy"]["payout_seed"]),
        ("ppe_life_years", rep["policy"]["ppe_life"]),
        ("inputs_on_template_default", len(rep["template_defaults"])),
        ("config_hash", cfg["config_hash"]),
        ("vintage", args.vintage),
    ]
    with open(os.path.join(args.out_dir, f"{tk}_status.csv"), "w", newline="") as _fh:
        _w = _csv.writer(_fh)
        _w.writerow(["field", "value"])
        for _k, _v in _status:
            _w.writerow([_k, _v])
    print(f"[status] wrote {tk}_status.csv (anchor-health + run verdicts)")

    # --- debt-feed guard (REPORT ONLY, 2026-08-09).
    #     The vendor "Total Debt" row changes definition partway through its own series
    #     for at least one committed name: it agrees with primary source to the dollar
    #     for Apple through fiscal 2021, then absorbs capitalized leases. That row is
    #     `in_debt`, which sets net financial obligations and — because net operating
    #     assets is plugged from common equity plus net financial obligations — reprices
    #     the whole forecast. Perturbation on 2026-08-09 moved Apple's tied equity from
    #     87.1659 to 89.8409 per share with the four-method tie green at 1.3e-14 BOTH
    #     times, so the tie cannot catch this class of error and something else has to.
    #
    #     This block does NOT change what the engine consumes. `in_debt` still comes off
    #     the vendor row exactly as before, so every valuation number is bit-identical
    #     with and without it. Which lease treatment the engine should VALUE on is a
    #     gated judgment that has not been made. All this does is refuse to let the
    #     disagreement stay invisible. It never aborts a run: a network failure, an
    #     unknown filer or an uncorroborated construction all report UNVERIFIED.
    if not args.no_debt_feed_check:
        try:
            import debt_feed as DFEED
            _bs = os.path.join(args.out_dir, f"{tk}_reported_bs.csv")
            if os.path.exists(_bs):
                _rep = DFEED.audit_debt_feed(
                    tk, _bs, anchor_year=results.get("anchor_year"),
                    cache_dir=args.sec_cache_dir)
                DFEED.write_report(_rep, os.path.join(args.out_dir, f"{tk}_debt_feed.csv"))
                print(DFEED.console_line(_rep))
                for _d in _rep["disagreements"]:
                    _lz = _d["lease_liabilities_musd"]
                    print(f"[debt-feed]   FY{_d['fiscal_year']}  vendor {_d['vendor_musd']:,.0f}"
                          f"  primary source {_d['primary_source_musd']:,.0f}"
                          f"  gap {_d['gap_musd']:+,.0f}m"
                          + (f"  = {_d['gap_equals']}" if _d["explained_by_leases"]
                             else "  UNEXPLAINED"))
                if _rep["verdict"] in ("AMBER", "RED"):
                    print("[debt-feed] NOTE: the engine is still valuing on the VENDOR row. "
                          "Changing that is gated and has not been authorized.")
            else:
                print("[debt-feed] skipped (no reported balance sheet written)")
        except Exception as _e:
            # Report-only work must never take down a valuation run.
            print(f"[debt-feed] skipped ({type(_e).__name__}: {_e})")

    # --- S2: input provenance register. One row per valuation-relevant Inputs cell with
    #     its value and whether that value came from this company's filings, was set
    #     deliberately by the analyst, or is still whatever MODEL_TEMPLATE.xlsx shipped
    #     with. The third category is the one that matters: it is how the template base
    #     company's dividend payout and plant life reached every valuation in the system.
    _reg_path = os.path.join(args.out_dir, f"{tk}_inputs_register.csv")
    with open(_reg_path, "w", newline="") as _fh:
        _w = _csv.DictWriter(_fh, fieldnames=["cell", "name", "class", "label", "value",
                                              "provenance", "source"])
        _w.writeheader()
        for _row in rep["inputs_register"]:
            _w.writerow({k: _row[k] for k in _w.fieldnames})
    print(f"[inputs] wrote {tk}_inputs_register.csv "
          f"({sum(1 for r in rep['inputs_register'] if r['provenance'] == 'filings')} from filings, "
          f"{sum(1 for r in rep['inputs_register'] if r['provenance'] == 'analyst')} analyst-set, "
          f"{len(rep['template_defaults'])} template default)")

    # --- S4: per-identity tie residual breakdown into the manifest. The standing gate
    #     reports one aggregate number (Audit!B5). When a live run ties at 1e-9 while the
    #     golden fixture ties at 1e-15, that aggregate cannot say WHICH identity carries
    #     the residual. This writes the component-level breakdown on every run, so the
    #     question is answered by the next refresh instead of by an investigation.
    try:
        import openpyxl as _oxl
        _vals = _oxl.load_workbook(out_xlsx, data_only=True)["Audit"]
        _IDENT = [
            ("bs_noa_cse_nfo", "B22"), ("bs_flev_cse_nfo", "B23"), ("bs_bps_shares", "B24"),
            ("bs_nfops_shares", "B25"), ("equity_enterprise_tie", "B27"),
            ("value_additivity", "B28"), ("buyback_vp6_invariance", "B29"),
            ("sum_balance_ties", "B31"), ("sum_anchor_reconciliation", "B44"),
            ("reform_partition_nominal", "B47"), ("reform_partition_real", "B48"),
            ("cap_engine_anchor", "B49"), ("sum_reformulation_ties", "B50"),
            ("forecast_bs", "B55"), ("forecast_eps_shares", "B56"), ("forecast_fcfe", "B57"),
            ("sum_forecast_ties", "B58"), ("dcf_fcfe_vs_ri", "B61"),
            ("dcf_fcff_vs_ri", "B62"), ("sum_dcf_recon", "B63"),
            ("crosstab_aeg_vs_ri", "B72"), ("ops_direct_vs_additive", "B75"),
            ("master_total", "B5"),
        ]
        _breakdown, _worst, _worstv = {}, None, -1.0
        for _nm, _cell in _IDENT:
            _v = _vals[_cell].value
            _v = float(_v) if isinstance(_v, (int, float)) else None
            _breakdown[_nm] = {"cell": f"Audit!{_cell}", "residual": _v}
            if _nm.startswith("sum_") or _nm == "master_total":
                continue
            if _v is not None and abs(_v) > _worstv:
                _worst, _worstv = _nm, abs(_v)
        import json as _json
        _mp = os.path.join(args.out_dir, f"{tk}_manifest.json")
        with open(_mp) as _fh:
            _mj = _json.load(_fh)
        _mj["tie_residuals"] = {"components": _breakdown, "worst_identity": _worst,
                                "worst_abs_residual": (None if _worst is None else _worstv)}
        with open(_mp, "w") as _fh:
            _json.dump(_mj, _fh, indent=2)
        print(f"[ties] residual breakdown -> manifest (worst identity: {_worst} "
              f"at {_worstv:.2e})")
    except Exception as _e:
        print(f"[ties] residual breakdown skipped ({_e})")
    # cost-of-debt provenance -> manifest (audit tab: cod_source / rating / coverage / audit / as_of)
    if feed is not None and feed.get("cod_provenance"):
        import json as _json
        _mp = os.path.join(args.out_dir, f"{tk}_manifest.json")
        try:
            with open(_mp) as _fh:
                _mj = _json.load(_fh)
            _mj["cost_of_debt"] = feed["cod_provenance"]
            if feed.get("erp_override") is not None:
                _mj["erp_override"] = {"value": feed["erp_override"], "method": "override",
                                       "reason": feed.get("erp_override_reason"),
                                       "note": "COE = model real_rf + erp_override (idio zeroed); "
                                               "this run did NOT use the company's own ERP curve"}
            with open(_mp, "w") as _fh:
                _json.dump(_mj, _fh, indent=2)
            print(f"[cod] provenance -> manifest: {feed['cod_provenance']['cod_source']} "
                  f"{feed['cod_provenance']['audit']} rating={feed['cod_provenance']['rating']}")
        except Exception as _e:
            print(f"[cod] manifest provenance skip ({_e})")
    # A run that reaches the headline produced a valuation, so any quarantine left by a
    # previous refusal is obsolete: clear the marker and the .STALE files.
    _clear_stale_markers(tk, args.out_dir)
    print(f"[done] {tk}  HEADLINE (convergence-corrected) "
          f"{convergence['headline_value_ps']:.4f}/sh  |  pre-convergence "
          f"{convergence['headline_value_pre_convergence_ps']:.4f}/sh  |  "
          f"tied engine equity={results.get('equity_value')}  tie={tie}")


if __name__ == "__main__":
    main()

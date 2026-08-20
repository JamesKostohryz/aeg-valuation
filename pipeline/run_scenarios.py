#!/usr/bin/env python3
"""run_scenarios.py — Phase-2.1 multi-scenario ("Run scenarios") engine path.

Gated behind payload.scenarios in run_company.py, so a normal single-scenario RUN /
RUN-unbonded dispatch never enters here and stays bit-identical.

The cockpit dispatches ONE payload carrying a `scenarios` array (base+bull+bear), each
entry FULLY self-contained (name, probability, mode, N, singles, drivers, and an OPTIONAL
erp_override). The cockpit already resolved every "blank = inherit base" rule before
dispatch, so the engine does NO inheritance — it values and TIES each scenario
independently, fail-closed per scenario, and publishes outputs/<TICKER>_scenarios.csv.

COE convention (COCKPIT 2026-07-30 16:07 correction): base and bear both run the
standard variance-v2 COE; ONLY a scenario that explicitly carries an erp_override gets a
flat COE override (bull's CAPM-when-lower, or a manual Control-tab override). We simply
APPLY whatever erp_override a scenario carries — no CAPM computation, no bear penalty.

Mechanics per scenario: start from a FRESH copy of the rate-repointed formulas workbook
(base_xlsx, produced by run_company after build -> deflator -> fy0 -> rate re-point), then
  apply_payload(scenario_as_single_payload)   # writes drivers/singles/mode/N
  if erp_override: repoint_rates.apply_erp_override(wb, erp)
  save -> recalc -> read_results -> gates + tie_check (fail-closed for THIS scenario)
and read intrinsic / real price / real COE / tie residual for the CSV row.

TRUNCATION/FUNDING/TERMINAL GATES (2026-08-14). Until this date, the ONLY scenario in a
`payload.scenarios` dispatch to get Gate A (terminal condition), Gate B (neutral level),
the unfunded-distribution guard, and the terminal-payout guard was the primary scenario —
because it runs through run_company.py's ordinary single-scenario path before this module
is ever imported. bull and bear went through _value_one() below, which checked only
completeness/provenance (AE.read_results) and the four-method tie (CK.tie_check). A
truncation that discarded real value, a distribution plan that silently issued equity to
fund a buyback it could not afford, or a continuing period with no forecaster-owned payout
policy would all publish for a non-primary scenario with every visible check green. Found
and fixed the same session as the Coca-Cola Round 3/4 guest-forecaster work: see
docs/KO-Round3-Bull-Bear-2026-08-14.md section 5 for the discovery and
docs/FORECASTER-KIT-v5-2026-08-13.md for the corrected description of this path. Every
scenario, primary or not, now goes through the identical four checks run_company.py's
primary path applies: read_results' completeness/provenance gates, the four-method tie,
convergence.converge_auto (Gate A + Gate B together), funding_check.funding_report, and
terminal_payout.terminal_payout_report — fail-closed per scenario, using the SAME
company-level review escape hatches (companies/<T>.yaml convergence.reviewed /
funding.reviewed / terminal.reviewed) the primary path already uses. There is no
per-scenario review flag: a reviewed:true in the company config clears that gate for every
scenario in the dispatch, exactly as it already did for whichever scenario happened to run
as primary. test_run_scenarios.py pins that a scenario whose own truncation, funding or
terminal-payout condition would refuse under run_company.py now also refuses here (the
"every scenario now gets the truncation/funding/terminal gates" block), and that a properly
reviewed scenario set still values and ties exactly as before (the gates are already proven
inert to the tie by test_horizon_gating.py and test_terminal_payout.py).
"""
import os
import shutil
import datetime as _dt

import openpyxl

import apply_payload as AP
import repoint_rates as RP
import aeg_engine as AE
import checks as CK
import convergence as CV
import funding_check as FCK
import terminal_payout as TP


# CSV column order (COCKPIT scenarios contract). One row per scenario + a summary row.
CSV_HEADER = [
    "ticker", "run_timestamp_utc", "commit_sha", "scenario", "probability",
    "coe_basis", "real_coe", "intrinsic_value_per_share_real",
    "current_real_price_per_share", "upside_downside_pct", "tie_residual",
]
PROB_TOL = 0.001            # cockpit-enforced; we only warn if it drifts
DEFAULT_CONVERGE_K = 3      # matches run_company.py's --converge-K default


class ScenariosError(Exception):
    """A scenario failed its gates/tie, or the payload's scenario array is malformed.
    Raised so run_company aborts non-zero (fail-closed) without writing a green CSV."""


def _named_scalar(wb_values, name):
    """Value of a single-cell defined name from a data_only workbook (or None)."""
    dn = wb_values.defined_names.get(name)
    if not dn:
        return None
    ref = str(dn.value).replace("$", "").replace("'", "")
    if "!" not in ref:
        return None
    sh, cell = ref.split("!")
    try:
        return wb_values[sh][cell].value
    except Exception:
        return None


def _as_single_payload(ticker, sc):
    """Reshape one scenario entry into the single-scenario payload apply_payload expects."""
    return {
        "ticker": ticker,
        "mode": sc.get("mode"),
        "N": sc.get("N"),
        "drivers": sc.get("drivers") or {},
        "singles": sc.get("singles") or {},
    }


def _truncation_funding_terminal_gates(work, cfg, converge_K):
    """The three checks run_company.py's primary path applies AFTER the tie, mirrored here so
    a non-primary scenario is held to the identical standard. Returns (conv, fund, term,
    reasons) — conv/fund/term are the raw reports (None if they could not be evaluated), and
    reasons is a list of human-readable failure strings, empty if everything passed or was
    cleared by the company config's reviewed:true escape hatch.

    cfg is the SAME company-level config dict run_company.py loads from companies/<T>.yaml.
    There is no per-scenario review flag: a reviewed:true in the config clears the gate for
    EVERY scenario in this dispatch, exactly as it already did for whichever scenario the
    primary happened to be.
    """
    cfg = cfg or {}
    reasons = []

    try:
        conv = CV.converge_auto(work, K=converge_K)
    except Exception as e:
        return None, None, None, [
            f"TRUNCATION GATES FAILED TO RUN: the stop year could not be judged against the "
            f"terminal and neutral-level conditions ({e})"]

    # A gate that fires must say whether its escape hatch was even available. On 2026-08-20
    # this fired on PepsiCo's bear case with convergence.reviewed: true sitting in the config,
    # and the message gave no way to tell whether the flag had not been read or had been read
    # and was False -- so the next hour went into reading code instead of the one fact that
    # settles it. The flags are now in the reason.
    _hatches = {k: bool(cfg.get(k)) for k in
                ("convergence_reviewed", "funding_reviewed", "terminal_reviewed")}
    if conv["verdict"] == "REVIEW" and not cfg.get("convergence_reviewed"):
        reasons.append(
            f"TRUNCATION REVIEW REQUIRED: {conv['verdict_reason']} (clear with "
            "convergence.reviewed: true in the company config, exactly as the primary "
            f"scenario's own truncation is cleared. Escape hatches as this run read them: "
            f"{_hatches}; cfg carried {len(cfg)} keys)")

    fund = FCK.funding_report(work)
    if fund["verdict"] == "REVIEW" and not cfg.get("funding_reviewed"):
        w = fund.get("worst") or {}
        reasons.append(
            f"UNFUNDED DISTRIBUTION: {fund['reason']}. Worst year {w.get('year')}: implied "
            f"dividend {w.get('implied_dps')!r}/sh (clear with funding.reviewed: true)")

    term = TP.terminal_payout_report(cfg.get("terminal_payout_ratio"), conv.get("norm_eps_N"))
    if term["verdict"] == "MISSING":
        reasons.append(f"NO TERMINAL DISTRIBUTION POLICY: {term['reason']} (set "
                        "terminal.payout_ratio in the company config — there is no "
                        "reviewed:true escape hatch for a ratio nobody set)")
    elif term["verdict"] == "REVIEW" and not cfg.get("terminal_reviewed"):
        reasons.append(f"TERMINAL PAYOUT REVIEW REQUIRED: {term['reason']} (clear with "
                        "terminal.reviewed: true)")

    return conv, fund, term, reasons


def _value_one(base_xlsx, work_dir, ticker, sc, price, recalc, cfg=None, converge_K=DEFAULT_CONVERGE_K):
    """Value + tie ONE scenario on a fresh copy of base_xlsx. Returns a result dict with
    ok/reasons and (when ok) the CSV fields. Never raises for a tie/gate failure — the
    caller aggregates and fail-closes; only a hard PayloadError bubbles as a failure."""
    name = sc.get("name")
    work = os.path.join(work_dir, f"{ticker}_scn_{name}.xlsx")
    shutil.copyfile(base_xlsx, work)
    single = _as_single_payload(ticker, sc)

    # inflation from the (already recalc'd) base copy, at THIS scenario's N
    vals0 = openpyxl.load_workbook(work, data_only=True)
    N = single["N"] if isinstance(single["N"], int) else 0
    infl = AP.engine_inflation(vals0, N or 1)

    # write drivers/singles/mode/N; apply_payload validates and raises PayloadError on
    # a bad scenario (out-of-range driver, wrong length, unknown key) -> hard fail-closed.
    wbp = openpyxl.load_workbook(work, data_only=False)
    try:
        AP.apply_payload(wbp, single, infl)
    except AP.PayloadError as e:
        return {"name": name, "ok": False, "reasons": [f"payload rejected: {e}"]}

    erp = sc.get("erp_override")
    if erp is not None:
        RP.apply_erp_override(wbp, float(erp))
    coe_basis = "override" if erp is not None else "variance_v2"

    wbp.save(work)
    recalc(work)

    results = AE.read_results(work, price=price)
    reasons = []
    if not results.get("ok"):
        reasons.append(f"gates failed: {results.get('gates')}")
    tie_ok, tie_detail = CK.tie_check(results)
    if not tie_ok:
        reasons += tie_detail["reasons"]

    # Truncation (Gate A + Gate B), funding, and terminal-payout — see module docstring.
    # Only worth judging a scenario that has already tied; a broken workbook has nothing
    # coherent to judge a stop year against.
    conv = fund = term = None
    if not reasons:
        conv, fund, term, gate_reasons = _truncation_funding_terminal_gates(work, cfg, converge_K)
        reasons += gate_reasons

    if reasons:
        return {"name": name, "ok": False, "reasons": reasons,
                "audit_ok": tie_detail["audit_ok"], "tie_ok": tie_detail["tie_ok"],
                "mode_ok": tie_detail["mode_ok"]}

    v = openpyxl.load_workbook(work, data_only=True)
    intrinsic = results.get("active_value")
    real_price = _named_scalar(v, "val_realprice")
    real_coe = _named_scalar(v, "val_rhoe_lr")
    tie_res = results.get("max_identity_tie")
    upside = (intrinsic / real_price - 1.0) if (isinstance(intrinsic, (int, float))
              and isinstance(real_price, (int, float)) and real_price) else None
    return {
        "name": name, "ok": True, "coe_basis": coe_basis,
        "probability": sc.get("probability"), "real_coe": real_coe,
        "intrinsic": intrinsic, "real_price": real_price,
        "upside": upside, "tie_residual": tie_res,
        "audit_ok": True, "tie_ok": True, "mode_ok": True,
        "gates": {
            "terminal": conv["verdict"] if conv else None,
            "funding": fund["verdict"] if fund else None,
            "terminal_payout": term["verdict"] if term else None,
        },
    }


def _validate_scenarios(scenarios):
    if not isinstance(scenarios, list) or not scenarios:
        raise ScenariosError("payload.scenarios must be a non-empty list")
    names = []
    for i, sc in enumerate(scenarios):
        if not isinstance(sc, dict):
            raise ScenariosError(f"scenario[{i}] must be an object")
        nm = sc.get("name")
        if not isinstance(nm, str) or not nm.strip():
            raise ScenariosError(f"scenario[{i}] missing a non-empty 'name'")
        p = sc.get("probability")
        if not isinstance(p, (int, float)) or isinstance(p, bool):
            raise ScenariosError(f"scenario '{nm}' has a non-numeric probability {p!r}")
        names.append(nm)
    total = sum(float(sc["probability"]) for sc in scenarios)
    return names, total


def _fmt(x):
    if x is None:
        return ""
    if isinstance(x, float):
        return repr(x)
    return str(x)


def run_scenarios(base_xlsx, scenarios, *, ticker, price, out_dir, recalc,
                  commit_sha="", work_dir=None, run_timestamp=None, cfg=None,
                  converge_K=DEFAULT_CONVERGE_K):
    """Value every scenario independently off base_xlsx and write
    outputs/<TICKER>_scenarios.csv (one row per scenario + an expected-value summary row).
    Fail-closed: raises ScenariosError if ANY scenario fails its gates/tie, so the CI job
    goes red and no misleading green CSV is committed.

    cfg is the company-level config dict (companies/<T>.yaml, as run_company.py loads it) —
    threaded through so every scenario is held to the SAME truncation/funding/terminal-payout
    standard the primary scenario already gets via run_company.py's own path, not just the
    tie. See _truncation_funding_terminal_gates and the module docstring."""
    names, total = _validate_scenarios(scenarios)
    if abs(total - 1.0) > PROB_TOL:
        print(f"[scenarios] WARNING: probabilities sum to {total:.6f}, not 1.0 "
              f"(cockpit enforces this; proceeding on the valuation gates)")
    work_dir = work_dir or os.path.dirname(os.path.abspath(base_xlsx))
    ts = run_timestamp or _dt.datetime.now(_dt.timezone.utc).isoformat()
    print(f"[scenarios] {ticker}: valuing {len(scenarios)} scenarios {names}")

    rows, failures = [], []
    for sc in scenarios:
        r = _value_one(base_xlsx, work_dir, ticker, sc, price, recalc, cfg=cfg,
                       converge_K=converge_K)
        if r["ok"]:
            g = r.get("gates") or {}
            print(f"[scenarios]   {r['name']}: TIE ok  iv={r['intrinsic']}  "
                  f"real_coe={r['real_coe']}  tie={r['tie_residual']:.2e}  basis={r['coe_basis']}  "
                  f"gates: terminal={g.get('terminal')} funding={g.get('funding')} "
                  f"terminal_payout={g.get('terminal_payout')}")
            rows.append(r)
        else:
            print(f"[scenarios]   {r['name']}: FAIL — {'; '.join(r['reasons'])}")
            failures.append(r)

    if failures:
        raise ScenariosError(
            "fail-closed: " + ", ".join(f"{f['name']} ({'; '.join(f['reasons'])})"
                                        for f in failures))

    _write_csv(out_dir, ticker, ts, commit_sha, rows)
    return {"ticker": ticker, "scenarios": names, "rows": len(rows)}


def _write_csv(out_dir, ticker, ts, commit_sha, rows):
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"{ticker}_scenarios.csv")

    # expected value = Σ prob·iv ; real price is anchor-based (identical across scenarios)
    prices = [r["real_price"] for r in rows if isinstance(r["real_price"], (int, float))]
    real_price = prices[0] if prices else None
    ev_ps = None
    if all(isinstance(r["probability"], (int, float)) and isinstance(r["intrinsic"], (int, float))
           for r in rows):
        ev_ps = sum(float(r["probability"]) * float(r["intrinsic"]) for r in rows)
    ev_upside = (ev_ps / real_price - 1.0) if (isinstance(ev_ps, (int, float))
                 and isinstance(real_price, (int, float)) and real_price) else None
    prob_sum = sum(float(r["probability"]) for r in rows
                   if isinstance(r["probability"], (int, float)))

    with open(path, "w", newline="") as fh:
        fh.write(",".join(CSV_HEADER) + "\n")
        for r in rows:
            fh.write(",".join(_fmt(x) for x in [
                ticker, ts, commit_sha, r["name"], r["probability"], r["coe_basis"],
                r["real_coe"], r["intrinsic"], r["real_price"], r["upside"],
                r["tie_residual"],
            ]) + "\n")
        # summary row: expected value across scenarios
        fh.write(",".join(_fmt(x) for x in [
            ticker, ts, commit_sha, "expected_value", prob_sum, "", "",
            ev_ps, real_price, ev_upside, "",
        ]) + "\n")
    print(f"[scenarios] wrote {ticker}_scenarios.csv  "
          f"expected_value_ps={ev_ps}  (real_price={real_price}, exp_upside={ev_upside})")
    return path

#!/usr/bin/env python3
"""extract_blue_ranges.py — compute the exact set of EDITABLE (blue-input) cells in the
Work Area, so the Google Sheets protection script can lock everything else without
guessing colors in-Sheet. Deterministic; run against the sealed Work Area workbook.

The Work Area's designated inputs are drawn in dark-blue font (the importer writes drivers
this way). Everything else — formulas, labels, the whole valuation area — is output and must
be protected so neither a human nor an AI can touch it. This emits the blue cells as compact
A1 ranges (contiguous runs per row collapsed) for `protect_workarea.gs` to consume.
"""
import json, sys, re
import openpyxl
from openpyxl.utils import get_column_letter

BLUE_FONT_RGB = {"FF0000FF", "000000FF", "FF0000FE", "FF0000FD"}
SHEET = "Forecast Work Area"


def blue_cells(path, sheet=SHEET):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    cells = []
    for row in ws.iter_rows():
        for c in row:
            f = c.font
            if f and f.color and f.color.rgb and str(f.color.rgb) in BLUE_FONT_RGB:
                cells.append((c.row, c.column))
    return cells


def collapse_ranges(cells):
    """Collapse per-row contiguous columns into A1 ranges (e.g. G7..AJ7 -> 'G7:AJ7')."""
    by_row = {}
    for r, col in cells:
        by_row.setdefault(r, []).append(col)
    ranges = []
    for r in sorted(by_row):
        cols = sorted(by_row[r])
        start = prev = cols[0]
        for col in cols[1:] + [None]:
            if col == prev + 1:
                prev = col
                continue
            a = f"{get_column_letter(start)}{r}"
            b = f"{get_column_letter(prev)}{r}"
            ranges.append(a if a == b else f"{a}:{b}")
            if col is not None:
                start = prev = col
    return ranges


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else "Forecast_WorkArea.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "sheet_integration/blue_ranges.json"
    cells = blue_cells(path)
    ranges = collapse_ranges(cells)
    payload = {"sheet": SHEET, "n_cells": len(cells), "editable_ranges": ranges}
    with open(out, "w") as fh:
        json.dump(payload, fh, indent=2)
    print(f"{len(cells)} editable cells -> {len(ranges)} ranges -> {out}")
    print("ranges:", ranges)


if __name__ == "__main__":
    main()

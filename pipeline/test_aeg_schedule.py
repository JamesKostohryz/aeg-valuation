#!/usr/bin/env python3
"""test_aeg_schedule.py — tests for aeg_schedule.py, which emits <TICKER>_aeg_schedule.csv
on every live build (the year-by-year AEG buildup behind the headline). The module already
self-verifies its own core tie (normal_value + sum(contrib_eps) == intrinsic, fail-loud); this
suite adds coverage for everything else in the file: the E4.1 driver columns (retention_rate,
rore, eps_real/dps_real), the forecast-vs-continuing-period blanking rule, and confirms the
self-verify guard actually fires on a broken workbook rather than just existing in principle.
Closes a gap flagged in AEG-Coverage-Map-2026-08-08.md.

Builds the golden AAPL engine and recalculates through real LibreOffice — no network access
needed.

Checks:
  1. `cum_contrib_eps` is an exact running sum of `contrib_eps` down the CSV.
  2. For every explicit forecast year (t=1..N, where contrib_eps is a nonzero part of the
     value build): retention_rate = retained_eps / eps_nominal recomputes exactly, where
     eps_nominal is independently reconstructed as eps_real x infl_index (the CSV's own
     documented deflation relationship) — not re-read from a column the module doesn't
     even emit, a genuine independent check.
  3. rore (return on retained equity) for t=1 uses the ANCHOR (t=0) EPS/retained as the
     prior-year base, per the module's own documented formula; for t>1 it uses the prior
     row. Recomputed independently from the Valuation tab's own anchor column and the
     CSV's own prior rows.
  4. Beyond the explicit forecast horizon (continuing period, contrib_eps==0), the four
     driver columns are blank — a cross-boundary RoRE would be meaningless there, per the
     module's own docstring, and this confirms the code actually withholds it rather than
     emitting a stale or wrong number.
  5. The self-verify tie guard actually raises ValueError on a workbook whose intrinsic
     value has been deliberately corrupted (not just documented as fail-loud in the
     docstring).
  6. THE AEG VALUE TEST, pinned against the workbook (added 2026-08-13). Two things:
     (a) the neutral line recomputes from the CSV's own columns as
         normal_eps_t = (1+pi_t)*eps_(t-1) + (coe_t - pi_t)*retained_(t-1), i.e.
         MODEL_TEMPLATE Valuation row 22, and aeg_eps_t = eps_t - normal_eps_t (row 23);
     (b) sign(aeg_eps_t) == sign(real_rore_t - real_coe_t), which is the comparison the
         module's docstring now tells a forecaster to make. Exists because the docstring
         previously said to compare NOMINAL rore against NOMINAL coe, and that test gave
         the opposite answer to the engine in 8 of Coca-Cola's 11 checkable forecast years.
         The point of this check is that the documented test cannot drift away from the
         workbook again without a build going red.

Usage: python3 test_aeg_schedule.py
"""
import csv
import os
import shutil
import sys

import openpyxl

_PIPE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_PIPE)
for p in (_ROOT, _PIPE):
    if p not in sys.path:
        sys.path.insert(0, p)
import aeg_engine as AE
import aeg_schedule as AS
from recalc_lo import recalc

GOLDEN = os.path.join(_ROOT, "tests", "golden", "AAPL")
TEMPLATE = os.path.join(_ROOT, "MODEL_TEMPLATE.xlsx")
WORK = os.path.join(_ROOT, "_aegschedtest")
os.makedirs(WORK, exist_ok=True)

_p = _f = 0
def check(c, m):
    global _p, _f
    if c: _p += 1; print(f"  PASS  {m}")
    else: _f += 1; print(f"  FAIL  {m}")


def build(out):
    files = {"is_csv": f"{GOLDEN}/REAL_IS.csv", "bs_csv": f"{GOLDEN}/REAL_BS.csv",
              "cf_csv": f"{GOLDEN}/REAL_CF.csv", "prices": f"{GOLDEN}/REAL_prices.csv",
              "dividends": f"{GOLDEN}/REAL_div.csv", "splits": f"{GOLDEN}/REAL_splits.csv"}
    cfg = {"company": "Apple Inc.", "ticker": "AAPL", "price": 315.0, "files": files,
           "fy_end_month": 9,
           "forecast_horizon_N": 4,   # P2: cfg_N is required and has no default; 4 is the
                                     # horizon these fixtures have always run at.
           "judgments": {"minority_include": False, "finlease": 0.0, "oi_adj_override": None,
                         "rd_capitalize": True, "rd_life": 5.0, "dps_override": None},
           "cost_of_debt": {"single_ytw": 0.05}}
    AE.build_model(cfg, TEMPLATE, out)
    recalc(out)


eng = os.path.join(WORK, "AAPL_sched.xlsx")
build(eng)
fn = AS.write_aeg_schedule(eng, "AAPL", WORK)
with open(os.path.join(WORK, fn)) as fh:
    rows = list(csv.DictReader(fh))

wb = openpyxl.load_workbook(eng, data_only=True)
V = wb["Valuation"]
anchor_eps = V.cell(7, 2).value
anchor_retained = V.cell(9, 2).value

print(f"== schedule has {len(rows)} rows; checking exact recomputation ==")

print("== cum_contrib_eps is an exact running sum ==")
running = 0.0
ok = True
for r in rows:
    ce = float(r["contrib_eps"])
    running += ce
    if abs(round(running, 8) - float(r["cum_contrib_eps"])) > 1e-6:
        ok = False
check(ok, "cum_contrib_eps matches an independently accumulated running sum, every row")

print("== forecast-year drivers recompute exactly (retention_rate, rore) ==")
forecast_rows = [r for r in rows if r["retention_rate"] != ""]
check(len(forecast_rows) >= 1, f"at least one explicit forecast year has non-blank drivers ({len(forecast_rows)} found)")
prev_eps, prev_retained = anchor_eps, anchor_retained
for i, r in enumerate(rows):
    if r["retention_rate"] == "":
        continue
    eps_nominal = float(r["eps_real"]) * float(r["infl_index"])
    retained = float(r["retained_eps"])
    want_retention = retained / eps_nominal
    got_retention = float(r["retention_rate"])
    check(abs(got_retention - want_retention) < 1e-3,
          f"t={r['t']}: retention_rate={got_retention} matches retained_eps/eps_nominal={want_retention:.6f} "
          f"(eps_nominal independently reconstructed as eps_real x infl_index)")
    want_rore = (eps_nominal - prev_eps) / prev_retained
    got_rore = float(r["rore"])
    check(abs(got_rore - want_rore) < 1e-3,
          f"t={r['t']}: rore={got_rore} matches (eps_t - eps_prev)/retained_prev={want_rore:.6f} "
          f"(prior year = {'anchor (t=0)' if i == 0 else 'previous row'})")
    prev_eps, prev_retained = eps_nominal, retained

print("== beyond the explicit forecast horizon, drivers are blank, not stale/wrong ==")
continuing_rows = [r for r in rows if r["contrib_eps"] == "0" and r not in forecast_rows]
check(len(continuing_rows) >= 1, f"schedule extends past the explicit forecast horizon ({len(continuing_rows)} continuing-period rows)")
for r in continuing_rows:
    check(r["retention_rate"] == "" and r["retained_eps"] == "" and r["coe"] == "" and r["rore"] == "",
          f"t={r['t']}: all four forecast-only drivers are blank in the continuing period (a cross-boundary "
          f"RoRE would be meaningless there, per the module's own docstring)")

print("== check 6: the AEG value test recomputes from the workbook's own definition ==")
# MODEL_TEMPLATE Valuation r22: normal_eps_t = (1+pi_t)*eps_(t-1) + (coe_t - pi_t)*retained_(t-1)
#                          r23: aeg_eps_t    = eps_t - normal_eps_t
# Everything below is rebuilt from the CSV's own columns plus the anchor column, so this is an
# independent recomputation of the identity rather than a re-read of it.
_prev_eps, _prev_ret = anchor_eps, anchor_retained
_sign_ok = _norm_ok = True
_naive_disagreements = 0
_checked = 0
for i, r in enumerate(rows):
    if r["retention_rate"] == "":
        continue
    pi = float(r["pi"])
    coe = float(r["coe"])
    eps_nominal = float(r["eps_real"]) * float(r["infl_index"])
    retained = float(r["retained_eps"])
    want_normal = (1.0 + pi) * _prev_eps + (coe - pi) * _prev_ret
    got_normal = float(r["normal_eps"])
    if abs(want_normal - got_normal) > max(1e-4, 1e-6 * abs(got_normal)):
        _norm_ok = False
        print(f"    t={r['t']}: normal_eps {got_normal} != recomputed {want_normal}")
    got_aeg = float(r["aeg_eps"])
    if abs((eps_nominal - got_normal) - got_aeg) > max(1e-4, 1e-6 * abs(eps_nominal)):
        _norm_ok = False
        print(f"    t={r['t']}: aeg_eps {got_aeg} != eps - normal_eps {eps_nominal - got_normal}")
    # (b) the documented comparison, in REAL terms, must agree in sign with the engine's own AEG
    real_rore = (eps_nominal / (1.0 + pi) - _prev_eps) / _prev_ret
    real_coe = (coe - pi) / (1.0 + pi)
    if abs(got_aeg) > 1e-9 and (got_aeg > 0) != (real_rore > real_coe):
        _sign_ok = False
        print(f"    t={r['t']}: aeg_eps={got_aeg:+.6f} but real_rore={real_rore:.6f} "
              f"vs real_coe={real_coe:.6f}")
    # informational only: how often the RETIRED nominal comparison would have been wrong
    if abs(got_aeg) > 1e-9 and (got_aeg > 0) != (float(r["rore"]) > coe):
        _naive_disagreements += 1
    _checked += 1
    _prev_eps, _prev_ret = eps_nominal, retained
check(_checked >= 1, f"check 6 examined {_checked} explicit forecast year(s)")
check(_norm_ok, "normal_eps and aeg_eps recompute from Valuation r22/r23 on the CSV's own columns")
check(_sign_ok, "sign(aeg_eps) == sign(real RoRE - real COE) in every explicit forecast year "
                "-- the comparison aeg_schedule.py's docstring tells a forecaster to make")
print(f"  note  the RETIRED nominal test (rore > coe) would have disagreed with the engine in "
      f"{_naive_disagreements} of {_checked} year(s) on this fixture; on KO it was 8 of 11")

print("== self-verify tie guard actually fires on a corrupted workbook ==")
bad = os.path.join(WORK, "AAPL_bad.xlsx")
wb_cached = openpyxl.load_workbook(eng, data_only=True)   # bakes every formula's cached value in as a literal
wb_cached["Valuation"]["B44"].value = 99999.0             # corrupt the intrinsic value (R_INTRINSIC)
wb_cached.save(bad)
try:
    AS.write_aeg_schedule(bad, "AAPL", WORK)
    check(False, "corrupted intrinsic value should raise ValueError, not silently ship a broken schedule")
except ValueError as e:
    check("does not tie" in str(e), f"raises ValueError naming the tie break: {str(e)[:120]}")

shutil.rmtree(WORK, ignore_errors=True)
print(f"\n{_p} passed, {_f} failed")
sys.exit(1 if _f else 0)

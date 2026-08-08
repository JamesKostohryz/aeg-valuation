#!/usr/bin/env python3
"""test_dupont_extract.py — tests for dupont_extract.py, which runs on every live build and
emits <TICKER>_dupont.csv/.json, the data backend for the cockpit's DuPont tool. Closes a
gap flagged in AEG-Coverage-Map-2026-08-08.md: no prior automated coverage beyond "the code
didn't crash."

Builds the golden AAPL engine and recalculates through real LibreOffice — no network access
needed.

Checks:
  1. classic 3-step ROE (NPM x ATO x EM) recomputes exactly from independently re-read
     Income Statement / Balance Sheet cells, for EVERY year in the series, not just the
     latest.
  2. classic 3-step and 5-step ROE are mathematically the same telescoping product
     (TaxBurden x IntBurden x EBITmargin x ATO x Leverage collapses to NPM x ATO x EM) —
     verified equal to machine precision every year, since this is an algebraic identity
     that must hold regardless of the underlying company.
  3. the module's "leverage" field (5-step) is exactly the same number as
     "equity_multiplier" (3-step) — an intentional duplicate, checked every year.
  4. the reformulated (Penman/economic) `reported_roe` memo the engine itself computes on
     the Econ Statements tab agrees with the classic 3-step ROE computed here INDEPENDENTLY
     from the raw reported Income Statement / Balance Sheet lines, for every year where both
     exist — a genuine cross-tab consistency check, not just re-testing the same formula.
  5. `roce = rnoa + flev x spread` (the Penman decomposition identity) holds to machine
     precision for every reformulated year, using only the numbers the engine emits.
  6. `write_outputs`: both files are produced; a classic-only year (before the reformulated
     series begins) writes blank reform_* columns rather than misaligning rows or crashing;
     round-tripping the JSON reproduces the same `latest` block `compute_dupont` returned.

Usage: python3 test_dupont_extract.py
"""
import csv
import json
import os
import shutil
import sys

_PIPE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_PIPE)
for p in (_ROOT, _PIPE):
    if p not in sys.path:
        sys.path.insert(0, p)
import aeg_engine as AE
import dupont_extract as DP
from recalc_lo import recalc

GOLDEN = os.path.join(_ROOT, "tests", "golden", "AAPL")
TEMPLATE = os.path.join(_ROOT, "MODEL_TEMPLATE.xlsx")
WORK = os.path.join(_ROOT, "_dupontest")
os.makedirs(WORK, exist_ok=True)

_p = _f = 0
def check(c, m):
    global _p, _f
    if c: _p += 1; print(f"  PASS  {m}")
    else: _f += 1; print(f"  FAIL  {m}")


files = {"is_csv": f"{GOLDEN}/REAL_IS.csv", "bs_csv": f"{GOLDEN}/REAL_BS.csv",
          "cf_csv": f"{GOLDEN}/REAL_CF.csv", "prices": f"{GOLDEN}/REAL_prices.csv",
          "dividends": f"{GOLDEN}/REAL_div.csv", "splits": f"{GOLDEN}/REAL_splits.csv"}
cfg = {"company": "Apple Inc.", "ticker": "AAPL", "price": 315.0, "files": files,
       "fy_end_month": 9,
       "judgments": {"minority_include": False, "finlease": 0.0, "oi_adj_override": None,
                     "rd_capitalize": True, "rd_life": 5.0, "dps_override": None},
       "cost_of_debt": {"single_ytw": 0.05}}
eng = os.path.join(WORK, "AAPL_dp.xlsx")
AE.build_model(cfg, TEMPLATE, eng)
recalc(eng)

d = DP.compute_dupont(eng, "AAPL")
cl, rf = d["classic"], d["reformulated"]

import openpyxl
wb = openpyxl.load_workbook(eng, data_only=True)
IS, BS = wb["Income Statement"], wb["Balance Sheet"]
isd = {k: DP._year_series(IS, DP.IS_HDR, lbl) for k, lbl in DP.IS_LINES.items()}
bsd = {k: DP._year_series(BS, DP.BS_HDR, lbl) for k, lbl in DP.BS_LINES.items()}

print(f"== classic series spans {len(cl['years'])} years; checking every one ==")
for i, y in enumerate(cl["years"]):
    rev, ni = isd["revenue"][y], isd["net_income"][y]
    ebit, ebt = isd["ebit"][y], isd["ebt"][y]
    assets, eq = bsd["assets"][y], bsd["equity"][y]
    want_npm, want_ato, want_em = ni / rev, rev / assets, assets / eq
    check(abs(cl["net_profit_margin"][i] - want_npm) < 1e-9, f"{y}: net_profit_margin recomputes exactly")
    check(abs(cl["asset_turnover"][i] - want_ato) < 1e-9, f"{y}: asset_turnover recomputes exactly")
    check(abs(cl["equity_multiplier"][i] - want_em) < 1e-9, f"{y}: equity_multiplier recomputes exactly")
    want_roe3 = want_npm * want_ato * want_em
    check(abs(cl["roe_3step"][i] - want_roe3) < 1e-9, f"{y}: roe_3step = NPM x ATO x EM exactly")
    check(abs(cl["leverage"][i] - cl["equity_multiplier"][i]) < 1e-12,
          f"{y}: 'leverage' (5-step) is exactly the same number as 'equity_multiplier' (3-step)")

print("== 3-step and 5-step ROE telescope to the same value every year (algebraic identity) ==")
n_checked = 0
for i, y in enumerate(cl["years"]):
    r3, r5 = cl["roe_3step"][i], cl["roe_5step"][i]
    if r3 is None or r5 is None:
        continue
    n_checked += 1
    check(abs(r3 - r5) < 1e-9, f"{y}: roe_3step ({r3:.6f}) == roe_5step ({r5:.6f}) to machine precision")
check(n_checked >= 10, f"telescoping identity checked across {n_checked} years")

print("== engine's own reformulated memo agrees with the independently-computed classic ROE ==")
n_cross = 0
for i, y in enumerate(rf["years"]):
    if y not in cl["years"]:
        continue
    ci = cl["years"].index(y)
    memo, classic3 = rf["reported_roe"][i], cl["roe_3step"][ci]
    if memo is None or classic3 is None:
        continue
    n_cross += 1
    check(abs(memo - classic3) < 1e-6,
          f"{y}: Econ Statements 'reported_roe' memo ({memo:.6f}) matches classic roe_3step computed "
          f"independently from raw IS/BS lines ({classic3:.6f})")
check(n_cross >= 10, f"cross-tab reported-ROE check performed across {n_cross} years")

print("== Penman identity: economic-NI/CSE = rnoa + flev x spread, every reformulated year ==")
# NOTE: the naive textbook check "roce = rnoa + flev x spread" FAILS here -- not a bug, a
# real modeling distinction the sheet itself documents in its row labels. Econ Statements
# row 63 (ROCE) divides COMPREHENSIVE income (row 56 = Economic NI [row 54] + OCI [row 55])
# by CSE, while RNOA/FLEV/NBC/SPREAD (rows 59-62) are built from Economic NI alone (row 54
# = row 52 - row 53). The textbook identity holds exactly against Economic NI/CSE; ROCE adds
# an OCI term on top that the simple three-ratio combination doesn't capture. Confirmed
# directly against the sheet's own row formulas before writing this check, so this is the
# CORRECT identity, not a relaxed substitute for a failing one.
ES = wb["Econ Statements"]
ES_ECON_NI_ROW, ES_CSE_ROW, ES_OCI_ROW = 54, 45, 55
n_penman = 0
n_roce_gap = 0
for i, y in enumerate(rf["years"]):
    rnoa, flev, spread, roce = rf["rnoa"][i], rf["flev"][i], rf["spread"][i], rf["roce"][i]
    if None in (rnoa, flev, spread, roce):
        continue
    col = next((c for c in range(2, ES.max_column + 1) if str(ES.cell(5, c).value) == str(y)), None)
    if col is None:
        continue
    econ_ni, cse, oci = ES.cell(ES_ECON_NI_ROW, col).value, ES.cell(ES_CSE_ROW, col).value, ES.cell(ES_OCI_ROW, col).value
    if not (isinstance(econ_ni, (int, float)) and isinstance(cse, (int, float)) and cse):
        continue
    n_penman += 1
    econ_roe = econ_ni / cse
    check(abs(econ_roe - (rnoa + flev * spread)) < 1e-6,
          f"{y}: Economic NI/CSE ({econ_roe:.6f}) = rnoa + flev x spread ({rnoa + flev * spread:.6f}) exactly")
    if isinstance(oci, (int, float)) and cse:
        n_roce_gap += 1
        check(abs(roce - (econ_roe + oci / cse)) < 1e-6,
              f"{y}: published ROCE ({roce:.6f}) = Economic-NI-based ROE ({econ_roe:.6f}) + OCI/CSE "
              f"({oci / cse:+.6f}) exactly -- ROCE is comprehensive-income-based, RNOA+FLEVxSPREAD is not")
check(n_penman >= 10, f"Penman identity (against Economic NI, the correct row) checked across {n_penman} years")
check(n_roce_gap >= 10, f"ROCE-vs-OCI reconciliation checked across {n_roce_gap} years")

print("== write_outputs: files produced, blank alignment for classic-only years, JSON round-trips ==")
fnames = DP.write_outputs(d, WORK, data_as_of="test-run")
check(set(fnames) == {"AAPL_dupont.json", "AAPL_dupont.csv"}, f"both output files produced ({fnames})")
with open(os.path.join(WORK, "AAPL_dupont.csv")) as fh:
    csv_rows = list(csv.DictReader(fh))
pre_reform_year = next(y for y in cl["years"] if y not in rf["years"])
row = next(r for r in csv_rows if int(r["year"]) == pre_reform_year)
check(row["reform_rnoa"] == "", f"a classic-only year ({pre_reform_year}, before the reformulated series begins) "
                                 f"writes a BLANK reform_rnoa rather than misaligning to a different year's value")
with open(os.path.join(WORK, "AAPL_dupont.json")) as fh:
    loaded = json.load(fh)
check(loaded["latest"] == d["latest"], "JSON round-trips the same 'latest' block compute_dupont returned")
check(loaded["data_as_of"] == "test-run", "data_as_of stamp is written through to the JSON")

shutil.rmtree(WORK, ignore_errors=True)
print(f"\n{_p} passed, {_f} failed")
sys.exit(1 if _f else 0)

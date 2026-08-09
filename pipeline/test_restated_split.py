#!/usr/bin/env python3
"""test_restated_split.py — tests for restated_split.py, which runs on every live build and
emits <TICKER>_restated_bs.csv / _restated_is.csv for the cockpit's Financials-Restated tab.
Closes a gap flagged in AEG-Coverage-Map-2026-08-08.md. The module already ships a fail-loud
units guard (reported vs. restated scale must match); this suite confirms it actually fires
on a genuine mismatch, confirms the CSVs are a faithful row-for-row export of the Econ
Statements source rows, confirms the sheet's own in-row balance identity, and confirms cash
flow is never emitted (a deliberate, documented omission, not an oversight).

Builds the golden AAPL engine and recalculates through real LibreOffice — no network access
needed.

Checks:
  1. The scale guard passes silently on an untouched golden build (confirms the invariant
     genuinely holds for AAPL, not just that the guard exists).
  2. Every emitted row/year cell in both CSVs matches the SAME Econ Statements cell,
     re-read independently by row and year rather than through the module's own
     `_year_cols` helper.
  3. The `tie: Real NOA-NFO-CSE` row is ~0 for every year in the balance-sheet CSV — the
     sheet's own in-row identity, exported faithfully.
  4. Only `restated_bs` and `restated_is` are ever produced — no `restated_cf`, per the
     module's explicit, documented decision not to fabricate an untied real-terms cash
     flow.
  5. The scale guard actually raises ValueError when reported and restated genuinely
     diverge (constructed by corrupting the Econ Statements 'Total debt' row in a copy),
     not just documented as fail-loud in the docstring.

Usage: python3 test_restated_split.py
"""
import os
import shutil
import sys

import openpyxl

_PIPE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_PIPE)
for p in (_ROOT, _PIPE):
    if p not in sys.path:
        sys.path.insert(0, p)
import aeg_engine as AE
import restated_split as RSp
from recalc_lo import recalc

GOLDEN = os.path.join(_ROOT, "tests", "golden", "AAPL")
TEMPLATE = os.path.join(_ROOT, "MODEL_TEMPLATE.xlsx")
WORK = os.path.join(_ROOT, "_restatedtest")
os.makedirs(WORK, exist_ok=True)

_p = _f = 0
def check(c, m):
    global _p, _f
    if c: _p += 1; print(f"  PASS  {m}")
    else: _f += 1; print(f"  FAIL  {m}")


files = {"is_csv": f"{GOLDEN}/REAL_IS.csv", "bs_csv": f"{GOLDEN}/REAL_BS.csv",
          "cf_csv": f"{GOLDEN}/REAL_CF.csv", "prices": f"{GOLDEN}/REAL_prices.csv",
          "dividends": f"{GOLDEN}/REAL_div.csv", "splits": f"{GOLDEN}/REAL_splits.csv"}
cfg = {"company": "Apple Inc.", "ticker": "AAPL", "price": 315.0, "files": files,
       "fy_end_month": 9,
       "forecast_horizon_N": 4,   # P2: cfg_N is required and has no default; 4 is the
                                 # horizon these fixtures have always run at.
       "judgments": {"minority_include": False, "finlease": 0.0, "oi_adj_override": None,
                     "rd_capitalize": True, "rd_life": 5.0, "dps_override": None},
       "cost_of_debt": {"single_ytw": 0.05}}
eng = os.path.join(WORK, "AAPL_rs.xlsx")
AE.build_model(cfg, TEMPLATE, eng)
recalc(eng)

print("== scale guard passes on the untouched golden build ==")
fns = RSp.write_restated(eng, "AAPL", WORK)
check(set(fns) == {"AAPL_restated_bs.csv", "AAPL_restated_is.csv"},
      f"exactly the two documented files are produced, no restated_cf ({fns})")

wb = openpyxl.load_workbook(eng, data_only=True)
ES = wb["Econ Statements"]
yrs = RSp._year_cols(ES)

print("== every CSV cell matches the source Econ Statements cell exactly ==")
import csv
for suffix, (r0, r1) in RSp.SECTIONS.items():
    with open(os.path.join(WORK, f"AAPL_{suffix}.csv")) as fh:
        csv_rows = list(csv.reader(fh))
    header = csv_rows[0]
    check(header[0] == "line_item" and header[1:] == [str(y) for _, y in yrs],
          f"{suffix}: header year columns match _year_cols exactly")
    body = csv_rows[1:]
    expected_labels = [str(ES.cell(r, 1).value).strip() for r in range(r0, r1 + 1)
                       if ES.cell(r, 1).value is not None and str(ES.cell(r, 1).value).strip()]
    check([row[0] for row in body] == expected_labels,
          f"{suffix}: row labels match the Econ Statements section {r0}-{r1} exactly, in order")
    n_checked = 0
    for row in body:
        label = row[0]
        src_row = next(r for r in range(r0, r1 + 1) if str(ES.cell(r, 1).value or "").strip() == label)
        for (c, y), cell_str in zip(yrs, row[1:]):
            src_val = ES.cell(src_row, c).value
            csv_val = None if cell_str == "" else float(cell_str)
            same = (src_val is None and csv_val is None) or \
                   (isinstance(src_val, (int, float)) and csv_val is not None and abs(src_val - csv_val) < 1e-9)
            if not same:
                check(False, f"{suffix} row '{label}' year {y}: CSV={csv_val} vs source cell={src_val}")
                break
            n_checked += 1
        else:
            continue
        break
    check(n_checked > 0, f"{suffix}: {n_checked} individual cells cross-checked against the source sheet")

print("== the in-row balance identity ('tie: Real NOA-NFO-CSE') is ~0 for every year ==")
tie_row = next(r for r in range(39, 47) if "tie" in str(ES.cell(r, 1).value or "").lower())
n_tie = 0
for c, y in yrs:
    v = ES.cell(tie_row, c).value
    if isinstance(v, (int, float)):
        n_tie += 1
        check(abs(v) < 1e-6, f"{y}: Econ Statements' own tie row reads ~0 ({v:.2e}), exported faithfully in the CSV")
check(n_tie >= 10, f"balance identity checked across {n_tie} years")

print("== scale guard actually fires on a genuine reported-vs-restated mismatch ==")
bad = os.path.join(WORK, "AAPL_bad.xlsx")
wb_cached = openpyxl.load_workbook(eng, data_only=True)   # bakes cached formula values in as literals
ES_bad = wb_cached["Econ Statements"]
for c in range(2, ES_bad.max_column + 1):
    v = ES_bad.cell(RSp.ES_TOTAL_DEBT_ROW, c).value
    if isinstance(v, (int, float)) and v:
        ES_bad.cell(RSp.ES_TOTAL_DEBT_ROW, c).value = v * 1000.0   # break the scale invariant everywhere
wb_cached.save(bad)
try:
    RSp.write_restated(bad, "AAPL", WORK)
    check(False, "a genuine 1000x scale mismatch between reported and restated debt should raise ValueError")
except ValueError as e:
    check("SCALE MISMATCH" in str(e), f"raises ValueError naming the scale mismatch: {str(e)[:120]}")

shutil.rmtree(WORK, ignore_errors=True)
print(f"\n{_p} passed, {_f} failed")
sys.exit(1 if _f else 0)

#!/usr/bin/env python3
"""
validate_load.py — fail-loud completeness + provenance gates for the AEG loader.

Run AFTER the model is populated and recalculated. Two guarantees:
  1. COMPLETENESS — every engine-critical line has a populated FY0 (hard abort if not);
     Cap-Engine history lines (depreciation, gross PP&E) are checked for recent-year
     coverage (loud WARN, surfaced in the report, never silent).
  2. PROVENANCE — no base-company (AT&T template) fingerprint survives: the cost-of-debt
     ladder must not still be AT&T's, no 'AT&T' text in per-company cells, price not the
     base default. Hard abort — the model must be fully repointed.

Design intent: turn "I trust it didn't change what it shouldn't" into "it proves it did,
and stops if it didn't." Nothing here is silent.
"""
import re
import openpyxl

def _norm(s): return re.sub(r"\s+", " ", str(s).strip()).lower()

# engine-critical lines whose FY0 MUST be populated (abort if blank)
CRITICAL_FY0 = [
    ("Income Statement", "Total Revenue"),
    ("Income Statement", "Cost of Revenue"),
    ("Income Statement", "Operating Income"),
    ("Income Statement", "Net Income Common Stockholders"),
    ("Income Statement", "Reconciled Depreciation"),
    ("Income Statement", "Tax Rate for Calcs"),
    ("Income Statement", "Diluted EPS"),
    ("Balance Sheet",    "Total Assets"),
    ("Balance Sheet",    "Cash And Cash Equivalents"),
    ("Balance Sheet",    "Net PPE"),
    ("Balance Sheet",    "Gross PPE"),
    ("Balance Sheet",    "Common Stock Equity"),
    ("Balance Sheet",    "Ordinary Shares Number"),
    ("Balance Sheet",    "Total Debt"),
    ("Cash Flow",        "Capital Expenditure"),
]
# Cap-Engine lines that need history; WARN if recent coverage is thin
HISTORY_LINES = {
    ("Income Statement", "Reconciled Depreciation"): 10,
    ("Balance Sheet",    "Gross PPE"): 10,
}
# AT&T template fingerprints (base model)
ATT_COD_LADDER = [0.04702, 0.0495, 0.05125, 0.0526, 0.05369]
ATT_PRICE_DEFAULT = 21.12


def _last_data_col(ws):
    cols = [c for c in range(2, 45) if ws.cell(3, c).value not in (None, "")]
    return max(cols) if cols else None

def _row_of(ws, label):
    n = _norm(label)
    for r in range(4, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and _norm(v) == n:
            return r
    return None


def check_completeness(wb):
    """Return (aborts:list, warns:list)."""
    aborts, warns = [], []
    for tab, label in CRITICAL_FY0:
        ws = wb[tab]; col = _last_data_col(ws); r = _row_of(ws, label)
        if r is None:
            aborts.append(f"{tab}: critical line '{label}' ROW MISSING"); continue
        if ws.cell(r, col).value in (None, ""):
            aborts.append(f"{tab}: critical line '{label}' has BLANK FY0 (col {col})")
    for (tab, label), need in HISTORY_LINES.items():
        ws = wb[tab]; col = _last_data_col(ws); r = _row_of(ws, label)
        if r is None or col is None:
            continue
        recent = [ws.cell(r, c).value for c in range(max(2, col - need + 1), col + 1)]
        filled = sum(1 for v in recent if v not in (None, ""))
        if filled < len(recent):
            warns.append(f"{tab}: '{label}' has {filled}/{len(recent)} of the most recent "
                         f"{len(recent)} years populated (Cap Engine leans on this history)")
    return aborts, warns


def check_provenance(wb, base_company="AT&T", price=None):
    """Return list of provenance violations (any → abort)."""
    v = []
    md = wb["Market Data"]
    cod = [md.cell(27, c).value for c in range(2, 7)]
    if cod == ATT_COD_LADDER:
        v.append(f"cost-of-debt ladder (Market Data row 27) is still the base company's: {cod}")
    for tab in ("Inputs", "Presentation", "Cap Engine"):
        if tab not in wb.sheetnames:
            continue  # Presentation is a display tab; absent in the slimmed engine
        ws = wb[tab]
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and base_company in c.value:
                    v.append(f"base-company text at {tab}!{c.coordinate}: '{c.value[:50]}'")
    if price is not None and abs(float(price) - ATT_PRICE_DEFAULT) < 1e-9:
        v.append(f"current price is still the base default {ATT_PRICE_DEFAULT}")
    return v


def validate(model_path, price=None, base_company="AT&T"):
    """Run all gates. Returns (ok:bool, report:dict). ok=False means the run must halt."""
    wb = openpyxl.load_workbook(model_path, data_only=True)
    aborts, warns = check_completeness(wb)
    prov = check_provenance(wb, base_company=base_company, price=price)
    ok = not aborts and not prov
    return ok, {"completeness_aborts": aborts, "completeness_warns": warns,
                "provenance_violations": prov}


def print_report(report):
    A, W, P = (report["completeness_aborts"], report["completeness_warns"],
               report["provenance_violations"])
    print("  COMPLETENESS")
    if A:
        for x in A: print(f"    [ABORT] {x}")
    else:
        print("    [OK] every critical line has a populated FY0")
    for x in W: print(f"    [WARN]  {x}")
    print("  PROVENANCE")
    if P:
        for x in P: print(f"    [ABORT] {x}")
    else:
        print("    [OK] no base-company fingerprint found")
    ok = not A and not P
    print(f"  => {'PASS' if ok else 'HALT'} ({len(A)+len(P)} aborts, {len(W)} warnings)")
    return ok


if __name__ == "__main__":
    import sys
    ok, rep = validate(sys.argv[1], price=(float(sys.argv[2]) if len(sys.argv) > 2 else None))
    print_report(rep)
    sys.exit(0 if ok else 1)

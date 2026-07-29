#!/usr/bin/env python3
"""repoint_fy0.py — point the reported-FY0 reconciliation anchors at the ACTUAL
anchor-year column (fail-closed).

The MODEL_TEMPLATE ships eight defined names — rep_debt_fy0, rep_cash_fy0,
rep_cse_fy0, rep_shares_fy0 (Balance Sheet) and rep_eps_fy0, rep_intexp_fy0,
rep_oi_fy0, rep_tax_fy0 (Income Statement) — HARDCODED to the last template
column ($AP$). They feed Audit CHECK-4b: in_* anchors == reported FY0.

That hardcode is only correct for an issuer whose statement history is long
enough that the newest fiscal year lands in the final column AP (AAPL/HD/T,
~1985..2025 -> AP). For a shorter-history issuer the newest year lands earlier
(POOL, ~1990..2025 -> column AK); AP is then blank, so every CHECK-4b residual
equals the full input value and the Audit master status reads "FAIL —
investigate" on a model that otherwise ties to ~1e-11.

This module repoints each of the eight names to the column whose year equals the
anchor year, on that name's own sheet (year rows rep_years_bs / rep_years_is).
No-op for issuers already anchored at AP (they stay bit-identical). Runs during
the build after build_model and before the first recalc (same slot as
deflator_extend), so the recalc computes the reconciliation against the right
column.
"""
import os, re, json, argparse
import openpyxl
from openpyxl.utils import get_column_letter


class Fy0Error(RuntimeError):
    pass


# name -> which statement year row governs its column
FY0_NAMES = {
    "rep_debt_fy0": "bs", "rep_cash_fy0": "bs",
    "rep_cse_fy0": "bs", "rep_shares_fy0": "bs",
    "rep_eps_fy0": "is", "rep_intexp_fy0": "is",
    "rep_oi_fy0": "is", "rep_tax_fy0": "is",
}
YEAR_RANGE = {"bs": "rep_years_bs", "is": "rep_years_is"}


def _dn(wb, name):
    dn = wb.defined_names.get(name)
    if dn is None:
        raise Fy0Error(f"expected defined name missing: {name}")
    return dn


def _as_year(v):
    """Coerce a year-header cell to int. Headers are written as TEXT ('2025') by
    build_model, but tolerate numeric too. Returns None if not a year."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return int(v)
    if isinstance(v, str):
        s = v.strip()
        if s.lstrip("-").isdigit():
            return int(s)
    return None


def _year_col_map(wb, years_name):
    """{int(year): column_index} built from a single-row year named range."""
    dn = _dn(wb, years_name)
    sheet, coord = list(dn.destinations)[0]
    ws = wb[sheet]
    out = {}
    for row in ws[coord]:
        for cell in row:
            y = _as_year(cell.value)
            if y is not None:
                out[y] = cell.column
    if not out:
        raise Fy0Error(f"no year headers found in {years_name} ({sheet}!{coord})")
    return out


def repoint_anchor_columns(path, anchor_year):
    """Repoint the eight rep_*_fy0 names to the anchor_year column on their own
    sheet. Idempotent. Returns a report dict. Fail-closed: raises Fy0Error if a
    name is missing, an anchor ref is malformed, or the anchor year is absent from
    a statement's year row."""
    anchor_year = int(anchor_year)
    wb = openpyxl.load_workbook(path, data_only=False)

    colmaps = {k: _year_col_map(wb, yrs) for k, yrs in YEAR_RANGE.items()}

    moved, unchanged = [], []
    for name, which in FY0_NAMES.items():
        colmap = colmaps[which]
        if anchor_year not in colmap:
            raise Fy0Error(f"{name}: anchor year {anchor_year} not in {YEAR_RANGE[which]}")
        new_letter = get_column_letter(colmap[anchor_year])
        dn = _dn(wb, name)
        ref = dn.value  # e.g. "'Balance Sheet'!$AP$100" or "Balance Sheet!$AP$100"
        m = re.match(r"^(.*)!\$([A-Za-z]+)\$(\d+)$", ref)
        if not m:
            raise Fy0Error(f"{name}: unexpected anchor ref {ref!r}")
        sheet_part, old_letter, row_part = m.group(1), m.group(2), m.group(3)
        if new_letter == old_letter:
            unchanged.append([name, old_letter])
            continue
        dn.value = f"{sheet_part}!${new_letter}${row_part}"
        moved.append([name, old_letter, new_letter])

    if moved:
        wb.save(path)
    return {"anchor_year": anchor_year, "moved": moved, "unchanged": unchanged}


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("anchor_year", type=int)
    a = ap.parse_args()
    print(json.dumps(repoint_anchor_columns(a.path, a.anchor_year)))

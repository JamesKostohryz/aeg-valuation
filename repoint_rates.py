#!/usr/bin/env python3
"""repoint_rates.py — bind the sealed AEG engine's Market Data rate rows to the
rate-infrastructure feed (rate_feed.load_all output). Called by the builder AFTER
statements/inputs are populated. Deterministic, openpyxl-only; caller recalcs.

What it touches on 'Market Data' (30 tenors, cols B..AE):

  row 21  Expected inflation SPOT (breakeven)  <- feed breakeven_spot   [values]
          -> engine's own row-22 formula derives the 1yr-forward breakeven.
  row 23  Real risk-free SPOT (TIPS)           <- feed real_rf_spot     [values]
          -> engine's own row-24 formula derives the real forward rf (== feed
             real_rf_fwd1y == coe real_rf, single-sourced from the spot curve).
  row 25  Equity risk premium (finrate_erp)    <- feed market_erp       [values]
          finrate_erp redefined B25 -> B25:AE25 (a TERM STRUCTURE, was a flat
          single). COE row 26 rewritten to add the per-tenor ERP instead of one
          flat value. This upgrades our engine from a flat ERP to the feed's
          decaying market-ERP term structure.
  row 28  ρD REAL forward [feed real_cod]       <- feed real_cod         [values]
          (row 28 repurposed as the real-COD holder). COD row 29 (finrate_cod)
          rewritten to select row 28 directly, preserving the cfg_rd_mode toggle,
          consuming real_cod with NO nominal step (per the locked contract).

Leverage stays entirely inside the engine (MM un/re-lever). credit_relative is
DROPPED here (never enters the sheet). idiosyncratic is a DISCLOSED premium added
by a separate step, not by re-pointing. Re-pointing must NOT move the four-method
tie — verified by recalc afterwards.
"""
import copy
import openpyxl

MD_SHEET = "Market Data"
COL0, COLN = 2, 31          # B..AE  = tenors 1..30
ROW = dict(tenor=20, be_spot=21, infl_fwd=22, rf_spot=23, rf_fwd=24,
           erp=25, coe=26, cod_nom_spot=27, cod_real=28, finrate_cod=29)


def _colname(c):
    return openpyxl.utils.get_column_letter(c)


def _check_tenors(MD):
    got = [MD.cell(ROW["tenor"], c).value for c in range(COL0, COLN + 1)]
    if got != list(range(1, 31)):
        raise ValueError(f"[repoint] Market Data tenor row is not 1..30: {got}")


def _write_row(MD, r, series, *, keep_font_from=None):
    """Write a length-30 series into row r across B..AE as plain values."""
    if len(series) != 30:
        raise ValueError(f"[repoint] row {r}: expected 30 values, got {len(series)}")
    font = copy.copy(MD.cell(r, keep_font_from).font) if keep_font_from else None
    for i, c in enumerate(range(COL0, COLN + 1)):
        cell = MD.cell(r, c)
        cell.value = float(series[i])
        if font is not None:
            cell.font = copy.copy(font)


def repoint(wb, feed, *, relabel=True):
    """Re-point the engine's rate rows from `feed` (rate_feed.load_all output).
    Returns a report dict. Raises ValueError on any structural mismatch."""
    if MD_SHEET not in wb.sheetnames:
        raise ValueError(f"[repoint] workbook has no '{MD_SHEET}' tab")
    MD = wb[MD_SHEET]
    _check_tenors(MD)

    req = ("breakeven_spot", "real_rf_spot", "market_erp")
    for k in req:
        if k not in feed:
            raise ValueError(f"[repoint] feed missing '{k}'")
    if "real_cod" not in feed:
        raise ValueError("[repoint] feed missing 'real_cod' "
                         "(bonded issuer required; use rating-curve fallback path otherwise)")

    # --- 1) breakeven spot (row 21) -> engine derives forward breakeven (row 22)
    _write_row(MD, ROW["be_spot"], feed["breakeven_spot"], keep_font_from=COL0)

    # --- 2) real rf spot (row 23) -> engine derives real forward rf (row 24)
    _write_row(MD, ROW["rf_spot"], feed["real_rf_spot"], keep_font_from=COL0)

    # --- 3) market ERP TERM STRUCTURE (row 25); redefine finrate_erp; rewrite COE
    _write_row(MD, ROW["erp"], feed["market_erp"], keep_font_from=COL0)
    _redefine_name(wb, "finrate_erp",
                   f"'{MD_SHEET}'!${_colname(COL0)}${ROW['erp']}:${_colname(COLN)}${ROW['erp']}")
    for c in range(COL0, COLN + 1):
        col = _colname(c)
        MD.cell(ROW["coe"], c).value = (
            f'=IF(cfg_coe_mode="Single",$AE${ROW["rf_fwd"]}+$AE${ROW["erp"]},'
            f'{col}{ROW["rf_fwd"]}+{col}{ROW["erp"]})')

    # --- 4) real COD direct (row 28 holds feed real_cod); rewrite finrate_cod (row 29)
    _write_row(MD, ROW["cod_real"], feed["real_cod"], keep_font_from=COL0)
    for c in range(COL0, COLN + 1):
        col = _colname(c)
        MD.cell(ROW["finrate_cod"], c).value = (
            f'=IF(cfg_rd_mode="Single",$AE${ROW["cod_real"]},{col}{ROW["cod_real"]})')

    # --- 5) install the idiosyncratic hook (default 0 -> no-op, tie preserved)
    install_idio_hook(wb)

    if relabel:
        MD.cell(ROW["be_spot"], 1).value = "Expected inflation — SPOT (breakeven) [feed:breakeven]"
        MD.cell(ROW["rf_spot"], 1).value = "Real risk-free — SPOT (TIPS) [feed:real]"
        MD.cell(ROW["erp"], 1).value = "Equity risk premium (real) TERM STRUCTURE [feed:market_erp; finrate_erp]"
        MD.cell(ROW["cod_nom_spot"], 1).value = "ρD nominal — SPOT [UNUSED: superseded by feed real_cod]"
        MD.cell(ROW["cod_real"], 1).value = "ρD REAL — 1yr FORWARD [feed:real_cod, consumed directly]"

    return {
        "ticker": feed.get("ticker"),
        "erp_termstructure": (round(feed["market_erp"][0], 5), round(feed["market_erp"][29], 5)),
        "real_cod_t1_t30": (round(feed["real_cod"][0], 5), round(feed["real_cod"][29], 5)),
        "breakeven_t1_t30": (round(feed["breakeven_spot"][0], 5), round(feed["breakeven_spot"][29], 5)),
        "real_rf_spot_t1_t30": (round(feed["real_rf_spot"][0], 5), round(feed["real_rf_spot"][29], 5)),
        "nfo_basis": feed.get("nfo_basis"),
        "market_nfo": feed.get("market_nfo"),
    }


FINRATE_IDIO_ROW = 31           # free Market Data row for the idiosyncratic series


def install_idio_hook(wb):
    """Install the V1-Plus idiosyncratic hook: a Market Data series `finrate_idio`
    (default 0) added onto the cost-of-equity term structure (COE, row 26) that the
    AEG/residual-income valuation discounts at. At 0 it is a NO-OP, so the tied headline
    valuation is untouched; the disclosure step populates it for a sensitivity run and
    reads the equity-value difference (the disclosed idiosyncratic haircut).

    COE — not the dormant DCF re-lever layer — is the lever the headline valuation
    actually consumes, and it is shared by the equity and enterprise sides, so bumping
    it preserves BOTH the four-method reconciliation tie AND equity=enterprise agreement
    (verified). Idempotent; must run AFTER the ERP rewrite of row 26."""
    MD = wb[MD_SHEET]
    r = FINRATE_IDIO_ROW
    MD.cell(r, 1).value = "Idiosyncratic premium (real) [feed:idiosyncratic; finrate_idio] — 0 in base"
    for c in range(COL0, COLN + 1):
        if not isinstance(MD.cell(r, c).value, (int, float)):
            MD.cell(r, c).value = 0.0
    _redefine_name(wb, "finrate_idio",
                   f"'{MD_SHEET}'!${_colname(COL0)}${r}:${_colname(COLN)}${r}")
    # append +<col>31 (the idio series, same-column) onto each COE cell, once
    for c in range(COL0, COLN + 1):
        cell = MD.cell(ROW["coe"], c)
        f = cell.value
        if not isinstance(f, str) or not f.startswith("="):
            continue
        if f"{_colname(c)}{r}" in f:
            continue
        cell.value = f + f"+{_colname(c)}{r}"


def set_idio(wb, series):
    """Write an idiosyncratic series (len 30, annual decimals) into finrate_idio for a
    disclosure sensitivity run. Pass zeros (or None) to restore the tied base."""
    MD = wb[MD_SHEET]
    r = FINRATE_IDIO_ROW
    if series is None:
        series = [0.0] * 30
    if len(series) != 30:
        raise ValueError(f"[set_idio] expected 30 values, got {len(series)}")
    for i, c in enumerate(range(COL0, COLN + 1)):
        MD.cell(r, c).value = float(series[i])


def _redefine_name(wb, name, ref):
    """Redefine (or create) a workbook-scoped defined name to `ref`."""
    from openpyxl.workbook.defined_name import DefinedName
    try:
        if name in wb.defined_names:
            del wb.defined_names[name]
    except Exception:
        pass
    wb.defined_names.add(DefinedName(name, attr_text=ref))

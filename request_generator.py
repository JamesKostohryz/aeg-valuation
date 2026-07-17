#!/usr/bin/env python3
"""
request_generator.py — build the Mailbox workbook (the 'nominal zone' the AI fills).

make_request(hist, company, ticker, N, out_path): writes a standalone workbook with
  - 10 years of NOMINAL history + the implied historical driver ratios (grey, read-only context)
  - empty YELLOW driver cells for Y1..YN, in NOMINAL terms, for the AI to fill
  - a mode toggle and horizon
No engine, no valuation formulas — a form, not the model. The importer reads it back.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

DRIVERS = ["revenue_growth","gross_margin","sga_ratio","da_rate","tax_rate",
           "buyback_rate","capex_ratio","noa_growth","target_flev"]
HDR_ROW = 8            # header row: 'driver | hist yrs | Y1..YN'
FIRST_DRIVER_ROW = 9

def implied_ratios(hist):
    """hist: {series: [(year, value)...]} nominal. Return {driver: [(year, ratio)...]}."""
    y = [yr for yr,_ in hist["revenue"]]
    def col(s): return {yr:v for yr,v in hist.get(s,[])}
    rev,gp,sga,da,netppe = col("revenue"),col("gross_profit"),col("sga"),col("da"),col("netppe")
    taxr,capex,shares = col("tax_rate"),col("capex"),col("shares")
    cse,debt,cash,sti = col("cse"),col("debt"),col("cash"),col("sti")
    def nfo(yr): return (debt.get(yr) or 0)-(cash.get(yr) or 0)-(sti.get(yr) or 0)
    def noa(yr): return (cse.get(yr) or 0)+nfo(yr)
    out={d:[] for d in DRIVERS}
    for i,yr in enumerate(y):
        pv = y[i-1] if i>0 else None
        def g(a,b): return (a/b-1) if (a is not None and b) else None
        out["revenue_growth"].append((yr, g(rev.get(yr), rev.get(pv)) if pv else None))
        out["gross_margin"].append((yr, (gp.get(yr)/rev[yr] if False else (gp.get(yr)/rev.get(yr) if rev.get(yr) else None))))
        out["sga_ratio"].append((yr, (sga.get(yr)/rev.get(yr)) if rev.get(yr) else None))
        out["da_rate"].append((yr, (da.get(yr)/netppe.get(pv)) if (pv and netppe.get(pv)) else None))
        out["tax_rate"].append((yr, taxr.get(yr)))
        out["capex_ratio"].append((yr, (abs(capex.get(yr))/rev.get(yr)) if (capex.get(yr) and rev.get(yr)) else None))
        out["buyback_rate"].append((yr, (-(shares.get(yr)-shares.get(pv))/shares.get(pv)) if (pv and shares.get(pv)) else None))
        out["noa_growth"].append((yr, g(noa(yr), noa(pv)) if pv else None))
        out["target_flev"].append((yr, (nfo(yr)/cse.get(yr)) if cse.get(yr) else None))
    return out

def make_request(hist, company, ticker, N, out_path, hist_years=10):
    ratios = implied_ratios(hist)
    yrs = [yr for yr,_ in hist["revenue"]][-hist_years:]
    wb = openpyxl.Workbook(); ws = wb.active; ws.title="Forecast Request"
    ws.sheet_view.showGridLines=False
    BLUE=Font(name="Arial",size=10,color="0000FF"); GREY=Font(name="Arial",size=10,color="808080")
    BOLD=Font(name="Arial",size=10,bold=True); TITLE=Font(name="Arial",size=13,bold=True,color="1F3864")
    SUB=Font(name="Arial",size=9,italic=True,color="595959")
    YEL=PatternFill("solid",fgColor="FFF2CC"); GRY=PatternFill("solid",fgColor="F0F0F0")
    thin=Side(style="thin",color="D9D9D9"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
    ws.column_dimensions['A'].width=22
    ws['A1']=f"AEG FORECAST REQUEST — {company} ({ticker})"; ws['A1'].font=TITLE
    ws['A2']=("Fill ONLY the yellow cells with your NOMINAL forecast for Y1…YN. Grey columns are the "
              "company's implied historical actuals (context) — do not edit. Do not touch anything else.")
    ws['A2'].font=SUB
    ws['A4']="Company"; ws['B4']=company; ws['A5']="Ticker"; ws['B5']=ticker
    ws['A6']="Forecast mode"; mode=ws['B6']; mode.value="Equity"; mode.font=BLUE; mode.fill=YEL
    ws['C6']="Horizon N"; hn=ws['D6']; hn.value=N; hn.font=BLUE; hn.fill=YEL
    dv=DataValidation(type="list",formula1='"Equity,Enterprise"'); ws.add_data_validation(dv); dv.add(mode)
    # header row
    nhist=len(yrs)
    ws.cell(HDR_ROW,1,"driver (nominal)").font=BOLD
    for j,yr in enumerate(yrs):
        c=ws.cell(HDR_ROW,2+j,str(yr)); c.font=GREY; c.alignment=Alignment(horizontal="center")
    for k in range(N):
        c=ws.cell(HDR_ROW,2+nhist+k,f"Y{k+1}"); c.font=BOLD; c.alignment=Alignment(horizontal="center")
    # driver rows
    for i,d in enumerate(DRIVERS):
        r=FIRST_DRIVER_ROW+i
        ws.cell(r,1,d).font=Font(name="Arial",size=10)
        for j,yr in enumerate(yrs):
            v=dict(ratios[d]).get(yr)
            cell=ws.cell(r,2+j, round(v,4) if isinstance(v,(int,float)) else None)
            cell.font=GREY; cell.fill=GRY; cell.number_format="0.000"
        for k in range(N):
            cell=ws.cell(r,2+nhist+k); cell.font=BLUE; cell.fill=YEL; cell.border=B; cell.number_format="0.000"
    # equity payout single
    pr=FIRST_DRIVER_ROW+len(DRIVERS)
    ws.cell(pr,1,"payout (equity)").font=Font(name="Arial",size=10)
    pc=ws.cell(pr,2+nhist); pc.font=BLUE; pc.fill=YEL; pc.number_format="0.000"
    ws.cell(pr+2,1,"Note: revenue_growth and noa_growth are NOMINAL; we convert to real on our side.").font=SUB
    wb.save(out_path)
    return {"drivers": len(DRIVERS), "hist_years": nhist, "N": N, "header_row": HDR_ROW, "first_driver_row": FIRST_DRIVER_ROW}

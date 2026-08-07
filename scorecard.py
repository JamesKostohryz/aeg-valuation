#!/usr/bin/env python3
"""scorecard.py — disclosed inflation scorecard + interest-tax-shield PVs.

ADDITIVE and TIE-SAFE. Computed from the recalced engine (in engine monetary units)
plus the rate feed (pi, real cost of debt). It never enters the four-method tie — it
is a disclosed output alongside disclose.py's valuation bridge.

Implements the tie-safe content of the new engine's sheets 10 (interest tax shield
under debt policy) and 15 (capital-intensity vs leverage scorecard):

  Inflation is a TAX on capital intensity — historical-cost depreciation understates
  the real cost of maintaining capacity, so the firm overpays tax by t x shortfall
  (the same shortfall Increment 1 put into the valuation).

  Inflation is a SUBSIDY to leverage — the real value of fixed nominal debt erodes,
  worth t x pi x net_debt per year in tax terms (interest deductible; principal repaid
  in cheaper dollars).

  A firm is a net beneficiary iff the leverage subsidy outweighs the capital-intensity
  tax; equivalently iff debt / annual D&A exceeds the breakeven [(1+pi)^age - 1] / pi.

The interest tax shield is reported under both debt policies (fixed-nominal vs
constant-real) and EXCLUDED from the headline by default (Miller 1977) — disclosure
only, exactly as the new engine treats it.
"""
import openpyxl


def _nm(wb, name):
    dn = wb.defined_names.get(name)
    if not dn:
        return None
    ref = str(dn.value).replace("$", "").replace("'", "")
    try:
        sh, cell = ref.split("!")
        return wb[sh][cell].value
    except Exception:
        return None


def _last(x):
    """Feed rate series -> long-run (last) value; scalar -> itself."""
    if isinstance(x, (list, tuple)):
        nums = [v for v in x if isinstance(v, (int, float))]
        return nums[-1] if nums else None
    if isinstance(x, dict):
        nums = [v for v in x.values() if isinstance(v, (int, float))]
        return nums[-1] if nums else None
    return x


def depreciation_penalty(engine):
    """Annual depreciation penalty  t x shortfall  = t x (economic - reported
    historical-cost) PP&E depreciation of the SAME live vintages, in engine monetary
    units. This is the measurement defect Increment 1 charges against the anchor: the
    single source of truth shared by the scorecard AND the disclose.py valuation bridge.

    `engine` is a recalced-workbook path or an already-loaded data_only workbook. Needs
    only the engine (no rate feed). None-safe: returns Nones when Cap Engine / tax / life
    are unavailable, so callers can treat it as a no-op."""
    wb = engine if hasattr(engine, "sheetnames") else openpyxl.load_workbook(engine, data_only=True)
    t = _nm(wb, "in_tax0")
    L = _nm(wb, "in_ppe_life")
    if "Cap Engine" not in wb.sheetnames or t is None or not L:
        return {"shortfall": None, "penalty_annual": None,
                "econ_dep": None, "reported_ppe_dep": None}
    CE = wb["Cap Engine"]
    real_gross = CE["B46"].value or 0.0
    nom_live_gross = sum((CE.cell(row=r, column=2).value or 0.0) * (CE.cell(row=r, column=6).value or 0.0)
                         for r in range(7, 44))
    econ_dep = real_gross / L
    reported_ppe_dep = nom_live_gross / L
    shortfall = econ_dep - reported_ppe_dep
    return {"shortfall": shortfall, "penalty_annual": t * shortfall,
            "econ_dep": econ_dep, "reported_ppe_dep": reported_ppe_dep}


def compute_scorecard(engine_path, feed, *, debt_policy="constant_real"):
    """Return the disclosed inflation scorecard for one recalced engine + its feed.
    All money quantities are in the engine's own monetary units (so the penalty and
    the leverage benefit are directly comparable); rates are unitless from the feed."""
    wb = openpyxl.load_workbook(engine_path, data_only=True)
    t = _nm(wb, "in_tax0")
    interest = _nm(wb, "in_intexp0")                      # nominal interest expense
    debt = _nm(wb, "in_debt") or 0.0
    net_debt = debt - (_nm(wb, "in_cash") or 0.0) - (_nm(wb, "in_sti") or 0.0)
    L = _nm(wb, "in_ppe_life")

    # --- depreciation penalty: t x (current-cost - historical-cost) PP&E depreciation
    #     of the SAME live vintages. Shared with disclose.py so the scorecard and the
    #     valuation bridge charge the identical shortfall (Increment 1).
    _dp = depreciation_penalty(wb)
    econ_dep = _dp["econ_dep"] or 0.0
    reported_ppe_dep = _dp["reported_ppe_dep"] or 0.0
    shortfall = _dp["shortfall"] or 0.0
    dep_penalty = _dp["penalty_annual"] or 0.0           # annual; >0 = penalty

    # --- rates from the feed
    pi = _last(feed.get("exp_inflation_fwd1y"))
    if pi is None:
        pi = _last(feed.get("exp_inflation_spot"))
    rd_real = _last(feed.get("real_cod"))
    rd_nom = ((1 + rd_real) * (1 + pi) - 1) if (rd_real is not None and pi is not None) else None

    # --- interest tax shield PV under both policies (disclosure only; Miller-excluded)
    shield_annual = (t * interest) if (t is not None and interest is not None) else None
    pv_fixednom = (shield_annual / rd_nom) if (shield_annual is not None and rd_nom) else None
    pv_constreal = (shield_annual / (rd_nom - pi)) if (shield_annual is not None and rd_nom is not None
                                                       and pi is not None and (rd_nom - pi) > 0) else None
    pv_adopted = {"fixed_nominal": pv_fixednom, "constant_real": pv_constreal,
                  "mixed": (None if (pv_fixednom is None or pv_constreal is None)
                            else 0.5 * (pv_fixednom + pv_constreal))}.get(debt_policy)

    # --- leverage subsidy + net verdict
    interest_benefit = (t * pi * net_debt) if (t is not None and pi is not None) else None
    net_position = (interest_benefit - dep_penalty) if (interest_benefit is not None) else None
    verdict = None if net_position is None else (
        "NET BENEFICIARY of inflation" if net_position > 0 else "NET LOSER from inflation")

    # --- breakeven leverage: debt / annual D&A vs [(1+pi)^age - 1]/pi
    #     reported (historical-cost) annual PP&E depreciation ~ nom_live_gross / L;
    #     average live-vintage age from the vintage table (real-gross weighted).
    CE = wb["Cap Engine"]
    real_gross = CE["B46"].value or 0.0
    num = sum((CE.cell(row=r, column=8).value or 0.0) * (CE.cell(row=r, column=5).value or 0.0)
              for r in range(7, 44))                     # H (real gross contrib) x E (age)
    den = real_gross
    avg_age = (num / den) if den else None
    debt_to_da = (net_debt / reported_ppe_dep) if reported_ppe_dep else None
    breakeven = (((1 + pi) ** avg_age - 1) / pi) if (avg_age is not None and pi) else None

    return {
        "debt_policy": debt_policy,
        "tax_rate": t, "expected_inflation": pi, "rd_nominal": rd_nom,
        "net_debt": net_debt, "nominal_interest": interest,
        "econ_depreciation": econ_dep, "reported_ppe_depreciation": reported_ppe_dep,
        "depreciation_shortfall": shortfall,
        "depreciation_penalty_annual": dep_penalty,
        "interest_benefit_annual": interest_benefit,
        "net_inflation_position_annual": net_position,
        "verdict": verdict,
        "avg_asset_age_yrs": avg_age,
        "debt_to_annual_da": debt_to_da,
        "breakeven_leverage": breakeven,
        "interest_tax_shield_pv_fixed_nominal": pv_fixednom,
        "interest_tax_shield_pv_constant_real": pv_constreal,
        "interest_tax_shield_pv_adopted": pv_adopted,
        "shield_treatment": "excluded from headline (Miller 1977); disclosure only",
        "verdict_basis": ("net_inflation_position (t*pi*net_debt vs t*shortfall), using the "
                          "engine's BEA capital-goods deflator for PP&E; the breakeven_leverage "
                          "column is v1.4's general-CPI rule of thumb and can differ when "
                          "capital-goods prices diverge from general inflation"),
    }


# ---- ordered field list for the disclosed CSV ---------------------------------------
_CSV_FIELDS = [
    "verdict", "net_inflation_position_annual", "depreciation_penalty_annual",
    "interest_benefit_annual", "tax_rate", "expected_inflation", "rd_nominal", "net_debt",
    "nominal_interest", "econ_depreciation", "reported_ppe_depreciation",
    "depreciation_shortfall", "avg_asset_age_yrs", "debt_to_annual_da", "breakeven_leverage",
    "interest_tax_shield_pv_fixed_nominal", "interest_tax_shield_pv_constant_real",
    "interest_tax_shield_pv_adopted", "debt_policy", "shield_treatment", "verdict_basis",
]


def write_scorecard_csv(path, d):
    """Write the disclosed scorecard as a two-column field,value CSV."""
    import csv
    with open(path, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["field", "value"])
        for k in _CSV_FIELDS:
            w.writerow([k, d.get(k)])
    return path


# ---------------------------------------------------------------- Increment 0 Stage B1
def depreciation_terminal_step(engine, ss_from=25, ss_to=30):
    """The TERMINAL half of the historical-cost depreciation penalty.

    Stage B1 puts the year-by-year penalty inside the nominal forecast, so years 1..N are
    priced by the engine's own AEG machinery. But the engine truncates at N and therefore
    assumes the year-N wedge persists for ever. The wedge is NOT flat: a firm adding capex
    at current cost dilutes it (new assets have a tax basis equal to replacement cost, so
    zero wedge) and it then rebuilds as those assets age. With cfg_N=4 the forecast can stop
    near the bottom of that transient.

    This returns the value of the missing permanent STEP from the year-N wedge to the
    steady-state wedge, read off the engine's own long-tenor path (so it uses the engine's
    depreciation construction rather than imposing an external formula). Closed form for a
    straight-line stack, wedge = 1 - S(n)/S(g) with S(r)=sum (1+r)^-k, k=0..L-1 and
    n=(1+g)(1+pi)-1, is a cross-check on magnitude; it is 0 when pi is 0, as it must be.

    Anchor-level and tie-preserving: it never enters the four-method identity.
    """
    wb = engine if hasattr(engine, "sheetnames") else openpyxl.load_workbook(engine, data_only=True)
    if "Forecast" not in wb.sheetnames or "Valuation" not in wb.sheetnames:
        return {"terminal_step_ps": None}
    F, V = wb["Forecast"], wb["Valuation"]
    t = _nm(wb, "in_tax0"); shares = _nm(wb, "anchor_shares0")
    N = _nm(wb, "cfg_N")
    rho = V["B20"].value                      # long-run REAL cost of equity
    if None in (t, shares, N, rho) or not shares or not rho:
        return {"terminal_step_ps": None}
    N = int(N)

    def cell(row, tt):                        # Forecast col G == t=1
        return F.cell(row, 6 + tt).value
    def num(x):
        return x if isinstance(x, (int, float)) else None

    econ_n, tax_n = num(cell(52, N)), num(cell(49, N))
    if not econ_n or tax_n is None:
        return {"terminal_step_ps": None}     # not a Stage B1 engine
    wedge_N = (econ_n - tax_n) / econ_n

    ws = []
    for tt in range(ss_from, ss_to + 1):
        e, x = num(cell(52, tt)), num(cell(49, tt))
        if e:
            ws.append((e - x) / e)
    if not ws:
        return {"terminal_step_ps": None}
    wedge_ss = sum(ws) / len(ws)

    econ_real_N = num(cell(13, N))             # economic depreciation, real spine
    dfE_N = num(V.cell(16, 2 + N).value)       # nominal cumulative DF to N
    idx_N = num(V.cell(57, 2 + N).value)       # cumulative inflation index to N
    if None in (econ_real_N, dfE_N, idx_N):
        return {"terminal_step_ps": None}
    df_real_N = dfE_N * idx_N                  # real discount factor to N

    step_annual_real = t * econ_real_N * (wedge_ss - wedge_N)
    step_ps = (step_annual_real / shares) / rho * df_real_N
    return {"wedge_N": wedge_N, "wedge_ss": wedge_ss, "N": N,
            "econ_dep_real_N": econ_real_N, "step_annual_real": step_annual_real,
            "df_real_N": df_real_N, "terminal_step_ps": step_ps}

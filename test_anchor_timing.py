#!/usr/bin/env python3
"""test_anchor_timing.py — the anchor-date convention, and the one thing the tie cannot see.

WHY THIS FILE EXISTS
--------------------
The abnormal-earnings-growth valuation anchors on FORWARD normal earnings — year one — not
on trailing year-zero earnings. Anchoring on trailing earnings while running the abnormal
stream from year one understates value by exactly the prior year's retained earnings.

THE FOUR-METHOD TIE PROVIDES NO PROTECTION HERE. The abnormal-earnings-growth, residual
income, free-cash-flow-to-equity and free-cash-flow-to-firm legs all share the timing
convention. Shift the anchor date and all four shift together, so all four would still tie
at 1e-15 while every published number was wrong by the anchor year's retained earnings. It
is the same blind spot that let the cross-tab gap survive until pull request #3: an identity
check can only catch errors that break the identity.

So this file checks the convention against an oracle that is NOT one of the four spokes —
the dividend discount model, and the perpetuity closed form it reduces to.

THE CONVENTION, STATED SO NOBODY LATER "FIXES" IT
-------------------------------------------------
With no abnormal growth, earnings E0, retention b and cost of equity r, normal growth is
g = b*r, so:

    E1 = E0*(1 + b*r)           D1 = E1*(1 - b)
    DDM value  = D1/(r - g)  =  E1/r      <-- the correct anchor, FORWARD
    trailing   = E0/r        =  E1/r - b*E0

The difference, b*E0, is the year-zero retained earnings. Capital retained at the end of
year zero earns its return during year one; an anchor on trailing earnings simply omits it.

Worked instance from the report-template chat, reproduced in the grid below: E0 = 100,
b = 0.60, r = 6.65% gives D1/(r-g) = 41.596/0.0266 = 1563.76 = E1/r = 103.99/0.0665, while
E0/r = 1503.76 — short by exactly 60.00 = 0.60 * 100.

WHAT THE ENGINE DOES
--------------------
Valuation!B43 is `=B10+(C22-C61*B10)*C62`: book value per share at t=0, plus year-ONE
normal residual income capitalised at the curve-implied annuity factor. Column C is t=1.
The earnings term is FORWARD. This is correct and must not be changed to B22 or B7.

A NOTE ON THE DISCOUNT FACTOR, because it is easy to get backwards. The annuity factor
A^E_1 (row 62) begins at DF^E_1, i.e. the first residual-income payment IS discounted one
period. That is right for a stream starting at t=1. What is undiscounted is the ANCHOR
itself, which is a time-zero quantity. Under a flat curve with zero inflation A^E_1
collapses to exactly 1/r, which is asserted below — if that ever stops holding, the
annuity construction has drifted.
"""
import os
import shutil
import sys

_ROOT = os.path.dirname(os.path.abspath(__file__))
for _p_ in (_ROOT, os.path.join(_ROOT, "pipeline")):
    if _p_ not in sys.path:
        sys.path.insert(0, _p_)

import openpyxl                                              # noqa: E402
import aeg_engine as AE                                      # noqa: E402
import checks as CK                                          # noqa: E402
from recalc_lo import recalc                                 # noqa: E402

GOLDEN = os.path.join(_ROOT, "tests", "golden", "AAPL")
TEMPLATE = os.path.join(_ROOT, "MODEL_TEMPLATE.xlsx")
WORK = os.environ.get("AAPL_ENG_WORK") or "/tmp/_anchor_timing_work"
os.makedirs(WORK, exist_ok=True)

_p = _f = 0


def ok(cond, msg):
    global _p, _f
    if cond:
        _p += 1
        print("  PASS", msg)
    else:
        _f += 1
        print("  FAIL", msg)


def near(a, b, tol, msg):
    got = (isinstance(a, (int, float)) and isinstance(b, (int, float)) and abs(a - b) <= tol)
    ok(got, f"{msg}  ({a!r} vs {b!r}, |diff| tol {tol:g})")


# ------------------------------------------------------------------ 1. the oracle
print("== the dividend discount model as an independent oracle ==")
# A grid, not the single worked example: a convention that only holds at one point is a
# coincidence. Includes the report chat's own case (100, 0.60, 0.0665) as the first entry.
CASES = [(100.0, 0.60, 0.0665), (100.0, 0.0, 0.08), (250.0, 0.90, 0.12),
         (12.5, 0.35, 0.045), (1.0, 0.99, 0.10), (80.0, 0.50, 0.0665)]
for E0, b, r in CASES:
    g = b * r
    E1 = E0 * (1.0 + g)
    D1 = E1 * (1.0 - b)
    ddm = D1 / (r - g) if r > g else None
    fwd = E1 / r
    trl = E0 / r
    tag = f"E0={E0:g} b={b:g} r={r:g}"
    near(ddm, fwd, 1e-9 * max(1.0, abs(fwd)),
         f"{tag}: forward anchor E1/r equals the dividend discount model D1/(r-g)")
    near(fwd - trl, b * E0, 1e-9 * max(1.0, abs(fwd)),
         f"{tag}: trailing anchor is short by exactly the year-zero retained earnings")
    # book value must drop out of the residual-income form entirely
    for B0 in (0.0, E0, 7 * E0):
        near(B0 + (E1 - r * B0) / r, fwd, 1e-9 * max(1.0, abs(fwd)),
             f"{tag}: residual-income form with B0={B0:g} still returns E1/r")
# the report chat's exact figures, so the numbers in the work order are locked too
near(100.0 * (1 + 0.60 * 0.0665) * 0.40 / (0.0665 - 0.60 * 0.0665), 1563.7593984962405,
     1e-6, "the worked example reproduces 1,563.76")
near(100.0 / 0.0665, 1503.7593984962405, 1e-6, "the trailing anchor gives 1,503.76")
near(1563.7593984962405 - 1503.7593984962405, 60.0, 1e-9, "the shortfall is exactly 60.00")

# ------------------------------------------------------- 2. the engine's formula shape
print("== the engine anchors on the FORWARD year (formula structure) ==")
FORM = openpyxl.load_workbook(TEMPLATE, data_only=False)
_b43 = str(FORM["Valuation"].cell(43, 2).value)
ok(_b43 == "=B10+(C22-C61*B10)*C62",
   f"Valuation!B43 is the forward-anchored normal value ({_b43})")
ok("C22" in _b43, "the earnings term is C22 — normal earnings at t=1, the FORWARD year")
ok("B22" not in _b43 and "B7" not in _b43,
   "the earnings term is NOT the trailing year (B22/B7) — this is the assertion that "
   "catches a well-meaning 'fix' back to trailing")
ok("B10" in _b43, "the book term is B10 — book value at t=0, which is correct: book is a "
                  "stock at the valuation date, earnings are a forward flow")
_c22 = str(FORM["Valuation"].cell(22, 3).value)
ok(_c22 == "=(1+C56)*B7+(C5-C56)*B9",
   f"normal EPS at t=1 is built from the ANCHOR year's earnings and retention ({_c22})")

# ------------------------------------------------------- 3. the engine, numerically
print("== the engine reproduces the closed form, measured (flat curve, zero inflation) ==")
CFG = {"company": "Apple Inc.", "ticker": "AAPL", "price": 315.0, "fy_end_month": 9,
       "forecast_horizon_N": 4,
       "files": {"is_csv": f"{GOLDEN}/REAL_IS.csv", "bs_csv": f"{GOLDEN}/REAL_BS.csv",
                 "cf_csv": f"{GOLDEN}/REAL_CF.csv", "prices": f"{GOLDEN}/REAL_prices.csv",
                 "dividends": f"{GOLDEN}/REAL_div.csv", "splits": f"{GOLDEN}/REAL_splits.csv"},
       "judgments": {"minority_include": False, "finlease": 0.0, "oi_adj_override": None,
                     "rd_capitalize": True, "rd_life": 5.0, "dps_override": None},
       "cost_of_debt": {"single_ytw": 0.05}}
BASE = os.path.join(WORK, "anchor_base.xlsx")
AE.build_model(CFG, TEMPLATE, BASE)
recalc(BASE)

# Flat curve + zero expected inflation is the regime where the engine's curve-implied
# annuity factor must collapse to the textbook 1/r, so the closed form applies exactly.
FLAT = os.path.join(WORK, "anchor_flat.xlsx")
shutil.copy(BASE, FLAT)
_wb = openpyxl.load_workbook(FLAT)
_wb["Inputs"]["B29"] = "Single"
for _c in range(2, 32):
    _wb["Market Data"].cell(22, _c).value = 0.0
_wb.save(FLAT)
recalc(FLAT)

_d = openpyxl.load_workbook(FLAT, data_only=True)
V = _d["Valuation"]
r_lr = V.cell(20, 2).value          # long-run real cost of equity
A1 = V.cell(62, 3).value            # A^E_1, the curve-implied annuity factor
nEPS1 = V.cell(22, 3).value         # normal EPS at t=1  (FORWARD)
EPS0 = V.cell(7, 2).value           # actual EPS at t=0  (trailing)
ret0 = V.cell(9, 2).value           # retained per share at t=0
B43 = V.cell(43, 2).value           # normal value

_res = AE.read_results(FLAT, price=315.0)
ok(CK.tie_check(_res)[0], f"the flat zero-inflation build still ties ({_res['max_identity_tie']:.1e})")
ok(V.cell(56, 3).value == 0 and V.cell(57, 3).value == 1,
   "expected inflation is zero and the index is 1, so nominal equals real")

near(A1, 1.0 / r_lr, 1e-12,
     "the curve-implied annuity factor collapses to exactly 1/r on a flat curve")
near(B43, nEPS1 / r_lr, 1e-9,
     "NORMAL VALUE EQUALS FORWARD NORMAL EARNINGS CAPITALISED — the anchor is forward")
ok(abs(B43 - EPS0 / r_lr) > 1e-6,
   f"and it is NOT trailing earnings capitalised ({EPS0 / r_lr:.6f})")
near((nEPS1 - EPS0) / r_lr, ret0, 1e-9,
     "the forward-versus-trailing gap is EXACTLY the anchor year's retained earnings — "
     "the same result the dividend discount model gives, now measured on the live engine")
print(f"     [measured] forward {B43:.6f} vs trailing {EPS0 / r_lr:.6f}; "
      f"gap {B43 - EPS0 / r_lr:.6f} against retained earnings per share {ret0:.6f}")

# ------------------------------------------------------- 4. the schedule's discount column
print("== the schedule must expose the factor its own contributions are built with ==")
_dv = openpyxl.load_workbook(BASE, data_only=True)["Valuation"]
_dfE = _dv.cell(16, 3).value        # DF^E, the equity discount factor
_dfF = _dv.cell(18, 3).value        # DF^F, the operating/blended factor
_aE = _dv.cell(62, 3).value
_aeg = _dv.cell(23, 3).value
_con = _dv.cell(24, 3).value
near(_con, _aeg * _aE, 1e-9,
     "the equity contribution is abnormal growth times the EQUITY annuity factor")
ok(abs(_dfE - _dfF) > 1e-9,
   f"the equity and operating discount factors genuinely differ (DF^E {_dfE:.9f} vs "
   f"DF^F {_dfF:.9f}) — so emitting the wrong one in the schedule is not harmless")
import aeg_schedule as SCH                                   # noqa: E402
ok("df_equity" in SCH.PERYEAR_ROWS and SCH.PERYEAR_ROWS["df_equity"] == 16,
   "the schedule emits the equity discount factor (Valuation row 16)")
ok("annuity_equity" in SCH.PERYEAR_ROWS and SCH.PERYEAR_ROWS["annuity_equity"] == 62,
   "the schedule emits the equity annuity factor (Valuation row 62)")
ok("dri_equity" in SCH.PERYEAR_ROWS and SCH.PERYEAR_ROWS["dri_equity"] == 63,
   "the schedule emits the residual-income increment (Valuation row 63) the contribution "
   "is actually built from")
# The reconciliation must close for EVERY explicit year, not just year one. Raw abnormal
# growth equals the residual-income increment only at t=1; using it beyond that is the
# mistake the emitted columns now make impossible.
import csv as _csv                                           # noqa: E402
SCH.write_aeg_schedule(BASE, "AAPL", WORK)
with open(os.path.join(WORK, "AAPL_aeg_schedule.csv"), newline="") as _fh:
    _rows = list(_csv.DictReader(_fh))
_bad, _explicit = [], 0
for _r in _rows:
    if _r["phase"] != "explicit":
        continue
    _explicit += 1
    _dri = float(_r["dri_equity"]); _ann = float(_r["annuity_equity"])
    _con = float(_r["contrib_eps"])
    if abs(_con - _dri * _ann) > 1e-9:
        _bad.append((_r["t"], _con, _dri * _ann))
ok(_explicit >= 4, f"the schedule carries the explicit forecast years ({_explicit})")
ok(not _bad,
   f"contrib_eps = dri_equity x annuity_equity for EVERY explicit year, so the file "
   f"reconciles by hand (mismatches: {_bad})")
# and demonstrate that the naive reconciliation, which the old columns invited, fails
_r2 = [r for r in _rows if r["phase"] == "explicit"][1]
ok(abs(float(_r2["contrib_eps"]) - float(_r2["aeg_eps"]) * float(_r2["annuity_equity"])) > 1e-6,
   "raw abnormal growth times the annuity factor does NOT reconcile beyond year one — "
   "which is exactly why the increment had to be emitted")

print(f"\n{_p} passed, {_f} failed")
raise SystemExit(1 if _f else 0)

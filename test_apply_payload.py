#!/usr/bin/env python3
"""test_apply_payload.py — S5, and the highest-priority of the three.

`apply_payload.py` had no automated coverage, triaged as low risk because it is
on-demand rather than every-build. That reasoning does not survive contact with P1 and
P2: apply_payload is the ONLY writer of cfg_N (Inputs B26) and the only writer of the
dividend payout seed (Inputs B39) outside the build itself. It is the single code path
through which the two most powerful economic judgments in the model can be set from the
cockpit at all, and it was the one module with no test.

The module's central safety claim is that a driver absent from the payload is not
touched, so a payload-free or partial run stays bit-identical to the legacy path. That
claim is asserted in the docstring; it is checked here.
"""
import os
import sys

_ROOT = os.path.dirname(os.path.abspath(__file__))
for _p in (_ROOT, os.path.join(_ROOT, "pipeline")):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import openpyxl                                              # noqa: E402
import aeg_engine as AE                                      # noqa: E402
import checks as CK                                          # noqa: E402
import apply_payload as AP                                   # noqa: E402
import import_forecast as IF                                 # noqa: E402
from recalc_lo import recalc                                 # noqa: E402

GOLDEN = os.path.join(_ROOT, "tests", "golden", "AAPL")
TEMPLATE = os.path.join(_ROOT, "MODEL_TEMPLATE.xlsx")
WORK = os.environ.get("AAPL_ENG_WORK") or "/tmp/_apply_payload_work"
os.makedirs(WORK, exist_ok=True)

_p = _f = 0


def ok(cond, msg):
    global _p, _f
    if cond:
        _p += 1
        print("  PASS", msg)
    else:
        _f += 1
        print("  FAIL", msg)


def expect_error(fn, needle, msg):
    global _p, _f
    try:
        fn()
        _f += 1
        print("  FAIL", msg, "(no error)")
    except AP.PayloadError as e:
        if needle.lower() in str(e).lower():
            _p += 1
            print("  PASS", msg)
        else:
            _f += 1
            print("  FAIL", msg, f"(wrong error: {e})")


CFG = {"company": "Apple Inc.", "ticker": "AAPL", "price": 315.0, "fy_end_month": 9,
       "forecast_horizon_N": 4,
       "files": {"is_csv": f"{GOLDEN}/REAL_IS.csv", "bs_csv": f"{GOLDEN}/REAL_BS.csv",
                 "cf_csv": f"{GOLDEN}/REAL_CF.csv", "prices": f"{GOLDEN}/REAL_prices.csv",
                 "dividends": f"{GOLDEN}/REAL_div.csv", "splits": f"{GOLDEN}/REAL_splits.csv"},
       "judgments": {"minority_include": False, "finlease": 0.0, "oi_adj_override": None,
                     "rd_capitalize": True, "rd_life": 5.0, "dps_override": None},
       "cost_of_debt": {"single_ytw": 0.05}}

print("== build + recalc the golden engine ==")
ENG = os.path.join(WORK, "payload_base.xlsx")
AE.build_model(CFG, TEMPLATE, ENG)
recalc(ENG)
BASE_RES = AE.read_results(ENG, price=315.0)
ok(CK.tie_check(BASE_RES)[0], f"base engine ties ({BASE_RES['max_identity_tie']:.1e})")

VALS = openpyxl.load_workbook(ENG, data_only=True)
INFL = AP.engine_inflation(VALS, 30)
ok(len(INFL) == 30 and all(isinstance(x, float) for x in INFL),
   "engine_inflation reads 30 forward-inflation values off the engine's own Market Data tab")

# snapshot every forecast-tab formula so "untouched" can be proved, not assumed
def snapshot(path):
    F = openpyxl.load_workbook(path, data_only=False)["Forecast"]
    return {(r, c): F.cell(r, c).value
            for r in range(1, 60) for c in range(AP.COL_P1, AP.COL_P1 + AP.MAX_N)}


BEFORE = snapshot(ENG)

print("== validation is fail-loud, and fails BEFORE writing anything ==")
GOOD = {"ticker": "AAPL", "mode": "Equity", "N": 4}
expect_error(lambda: AP.validate_payload({"mode": "Equity", "N": 4}), "ticker",
             "missing ticker aborts")
expect_error(lambda: AP.validate_payload({**GOOD, "mode": "Sideways"}), "mode",
             "unknown mode aborts")
expect_error(lambda: AP.validate_payload({**GOOD, "N": 0}), "'N'", "N=0 aborts")
expect_error(lambda: AP.validate_payload({**GOOD, "N": 31}), "'N'", "N=31 aborts")
expect_error(lambda: AP.validate_payload({**GOOD, "N": True}), "'N'",
             "a boolean N aborts (bool is a Python int; the guard must exclude it)")
expect_error(lambda: AP.validate_payload({**GOOD, "drivers": {"vibes": [0.1] * 4}}),
             "unknown driver", "unknown driver aborts")
expect_error(lambda: AP.validate_payload({**GOOD, "drivers": {"tax_rate": [0.2] * 3}}),
             "exactly n=4", "wrong-length driver aborts")
expect_error(lambda: AP.validate_payload({**GOOD, "drivers": {"tax_rate": [0.2, 0.2, "x", 0.2]}}),
             "not numeric", "non-numeric driver value aborts")
_hi = IF.DRIVER_SPEC["tax_rate"]["hi"]
expect_error(lambda: AP.validate_payload({**GOOD, "drivers": {"tax_rate": [_hi + 1] * 4}}),
             "out of range", "out-of-range driver aborts")
expect_error(lambda: AP.validate_payload({**GOOD, "singles": {"nonsense": 1.0}}),
             "unknown single", "unknown single aborts")
expect_error(lambda: AP.validate_payload("not a dict"), "json object",
             "a non-object payload aborts")
# nothing above should have touched the workbook
ok(snapshot(ENG) == BEFORE, "no rejected payload wrote a single cell")

print("== absent drivers are not touched (the module's core safety claim) ==")
W1 = os.path.join(WORK, "partial.xlsx")
wb = openpyxl.load_workbook(ENG)
rep = AP.apply_payload(wb, {**GOOD, "drivers": {"tax_rate": [0.21, 0.21, 0.21, 0.21]}}, INFL)
wb.save(W1)
AFTER = snapshot(W1)
_touched = {k for k in BEFORE if BEFORE[k] != AFTER[k]}
_tax_cells = {(AP.DRIVER_ROWS["tax_rate"], AP.COL_P1 + t) for t in range(4)}
ok(_touched == _tax_cells,
   f"exactly the four supplied tax-rate cells changed (changed: {sorted(_touched)})")
ok(sorted(rep["held_at_anchor"]) == sorted(set(AP.DRIVER_ROWS) - {"tax_rate"}),
   "every other driver is reported held at anchor")
_F = openpyxl.load_workbook(W1)["Forecast"]
ok(_F.cell(AP.DRIVER_ROWS["tax_rate"], AP.COL_P1 + 4).value == "=$F$17",
   "year 5 of a driver written for N=4 keeps its anchor-hold formula")

print("== cfg_N and the canonical closure: the first-order judgments ==")
# CONTRACT CHANGE 2026-08-10: the operating closure is canonical. cfg_mode is forced to
# Enterprise regardless of what the cockpit sends, and the payout seed is REJECTED because
# it is inert there -- distributions are implied. See apply_payload.CANONICAL_MODE.
W2 = os.path.join(WORK, "policy.xlsx")
wb = openpyxl.load_workbook(ENG)
AP.apply_payload(wb, {"ticker": "AAPL", "mode": "Equity", "N": 9,
                      "singles": {"target_flev": 0.55}}, INFL)
wb.save(W2)
_wb2 = openpyxl.load_workbook(W2)
ok(_wb2["Inputs"]["B26"].value == 9, "payload N lands on cfg_N (Inputs B26)")
ok(_wb2["Inputs"]["B37"].value == "Enterprise",
   "a payload asking for Equity is forced to the canonical operating closure")
_f51 = _wb2["Forecast"].cell(AP.SINGLE_ROWS["target_flev"], AP.COL_P1).value
ok(_f51 == 0.55, "payload target_flev lands on the PERIOD leverage cell, not the anchor")
expect_error(lambda: AP.apply_payload(openpyxl.load_workbook(ENG),
                                      {"ticker": "AAPL", "mode": "Enterprise", "N": 4,
                                       "singles": {"payout": 0.44}}, INFL),
             "cannot be set under the canonical operating closure",
             "a payout seed is rejected loudly rather than silently ignored")
# a payload horizon must actually be honoured by the valuation, not just written
recalc(W2)
_r2 = AE.read_results(W2, price=315.0)
ok(CK.tie_check(_r2)[0], f"the engine still ties after a payload ({_r2['max_identity_tie']:.1e})")
# Change ONE judgment at a time. The combined payload above moves the horizon and the
# leverage together, so a combined test could pass or fail for reasons that have nothing
# to do with whether the payload was honoured. The leverage leg is also the regression
# test for the 2026-08-10 silent-ignore bug: before that fix, Forecast row 25 read the
# ANCHOR leverage cell in both default branches, so a submitted target_flev was written
# and never read, and this assertion would fail with the valuation completely unmoved.
for _label, _pl in (("horizon", {"ticker": "AAPL", "mode": "Equity", "N": 12}),
                    ("leverage", {"ticker": "AAPL", "mode": "Equity", "N": 4,
                                  "singles": {"target_flev": 2.0}})):
    _w = os.path.join(WORK, f"iso_{_label}.xlsx")
    _wb = openpyxl.load_workbook(ENG)
    AP.apply_payload(_wb, _pl, INFL)
    _wb.save(_w)
    recalc(_w)
    _ri = AE.read_results(_w, price=315.0)
    ok(CK.tie_check(_ri)[0], f"{_label}-only payload still ties ({_ri['max_identity_tie']:.1e})")
    ok(abs(_ri["equity_value"] - BASE_RES["equity_value"]) > 1.0,
       f"{_label}-only payload moves the valuation "
       f"({BASE_RES['equity_value']:.4f} -> {_ri['equity_value']:.4f})")
_spec = IF.SINGLES["payout"]
expect_error(lambda: AP.validate_payload({**GOOD, "singles": {"payout": _spec["hi"] + 1}}),
             "out of range", "an out-of-range payout single aborts")

print("== nominal drivers are deflated with the engine's own inflation ==")
W3 = os.path.join(WORK, "deflate.xlsx")
wb = openpyxl.load_workbook(ENG)
_nom = [0.06, 0.05, 0.04, 0.03]
AP.apply_payload(wb, {**GOOD, "drivers": {"revenue_growth": _nom}}, INFL)
wb.save(W3)
_F3 = openpyxl.load_workbook(W3)["Forecast"]
_row = AP.DRIVER_ROWS["revenue_growth"]
_bad = []
for t in range(4):
    want = (1 + _nom[t]) / (1 + INFL[t]) - 1
    got = _F3.cell(_row, AP.COL_P1 + t).value
    if abs(got - want) > 1e-15:
        _bad.append((t + 1, got, want))
ok(not _bad, f"revenue growth deflated exactly year by year (mismatches: {_bad})")
ok(IF.DRIVER_SPEC["revenue_growth"]["deflate"] is True and
   IF.DRIVER_SPEC["tax_rate"]["deflate"] is False,
   "only the growth drivers carry the deflate flag; ratios pass through")
_F1 = openpyxl.load_workbook(W1)["Forecast"]
ok(_F1.cell(AP.DRIVER_ROWS["tax_rate"], AP.COL_P1).value == 0.21,
   "a unit-free ratio is written through unchanged")

print("== inflation guard ==")
expect_error(lambda: AP.apply_payload(openpyxl.load_workbook(ENG), {**GOOD, "N": 4}, INFL[:2]),
             "inflation series", "too-short inflation series aborts")
expect_error(lambda: AP.engine_inflation(openpyxl.load_workbook(ENG, data_only=False), 4),
             "recalc", "an un-recalculated workbook aborts with a message naming recalc")

print(f"\n{_p} passed, {_f} failed")
raise SystemExit(1 if _f else 0)

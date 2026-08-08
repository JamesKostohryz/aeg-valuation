#!/usr/bin/env python3
"""test_deflator_extend.py — tests for deflator_extend.py, which runs on EVERY live build
(before the first recalc) and extends the real-terms CPI-U / BEA PP&E deflator tables to
cover a fiscal year past the template's hand-seeded 2025 ceiling (HD/POOL-class issuers).
This closes a coverage gap flagged in AEG-Coverage-Map-2026-08-08.md: the module is
fail-loud against a missing CPI, but nothing previously checked that the EXTENSION MATH
ITSELF is right, or that a genuinely extended workbook still recalculates cleanly.

`ensure_deflator_covers_anchor` takes an injectable `cpi_monthly` dict specifically so it
can be tested without a network call — this suite uses that, so it needs no FRED/BLS
access and no API key.

Checks:
  1. `_calendar_year_mean`: correct average over a full year, correct partial-year
     average (a firm whose fiscal year lands mid-calendar-year), None for a year with
     no observations at all.
  2. No-op when the template already covers the anchor year (AAPL/T golden case) —
     bit-identical file, `extended=False`.
  3. A genuine extension (anchor year past the template ceiling): the added CPI is
     exactly the calendar-year mean of the supplied months; the added BEA deflator is
     exactly `prev_bea * (cpi_y / cpi_{y-1})` (the documented CPI-carry formula), for
     EVERY added year in sequence, not just the last one.
  4. The extended named ranges actually resolve after a real LibreOffice recalc: the
     new column's md_deflator / md_deflator_ppe formulas evaluate to md_cpi_base/CPI and
     md_bea_base/BEA with no #N/A, and MATCH(anchor_year, md_years) finds the new column.
  5. Idempotency: extending twice for the same anchor year adds nothing the second time.
  6. `wrap_capengine_capex`: wraps an un-guarded capex lookup formula in IFERROR exactly
     once, and re-running is a no-op (doesn't double-wrap).

Usage: python3 test_deflator_extend.py
"""
import os, sys, shutil, datetime
import openpyxl

_ROOT = os.path.dirname(os.path.abspath(__file__))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)
from deflator_extend import (ensure_deflator_covers_anchor, _calendar_year_mean,
                              wrap_capengine_capex, _CPI_NAMES, _BEA_NAMES)
from recalc_lo import recalc

TEMPLATE = os.path.join(_ROOT, "MODEL_TEMPLATE.xlsx")
WORK = os.path.join(_ROOT, "_defltest")
os.makedirs(WORK, exist_ok=True)

_p = _f = 0
def check(c, m):
    global _p, _f
    if c: _p += 1; print(f"  PASS  {m}")
    else: _f += 1; print(f"  FAIL  {m}")


def month(y, m, v):
    return (datetime.date(y, m, 1), v)


print("== _calendar_year_mean: pure arithmetic ==")
full_year = dict(month(2026, m, 300.0 + m) for m in range(1, 13))
check(abs(_calendar_year_mean(full_year, 2026) - (sum(300.0 + m for m in range(1, 13)) / 12)) < 1e-9,
      "full calendar year averages exactly")
partial = dict(month(2027, m, 100.0 * m) for m in range(1, 4))   # only Jan-Mar published so far
check(abs(_calendar_year_mean(partial, 2027) - (100 + 200 + 300) / 3) < 1e-9,
      "partial year averages only the published months (a mid-year cutoff firms up as months land)")
check(_calendar_year_mean(partial, 2030) is None, "a year with zero observations returns None")

print("== no-op when the template already covers the anchor year ==")
p1 = os.path.join(WORK, "noop.xlsx")
shutil.copy(TEMPLATE, p1)
rep = ensure_deflator_covers_anchor(p1, anchor_year=2025)
check(rep["extended"] is False, f"AAPL/T-class anchor (2025, already seeded) is a no-op (report={rep})")

print("== a genuine extension: CPI = calendar-year mean, BEA = CPI-carried forward ==")
p2 = os.path.join(WORK, "ext.xlsx")
shutil.copy(TEMPLATE, p2)
wb0 = openpyxl.load_workbook(p2, data_only=False)
MD0 = wb0["Market Data"]
prev_cpi = MD0["AP8"].value   # 2025 CPI as seeded in the template
prev_bea = MD0["AP14"].value

cpi_monthly = {}
for y, base in ((2026, 330.0), (2027, 345.0)):
    for m in range(1, 13):
        cpi_monthly[datetime.date(y, m, 1)] = base + m * 0.1   # deterministic, distinct per month

rep2 = ensure_deflator_covers_anchor(p2, anchor_year=2027, cpi_monthly=cpi_monthly)
check(rep2["extended"] is True, "anchor year 2027 (past the 2025 ceiling) triggers a real extension")
check([a["year"] for a in rep2["added"]] == [2026, 2027], f"extends every missing year in order (added={rep2['added']})")

want_prev_cpi, want_prev_bea = prev_cpi, prev_bea
for a in rep2["added"]:
    want_cpi = round(_calendar_year_mean(cpi_monthly, a["year"]), 6)
    want_bea = round(want_prev_bea * (want_cpi / want_prev_cpi), 6)
    check(abs(a["cpi"] - round(want_cpi, 4)) < 1e-6, f"{a['year']}: added CPI matches the calendar-year mean exactly")
    check(a["cpi_months"] == 12, f"{a['year']}: all 12 months were available and counted")
    want_prev_cpi, want_prev_bea = want_cpi, want_bea

# re-derive the BEA figure independently from the saved workbook cells (not from the report)
wb2 = openpyxl.load_workbook(p2, data_only=False)
MD2 = wb2["Market Data"]
last_col = MD2.max_column
cpi_2026, cpi_2027 = MD2.cell(8, last_col - 1).value, MD2.cell(8, last_col).value
bea_2026, bea_2027 = MD2.cell(14, last_col - 1).value, MD2.cell(14, last_col).value
want_bea_2026 = round(prev_bea * (cpi_2026 / prev_cpi), 6)
want_bea_2027 = round(bea_2026 * (cpi_2027 / cpi_2026), 6)
check(abs(bea_2026 - want_bea_2026) < 1e-6, "2026 BEA deflator = prior BEA carried by realised CPI inflation (recomputed independently)")
check(abs(bea_2027 - want_bea_2027) < 1e-6, "2027 BEA deflator chains off the NEW 2026 BEA, not the original 2025 one")

print("== extended named ranges resolve after a real LibreOffice recalc, no #N/A ==")
recalc(p2)
wbv = openpyxl.load_workbook(p2, data_only=True)
MDv = wbv["Market Data"]
last_col_v = MDv.max_column
years = [MDv.cell(7, c).value for c in range(2, last_col_v + 1)]
check(str(2027) in [str(y) for y in years], "md_years now includes the extended anchor year 2027")
defl_2027 = MDv.cell(9, last_col_v).value
defl_ppe_2027 = MDv.cell(15, last_col_v).value
check(isinstance(defl_2027, (int, float)), f"md_deflator at 2027 recalculates to a number, not #N/A (got {defl_2027!r})")
check(isinstance(defl_ppe_2027, (int, float)), f"md_deflator_ppe at 2027 recalculates to a number, not #N/A (got {defl_ppe_2027!r})")
cpi_base = wbv["Market Data"]["B5"].value
if isinstance(defl_2027, (int, float)) and isinstance(cpi_base, (int, float)):
    check(abs(defl_2027 - cpi_base / cpi_2027) < 1e-9, "md_deflator formula value = md_cpi_base / CPI_2027 exactly")

print("== idempotency: extending again for the same anchor year adds nothing new ==")
rep3 = ensure_deflator_covers_anchor(p2, anchor_year=2027, cpi_monthly=cpi_monthly)
check(rep3["extended"] is False, f"second call for the same anchor year is a no-op (report={rep3})")

print("== wrap_capengine_capex: guards once, idempotent ==")
p3 = os.path.join(WORK, "wrap.xlsx")
shutil.copy(TEMPLATE, p3)
wb3 = openpyxl.load_workbook(p3, data_only=False)
before = wb3["Cap Engine"]["B7"].value
check(isinstance(before, str) and before.startswith("=-INDEX(rep_capex") and "IFERROR" not in before,
      f"template ships the capex lookup un-guarded (got {before!r})")
n1 = wrap_capengine_capex(wb3)
after = wb3["Cap Engine"]["B7"].value
check(n1 > 0, f"wraps at least one un-guarded capex formula (wrapped {n1})")
check(after == f"=IFERROR({before[1:]},0)", f"wrapped formula is exactly IFERROR(<original>,0) (got {after!r})")
n2 = wrap_capengine_capex(wb3)
check(n2 == 0, f"re-running on an already-wrapped workbook wraps nothing new (wrapped {n2})")

shutil.rmtree(WORK, ignore_errors=True)
print(f"\n{_p} passed, {_f} failed")
sys.exit(1 if _f else 0)

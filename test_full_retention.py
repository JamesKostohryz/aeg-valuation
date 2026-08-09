#!/usr/bin/env python3
"""test_full_retention.py — Phase 1, Property 5: the degenerate retention cases.

Property 5 asks the engine to hold up at both ends of the retention range, where the usual
intuitions stop working.

ZERO RETENTION is already covered: test_zero_growth.py runs the payout=1.00 case, where
normal growth is exactly zero and value collapses to a flat perpetuity. This file adds the
other end.

FULL RETENTION — payout zero, every dollar retained, no dividend ever paid. It is the case
that breaks the dividend discount model outright: with no dividends there is nothing to
discount, the Gordon tail divides by (r - g) = 0, and a naive reading says the shares are
worthless. The abnormal-earnings-growth form must still return forward normal earnings
capitalized, because value comes from the book value the retained earnings build, not from
the cash handed out. If the engine cannot do that, it is not implementing the framework.

THE CLOSED FORM, computed outside the model. With every abnormal earnings growth term zeroed
and full retention:

    g       = r * b = r * 1 = r        (retained capital earns exactly the cost of equity)
    EPS_t   = EPS_1 * (1 + r)^(t-1)
    value   = EPS_1 / r                (unchanged — abnormal growth is what is absent)

Note what this says: earnings compound at the full cost of equity forever and the value is
STILL just E1/r. Growth funded at the cost of capital adds nothing. That is the framework's
central claim, and full retention is where it is most visibly true.

MECHANISM. The same one test_zero_growth.py established: overwrite real net income on
Forecast row 19 for the explicit years and re-point operating income on row 16 to `NI + NFE`
so the operating leg stays exactly consistent. Every other formula is untouched, which is why
the four-method tie stays green and is asserted.

Needs LibreOffice.
"""
import os
import shutil
import sys

_ROOT = os.path.dirname(os.path.abspath(__file__))
for _p_ in (_ROOT, os.path.join(_ROOT, "pipeline")):
    if _p_ not in sys.path:
        sys.path.insert(0, _p_)

import openpyxl                                              # noqa: E402
import aeg_engine as AE                                      # noqa: E402
import checks as CK                                          # noqa: E402
from recalc_lo import recalc                                 # noqa: E402

GOLDEN = os.path.join(_ROOT, "tests", "golden", "AAPL")
TEMPLATE = os.path.join(_ROOT, "MODEL_TEMPLATE.xlsx")
WORK = os.environ.get("AAPL_ENG_WORK") or "/tmp/_full_retention_work"
PRICE = 315.0

FC_COL = lambda t: 6 + t          # noqa: E731
VL_COL = lambda t: 2 + t          # noqa: E731

CFG = {"company": "Apple Inc.", "ticker": "AAPL", "price": PRICE, "fy_end_month": 9,
       "forecast_horizon_N": 4,
       "files": {"is_csv": f"{GOLDEN}/REAL_IS.csv", "bs_csv": f"{GOLDEN}/REAL_BS.csv",
                 "cf_csv": f"{GOLDEN}/REAL_CF.csv", "prices": f"{GOLDEN}/REAL_prices.csv",
                 "dividends": f"{GOLDEN}/REAL_div.csv", "splits": f"{GOLDEN}/REAL_splits.csv"},
       "judgments": {"minority_include": False, "finlease": 0.0, "oi_adj_override": None,
                     "rd_capitalize": True, "rd_life": 5.0, "dps_override": None},
       "cost_of_debt": {"single_ytw": 0.05}}

_p = _f = 0


def ok(cond, msg):
    global _p, _f
    if cond:
        _p += 1
        print("  PASS", msg)
    else:
        _f += 1
        print("  FAIL", msg)


def close(a, b, tol, msg):
    got = (isinstance(a, (int, float)) and isinstance(b, (int, float))
           and abs(a - b) <= tol * max(1.0, abs(b)))
    ok(got, f"{msg}  ({a!r} vs {b!r}, rel tol {tol:g})")


os.makedirs(WORK, exist_ok=True)
BASE = os.path.join(WORK, "base.xlsx")
AE.build_model(CFG, TEMPLATE, BASE)
recalc(BASE)
d0 = openpyxl.load_workbook(BASE, data_only=True)
V0 = d0["Valuation"]
EPS0, RET0 = V0.cell(7, 2).value, V0.cell(9, 2).value
SHARES0 = d0["Inputs"]["B9"].value
print(f"== anchor: EPS0 {EPS0:.9f}, retained0 {RET0:.9f}, "
      f"anchor retention {RET0 / EPS0:.6f} ==\n")


def run(tag, N, r):
    """Full retention: payout seed zero, flat real curve, zero expected inflation."""
    eps = [EPS0, EPS0 + r * RET0]                 # year one grows on the ANCHOR's retention
    for t in range(2, 31):
        eps.append(eps[t - 1] * (1.0 + r))        # thereafter b = 1, so g = r
    path = os.path.join(WORK, f"{tag}.xlsx")
    shutil.copy(BASE, path)
    wb = openpyxl.load_workbook(path)
    IN, MD, F = wb["Inputs"], wb["Market Data"], wb["Forecast"]
    IN["B37"], IN["B29"], IN["B26"], IN["B39"] = "Equity", "Single", N, 0.0
    for c in range(2, 32):
        MD.cell(22, c).value = 0.0                # zero expected inflation
        MD.cell(26, c).value = r                  # flat real cost of equity
    for t in range(1, N + 1):
        col = F.cell(19, FC_COL(t)).column_letter
        F.cell(19, FC_COL(t)).value = eps[t] * SHARES0
        F.cell(16, FC_COL(t)).value = f"={col}19+{col}18"
    wb.save(path)
    recalc(path)
    d = openpyxl.load_workbook(path, data_only=True)
    V = d["Valuation"]
    res = AE.read_results(path, price=PRICE)
    return {"eps": eps, "A1": V.cell(62, 3).value, "EPS1": V.cell(7, 3).value,
            "DPS1": V.cell(8, 3).value,
            "value": V.cell(36, 2).value, "normal": V.cell(43, 2).value,
            "intrinsic": V.cell(44, 2).value,
            "aeg": [V.cell(23, VL_COL(t)).value for t in range(1, N + 1)],
            "tie": res["max_identity_tie"], "audit": res["audit_status"],
            "tie_ok": CK.tie_check(res)[0]}


for tag, N, r in (("full_ret_r0665_N4", 4, 0.0665),
                  ("full_ret_r09_N12", 12, 0.09),
                  ("full_ret_r05_N30", 30, 0.05)):
    print(f"== {tag}: N={N}, cost of equity {r:.4%}, payout 0.0 ==")
    m = run(tag, N, r)
    eps = m["eps"]

    close(m["A1"], 1.0 / r, 1e-12,
          "the curve-implied annuity factor collapses to 1/r on a flat curve")
    close(m["EPS1"], eps[1], 1e-12, "the engine is running the path this file constructed")
    ok(isinstance(m["DPS1"], (int, float)) and abs(m["DPS1"]) <= 1e-12,
       f"no dividend is paid in any year (year one DPS {m['DPS1']!r})")

    worst = max(abs(a) for a in m["aeg"])
    ok(worst <= 1e-12 * max(1.0, abs(eps[1])),
       f"abnormal earnings growth is zero at every explicit year (worst {worst:.3e})")

    # THE CLOSED FORM. Earnings compound at the full cost of equity and value is still E1/r.
    close(m["value"], eps[1] / r, 1e-9,
          "VALUE EQUALS FORWARD NORMAL EARNINGS OVER THE COST OF EQUITY, with no dividend "
          "ever paid")
    close(m["normal"], m["value"], 1e-12, "the whole value is the normal value")
    close(m["intrinsic"], m["value"], 1e-12, "the published readout agrees")

    # And the point of the case: compounding at r for N years added nothing.
    grew = eps[N] / eps[1]
    ok(grew > 1.05, f"earnings really did compound over the horizon (x{grew:.4f})")
    close(m["value"], eps[1] / r, 1e-9,
          f"yet the value is unchanged by that growth — reinvestment at the cost of capital "
          f"creates no value")

    ok(m["tie_ok"] and m["audit"].startswith("PASS"),
       f"and the four-method tie still holds ({m['tie']:.1e})")
    print(f"     [measured] value {m['value']:.9f} = E1/r {eps[1] / r:.9f}\n")

if _f:
    print(f"FAIL  test_full_retention.py  ({_p} passed, {_f} failed)")
    sys.exit(1)
print(f"{_p} passed, 0 failed")

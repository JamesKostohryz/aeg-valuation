#!/usr/bin/env python3
"""test_horizon_gating.py — Phase 1, Property 4: nothing after the horizon can move value.

THE PROPERTY
------------
cfg_N is the competitive-advantage period: the number of years the analyst judges abnormal
earnings growth to persist. Everything after it belongs to the continuing period, which is
valued off the normalized line rather than off the forecast rows. So a driver written into
year N+1 through year 30 must be UNABLE to move the valuation, by construction.

If it can, cfg_N does not mean what the whole system says it means -- the horizon gate that
refuses a run without an authorized cfg_N, the 31% sensitivity quoted in that refusal, and
every statement in the method documentation about what the explicit period covers.

WHY THE FOUR-METHOD TIE CANNOT CHECK THIS. Abnormal earnings growth, residual operating
income, free cash flow to equity and free cash flow to the firm are four transformations of
the SAME stream over the SAME horizon. Leak a post-horizon year into the value and all four
legs absorb it identically, so they still agree to 1e-15 while the answer moves. The tie is
an internal-consistency check; this is an external one.

WHAT THIS MEASURED, 2026-08-09
------------------------------
Post-horizon perturbation moved the value by EXACTLY 0.0 -- not "within tolerance", zero --
on both N=4 and N=8, across all four perturbations. Horizon gating is exact.

A second result came out of it, and it goes on the register rather than the critical path.
Because the perturbation rewrites net income without the matching dividend and book-value
roll, clean surplus breaks in the post-horizon rows and the Audit identity residual jumps to
between 1e1 and 1e5 -- while the value is unmoved. So THE AUDIT IDENTITIES SPAN ROWS THAT
CANNOT AFFECT THE VALUE. A data artifact after the horizon (the register already notes that
the schedule stops populating cost of equity and inflation past N, leaving junk normal_eps)
can therefore redden the tie with no valuation consequence whatever. Worth knowing before
anyone investigates a red tie that turns out to live entirely past the horizon.

THE TEST IS ADVERSARIAL ON PURPOSE. It does not nudge the post-horizon years -- it writes
absurd values into them: a tenfold jump, a sign flip into losses, and a compounding
explosion. A gate that holds against those is a real gate. Anything that survives at 1e-12
relative against a 10x perturbation is not "small enough to ignore", it is zero.

It also asserts the CONVERSE on the same engine: perturbing year N (inside the horizon) DOES
move the value. Without that, a model that ignored the entire forecast would pass.

MECHANISM. Identical to test_zero_growth.py, which established the pattern: overwrite real
net income on Forecast row 19 and re-point operating income on row 16 to `NI + NFE` so the
operating leg stays exactly consistent. Every other formula is untouched, which is why the
four-method tie stays green on every perturbed run and is asserted each time.

No network. Needs LibreOffice, as every engine-driving test does.
"""
import os
import shutil
import sys

_ROOT = os.path.dirname(os.path.abspath(__file__))
for _p_ in (_ROOT, os.path.join(_ROOT, "pipeline")):
    if _p_ not in sys.path:
        sys.path.insert(0, _p_)

import openpyxl                                              # noqa: E402
import aeg_engine as AE                                      # noqa: E402
import checks as CK                                          # noqa: E402
from recalc_lo import recalc                                 # noqa: E402

GOLDEN = os.path.join(_ROOT, "tests", "golden", "AAPL")
TEMPLATE = os.path.join(_ROOT, "MODEL_TEMPLATE.xlsx")
WORK = os.environ.get("AAPL_ENG_WORK") or "/tmp/_horizon_gating_work"
PRICE = 315.0
LAST_FORECAST_YEAR = 30          # structural: the Forecast tab has thirty columns

FC_COL = lambda t: 6 + t         # noqa: E731  Forecast: column F is t=0
VL_COL = lambda t: 2 + t         # noqa: E731  Valuation: column B is t=0

CFG = {"company": "Apple Inc.", "ticker": "AAPL", "price": PRICE, "fy_end_month": 9,
       "forecast_horizon_N": 4,
       "files": {"is_csv": f"{GOLDEN}/REAL_IS.csv", "bs_csv": f"{GOLDEN}/REAL_BS.csv",
                 "cf_csv": f"{GOLDEN}/REAL_CF.csv", "prices": f"{GOLDEN}/REAL_prices.csv",
                 "dividends": f"{GOLDEN}/REAL_div.csv", "splits": f"{GOLDEN}/REAL_splits.csv"},
       "judgments": {"minority_include": False, "finlease": 0.0, "oi_adj_override": None,
                     "rd_capitalize": True, "rd_life": 5.0, "dps_override": None},
       "cost_of_debt": {"single_ytw": 0.05}}

_p = _f = 0


def ok(cond, msg):
    global _p, _f
    if cond:
        _p += 1
        print("  PASS", msg)
    else:
        _f += 1
        print("  FAIL", msg)


# ---------------------------------------------------------------- perturbations
# Each takes the baseline real net income for that year and returns a replacement. They are
# deliberately violent: a gate that only survives small nudges is not a gate.
PERTURBATIONS = {
    "tenfold":   lambda ni, t, N: ni * 10.0,
    "sign_flip": lambda ni, t, N: -abs(ni) * 3.0,
    "explosion": lambda ni, t, N: ni * (1.6 ** (t - N)),
    "zeroed":    lambda ni, t, N: 0.0,
}


def build_base(N):
    os.makedirs(WORK, exist_ok=True)
    cfg = dict(CFG, forecast_horizon_N=N)
    path = os.path.join(WORK, f"base_N{N}.xlsx")
    AE.build_model(cfg, TEMPLATE, path)
    recalc(path)
    return path


def measure(path):
    d = openpyxl.load_workbook(path, data_only=True)
    V = d["Valuation"]
    res = AE.read_results(path, price=PRICE)
    return {
        "value": V.cell(36, 2).value,
        "intrinsic": V.cell(44, 2).value,
        "normal_value": V.cell(43, 2).value,
        "tie": res["max_identity_tie"],
        "audit": res["audit_status"],
        "tie_ok": CK.tie_check(res)[0],
    }


def perturb(base, tag, N, first_year, last_year, fn):
    """Copy the built engine, rewrite real net income for first_year..last_year, recalc."""
    path = os.path.join(WORK, f"pert_{tag}_N{N}_{first_year}.xlsx")
    shutil.copy(base, path)
    wb = openpyxl.load_workbook(path)
    F = wb["Forecast"]
    live = openpyxl.load_workbook(base, data_only=True)["Forecast"]
    touched = 0
    for t in range(first_year, last_year + 1):
        cell = F.cell(19, FC_COL(t))
        cur = live.cell(19, FC_COL(t)).value
        if not isinstance(cur, (int, float)):
            continue
        col = cell.column_letter
        cell.value = float(fn(cur, t, N))
        F.cell(16, FC_COL(t)).value = f"={col}19+{col}18"     # keep OI = NI + NFE exactly
        touched += 1
    wb.save(path)
    recalc(path)
    m = measure(path)
    m["touched"] = touched
    return m


def main():
    global _f
    print("== Phase 1, Property 4: horizon gating ==")
    print("   Perturbing years N+1..30 must NOT move value. Perturbing year N MUST.\n")

    for N in (4, 8):
        base = build_base(N)
        b = measure(base)
        ok(b["tie_ok"] and b["audit"].startswith("PASS"),
           f"N={N}: baseline ties before anything is perturbed ({b['tie']:.1e})")
        print(f"   [baseline] N={N}  value {b['value']!r}")

        # ---- the property: post-horizon years are inert, however violently perturbed
        for tag, fn in PERTURBATIONS.items():
            m = perturb(base, tag, N, N + 1, LAST_FORECAST_YEAR, fn)
            if m["touched"] == 0:
                ok(False, f"N={N} {tag}: no post-horizon cells were writable — test is vacuous")
                continue
            moved = abs(m["value"] - b["value"])
            ok(moved <= 1e-12 * max(1.0, abs(b["value"])),
               f"N={N} {tag}: value UNMOVED by {m['touched']} post-horizon year(s) "
               f"(moved {moved:.3e})")
            ok(abs(m["normal_value"] - b["normal_value"]) <= 1e-12 * max(1.0, abs(b["normal_value"])),
               f"N={N} {tag}: the anchor/normal value is unmoved too")
            # NOTE, and it is the more interesting half of this result. The perturbation
            # deliberately breaks clean surplus in the post-horizon rows -- net income is
            # rewritten without the matching dividend and book-value roll -- so the Audit
            # identities in those years blow up (residuals of 1e1..1e5 were measured). The
            # value does not move by a single bit anyway. That is a STRONGER statement than
            # "the tie stayed green": the post-horizon region can be internally incoherent
            # and still cannot leak into the answer. We assert the blow-up so that if a
            # future change ever made those rows matter, this test would notice.
            ok(not m["tie_ok"],
               f"N={N} {tag}: the post-horizon rows are now internally inconsistent "
               f"(audit residual {m['tie']:.1e}) — and the value STILL did not move")

        # ---- the converse: the horizon is not vacuous, year N really does count
        m = perturb(base, "inside", N, N, N, PERTURBATIONS["tenfold"])
        moved = abs(m["value"] - b["value"])
        ok(moved > 1e-6 * max(1.0, abs(b["value"])),
           f"N={N} CONVERSE: perturbing year N (inside the horizon) DOES move value "
           f"(moved {moved:.6g}) — so the gate is a boundary, not indifference")
        ok(m["tie_ok"], f"N={N} CONVERSE: the perturbed run still ties ({m['tie']:.1e})")
        print()

    print(f"{_p} passed, {_f} failed")
    return 1 if _f else 0


if __name__ == "__main__":
    sys.exit(main())

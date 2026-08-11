#!/usr/bin/env python3
"""test_policy_inputs.py — P1..P5 plus the discarded-dividend fix.

Builds and recalculates the golden Apple engine through real LibreOffice and checks the
five fixes against independently re-derived numbers rather than against the code's own
arithmetic.

What is under test, and why each one exists:

P1  in_payout_seed (Inputs B39). Shipped as a template constant of 0.365 — the base
    company's dividend payout (AT&T: 1.112 / 3.04) — and was never written by the build,
    so every issuer in the system forecast AT&T's dividend policy. Measured on this
    fixture the seed moves intrinsic value from 110.33 at 0.1377 to 130.08 at 0.95.

P2  cfg_N (Inputs B26). The competitive-advantage period. Same defect class: a template
    constant of 4 that nobody chose, worth 31% between horizons 4 and 30. Now required,
    with no default. Choosing 4 deliberately must remain completely ordinary.

P3  in_ppe_life (Inputs B42). AT&T's 18-year composite plant life, inherited by every
    company. Worth about 2.9% on this fixture between 8 and 18 years.

P4  Valuation row 11. The one row that still froze net financial obligations per share at
    the anchor-year real value regardless of cfg_mode, while the Forecast tab grew debt
    with the equity base under a leverage target.

P5  The operations wedge. rho_F is BOOK-weighted, which makes residual operating income
    identical to residual equity income year for year; discounting that identical stream
    at rho_F rather than rho_E then guarantees a gap, measured at 12.0% of equity value
    and excluded from the master gate. Now discounted at a value-weighted rate and gated.

DIV The dividends-file figure resolved by apply_market_data never reached the sheet,
    because write_inputs had already run. The documented precedence "manual > dividends
    file > cash-flow fallback" was defeated by ordering: Apple filed $0.96 and the sheet
    carried the $1.027724 fallback.
"""
import os
import sys

_ROOT = os.path.dirname(os.path.abspath(__file__))
for _p in (_ROOT, os.path.join(_ROOT, "pipeline")):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import openpyxl                                              # noqa: E402
import aeg_engine as AE                                      # noqa: E402
import loader_core as LC                                     # noqa: E402
import checks as CK                                          # noqa: E402
from recalc_lo import recalc                                 # noqa: E402

GOLDEN = os.path.join(_ROOT, "tests", "golden", "AAPL")
TEMPLATE = os.path.join(_ROOT, "MODEL_TEMPLATE.xlsx")
WORK = os.environ.get("AAPL_ENG_WORK") or "/tmp/_policy_inputs_work"
os.makedirs(WORK, exist_ok=True)

_p = _f = 0


def ok(cond, msg):
    global _p, _f
    if cond:
        _p += 1
        print("  PASS", msg)
    else:
        _f += 1
        print("  FAIL", msg)


def near(a, b, tol, msg):
    got = (a is not None and b is not None and abs(a - b) <= tol)
    ok(got, f"{msg}  (got {a!r} vs {b!r}, tol {tol:g})")


def cfg(**over):
    files = {"is_csv": f"{GOLDEN}/REAL_IS.csv", "bs_csv": f"{GOLDEN}/REAL_BS.csv",
             "cf_csv": f"{GOLDEN}/REAL_CF.csv", "prices": f"{GOLDEN}/REAL_prices.csv",
             "dividends": f"{GOLDEN}/REAL_div.csv", "splits": f"{GOLDEN}/REAL_splits.csv"}
    c = {"company": "Apple Inc.", "ticker": "AAPL", "price": 315.0, "files": files,
         "fy_end_month": 9, "forecast_horizon_N": 4,
         "judgments": {"minority_include": False, "finlease": 0.0, "oi_adj_override": None,
                       "rd_capitalize": True, "rd_life": 5.0, "dps_override": None},
         "cost_of_debt": {"single_ytw": 0.05}}
    c["judgments"].update(over.pop("judgments", {}))
    c.update(over)
    return c


def build(name, **over):
    path = os.path.join(WORK, f"{name}.xlsx")
    rep = AE.build_model(cfg(**over), TEMPLATE, path)
    return path, rep


# ---------------------------------------------------------------- build once
print("== building golden AAPL (P1..P5 applied) ==")
ENG, REP = build("base")
recalc(ENG)
VALS = openpyxl.load_workbook(ENG, data_only=True)
FORM = openpyxl.load_workbook(ENG, data_only=False)
V, D, A, I = VALS["Valuation"], VALS["DCF Reconciliation"], VALS["Audit"], VALS["Inputs"]

print("== the engine still ties ==")
RES = AE.read_results(ENG, price=315.0)
_tie_ok, _ = CK.tie_check(RES)
ok(_tie_ok, f"standing tie check passes (tie={RES['max_identity_tie']:.2e})")
ok(RES["audit_status"].startswith("PASS"), f"audit status {RES['audit_status']!r}")

print("== DIV: the dividends file reaches the sheet ==")
import market_data as MDX                                    # noqa: E402
_drows = MDX.parse_dividends(f"{GOLDEN}/REAL_div.csv")
_splits = MDX.parse_splits(f"{GOLDEN}/REAL_splits.csv")
_filed_dps, _comps = MDX.fiscal_year_dps(_drows, _splits, 2025, 9)
near(I["B15"].value, _filed_dps, 1e-12,
     f"Inputs B15 carries the dividends-file figure ({len(_comps)} ex-dates)")
ok(abs(I["B15"].value - 1.027724) > 1e-9,
   "Inputs B15 is NOT the cash-flow-derived fallback that used to win by ordering")

print("== P1: dividend payout seed ==")
seed = I["B39"].value
ok(abs(seed - 0.365) > 1e-9, "payout seed is no longer the template's 0.365")
near(seed, I["B15"].value / I["B13"].value, 1e-6,
     "payout seed re-derives as filed DPS / filed diluted EPS")
ok(0.0 <= seed <= 1.0, f"payout seed in [0,1] for a dividend payer (got {seed})")
near(REP["policy"]["payout_seed"], seed, 1e-12, "build report agrees with the sheet")
# the seed is a DIVIDEND payout: it must not silently absorb buybacks
ok(seed < 0.5, "seed is a dividend payout, not total shareholder distribution")
# override path
_po, _rp = build("payout_override", judgments={"payout_override": 0.42})
ok(openpyxl.load_workbook(_po)["Inputs"]["B39"].value == 0.42, "payout_override wins")
ok(_rp["inputs_register"] and any(r["name"] == "in_payout_seed" and r["provenance"] == "analyst"
                                  for r in _rp["inputs_register"]),
   "an overridden payout is registered as analyst-set, not filings-derived")
# fail-loud on an underivable payout
try:
    LC.resolve_policy_inputs(FORM, {13: {"value": -1.0}, 15: {"value": 0.5}})
    ok(False, "a loss-year payout aborts")
except LC.PolicyInputError as e:
    ok("payout_override" in str(e), "a loss-year payout aborts and names the config key")

print("== P2: cfg_N is required, explicit, and 4 is ordinary ==")
ok(I["B26"].value == 4, "cfg_N landed on the sheet as the chosen 4")
near(REP["forecast_horizon_N"], 4, 0, "build report records the chosen horizon")
for _bad in (None, 0, 31, "four"):
    _c = cfg()
    _c["forecast_horizon_N"] = _bad
    if _bad is None:
        _c.pop("forecast_horizon_N")
    try:
        AE.build_model(_c, TEMPLATE, os.path.join(WORK, "bad.xlsx"))
        ok(False, f"horizon {_bad!r} rejected")
    except ValueError as e:
        ok("forecast_horizon_N" in str(e) or "horizon" in str(e).lower(),
           f"horizon {_bad!r} rejected loudly")
# choosing 4 raises nothing; a different horizon is equally accepted and MOVES the value
_p8, _ = build("h8", forecast_horizon_N=8)
recalc(_p8)
_r8 = AE.read_results(_p8, price=315.0)
_ok8, _ = CK.tie_check(_r8)
ok(_ok8, f"horizon 8 builds and ties (tie={_r8['max_identity_tie']:.1e})")
ok(abs(_r8["equity_value"] - RES["equity_value"]) > 1.0,
   f"horizon is first-order: N=4 {RES['equity_value']:.4f} vs N=8 {_r8['equity_value']:.4f}")

print("== P3: composite depreciable life ==")
life = I["B42"].value
ok(abs(life - 18) > 1e-9, "plant life is no longer the template's 18 years")
_gross = LC._lastn(VALS["Balance Sheet"], "Gross PPE", 3)
_dep = LC._lastn(VALS["Income Statement"], "Reconciled Depreciation", 3)
near(life, round((sum(_gross) / len(_gross)) / (sum(_dep) / len(_dep)), 2), 1e-9,
     "plant life re-derives as mean gross PP&E / mean reported depreciation")
ok(2.0 <= life <= 50.0, f"plant life inside the plausible band (got {life})")
_pl, _ = build("life_override", judgments={"ppe_life_override": 12.0})
ok(openpyxl.load_workbook(_pl)["Inputs"]["B42"].value == 12.0, "ppe_life_override wins")

print("== S2: provenance register ==")
_reg = {r["name"]: r for r in REP["inputs_register"]}
ok(_reg["in_payout_seed"]["provenance"] == "filings", "payout seed registers as filings-derived")
ok(_reg["in_ppe_life"]["provenance"] == "filings", "plant life registers as filings-derived")
ok(_reg["cfg_N"]["provenance"] == "analyst", "cfg_N registers as analyst-set")
_stragglers = [r["name"] for r in REP["inputs_register"]
               if r["provenance"] == "template" and r["class"] == "company"]
ok(not _stragglers,
   f"no company-level input is still on a template default (found {_stragglers})")
ok(_reg["in_g_terminal"]["provenance"] == "inert",
   "the dead terminal-growth input is reported inert rather than as an exposure")

print("== P4: Valuation row 11 follows the modelled financing policy ==")
_f11 = FORM["Valuation"].cell(11, 3).value
ok('cfg_mode="Enterprise"' in _f11, "row 11 now branches on cfg_mode like rows 7 and 10")
ok("fc_nfo" in _f11, "row 11 reads the live forecast NFO path")
ok("anchor_real_nfo0" not in _f11, "row 11 no longer freezes NFO at the anchor")
# The per-share basis depends on the closure, and row 11's own formula says so:
# "/anchor_shares0" under Enterprise, "/INDEX(fc_shares,C4)" under Equity. Under the
# canonical operating closure (2026-08-10) share repurchases are live, so the two
# denominators genuinely differ -- by exactly the retirement rate -- and the assertion has
# to pick the denominator the active closure actually uses. Comparing against the live
# share count under Enterprise fails by the first-year buyback (about 3%), which is the
# model behaving as designed, not a defect.
_fc_nfo_1 = VALS["Forecast"].cell(44, 7).value        # G44, t=1
_mode = FORM["Inputs"]["B37"].value
if _mode == "Enterprise":
    _den_1 = VALS["Forecast"].cell(20, 6).value       # F20, ANCHOR shares
    _basis = "per ANCHOR share (Enterprise closure)"
else:
    _den_1 = VALS["Forecast"].cell(20, 7).value       # G20, live t=1 shares
    _basis = "per live share (Equity closure)"
near(V.cell(11, 3).value, _fc_nfo_1 / _den_1, 1e-9,
     f"row 11 t=1 equals the Forecast tab's nominal NFO {_basis}")
# and it must now actually move with the leverage target, which the frozen row could not do
ok(abs(V.cell(11, 3).value - V.cell(11, 2).value) > 1e-6,
   "row 11 responds to the forecast financing path instead of tracking only inflation")

print("== P5: the operations wedge ==")
# the cause, asserted on the live engine rather than argued: book weights make ReOI == RI
_worst = max(abs((D.cell(28, c).value or 0.0) - (D.cell(27, c).value or 0.0))
             for c in range(3, 33))
ok(_worst < 1e-12,
   f"book-weighted rho_F makes residual operating income identical to residual equity "
   f"income (max |ReOI-RI| = {_worst:.1e}) — which is why the direct route could never tie")
# the value-weighted rate is strictly above the book-weighted one whenever debt is cheaper
ok(D.cell(51, 3).value > D.cell(14, 3).value,
   f"value-weighted rho_F* ({D.cell(51,3).value:.6f}) exceeds book-weighted rho_F "
   f"({D.cell(14,3).value:.6f}) — the old rate was too low, so the old route read too high")
# the tie itself
near(D.cell(55, 2).value, D.cell(37, 2).value, 1e-12,
     "V(ops) direct at rho_F* equals V(ops) additive")
ok(abs(D.cell(56, 2).value) < 1e-12, f"CHECK 7 operations tie is zero (B56={D.cell(56,2).value!r})")
ok(abs(D.cell(57, 2).value) < 1e-9,
   f"flow additivity FCFF = FCFE + FCFD holds (B57={D.cell(57,2).value:.1e})")
# and the old book-weighted diagnostic is still materially non-zero, which is the whole point
ok(abs(D.cell(40, 2).value) > 1e-6,
   f"the old book-weighted diagnostic remains non-zero ({D.cell(40,2).value:.4f}) and is "
   f"now labelled as unable to tie rather than left looking broken")

print("== the master gate covers the new checks and drops the tautologies ==")
_b5 = FORM["Audit"].cell(5, 2).value
for _t in ("B31", "B44", "B50", "B58", "B29", "B63", "B72", "B77"):
    ok(_t in _b5, f"Audit B5 sums {_t}")
for _t in ("B27", "B28"):
    ok(_t not in _b5, f"Audit B5 no longer sums the tautological {_t}")
ok(abs(A["B27"].value) < 1e-15 and abs(A["B28"].value) < 1e-15,
   "the tautological cells are indeed identically zero (kept, relabelled, not counted)")
near(A["B77"].value, abs(D.cell(56, 2).value) + abs(D.cell(57, 2).value), 1e-15,
     "CHECK 7 total is the sum of its two parts")
ok(A["B6"].value.startswith("PASS"), f"in-sheet status {A['B6'].value!r}")
# the Python-side metric must mirror the in-sheet gate
ok(RES["max_identity_tie"] >= abs(A["B77"].value) - 1e-18,
   "max_identity_tie now sees CHECK 7")

print(f"\n{_p} passed, {_f} failed")
raise SystemExit(1 if _f else 0)

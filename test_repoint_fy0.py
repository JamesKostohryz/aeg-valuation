#!/usr/bin/env python3
"""test_repoint_fy0.py — tests for repoint_fy0.py, which runs on EVERY live build
(before the first recalc) and repoints the eight rep_*_fy0 reconciliation-anchor names
away from the template's hardcoded last column ($AP$) to whatever column actually holds
the anchor fiscal year. Closes a coverage gap flagged in AEG-Coverage-Map-2026-08-08.md:
the module is fail-closed on a missing name/malformed ref/absent year, but nothing
previously checked that a genuine repoint (a shorter-history issuer, e.g. POOL-class)
actually lands on the RIGHT column and RIGHT value, only that AAPL/T's no-op case exists.

Uses the golden AAPL build so it needs no network access. AAPL's own anchor (2025) is
already at column AP by construction (the documented no-op case) — to exercise the real
repoint logic this suite deliberately repoints the SAME AAPL workbook to an EARLIER year
that still exists in its 40-year history (2020), which is a legitimate, fully-general
test of "does repoint_anchor_columns move these eight names to the column whose header
equals the given year, and read the value that's actually THERE" — independent of
whether 2020 is the model's real build anchor.

Checks:
  1. The no-op case: AAPL's own anchor year (2025) is already at column AP -> moved=[].
  2. A genuine repoint (anchor_year=2020, not the last column): all eight names move,
     and after the move each name's cell VALUE matches the Balance Sheet / Income
     Statement's own reported 2020 figure exactly (read independently via the year row,
     not via the repointed name itself — an independent cross-check).
  3. Idempotency: repointing to the same year twice reports the second call as fully
     unchanged.
  4. Fail-closed: an anchor year absent from the statement's year row raises Fy0Error
     instead of silently doing nothing.
  5. After a genuine repoint + recalc, Audit CHECK-4b (the anchor-vs-reported
     reconciliation, B35:B42) reads exactly what it should: the *reconciliation* residuals
     (in_debt/anchor_cse0/etc vs. the now-2020-pointed rep_*_fy0) are large (because the
     model's real anchor is still 2025) — proving the repoint changed what CHECK-4b
     compares against, not that it silently no-ops.

Usage: python3 test_repoint_fy0.py
"""
import os, sys, shutil
import openpyxl

_ROOT = os.path.dirname(os.path.abspath(__file__))
for p in (_ROOT, os.path.join(_ROOT, "pipeline")):
    if p not in sys.path:
        sys.path.insert(0, p)
from repoint_fy0 import repoint_anchor_columns, Fy0Error, FY0_NAMES, YEAR_RANGE
import aeg_engine as AE
from recalc_lo import recalc

GOLDEN = os.path.join(_ROOT, "tests", "golden", "AAPL")
TEMPLATE = os.path.join(_ROOT, "MODEL_TEMPLATE.xlsx")
WORK = os.path.join(_ROOT, "_fy0test")
os.makedirs(WORK, exist_ok=True)

_p = _f = 0
def check(c, m):
    global _p, _f
    if c: _p += 1; print(f"  PASS  {m}")
    else: _f += 1; print(f"  FAIL  {m}")


def build_golden(out):
    files = {"is_csv": f"{GOLDEN}/REAL_IS.csv", "bs_csv": f"{GOLDEN}/REAL_BS.csv",
              "cf_csv": f"{GOLDEN}/REAL_CF.csv", "prices": f"{GOLDEN}/REAL_prices.csv",
              "dividends": f"{GOLDEN}/REAL_div.csv", "splits": f"{GOLDEN}/REAL_splits.csv"}
    cfg = {"company": "Apple Inc.", "ticker": "AAPL", "price": 315.0, "files": files,
           "fy_end_month": 9,
           "forecast_horizon_N": 4,   # P2: cfg_N is required and has no default; 4 is the
                                     # horizon these fixtures have always run at.
           "judgments": {"minority_include": False, "finlease": 0.0, "oi_adj_override": None,
                         "rd_capitalize": True, "rd_life": 5.0, "dps_override": None},
           "cost_of_debt": {"single_ytw": 0.05}}
    AE.build_model(cfg, TEMPLATE, out)


def defined_name_value(wb, name):
    dn = wb.defined_names.get(name)
    ref = dn.value.replace("$", "").replace("'", "")
    sh, cell = ref.split("!")
    return wb[sh][cell].value


def year_col(wb, sheet, row, year):
    ws = wb[sheet]
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row, c).value
        if v is not None and str(v).strip() == str(year):
            return c
    return None


print("== case 1: AAPL's own anchor year (2025) is already at column AP -> no-op ==")
p1 = os.path.join(WORK, "noop.xlsx")
build_golden(p1)
rep1 = repoint_anchor_columns(p1, anchor_year=2025)
check(rep1["moved"] == [], f"2025 is already the AP column for AAPL/T-class history -> nothing moves (moved={rep1['moved']})")
check(len(rep1["unchanged"]) == len(FY0_NAMES), "all eight names report unchanged")

print("== case 2: a genuine repoint to an earlier year (2020) that exists in the history ==")
p2 = os.path.join(WORK, "repoint.xlsx")
build_golden(p2)
wb_before = openpyxl.load_workbook(p2, data_only=False)
bs_col_2020 = year_col(wb_before, "Balance Sheet", 3, 2020)
is_col_2020 = year_col(wb_before, "Income Statement", 3, 2020)
check(bs_col_2020 is not None and is_col_2020 is not None, "2020 exists in AAPL's 40-year statement history")

# independently read what 2020's reported figures actually are, BEFORE repointing
from openpyxl.utils import get_column_letter
want = {}
BS, IS = wb_before["Balance Sheet"], wb_before["Income Statement"]
row_of = {"rep_debt_fy0": ("Balance Sheet", 100), "rep_cash_fy0": ("Balance Sheet", 7),
          "rep_cse_fy0": ("Balance Sheet", 94), "rep_shares_fy0": ("Balance Sheet", 103),
          "rep_eps_fy0": ("Income Statement", 41), "rep_intexp_fy0": ("Income Statement", 49),
          "rep_oi_fy0": ("Income Statement", 13), "rep_tax_fy0": ("Income Statement", 59)}
for nm, (sheet, row) in row_of.items():
    col = bs_col_2020 if sheet == "Balance Sheet" else is_col_2020
    want[nm] = wb_before[sheet].cell(row, col).value

rep2 = repoint_anchor_columns(p2, anchor_year=2020)
check(len(rep2["moved"]) == len(FY0_NAMES), f"all eight names move for a genuinely different anchor year (moved={len(rep2['moved'])})")
check(rep2["unchanged"] == [], "nothing is reported unchanged when every name actually moves")

wb_after = openpyxl.load_workbook(p2, data_only=False)
letter_2020_bs, letter_2020_is = get_column_letter(bs_col_2020), get_column_letter(is_col_2020)
for name, (moved_name, old_l, new_l) in zip(rep2["moved"], rep2["moved"]):
    pass
for entry in rep2["moved"]:
    nm, old_l, new_l = entry
    sheet, _ = row_of[nm]
    expected_letter = letter_2020_bs if sheet == "Balance Sheet" else letter_2020_is
    check(new_l == expected_letter, f"{nm}: repointed to column {expected_letter} (the 2020 column), got {new_l}")

for nm in row_of:
    got = defined_name_value(wb_after, nm)
    w = want[nm]
    same = (got == w) or (isinstance(got, (int, float)) and isinstance(w, (int, float)) and abs(got - w) < 1e-9)
    check(same, f"{nm}: repointed cell VALUE equals 2020's own reported figure independently re-read ({got!r} vs {w!r})")

print("== idempotency: repointing to the same year again changes nothing further ==")
rep3 = repoint_anchor_columns(p2, anchor_year=2020)
check(rep3["moved"] == [], f"second call for the same anchor year moves nothing (moved={rep3['moved']})")
check(len(rep3["unchanged"]) == len(FY0_NAMES), "second call reports everything unchanged")

print("== fail-closed: an anchor year absent from the history raises, never silently no-ops ==")
p4 = os.path.join(WORK, "badyear.xlsx")
build_golden(p4)
try:
    repoint_anchor_columns(p4, anchor_year=1899)
    check(False, "a nonexistent anchor year should raise Fy0Error")
except Fy0Error:
    check(True, "nonexistent anchor year raises Fy0Error (fail-closed, not a silent no-op)")

print("== after a genuine repoint, Audit CHECK-4b's residuals recompute exactly against the NEW (2020) target ==")
# NOTE: a naive "residual should be large" check is the wrong test here -- these engine
# cells are internally scaled (~1e-6 of the raw statement figures), so an absolute
# magnitude threshold doesn't mean what it would in dollars, and some AAPL lines (total
# debt, in particular) are genuinely similar between 2020 and 2025 so a small residual is
# CORRECT, not a sign the repoint no-op'd. The real proof is that CHECK-4b's formula now
# recomputes exactly against the independently-captured 2020 figure -- i.e. the repoint
# changed the comparison target, and Audit picked it up on recalc.
recalc(p2)
wbv = openpyxl.load_workbook(p2, data_only=True)
A = wbv["Audit"]
AUDIT_ROW = {"rep_debt_fy0": (35, "in_debt"), "rep_cash_fy0": (36, "in_cash"),
             "rep_cse_fy0": (37, "anchor_cse0"), "rep_shares_fy0": (38, "anchor_shares0"),
             "rep_eps_fy0": (39, "anchor_eps0"), "rep_intexp_fy0": (40, "in_intexp0"),
             "rep_oi_fy0": (41, "in_oiadj0"), "rep_tax_fy0": (42, "in_tax0")}

def nm_value(wb, name):
    dn = wb.defined_names.get(name)
    ref = dn.value.replace("$", "").replace("'", "")
    sh, cell = ref.split("!")
    return wb[sh][cell].value

for rep_name, (row, in_name) in AUDIT_ROW.items():
    in_val = nm_value(wbv, in_name)
    audit_resid = A[f"B{row}"].value
    expected_resid = in_val - want[rep_name]   # want[] captured BEFORE repoint, straight off the 2020 column
    same = isinstance(audit_resid, (int, float)) and abs(audit_resid - expected_resid) < 1e-6
    check(same, f"Audit!B{row} ({in_name} vs {rep_name}) = {audit_resid!r}, matches "
                f"{in_name}-independently_read_2020_value = {expected_resid!r} exactly")

shutil.rmtree(WORK, ignore_errors=True)
print(f"\n{_p} passed, {_f} failed")
sys.exit(1 if _f else 0)

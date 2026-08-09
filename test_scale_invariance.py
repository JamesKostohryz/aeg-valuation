#!/usr/bin/env python3
"""test_scale_invariance.py — Phase 1, Property 1: unit-scale invariance.

WHY THIS FILE EXISTS
--------------------
Multiply every aggregate currency figure the engine is fed — income statement, balance
sheet, cash flow statement, and the share count — by the same constant k, and leave the
per-share and rate rows alone. Earnings per share, dividends per share and price are then
unchanged by construction, so EVERY per-share output of the engine must be identical.

That is an oracle OUTSIDE the four spokes. The abnormal-earnings-growth, residual income,
free-cash-flow-to-equity and free-cash-flow-to-firm legs are four transformations of one
restated stream, so they agree with each other at any scale, including a broken one. A
units bug scales all four together and the tie reads 1e-15 while every published number is
wrong. This test compares the engine against arithmetic instead.

It is the company-independent generalisation of the scale defect found on 2026-08-09, and
it would have caught it.

WHY k=1 IS NOT IN THE GRID
--------------------------
The golden fixtures are already denominated in millions and the loader divides by 1e6
again, so at k=1 the engine runs at one millionth of production scale, where the loader's
round(x/1e6, 6) quantisation leaves only four to six significant figures. That costs
$0.00215 per share on Apple — two parts in a hundred thousand. It is a real weakness in the
regression fixtures, it is on the defect register, and it is NOT a Phase 1 failure: live
runs consume raw dollars and land at k=1e6, inside the converged regime. Including k=1 here
would paint the property red for a fixture defect. See
claude/AEG-Phase1-Property-1-Scale-Invariance-RESULT-2026-08-09.md.

RESULT WHEN THIS WAS FIRST RUN (2026-08-09)
-------------------------------------------
Across a factor of one billion in input magnitude the intrinsic value per share is
106.375223274483 at every k — not to a tolerance, bit-identical — while the anchor scales
by exactly k.
"""
import csv
import os
import re
import shutil
import sys

_ROOT = os.path.dirname(os.path.abspath(__file__))
for _p_ in (_ROOT, os.path.join(_ROOT, "pipeline")):
    if _p_ not in sys.path:
        sys.path.insert(0, _p_)

import openpyxl                                              # noqa: E402
import aeg_engine as AE                                      # noqa: E402
import checks as CK                                          # noqa: E402
from recalc_lo import recalc                                 # noqa: E402

GOLDEN = os.path.join(_ROOT, "tests", "golden", "AAPL")
TEMPLATE = os.path.join(_ROOT, "MODEL_TEMPLATE.xlsx")
WORK = os.environ.get("AAPL_ENG_WORK") or "/tmp/_scale_invariance_work"
PRICE = 315.0

# The same rule the loader applies: a row whose label marks it per-share or a rate is not a
# currency aggregate and must NOT be scaled. Scaling those would change earnings per share
# and the test would be asserting nothing.
NOSCALE = re.compile(r"\b(EPS|PER SHARE|RATE|MARGIN|RATIO|YIELD)\b|%", re.I)

_p = _f = 0


def ok(cond, msg):
    global _p, _f
    if cond:
        _p += 1
        print("  PASS", msg)
    else:
        _f += 1
        print("  FAIL", msg)


def scale_statement(src, dst, k):
    """Scale aggregate rows by k; leave NOSCALE rows (per-share, rates) untouched."""
    with open(src, newline="") as fh:
        rows = list(csv.reader(fh))
    out = [rows[0]]
    for r in rows[1:]:
        if not r:
            continue
        if NOSCALE.search(r[0]):
            out.append(r)
            continue
        nr = [r[0]]
        for v in r[1:]:
            try:
                nr.append(repr(float(v) * k))
            except (TypeError, ValueError):
                nr.append(v)
        out.append(nr)
    with open(dst, "w", newline="") as fh:
        csv.writer(fh).writerows(out)


def build_at(k):
    """Build + recalc the golden Apple engine with every currency aggregate times k."""
    d = os.path.join(WORK, f"k{k:.0e}")
    os.makedirs(d, exist_ok=True)
    files = {}
    for key, name in (("is_csv", "REAL_IS.csv"), ("bs_csv", "REAL_BS.csv"),
                      ("cf_csv", "REAL_CF.csv")):
        files[key] = os.path.join(d, name)
        scale_statement(os.path.join(GOLDEN, name), files[key], k)
    # Prices, dividends per share and splits are per-share series. They are what makes this
    # a real test: leaving them alone is why earnings per share and price are invariant.
    for key, name in (("prices", "REAL_prices.csv"), ("dividends", "REAL_div.csv"),
                      ("splits", "REAL_splits.csv")):
        files[key] = os.path.join(GOLDEN, name)
    cfg = {"company": "Apple Inc.", "ticker": "AAPL", "price": PRICE, "fy_end_month": 9,
           "forecast_horizon_N": 4, "files": files,
           "judgments": {"minority_include": False, "finlease": 0.0, "oi_adj_override": None,
                         "rd_capitalize": True, "rd_life": 5.0, "dps_override": None},
           "cost_of_debt": {"single_ytw": 0.05}}
    path = os.path.join(d, "engine.xlsx")
    AE.build_model(cfg, TEMPLATE, path)
    recalc(path)
    res = AE.read_results(path, price=PRICE)
    V = openpyxl.load_workbook(path, data_only=True)["Valuation"]
    return {
        "value": res["equity_value"],
        "enterprise": res["enterprise_value"],
        "tie": res["max_identity_tie"],
        "audit": res["audit_status"],
        "anchor_cse0": res["anchors"]["anchor_cse0"],
        "eps0": V.cell(7, 2).value,
        "dps0": V.cell(8, 2).value,
        "bps0": V.cell(10, 2).value,
        "normal_value": V.cell(43, 2).value,
        "intrinsic": V.cell(44, 2).value,
        "price_real": V.cell(45, 2).value,
        "pvgo": V.cell(46, 2).value,
    }


def main():
    global _f
    shutil.rmtree(WORK, ignore_errors=True)
    os.makedirs(WORK, exist_ok=True)

    KS = [1e3, 1e6, 1e9]
    print("== building the golden Apple engine at three input scales ==")
    runs = {}
    for k in KS:
        runs[k] = build_at(k)
        r = runs[k]
        print(f"  k={k:.0e}: value/sh {r['value']!r}  anchor CSE {r['anchor_cse0']!r}  "
              f"tie {r['tie']:.1e}  audit {r['audit'][:4]}")

    print("== every build must be internally sound at every scale ==")
    for k in KS:
        ok(runs[k]["audit"].startswith("PASS"), f"k={k:.0e}: audit PASS ({runs[k]['audit'][:30]})")

    print("== per-share outputs must be IDENTICAL across a factor of one billion ==")
    # Bit-identical, not within a tolerance. A tolerance here would hide exactly the class
    # of defect this property exists to catch.
    PERSHARE = ["value", "enterprise", "eps0", "dps0", "bps0",
                "normal_value", "intrinsic", "price_real", "pvgo"]
    ref = runs[KS[0]]
    for name in PERSHARE:
        vals = [runs[k][name] for k in KS]
        ok(all(v == ref[name] for v in vals),
           f"{name}: identical at every scale ({ref[name]!r})")

    print("== and the anchor must scale by exactly k, proving the inputs really changed ==")
    base = runs[KS[0]]["anchor_cse0"]
    for k in KS[1:]:
        ratio = runs[k]["anchor_cse0"] / base
        expect = k / KS[0]
        ok(abs(ratio / expect - 1.0) < 1e-12,
           f"k={k:.0e}: anchor CSE scaled by {ratio:.6g} (expected {expect:.6g})")
    ok(runs[KS[-1]]["anchor_cse0"] != runs[KS[0]]["anchor_cse0"],
       "the scaled builds are genuinely different builds, not the same file read twice")

    # NOT asserted: that the tie residual is stable across scales. It is floating-point
    # noise at every scale, so comparing noise here against noise there tests nothing. The
    # first draft of this file asserted it and the assertion was meaningless.
    _res = ", ".join("%.1e" % runs[k]["tie"] for k in KS)
    print("     [not asserted] tie residuals " + _res + " — floating-point noise by nature")

    shutil.rmtree(WORK, ignore_errors=True)
    print(f"\n{_p} passed, {_f} failed")
    raise SystemExit(1 if _f else 0)


if __name__ == "__main__":
    main()

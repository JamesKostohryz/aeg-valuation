#!/usr/bin/env python3
"""test_zero_growth.py — Phase 1, Property 2: the zero-abnormal-growth reduction.

WHY THIS FILE EXISTS
--------------------
Strip the abnormal growth out of a forecast and the abnormal-earnings-growth valuation must
collapse to the oldest closed form there is: value equals FORWARD normal earnings divided by
the cost of equity, which is also what the dividend discount model returns on the same
inputs. If the engine's anchor, its curve-implied annuity factor and its per-year
contributions do not compose back into that one number, something in the valuation
machinery is wrong no matter how prettily the four legs agree with each other.

THE FOUR-METHOD TIE IS STRUCTURALLY BLIND TO THIS. Abnormal earnings growth, residual
operating income, free cash flow to equity and free cash flow to the firm are four
transformations of the SAME restated stream and the SAME timing convention, so they tie for
any forecast and any anchor, including wrong ones. This file checks the engine against
arithmetic done outside the workbook instead — the same discipline as test_anchor_timing.py,
carried from the anchor term to the whole valuation.

James verified the closed form by hand: earnings 100, retention 0.60, cost of equity 6.65%
gives 1,563.76 by both routes, while anchoring on TRAILING earnings gives 1,503.76 — short
by exactly the year-zero retained earnings of 60.00. Those arithmetic identities are pinned
in test_anchor_timing.py. What this file adds is the measurement on the live engine, driven
on a real company's real anchor.

HOW THE FORECAST IS ZEROED, AND WHY IT IS NOT CIRCULAR
------------------------------------------------------
The path written into the sheet is constructed HERE, in Python, from the textbook
definition and nothing else:

    b  = retained0 / EPS0        (the engine's own anchor retention)
    g  = b * r                   (normal growth: retained capital earns the cost of equity)
    EPS_t = EPS_0 * (1+g)^t      DPS_t = (1-b) * EPS_t

That is the Gordon path. It is NOT read back out of the engine's own normal-earnings row,
so "abnormal earnings growth is zero at every year" is a PREDICTION this file makes about
the engine, not a construction it imposes. The engine then has to agree — and the value it
produces has to equal both E1/r and a dividend discount model summed in plain Python.

Two cases deliberately use a payout that does NOT match the anchor retention, so the path
is two-stage rather than pure Gordon. Abnormal growth must still be zero at every year and
the value must still be E1/r: the reduction is a property of the valuation, not of a
special case.

MECHANISM. Only real net income (Forecast row 19) is overwritten, for the explicit years
t=1..N, with operating income re-pointed to `NI + NFE` so the operating leg stays exactly
consistent with it. Dividends per share come out of the engine's own payout seed, book
value rolls by the engine's own clean-surplus recursion, and every other formula in the
workbook is left alone. That is why the four-method tie stays green on every case below and
is asserted: this reduction is achieved by feeding the engine a different forecast, not by
breaking it.

The regime is a flat real cost-of-equity curve and zero expected inflation, which is the
only regime where the textbook closed form applies exactly — the engine's curve-implied
annuity factor must collapse to precisely 1/r there, and that collapse is asserted first,
because if it fails everything after it is measuring the wrong thing.

RESULT WHEN THIS WAS FIRST RUN (2026-08-09)
-------------------------------------------
Abnormal earnings growth was exactly 0.0 at every explicit year in every case, every
per-year contribution was exactly 0.0, and the engine's intrinsic value matched E1/r and
the independently summed dividend discount model to about 1e-15 relative — while the
four-method tie stayed at 1e-15 and the audit read PASS.
"""
import os
import shutil
import sys

_ROOT = os.path.dirname(os.path.abspath(__file__))
for _p_ in (_ROOT, os.path.join(_ROOT, "pipeline")):
    if _p_ not in sys.path:
        sys.path.insert(0, _p_)

import openpyxl                                              # noqa: E402
import aeg_engine as AE                                      # noqa: E402
import checks as CK                                          # noqa: E402
from recalc_lo import recalc                                 # noqa: E402

GOLDEN = os.path.join(_ROOT, "tests", "golden", "AAPL")
TEMPLATE = os.path.join(_ROOT, "MODEL_TEMPLATE.xlsx")
WORK = os.environ.get("AAPL_ENG_WORK") or "/tmp/_zero_growth_work"
PRICE = 315.0

# Sheet geography. Forecast: column F is t=0, so t sits at column 6+t (t=30 -> AJ, the last
# forecast column). Valuation: column B is t=0, so t sits at column 2+t.
FC_COL = lambda t: 6 + t          # noqa: E731
VL_COL = lambda t: 2 + t          # noqa: E731

CFG = {"company": "Apple Inc.", "ticker": "AAPL", "price": PRICE, "fy_end_month": 9,
       "forecast_horizon_N": 4,
       "files": {"is_csv": f"{GOLDEN}/REAL_IS.csv", "bs_csv": f"{GOLDEN}/REAL_BS.csv",
                 "cf_csv": f"{GOLDEN}/REAL_CF.csv", "prices": f"{GOLDEN}/REAL_prices.csv",
                 "dividends": f"{GOLDEN}/REAL_div.csv", "splits": f"{GOLDEN}/REAL_splits.csv"},
       "judgments": {"minority_include": False, "finlease": 0.0, "oi_adj_override": None,
                     "rd_capitalize": True, "rd_life": 5.0, "dps_override": None},
       "cost_of_debt": {"single_ytw": 0.05}}

_p = _f = 0


def ok(cond, msg):
    global _p, _f
    if cond:
        _p += 1
        print("  PASS", msg)
    else:
        _f += 1
        print("  FAIL", msg)


def rel(a, b, tol, msg):
    got = (isinstance(a, (int, float)) and isinstance(b, (int, float))
           and abs(a - b) <= tol * max(1.0, abs(b)))
    ok(got, f"{msg}  ({a!r} vs {b!r}, rel tol {tol:g})")


def eps_path(EPS0, ret0, r, payout):
    """The zero-abnormal-growth earnings path, built from the textbook definition only.

    Year one grows the anchor by the return on the ANCHOR year's retained capital; from then
    on retention is the engine's payout seed, so the path is geometric at g = b*r. When the
    payout seed is set to the anchor's own retention the two stages coincide and the whole
    path is the pure Gordon path EPS_0*(1+g)^t."""
    b = 1.0 - payout
    eps = [EPS0, EPS0 + r * ret0]
    for t in range(2, 31):
        eps.append(eps[t - 1] * (1.0 + r * b))
    return eps


def ddm(eps, r, payout, M=30):
    """Dividend discount model, summed in plain Python: M explicit years of dividends plus a
    Gordon tail. Never touches the workbook. This is the external oracle."""
    g = r * (1.0 - payout)
    pv = sum(eps[t] * payout / (1.0 + r) ** t for t in range(1, M + 1))
    return pv + eps[M] * payout * (1.0 + g) / ((r - g) * (1.0 + r) ** M)


def run_case(tag, base, N, r, payout, anchor, note):
    """Copy the built engine, force the flat zero-inflation regime, overwrite the explicit
    forecast with the zero-abnormal-growth path, recalculate, and measure."""
    EPS0, ret0, shares0 = anchor["EPS0"], anchor["ret0"], anchor["shares0"]
    eps = eps_path(EPS0, ret0, r, payout)

    path = os.path.join(WORK, f"zero_{tag}.xlsx")
    shutil.copy(base, path)
    wb = openpyxl.load_workbook(path)
    IN, MD, F = wb["Inputs"], wb["Market Data"], wb["Forecast"]
    IN["B37"] = "Equity"          # cfg_mode
    IN["B29"] = "Single"          # cfg_coe_mode — flat curve
    IN["B26"] = N                 # cfg_N
    IN["B39"] = payout            # in_payout_seed
    for c in range(2, 32):
        MD.cell(22, c).value = 0.0   # finrate_infl — zero expected inflation
        MD.cell(26, c).value = r     # finrate_coe  — flat real cost of equity
    for t in range(1, N + 1):
        col = F.cell(19, FC_COL(t)).column_letter
        F.cell(19, FC_COL(t)).value = eps[t] * shares0        # real net income
        F.cell(16, FC_COL(t)).value = f"={col}19+{col}18"     # keep OI = NI + NFE exactly
    wb.save(path)
    recalc(path)

    d = openpyxl.load_workbook(path, data_only=True)
    V = d["Valuation"]
    res = AE.read_results(path, price=PRICE)
    m = {
        "r_sheet": V.cell(5, 3).value,
        "pi": V.cell(56, 3).value,
        "index": V.cell(57, 3).value,
        "A1": V.cell(62, 3).value,
        "nEPS1": V.cell(22, 3).value,
        "EPS1": V.cell(7, 3).value,
        "value": V.cell(36, 2).value,
        "normal_value": V.cell(43, 2).value,
        "intrinsic": V.cell(44, 2).value,
        "aeg": [V.cell(23, VL_COL(t)).value for t in range(1, N + 1)],
        "contrib": [V.cell(24, VL_COL(t)).value for t in range(1, N + 1)],
        "tie": res["max_identity_tie"],
        "audit": res["audit_status"],
        "tie_ok": CK.tie_check(res)[0],
        "eps": eps,
    }
    print(f"\n== {tag}: N={N}, cost of equity {r:.4%}, payout {payout:.4f} — {note} ==")
    return m


def assert_case(m, N, r, payout, anchor):
    EPS0, ret0 = anchor["EPS0"], anchor["ret0"]
    eps = m["eps"]
    g = r * (1.0 - payout)

    # 0. the regime really is the one the closed form applies to
    ok(m["pi"] == 0 and m["index"] == 1, "expected inflation is zero, so nominal equals real")
    rel(m["r_sheet"], r, 1e-12, "the sheet's nominal cost of equity is the flat rate given")
    rel(m["A1"], 1.0 / r, 1e-12,
        "the curve-implied annuity factor collapses to exactly 1/r on a flat curve")

    # 1. the prediction: abnormal earnings growth is zero at EVERY explicit year
    worst = max(abs(a) for a in m["aeg"])
    ok(worst <= 1e-12 * max(1.0, abs(eps[1])),
       f"abnormal earnings growth is zero at every explicit year (worst {worst:.3e})")
    worst_c = max(abs(c) for c in m["contrib"])
    ok(worst_c <= 1e-12 * max(1.0, abs(eps[1])),
       f"every per-year contribution is therefore zero (worst {worst_c:.3e})")
    rel(m["EPS1"], eps[1], 1e-12, "the engine is running the path this file constructed")

    # 2. THE REDUCTION. Value equals forward normal earnings capitalised.
    rel(m["value"], eps[1] / r, 1e-9,
        "INTRINSIC VALUE EQUALS FORWARD NORMAL EARNINGS OVER THE COST OF EQUITY")
    rel(m["normal_value"], m["value"], 1e-12,
        "with no abnormal growth the whole value IS the normal value — nothing is added")
    rel(m["intrinsic"], m["value"], 1e-12, "the published intrinsic-value readout agrees")

    # 3. and it equals the dividend discount model, summed outside the workbook
    rel(m["value"], ddm(eps, r, payout), 1e-9,
        "AND EQUALS THE DIVIDEND DISCOUNT MODEL ON THE SAME INPUTS (summed in Python)")

    # 4. the anchor is FORWARD, not trailing — the error the four-method tie cannot see
    trailing = EPS0 / r
    ok(abs(m["value"] - trailing) > 1e-6,
       f"it is NOT trailing earnings capitalised ({trailing!r})")
    rel(m["value"] - trailing, ret0, 1e-9,
        "and the shortfall of the trailing anchor is EXACTLY the anchor year's retained "
        "earnings — James's 60.00, measured on the live engine")

    # 5. the reduction was achieved by changing the forecast, not by breaking the model
    ok(m["tie_ok"] and m["audit"].startswith("PASS"),
       f"the four-method tie still holds ({m['tie']:.1e}, audit {m['audit'][:24]})")

    print(f"     [measured] value {m['value']:.9f} = E1/r {eps[1] / r:.9f} = DDM "
          f"{ddm(eps, r, payout):.9f}; trailing {trailing:.9f}; g={g:.6%}")


def main():
    global _f
    shutil.rmtree(WORK, ignore_errors=True)
    os.makedirs(WORK, exist_ok=True)

    print("== build the golden Apple engine once; every case starts from it ==")
    base = os.path.join(WORK, "base.xlsx")
    AE.build_model(CFG, TEMPLATE, base)
    recalc(base)
    d = openpyxl.load_workbook(base, data_only=True)
    V0 = d["Valuation"]
    anchor = {"EPS0": V0.cell(7, 2).value, "ret0": V0.cell(9, 2).value,
              "BPS0": V0.cell(10, 2).value, "shares0": d["Inputs"]["B9"].value}
    b0 = anchor["ret0"] / anchor["EPS0"]
    print(f"  anchor: EPS0 {anchor['EPS0']:.9f}  retained0 {anchor['ret0']:.9f}  "
          f"retention {b0:.6f}  BPS0 {anchor['BPS0']:.9f}")

    # NEGATIVE CONTROL, first, so nothing below can pass vacuously. On the real Apple
    # forecast the abnormal growth is NOT zero and the value is NOT the normal value.
    print("\n== negative control: the untouched Apple forecast ==")
    base_res = AE.read_results(base, price=PRICE)
    aeg0 = [V0.cell(23, VL_COL(t)).value for t in range(1, 5)]
    ok(max(abs(a) for a in aeg0) > 1e-3,
       f"the real forecast genuinely has abnormal growth (max |AEG| {max(abs(a) for a in aeg0):.4f})")
    ok(abs(V0.cell(36, 2).value - V0.cell(43, 2).value) > 1e-3,
       f"and its value differs from its normal value ({V0.cell(36, 2).value:.6f} vs "
       f"{V0.cell(43, 2).value:.6f}) — so a zero reading below means something")
    ok(base_res["audit_status"].startswith("PASS"), "the untouched build is sound to start with")

    # The grid. The first three are the PURE Gordon path (payout seed set to the anchor's
    # own retention, so the two stages coincide); the last three deliberately do not match,
    # including the degenerate zero-retention case where normal growth is exactly zero.
    CASES = [
        ("gordon_r665_N4",  4,  0.0665, 1.0 - b0, "pure Gordon path, James's 6.65% rate"),
        ("gordon_r09_N12", 12,  0.09,   1.0 - b0, "pure Gordon path, longer horizon"),
        ("gordon_r05_N30", 30,  0.05,   1.0 - b0, "pure Gordon path, the full 30-year sheet"),
        ("twostage_p40_N4", 4,  0.0665, 0.40,     "payout differs from anchor retention"),
        ("twostage_p20_N8", 8,  0.09,   0.20,     "high retention, growth close to the rate"),
        ("noretention_N6",  6,  0.0665, 1.00,     "degenerate: zero retention, zero growth"),
    ]
    seen = {}
    for tag, N, r, payout, note in CASES:
        m = run_case(tag, base, N, r, payout, anchor, note)
        assert_case(m, N, r, payout, anchor)
        seen.setdefault(r, []).append((tag, N, payout, m["value"]))

    # A SECOND ANCHOR. Everything above rides one anchor, so a reduction that only worked at
    # Apple's 87% retention would still read green. Rebuild the engine with the anchor
    # dividend overridden through the engine's own supported judgment (dps_override), which
    # lands a materially different anchor retention — near the 0.60 of James's hand-worked
    # case — and re-run the pure Gordon path on it.
    print("\n== a second, materially different anchor ==")
    cfg2 = dict(CFG)
    cfg2["judgments"] = dict(CFG["judgments"], dps_override=3.0)
    base2 = os.path.join(WORK, "base_anchor2.xlsx")
    AE.build_model(cfg2, TEMPLATE, base2)
    recalc(base2)
    d2 = openpyxl.load_workbook(base2, data_only=True)
    V2 = d2["Valuation"]
    anchor2 = {"EPS0": V2.cell(7, 2).value, "ret0": V2.cell(9, 2).value,
               "BPS0": V2.cell(10, 2).value, "shares0": d2["Inputs"]["B9"].value}
    b2 = anchor2["ret0"] / anchor2["EPS0"]
    print(f"  anchor 2: EPS0 {anchor2['EPS0']:.9f}  retained0 {anchor2['ret0']:.9f}  "
          f"retention {b2:.6f}  (anchor 1 was {b0:.6f})")
    ok(abs(b2 - b0) > 0.15, f"the second anchor really is different (retention {b2:.4f} "
                            f"against {b0:.4f})")
    for tag, N, r, payout, note in [
            ("anchor2_gordon_N4", 4, 0.0665, 1.0 - b2, "second anchor, pure Gordon path"),
            ("anchor2_p50_N10", 10, 0.075, 0.50, "second anchor, payout unrelated to it")]:
        m = run_case(tag, base2, N, r, payout, anchor2, note)
        assert_case(m, N, r, payout, anchor2)

    # DIVIDEND-POLICY IRRELEVANCE, which falls out of the grid and is worth asserting on its
    # own. With no abnormal growth, retaining a dollar earns exactly the cost of equity, so
    # how much is paid out cannot change the value — and neither can the horizon, because
    # there is nothing for the explicit period to add. Cases sharing a cost of equity must
    # therefore land on the SAME value despite different payouts and different cfg_N. It is
    # a real engine property, not a restatement of the closed form: it would break if the
    # annuity factor or the horizon gate leaked payout or N into the anchor.
    print("\n== dividend policy and horizon must not move the value when growth is normal ==")
    for r, group in sorted(seen.items()):
        if len(group) < 2:
            continue
        ref = group[0][3]
        desc = ", ".join(f"{t}(N={n}, payout={p:.4f})" for t, n, p, _ in group)
        ok(all(abs(v - ref) <= 1e-9 * max(1.0, abs(ref)) for _, _, _, v in group),
           f"cost of equity {r:.4%}: identical value {ref:.9f} across {desc}")

    shutil.rmtree(WORK, ignore_errors=True)
    print(f"\n{_p} passed, {_f} failed")
    raise SystemExit(1 if _f else 0)


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""gfd_sector_extract.py — pull the ten top-level S&P 500 GICS sector price series out of the
Global Financial Data workbooks into one tidy CSV.

The workbooks hold ~60 series each, almost all sub-industries. These ten are the top-level GICS
sectors, identified by their two-digit GICS code in the series description and checked against
the symbol, not guessed from the file name.

  python3 tools/gfd_sector_extract.py

Writes outputs/2026-08-20-sectors/gfd_sector_daily.csv. NOT A VALUATION.
"""
from __future__ import annotations

import csv
import datetime as dt
import os

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "outputs", "gfd_sector_price_raw", "extracted")
OUT = os.path.join(ROOT, "outputs", "2026-08-20-sectors")

# WHAT IS ACTUALLY IN THE DOWNLOAD, AND WHAT IS NOT.
#
# Each workbook's "Index" sheet is GFD's CATALOGUE of series that exist in their database, not a
# list of what was downloaded. Ten top-level S&P 500 GICS sector indices are catalogued; only
# FIVE have a data sheet. Both zips were checked and both are fully extracted, so this is a
# partial download and not a missing file. Taking the catalogue for the contents is exactly the
# kind of thing that looks complete and is not.
#
# symbol -> (GICS code, name, workbook)
SECTORS = {
    "_SPLRCI": ("20", "Industrials", "S&P Sector-Industrials.xlsx"),
    "_SPLRCD": ("25", "Consumer Discretionary", "S&P Sector-Consumer Discretionary.xlsx"),
    "_SPLRCS": ("30", "Consumer Staples", "S&P Sector-Consumer Staples.xlsx"),
    "_SPLRCF": ("40", "Financials", "S&P Sector-Finance.xlsx"),
    "_SPLRCU": ("55", "Utilities", "S&P Sector-Utilities.xlsx"),
}

# CATALOGUED BUT NOT DOWNLOADED. These are the five to pull from GFD, by exact symbol. The
# "best sub-industry present" column is what is in the download today and is NOT a substitute --
# a sub-industry is more volatile than its parent sector, so using one would overstate that
# sector's premium. Recorded so the gap is visible, not so it can be papered over.
MISSING = {
    "_SPLRCE": ("10", "Energy", "weekly 1986, daily Sep 1989", "_SPLRCOIG Oil, Gas & Consumable Fuels"),
    "_SPLRCM": ("15", "Materials", "daily Sep 1989", "_SPLRCPM Chemicals"),
    "_SPLRCA": ("35", "Health Care", "weekly 1987, daily Sep 1989", "_SPLRCCARG Pharmaceuticals"),
    "_SPLRCT": ("45", "Information Technology", "weekly 1986, daily Sep 1989", "_SPLRCSOFW Software"),
    "_SPLRCL": ("50", "Telecommunication Services", "daily Sep 1989", "_SPLRCTELP Integrated Telecom"),
}
# Real Estate (60) is absent everywhere: it became a GICS sector in 2016 and this data ends 2014.


def parse_date(v):
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    s = str(v).strip()
    if "/" in s:
        m, d, y = s.split("/")
        return "%s-%s-%s" % (y, m.zfill(2), d.zfill(2))
    return s[:10]


def main():
    import openpyxl
    os.makedirs(OUT, exist_ok=True)
    rows = []
    for sym, (code, name, wbname) in SECTORS.items():
        path = os.path.join(SRC, wbname)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sym not in wb.sheetnames:
            raise SystemExit("%s not a sheet in %s — refusing rather than guessing" % (sym, wbname))
        ws = wb[sym]
        n = 0
        prev = None
        gaps = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r[0] or r[5] in (None, ""):
                continue
            d = parse_date(r[0])
            try:
                c = float(r[5])
            except (TypeError, ValueError):
                continue
            if c <= 0:
                continue
            if prev:
                gap = (dt.date.fromisoformat(d) - dt.date.fromisoformat(prev)).days
                if gap > 10:
                    gaps += 1
            prev = d
            rows.append((code, name, sym, d, c))
            n += 1
        wb.close()
        first = min(x[3] for x in rows if x[2] == sym)
        last = max(x[3] for x in rows if x[2] == sym)
        # where does it become daily? first month with >= 15 observations
        bym = {}
        for x in rows:
            if x[2] == sym:
                bym[x[3][:7]] = bym.get(x[3][:7], 0) + 1
        daily_from = next((m for m in sorted(bym) if bym[m] >= 15), None)
        print("%-8s %-2s %-26s n=%6d  %s .. %s   daily from %s"
              % (sym, code, name, n, first, last, daily_from))

    print("\nCATALOGUED BUT NOT IN THE DOWNLOAD — the shopping list, by exact GFD symbol:")
    for sym, (code, name, cov, proxy) in sorted(MISSING.items(), key=lambda x: x[1][0]):
        print("   %-10s %-2s %-26s  %-28s  present instead: %s"
              % (sym, code, name, cov, proxy))

    rows.sort(key=lambda x: (x[3], x[0]))
    p = os.path.join(OUT, "gfd_sector_daily.csv")
    with open(p, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["gics", "sector", "symbol", "date", "close"])
        w.writerows(rows)
    print("\nwrote %s (%d rows, %d sectors)" % (p, len(rows), len(SECTORS)))


if __name__ == "__main__":
    main()
